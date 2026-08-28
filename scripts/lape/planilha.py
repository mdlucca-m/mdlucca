"""A planilha que se atualiza sozinha, conforme o cadastro anda.

O laboratorio trabalha em Excel. O sistema nasceu para ler essa planilha,
mas a mao inversa faltava: quem cadastra pelo navegador nao via nada mudar
no arquivo, e a planilha ia envelhecendo em silencio.

Aqui a planilha e reescrita a partir do banco -- as mesmas abas, os mesmos
cabecalhos que o laboratorio ja usa. Duas propriedades importam:

  1. **Ida e volta.** O arquivo gerado pode ser reimportado pelo proprio
     sistema e produzir os mesmos dados. Isso nao e enfeite: e a prova de
     que nenhuma coluna se perdeu no caminho. Ha teste para isso.
  2. **Espelho, nao caderno.** O arquivo e reescrito inteiro a cada
     atualizacao. Quem editar a mao perde a edicao na proxima passagem --
     por isso a primeira aba avisa isso em letras grandes, e o cadastro
     continua sendo pelo sistema (ou pela planilha de entrada em
     `data/raw/`, que este modulo nunca toca).

A regeneracao segue o mesmo criterio da copia de seguranca: quem manda e o
`change_log`. Cadastrou, a planilha muda. Ninguem mexeu, nao ha o que
reescrever. O registro fica em `ingest_log` com `source = 'planilha'`.
"""
from __future__ import annotations

import os
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

from . import config, mapping
from .db import Database

# Espacar as reescritas: numa tarde de cadastro em massa, um arquivo por
# registro so gasta disco e trava o Excel de quem estiver com ele aberto.
INTERVALO_MINIMO_MIN = int(os.environ.get("LAPE_PLANILHA_INTERVALO_MIN", "10"))
MUDANCAS_PARA_GERAR = int(os.environ.get("LAPE_PLANILHA_MUDANCAS", "1"))
# Mesmo sem cadastro novo, uma planilha por dia -- assim o "atualizada em"
# na primeira aba nunca fica velho a ponto de gerar duvida.
INTERVALO_DIARIO_H = int(os.environ.get("LAPE_PLANILHA_DIARIO_H", "24"))

NOME_ARQUIVO = os.environ.get("LAPE_PLANILHA_NOME", "LAPE_atualizada.xlsx")

AVISO = (
    "Esta planilha é gerada pelo sistema e reescrita a cada atualização. "
    "Anotação feita aqui à mão se perde na próxima vez. "
    "Para mudar dados, use o sistema (aba Cadastro) — é de lá que este "
    "arquivo é montado."
)


def caminho(db_path: Path | None = None) -> Path:
    """Onde a planilha mora. Ao lado do banco, salvo indicacao em contrario."""
    indicado = os.environ.get("LAPE_PLANILHA_DIR")
    pasta = Path(indicado) if indicado else Path(db_path or config.DB_PATH).parent / "planilha"
    return pasta / NOME_ARQUIVO


# ----------------------------------------------------------------------
# As abas
# ----------------------------------------------------------------------
# Cada aba: (nome da aba, SQL, [(cabecalho, coluna do SELECT)]).
# Os cabecalhos sao os que o laboratorio ja usa, e sao tambem os que o
# `mapping` reconhece na volta -- e o que faz a ida e volta fechar.
SIM_NAO = ("Sim", "Não")

ABAS: tuple[tuple[str, str, tuple[tuple[str, str], ...]], ...] = (
    ("Integrantes", """
        SELECT m.full_name, m.short_name, m.role, m.degree,
               rl.name AS research_line, i.name AS institution,
               m.lattes_id, m.orcid, m.email, m.phone,
               m.joined_on, m.left_on, m.is_external, m.active,
               o.full_name AS advisor, c.full_name AS co_advisor,
               m.thesis_title, m.thesis_kind, m.thesis_status, m.thesis_due_on
          FROM members m
          LEFT JOIN research_lines rl ON rl.id = m.research_line_id
          LEFT JOIN institutions  i  ON i.id  = m.institution_id
          LEFT JOIN members o ON o.id = m.advisor_id
          LEFT JOIN members c ON c.id = m.co_advisor_id
         ORDER BY m.full_name
     """, (
        ("Nome completo", "full_name"), ("Nome curto", "short_name"),
        ("Função", "role"), ("Titulação", "degree"),
        ("Linha de pesquisa", "research_line"), ("Instituição", "institution"),
        ("Lattes", "lattes_id"), ("ORCID", "orcid"),
        ("E-mail", "email"), ("Telefone", "phone"),
        ("Data de entrada", "joined_on"), ("Data de saída", "left_on"),
        ("Externo", "is_external"), ("Ativo", "active"),
        ("Orientador", "advisor"), ("Coorientador", "co_advisor"),
        ("Título da tese", "thesis_title"), ("Tipo de trabalho", "thesis_kind"),
        ("Situação da tese", "thesis_status"), ("Prazo para conclusão", "thesis_due_on"),
    )),

    ("Publicações", """
        SELECT a.internal_code, a.title,
               (SELECT group_concat(x.author_name, '; ') FROM
                  (SELECT author_name FROM article_authors
                    WHERE article_id = a.id ORDER BY author_order) x) AS authors,
               a.lead_name, a.status, rl.name AS research_line, a.study_type, a.language,
               a.started_on, a.first_submission_on, a.accepted_on, a.published_on,
               a.year_published, a.journal, a.issn, a.qualis, a.impact_factor,
               a.doi, a.url, a.wos_id, a.scopus_id,
               a.wos_citations, a.scopus_citations, a.open_access, a.notes,
               a.internal_review_on,
               (SELECT occurred_on FROM article_milestones
                 WHERE article_id = a.id AND milestone = 'version' AND seq = 1) AS version_1,
               (SELECT occurred_on FROM article_milestones
                 WHERE article_id = a.id AND milestone = 'version' AND seq = 2) AS version_2,
               (SELECT occurred_on FROM article_milestones
                 WHERE article_id = a.id AND milestone = 'version' AND seq = 3) AS version_3,
               (SELECT occurred_on FROM article_milestones
                 WHERE article_id = a.id AND milestone = 'version' AND seq = 4) AS version_4,
               (SELECT occurred_on FROM article_milestones
                 WHERE article_id = a.id AND milestone = 'version_final') AS version_final
          FROM articles a
          LEFT JOIN research_lines rl ON rl.id = a.research_line_id
         ORDER BY COALESCE(a.year_published, 9999) DESC, a.title
     """, (
        ("Código", "internal_code"), ("Título", "title"), ("Autores", "authors"),
        ("Responsável", "lead_name"), ("Status", "status"),
        ("Linha de pesquisa", "research_line"), ("Tipo de estudo", "study_type"),
        ("Idioma", "language"), ("Data de início", "started_on"),
        ("Data de submissão", "first_submission_on"), ("Data do aceite", "accepted_on"),
        ("Data de publicação", "published_on"), ("Ano", "year_published"),
        ("Periódico", "journal"), ("ISSN", "issn"), ("Qualis", "qualis"),
        ("Fator de impacto", "impact_factor"), ("DOI", "doi"), ("Link", "url"),
        ("WoS ID", "wos_id"), ("Scopus ID", "scopus_id"),
        ("Citações WoS", "wos_citations"), ("Citações Scopus", "scopus_citations"),
        ("Acesso aberto", "open_access"), ("Observações", "notes"),
        ("Revisão interna", "internal_review_on"),
        ("Data da primeira versão", "version_1"), ("Data segunda versão", "version_2"),
        ("Data terceira versão", "version_3"), ("Data quarta versão", "version_4"),
        ("Data versão final", "version_final"),
    )),

    ("Autoria", """
        SELECT a.title AS article, aa.author_name, aa.author_order,
               aa.is_corresponding, aa.is_external
          FROM article_authors aa
          JOIN articles a ON a.id = aa.article_id
         ORDER BY a.title, aa.author_order
     """, (
        ("Artigo", "article"), ("Autor", "author_name"), ("Ordem", "author_order"),
        ("Correspondente", "is_corresponding"), ("Externo", "is_external"),
    )),

    ("Submissões", """
        SELECT a.title AS article, s.attempt_no, s.journal, s.issn,
               s.submitted_on, s.decision, s.decision_on,
               r.label AS rejection_reason, s.rejection_notes,
               s.desk_reject, s.review_rounds
          FROM submissions s
          JOIN articles a ON a.id = s.article_id
          LEFT JOIN rejection_reasons r ON r.id = s.rejection_reason_id
         ORDER BY a.title, s.attempt_no
     """, (
        ("Artigo", "article"), ("Tentativa", "attempt_no"), ("Periódico", "journal"),
        ("ISSN", "issn"), ("Data de submissão", "submitted_on"),
        ("Resultado", "decision"), ("Data de decisão", "decision_on"),
        ("Motivo da recusa", "rejection_reason"), ("Observações", "rejection_notes"),
        ("Recusa direta", "desk_reject"), ("Rodadas de revisão", "review_rounds"),
    )),

    ("Projetos", """
        SELECT p.code, p.name, p.description, rl.name AS research_line,
               p.coordinator_name, p.kind, p.funder, p.grant_number, p.amount,
               p.started_on, p.ended_on, p.status, p.ethics_approval, p.url
          FROM projects p
          LEFT JOIN research_lines rl ON rl.id = p.research_line_id
         ORDER BY p.name
     """, (
        ("Código", "code"), ("Projeto", "name"), ("Descrição", "description"),
        ("Linha de pesquisa", "research_line"), ("Coordenador", "coordinator_name"),
        ("Tipo", "kind"), ("Financiador", "funder"), ("Processo", "grant_number"),
        ("Valor", "amount"), ("Início", "started_on"), ("Término", "ended_on"),
        ("Situação", "status"), ("Parecer ético", "ethics_approval"), ("Link", "url"),
    )),

    ("Equipe do projeto", """
        SELECT p.name AS project, m.full_name AS member, pm.role, pm.joined_on
          FROM project_members pm
          JOIN projects p ON p.id = pm.project_id
          JOIN members  m ON m.id = pm.member_id
         ORDER BY p.name, m.full_name
     """, (
        ("Projeto", "project"), ("Integrante", "member"),
        ("Função", "role"), ("Entrada", "joined_on"),
    )),

    ("Eventos", """
        SELECT COALESCE(e.external_key, 'EV-' || printf('%04d', e.id)) AS external_key,
               e.kind, e.title, e.description,
               e.start_at, e.end_at, e.all_day, e.status, e.location_name,
               i.name AS institution, e.city, e.state, e.country,
               e.latitude, e.longitude, rl.name AS research_line, e.url
          FROM events e
          LEFT JOIN institutions   i  ON i.id  = e.institution_id
          LEFT JOIN research_lines rl ON rl.id = e.research_line_id
         ORDER BY e.start_at DESC
     """, (
        ("Código", "external_key"), ("Tipo", "kind"), ("Título", "title"),
        ("Descrição", "description"), ("Data", "start_at"),
        ("Data de término", "end_at"), ("Dia inteiro", "all_day"),
        ("Status", "status"), ("Local", "location_name"),
        ("Instituição", "institution"), ("Cidade", "city"), ("Estado", "state"),
        ("País", "country"), ("Latitude", "latitude"), ("Longitude", "longitude"),
        ("Linha de pesquisa", "research_line"), ("Link", "url"),
    )),

    ("Participação", """
        -- Pelo codigo, nunca pelo titulo: "Reunião semanal" se repete dezenas
        -- de vezes na agenda, e o titulo faria toda a presenca do ano cair na
        -- primeira reuniao. O codigo e o do proprio evento, ou um derivado do
        -- id quando o evento nao tem um.
        SELECT COALESCE(e.external_key, 'EV-' || printf('%04d', e.id)) AS event,
               m.full_name AS member, ep.role, ep.attended
          FROM event_participants ep
          JOIN events  e ON e.id = ep.event_id
          JOIN members m ON m.id = ep.member_id
         ORDER BY e.start_at DESC, m.full_name
     """, (
        ("Evento", "event"), ("Integrante", "member"),
        ("Função", "role"), ("Presente", "attended"),
    )),

    ("Linhas de Pesquisa", """
        SELECT code, name, description, coordinator, started_on, keywords, active
          FROM research_lines ORDER BY name
     """, (
        ("Código", "code"), ("Linha de pesquisa", "name"), ("Descrição", "description"),
        ("Coordenador", "coordinator"), ("Data de início", "started_on"),
        ("Palavras-chave", "keywords"), ("Ativa", "active"),
    )),

    ("Instituições", """
        SELECT name, acronym, city, state, country, latitude, longitude
          FROM institutions ORDER BY name
     """, (
        ("Instituição", "name"), ("Sigla", "acronym"), ("Cidade", "city"),
        ("Estado", "state"), ("País", "country"),
        ("Latitude", "latitude"), ("Longitude", "longitude"),
    )),

    ("Motivos de recusa", """
        SELECT code, label, category FROM rejection_reasons ORDER BY code
     """, (
        ("Código", "code"), ("Motivo", "label"), ("Categoria", "category"),
    )),
)

# ----------------------------------------------------------------------
# Vocabulario: o banco guarda `em_producao`, a planilha mostra "Em produção"
# ----------------------------------------------------------------------
# Quem abre a planilha e uma pessoa, nao um programa: `desk_reject` numa
# celula e ruido. Cada rotulo aqui volta ao mesmo codigo quando o arquivo e
# reimportado -- ha teste passando cada um destes pelos mapas do `mapping`,
# porque um rotulo que nao volta perde o dado em silencio na reimportacao.
ROTULOS: dict[str, dict[str, str]] = {
    "status": {
        "em_producao": "Em produção", "submetido": "Submetido",
        "em_revisao": "Em revisão", "aceito": "Aceito",
        "publicado": "Publicado", "rejeitado": "Rejeitado",
        "arquivado": "Arquivado",
    },
    "decision": {
        "em_avaliacao": "Em avaliação", "revisao_solicitada": "Revisão solicitada",
        "aceito": "Aceito", "rejeitado": "Rejeitado",
        "desk_reject": "Recusa direta", "retirado": "Retirado",
    },
    "project_status": {
        "em_andamento": "Em andamento", "concluido": "Concluído",
        "planejado": "Planejado", "suspenso": "Suspenso",
    },
    "thesis_kind": {
        "tese": "Tese", "dissertacao": "Dissertação", "tcc": "TCC",
        "projeto": "Projeto", "relatorio": "Relatório",
    },
    "thesis_status": {
        "em_andamento": "Em andamento", "coleta": "Coleta",
        "analise": "Análise", "qualificacao": "Qualificação",
        "defesa_marcada": "Defesa marcada", "concluida": "Concluída",
        "trancada": "Trancada",
    },
    "kind_evento": {
        "reuniao": "Reunião", "seminario": "Seminário", "congresso": "Congresso",
        "curso": "Curso", "defesa": "Defesa", "qualificacao": "Qualificação",
        "coleta": "Coleta", "extensao": "Extensão", "visita_tecnica": "Visita técnica",
    },
    "role": dict(mapping.ROLE_LABEL),
}

# Qual campo usa qual dicionario de rotulos. O nome do campo vem do SELECT,
# e por isso a mesma coluna `status` precisa ser desambiguada por aba.
ROTULO_DO_CAMPO: dict[tuple[str, str], str] = {
    ("Publicações", "status"): "status",
    ("Submissões", "decision"): "decision",
    ("Projetos", "status"): "project_status",
    ("Integrantes", "role"): "role",
    ("Integrantes", "thesis_kind"): "thesis_kind",
    ("Integrantes", "thesis_status"): "thesis_status",
    ("Eventos", "kind"): "kind_evento",
}

# Colunas que o banco guarda como 0/1 e a planilha mostra como Sim/Nao.
BOOLEANAS = {"is_external", "active", "is_corresponding", "open_access",
             "desk_reject", "all_day", "attended"}


def _valor(aba: str, campo: str, bruto: Any) -> Any:
    if bruto is None:
        return None
    if campo in BOOLEANAS:
        return SIM_NAO[0] if int(bruto) else SIM_NAO[1]
    dicionario = ROTULO_DO_CAMPO.get((aba, campo))
    if dicionario:
        # sem rotulo conhecido, sai o codigo cru: melhor feio do que vazio
        return ROTULOS[dicionario].get(str(bruto), bruto)
    return bruto


def dados(db: Database) -> list[tuple[str, list[str], list[list[Any]]]]:
    """(aba, cabecalhos, linhas) para cada aba, na ordem."""
    saida = []
    for nome, sql, colunas in ABAS:
        # Sem rede de seguranca aqui de proposito. A primeira versao
        # engolia o erro e escrevia a aba vazia -- e uma planilha com a aba
        # Publicacoes em branco parece um laboratorio sem producao, nao um
        # defeito. Erro de consulta tem de aparecer.
        registros = db.dicts(sql)
        cabecalhos = [rotulo for rotulo, _ in colunas]
        linhas = [[_valor(nome, campo, r.get(campo)) for _, campo in colunas]
                  for r in registros]
        saida.append((nome, cabecalhos, linhas))
    return saida


def gerar(db: Database, destino: Path | None = None,
          db_path: Path | None = None) -> Path:
    """Escreve a planilha e devolve o caminho.

    Escreve num arquivo ao lado e so entao troca pelo definitivo: se o
    processo morrer no meio, o que sobra e o arquivo velho inteiro, nunca
    um .xlsx pela metade. E se o Excel estiver com ele aberto (o Windows
    tranca), a troca falha -- entao a versao nova fica gravada ao lado,
    com aviso, em vez de se perder.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    alvo = Path(destino) if destino else caminho(db_path)
    alvo.parent.mkdir(parents=True, exist_ok=True)

    livro = Workbook()
    livro.remove(livro.active)

    # A primeira aba e um aviso, nao dado. O nome "Instruções" esta na lista
    # de abas ignoradas pela importacao, entao ela nao atrapalha a volta.
    guia = livro.create_sheet("Instruções")
    guia["A1"] = "LAPE — planilha gerada pelo sistema"
    guia["A1"].font = Font(bold=True, size=14)
    guia["A3"] = "Atualizada em"
    guia["A3"].font = Font(bold=True)
    guia["B3"] = datetime.now().strftime("%d/%m/%Y %H:%M")
    guia["A5"] = AVISO
    guia["A5"].alignment = Alignment(wrap_text=True, vertical="top")
    guia.merge_cells("A5:F9")
    guia.column_dimensions["A"].width = 22
    for letra in "BCDEF":
        guia.column_dimensions[letra].width = 18

    titulo = Font(bold=True, color="FFFFFF")
    fundo = PatternFill("solid", fgColor="12799F")
    resumo: dict[str, int] = {}

    for nome, cabecalhos, linhas in dados(db):
        aba = livro.create_sheet(nome)
        aba.append(cabecalhos)
        for linha in linhas:
            aba.append(linha)
        for celula in aba[1]:
            celula.font = titulo
            celula.fill = fundo
        aba.freeze_panes = "A2"
        if linhas:
            aba.auto_filter.ref = aba.dimensions
        for i, rotulo in enumerate(cabecalhos, start=1):
            maior = max([len(str(rotulo))] +
                        [len(str(l[i - 1])) for l in linhas[:200] if l[i - 1] is not None] or [0])
            aba.column_dimensions[get_column_letter(i)].width = min(max(maior + 2, 10), 52)
        resumo[nome] = len(linhas)

    guia["A11"] = "O que há nesta planilha"
    guia["A11"].font = Font(bold=True)
    for i, (nome, quantas) in enumerate(resumo.items(), start=12):
        guia[f"A{i}"] = nome
        guia[f"B{i}"] = quantas

    provisorio = alvo.with_suffix(".novo.xlsx")
    livro.save(provisorio)
    try:
        os.replace(provisorio, alvo)
    except OSError:
        # Arquivo aberto no Excel: a versao nova fica ao lado, e a proxima
        # passagem tenta de novo. Melhor um arquivo a mais do que uma
        # atualizacao perdida sem ninguem saber.
        return provisorio
    return alvo


# ----------------------------------------------------------------------
# Quando gerar
# ----------------------------------------------------------------------
def _agora_do_banco(db: Database) -> datetime:
    """O relogio do SQLite e UTC; o do Python e local. Comparar os dois
    misturados faz a conta de "quanto tempo desde a ultima" dar negativo,
    e nunca chegar a hora."""
    return datetime.strptime(db.scalar("SELECT datetime('now')"), "%Y-%m-%d %H:%M:%S")


def marca_atual(db: Database) -> int:
    """O maior id do change_log agora: a marca de ate onde a planilha viu."""
    return int(db.scalar("SELECT COALESCE(MAX(id), 0) FROM change_log") or 0)


def ultima(db: Database) -> dict[str, Any] | None:
    linhas = db.dicts(
        "SELECT * FROM ingest_log WHERE source = 'planilha' AND status = 'ok'"
        " ORDER BY id DESC LIMIT 1")
    return linhas[0] if linhas else None


def pendente(db: Database) -> tuple[bool, str]:
    """(gerar?, motivo). O motivo vai para o log -- e para a conferencia."""
    anterior = ultima(db)
    if anterior is None:
        return True, "primeira planilha"
    marca = int(anterior.get("rows_read") or 0)
    novas = int(db.scalar("SELECT COUNT(*) FROM change_log WHERE id > ?", (marca,)) or 0)
    quando = anterior.get("run_at")
    idade = None
    if quando:
        try:
            idade = _agora_do_banco(db) - datetime.strptime(str(quando)[:19], "%Y-%m-%d %H:%M:%S")
        except ValueError:
            idade = None
    if idade is not None and idade < timedelta(minutes=INTERVALO_MINIMO_MIN):
        return False, f"planilha de {int(idade.total_seconds() // 60)} min atrás"
    if novas >= MUDANCAS_PARA_GERAR:
        return True, f"{novas} mudança(s) no cadastro"
    if idade is None or idade >= timedelta(hours=INTERVALO_DIARIO_H):
        return True, "planilha diária"
    return False, "nada mudou desde a última"


def rodar(db: Database, forcar: bool = False,
          db_path: Path | None = None) -> dict[str, Any]:
    """Gera a planilha se houver motivo. Devolve o que foi feito."""
    gerar_agora, motivo = (True, "pedido explícito") if forcar else pendente(db)
    if not gerar_agora:
        return {"gerou": False, "motivo": motivo}
    marca = marca_atual(db)
    inicio = _agora_do_banco(db)
    try:
        alvo = gerar(db, db_path=db_path)
    except Exception as exc:
        db.log_ingest("planilha", target=NOME_ARQUIVO, status="erro", message=str(exc))
        raise
    # `rows_read` guarda a marca do change_log: e assim que a proxima
    # passagem sabe se houve cadastro novo desde esta.
    db.log_ingest("planilha", target=str(alvo), status="ok", rows_read=marca,
                  message=motivo)
    return {"gerou": True, "motivo": motivo, "arquivo": str(alvo),
            "marca": marca, "segundos": (_agora_do_banco(db) - inicio).total_seconds()}


def resumo(db: Database, db_path: Path | None = None) -> dict[str, Any]:
    """Estado da planilha, para a conferencia e para o painel."""
    anterior = ultima(db)
    alvo = caminho(db_path)
    return {
        "arquivo": str(alvo),
        "existe": alvo.exists(),
        "bytes": alvo.stat().st_size if alvo.exists() else 0,
        "atualizada_em": (anterior or {}).get("run_at"),
        "motivo": (anterior or {}).get("message"),
        "pendente": pendente(db)[1],
    }
