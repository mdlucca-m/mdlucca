#!/usr/bin/env python3
"""Gera a planilha complementar de cadastros do LAPE.

A planilha 'LAPE_Gestao_Indicadores_Cientificos' ja cobre o pipeline de
artigos e as tentativas de submissao. Faltam os cadastros que sustentam as
analises de rede, linhas de pesquisa, calendario e citacoes -- e e isso que
este script cria, em data/raw/LAPE_cadastros.xlsx.

A aba 'Integrantes' ja vem pre-preenchida com os nomes encontrados nas
planilhas existentes, para o laboratorio apenas completar e consolidar as
variacoes de grafia (ex.: 'Alexandro' e 'Andrade' sao a mesma pessoa).

Uso:  python3 scripts/make_templates.py [--force]
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from lape import config
from lape.ingest_excel import discover_sources, read_source, rows_of
from lape.mapping import resolve_sheet
from lape.util import clean_text, split_authors

TEMPLATE_PATH = config.RAW_DIR / "LAPE_cadastros.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)

SHEETS: dict[str, list[str]] = {
    "Integrantes": [
        "Nome completo", "Nome curto", "Variações", "Função", "Linha de pesquisa",
        "Instituição", "Lattes", "ORCID", "E-mail", "Data de entrada", "Data de saída",
        "Externo", "Ativo",
    ],
    "Linhas de Pesquisa": [
        "Código", "Linha de pesquisa", "Descrição", "Coordenador",
        "Data de início", "Palavras-chave", "Ativa",
    ],
    "Instituições": [
        "Instituição", "Sigla", "Cidade", "Estado", "País", "Latitude", "Longitude",
    ],
    "Publicações": [
        "Título", "Autores", "Linha de pesquisa", "Status", "Data de início",
        "Data de submissão", "Data do aceite", "Data de publicação", "Ano",
        "Periódico", "ISSN", "Qualis", "Fator de impacto", "DOI", "Link",
        "Citações WoS", "Citações Scopus", "Acesso aberto", "Observações",
    ],
    "Eventos": [
        "Tipo", "Título", "Descrição", "Data", "Data de término", "Local",
        "Instituição", "Cidade", "Estado", "País", "Latitude", "Longitude",
        "Linha de pesquisa", "Participantes", "Link",
    ],
    "Motivos de recusa": ["Código", "Motivo", "Categoria"],
}

REJECTION_CATALOG = [
    ("fora_do_escopo", "Fora do escopo da revista", "Editorial"),
    ("desk_reject_prioridade", "Baixa prioridade editorial", "Editorial"),
    ("metodo_fragil", "Fragilidade metodológica", "Método"),
    ("amostra_insuficiente", "Amostra insuficiente ou não representativa", "Método"),
    ("analise_inadequada", "Análise estatística inadequada", "Método"),
    ("originalidade", "Contribuição pouco original", "Mérito"),
    ("redacao_ingles", "Qualidade da redação em inglês", "Forma"),
    ("formatacao", "Não atende às normas da revista", "Forma"),
    ("etica", "Pendência ética ou de registro do estudo", "Conformidade"),
    ("revisao_negativa", "Pareceres negativos após revisão por pares", "Revisão por pares"),
    ("outro", "Outro motivo", "Outro"),
]

EXAMPLE_ROWS: dict[str, list[list]] = {
    "Linhas de Pesquisa": [
        ["psicologia_esporte", "Psicologia do Esporte e do Exercício",
         "Aspectos psicológicos do desempenho, treinamento e competição.",
         "Alexandro Andrade", "2000-01-01", "ansiedade; motivação; humor; atletas", "Sim"],
        ["saude_mental_exercicio", "Saúde Mental e Exercício Físico",
         "Efeitos do exercício sobre depressão, ansiedade e qualidade de vida.",
         "Alexandro Andrade", "2005-01-01", "depressão; qualidade de vida; exercício", "Sim"],
        ["dor_cronica", "Dor Crônica e Fibromialgia",
         "Intervenções psicológicas e de exercício no manejo da dor crônica.",
         "", "2012-01-01", "fibromialgia; dor; treinamento resistido", "Sim"],
    ],
    "Instituições": [
        ["Universidade do Estado de Santa Catarina", "UDESC", "Florianópolis", "SC",
         "Brasil", -27.5954, -48.5480],
        ["Centro de Ciências da Saúde e do Esporte", "CEFID/UDESC", "Florianópolis", "SC",
         "Brasil", -27.5954, -48.5480],
    ],
    "Eventos": [
        ["Reunião", "Reunião geral do LAPE", "Alinhamento do pipeline de artigos",
         "2026-09-02 14:00", "", "Sala de reuniões", "CEFID/UDESC", "Florianópolis", "SC",
         "Brasil", -27.5954, -48.5480, "", "Andrade; Vilarino; Loiane", ""],
    ],
}

NOTES = [
    ("Como usar esta planilha", ""),
    ("", ""),
    ("Integrantes", "Uma linha por pessoa. Em 'Variações' liste, separadas por ponto e vírgula, "
                    "todas as grafias usadas nas outras planilhas (ex.: Alexandro; Andrade; ANDRADE, A.). "
                    "É isso que consolida a rede de colaboração."),
    ("Linhas de Pesquisa", "O índice de linhas do painel vem daqui. Use o mesmo texto da coluna "
                           "'Linha de pesquisa' nas abas Publicações e Eventos."),
    ("Instituições", "Latitude/longitude alimentam o mapa de distribuição espacial. "
                     "Sem coordenadas o local ainda aparece nas tabelas, mas não no mapa."),
    ("Publicações", "Artigos já publicados/aceitos, com DOI. O DOI é o que permite buscar as "
                    "citações no Scopus e na Web of Science automaticamente. Artigos importados "
                    "do XML do Lattes são mesclados por título."),
    ("Eventos", "Reuniões, coletas, defesas, congressos, cursos. Alimentam o calendário, "
                "a análise temporal e o mapa."),
    ("Motivos de recusa", "Catálogo controlado. Use exatamente o mesmo texto na coluna "
                          "'Motivo/observação' da aba de tentativas de submissão."),
    ("", ""),
    ("Importante", "Não renomeie as abas. Colunas podem ser reordenadas e novas colunas podem "
                   "ser acrescentadas — o importador reconhece as colunas pelo nome."),
]


def detected_members(raw_dir: Path) -> list[str]:
    """Coleta os nomes de pessoas que ja aparecem nas planilhas existentes."""
    names: set[str] = set()
    for path in discover_sources(raw_dir):
        if path.name == TEMPLATE_PATH.name:
            continue
        try:
            sheets = read_source(path)
        except Exception:
            continue
        for sheet_name, frame in sheets.items():
            if resolve_sheet(str(sheet_name)) != "articles":
                continue
            for row in rows_of(frame, "articles"):
                names.update(split_authors(row.get("authors")))
                lead = clean_text(row.get("lead"))
                if lead:
                    names.add(lead)
    return sorted(names, key=str.lower)


def build_frames(raw_dir: Path) -> dict[str, pd.DataFrame]:
    frames: dict[str, pd.DataFrame] = {}
    for sheet, columns in SHEETS.items():
        rows = EXAMPLE_ROWS.get(sheet, [])
        if sheet == "Motivos de recusa":
            rows = [list(item) for item in REJECTION_CATALOG]
        if sheet == "Integrantes":
            rows = [
                [name, "", name, "", "", "", "", "", "", "", "", "Não", "Sim"]
                for name in detected_members(raw_dir)
            ]
        frames[sheet] = pd.DataFrame(rows, columns=columns)
    frames["Instruções"] = pd.DataFrame(NOTES, columns=["Aba", "Orientação"])
    return frames


def style_workbook(path: Path) -> None:
    from openpyxl import load_workbook

    book = load_workbook(path)
    for sheet in book.worksheets:
        sheet.freeze_panes = "A2"
        for cell in sheet[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        sheet.row_dimensions[1].height = 28
        for index, column in enumerate(sheet.columns, start=1):
            width = max((len(str(c.value)) for c in column if c.value is not None), default=10)
            sheet.column_dimensions[get_column_letter(index)].width = min(max(width + 3, 14), 52)
    book.save(path)


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--force", action="store_true", help="sobrescreve a planilha existente")
    parser.add_argument("--output", type=Path, default=TEMPLATE_PATH)
    args = parser.parse_args()

    if args.output.exists() and not args.force:
        print(f"! {args.output} ja existe. Use --force para sobrescrever.")
        return 1

    args.output.parent.mkdir(parents=True, exist_ok=True)
    frames = build_frames(config.RAW_DIR)
    with pd.ExcelWriter(args.output, engine="openpyxl") as writer:
        for sheet, frame in frames.items():
            frame.to_excel(writer, sheet_name=sheet, index=False)
    style_workbook(args.output)

    members = len(frames["Integrantes"])
    print(f"Planilha de cadastros criada: {args.output}")
    print(f"  Integrantes pre-preenchidos a partir das planilhas existentes: {members}")
    print("  Preencha e rode: python3 scripts/run_pipeline.py")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
