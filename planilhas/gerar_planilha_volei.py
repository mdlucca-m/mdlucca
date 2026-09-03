# -*- coding: utf-8 -*-
"""
Gerador da Planilha de Controle e Prescricao de Treinamento - VOLEIBOL
Estrutura: Macrociclo -> Mesociclo -> Microciclo -> Sessao -> Exercicio
Controles: PSE da sessao (Foster), ACWR, Monotonia/Strain, Indice de Hooper,
           presenca, testes fisicos, area do atleta e painel interativo.
"""
import random
from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule, ColorScaleRule
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.comments import Comment

# ----------------------------------------------------------------------------
# IDENTIDADE VISUAL
# ----------------------------------------------------------------------------
F      = "Arial"
NAVY   = "13315C"   # cabecalho principal
NAVY2  = "1F4E96"   # sub-cabecalho
BLUE3  = "2E75B6"   # cabecalho de tabela
LIGHT  = "EAF1F8"   # faixa clara
LIGHT2 = "F5F8FC"
GOLD   = "E8A33D"   # acento (bola de volei)
GOLD_L = "FFF2CC"   # celula de entrada
GREEN  = "C6EFCE"; GREEN_T = "0B6B2E"
YELL   = "FFEB9C"; YELL_T  = "8A6D0B"
RED    = "FFC7CE"; RED_T   = "9C0006"
GREY_T = "7F8C9A"
WHITE  = "FFFFFF"

thin = Side(style="thin", color="BFCEDD")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

# formatos numericos (zero exibido em branco)
NF_UA   = '#,##0;-#,##0;""'
NF_DEC  = '0.00;-0.00;""'
NF_INT  = '#,##0;-#,##0;""'
NF_PCT  = '0%;-0%;""'
NF_PCT1 = '0.0%;-0.0%;""'
NF_DATE = 'DD/MM/YYYY'
NF_CM   = '0.0" cm";;""'

# ----------------------------------------------------------------------------
# LIMITES DAS TABELAS
# ----------------------------------------------------------------------------
CAD_F, CAD_L   = 7, 66      # Cadastro (60 atletas)
EX_F,  EX_L    = 7, 206     # Biblioteca de exercicios
PRE_F, PRE_L   = 8, 507     # Prescricao
CAR_F, CAR_L   = 8, 1007    # Controle de carga (PSE)
WEL_F, WEL_L   = 8, 1007    # Wellness
PRS_F, PRS_L   = 8, 1007    # Presenca
TST_F, TST_L   = 8, 407     # Testes fisicos
N_SEM          = 48         # semanas do macrociclo nos paineis

wb = Workbook()

# ----------------------------------------------------------------------------
# HELPERS DE ESTILO
# ----------------------------------------------------------------------------
def banner(ws, titulo, subtitulo, last_col, tab=NAVY):
    ws.sheet_properties.tabColor = tab
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    c = ws.cell(1, 1, titulo)
    c.font = Font(name=F, size=15, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.row_dimensions[1].height = 30
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
    c = ws.cell(2, 1, subtitulo)
    c.font = Font(name=F, size=9, color=WHITE, italic=True)
    c.fill = PatternFill("solid", fgColor=GOLD)
    c.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.row_dimensions[2].height = 16
    ws.sheet_view.showGridLines = False

def secao(ws, row, texto, last_col, col=1):
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=last_col)
    c = ws.cell(row, col, texto)
    c.font = Font(name=F, size=10, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=NAVY2)
    c.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.row_dimensions[row].height = 20

def cab_tabela(ws, row, headers, col0=1, wrap=True):
    for i, h in enumerate(headers):
        c = ws.cell(row, col0 + i, h)
        c.font = Font(name=F, size=9, bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=BLUE3)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
        c.border = BORDER
    ws.row_dimensions[row].height = 32

def rotulo(ws, row, col, texto):
    c = ws.cell(row, col, texto)
    c.font = Font(name=F, size=9, bold=True, color=NAVY)
    c.alignment = Alignment(horizontal="right", vertical="center")
    return c

def entrada(ws, row, col, valor=None, nf=None, largura_merge=None):
    c = ws.cell(row, col, valor)
    c.font = Font(name=F, size=10, color="0000FF")
    c.fill = PatternFill("solid", fgColor=GOLD_L)
    c.border = BORDER
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    if nf: c.number_format = nf
    if largura_merge:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + largura_merge - 1)
    return c

def calc(ws, row, col, formula, nf=None):
    c = ws.cell(row, col, formula)
    c.font = Font(name=F, size=10, bold=True, color=NAVY2)
    c.fill = PatternFill("solid", fgColor=LIGHT)
    c.border = BORDER
    c.alignment = Alignment(horizontal="center", vertical="center")
    if nf: c.number_format = nf
    return c

def corpo_tabela(ws, r1, r2, c1, c2, zebra=True):
    """Aplica borda, fonte e zebra ao corpo de uma tabela."""
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cel = ws.cell(r, c)
            cel.font = Font(name=F, size=9)
            cel.border = BORDER
            cel.alignment = Alignment(vertical="center", horizontal="center")
            if zebra and (r - r1) % 2 == 1:
                cel.fill = PatternFill("solid", fgColor=LIGHT2)

def larguras(ws, mapa):
    for col, w in mapa.items():
        ws.column_dimensions[col].width = w

def nota(ws, row, col, texto, last_col=None):
    c = ws.cell(row, col, texto)
    c.font = Font(name=F, size=8, italic=True, color=GREY_T)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    if last_col:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=last_col)
    return c

def dv(ws, formula1, ref, allow_blank=True):
    d = DataValidation(type="list", formula1=formula1, allow_blank=allow_blank, showErrorMessage=False)
    ws.add_data_validation(d)
    d.add(ref)
    return d

def kpi(ws, row, col, titulo, formula, nf=NF_INT, cor=NAVY2, w=3):
    """Cartao de KPI ocupando w colunas x 2 linhas."""
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + w - 1)
    t = ws.cell(row, col, titulo)
    t.font = Font(name=F, size=8, bold=True, color=WHITE)
    t.fill = PatternFill("solid", fgColor=cor)
    t.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 26
    ws.merge_cells(start_row=row + 1, start_column=col, end_row=row + 1, end_column=col + w - 1)
    v = ws.cell(row + 1, col, formula)
    v.font = Font(name=F, size=16, bold=True, color=cor)
    v.fill = PatternFill("solid", fgColor=LIGHT)
    v.alignment = Alignment(horizontal="center", vertical="center")
    v.number_format = nf
    v.border = BORDER
    ws.row_dimensions[row + 1].height = 30
    return v

# ============================================================================
# 1) LISTAS  (fonte das validacoes de dados)
# ============================================================================
LISTAS = [
 ("Posição",            ["Levantador","Oposto","Ponteiro (Ponta)","Central","Líbero","Defensor Específico"]),
 ("Categoria",          ["Sub-13","Sub-15","Sub-17","Sub-19","Sub-21","Adulto","Master"]),
 ("Sexo",               ["Masculino","Feminino"]),
 ("Dominância",         ["Destro","Canhoto","Ambidestro"]),
 ("Status do Atleta",   ["Ativo","Lesionado","Departamento Médico","Em transição","Afastado","Inativo"]),
 ("Período do Macro",   ["Preparatório Geral","Preparatório Específico","Pré-Competitivo","Competitivo I","Competitivo II","Transição"]),
 ("Tipo de Microciclo", ["Incorporação","Ordinário","Choque","Recuperativo","Pré-Competitivo","Competitivo","Polimento (Taper)"]),
 ("Bloco da Sessão",    ["Aquecimento","Ativação / Prevenção","Parte Principal","Complementar","Volta à Calma"]),
 ("Categoria Exerc.",   ["Técnico","Tático","Físico","Preventivo/Compensatório","Recuperação","Cognitivo/Visual"]),
 ("Fundamento",         ["Saque","Recepção (Passe)","Levantamento","Ataque","Bloqueio","Defesa","Deslocamento","Jogo/Coletivo","N/A"]),
 ("Capacidade Física",  ["Força Máxima","Força Explosiva/Potência","Pliometria","Velocidade","Agilidade/COD","Resistência Aeróbia",
                         "Resistência Anaeróbia","Mobilidade/Flexibilidade","Core/Estabilidade","Coordenação","N/A"]),
 ("Presença",           ["Presente","Falta Justificada","Falta Não Justificada","Lesionado","Departamento Médico","Liberado","Seleção/Convocado"]),
 ("Intensidade",        ["Muito Baixa","Baixa","Moderada","Alta","Muito Alta","Máxima"]),
 ("PSE (0-10)",         [0,1,2,3,4,5,6,7,8,9,10]),
 ("Escala 1-7",         [1,2,3,4,5,6,7]),
 ("Turno",              ["Manhã","Tarde","Noite"]),
 ("Tipo de Sessão",     ["Técnico-Tático","Físico (Força)","Físico (Potência)","Físico (Condicionamento)","Coletivo/Jogo",
                         "Amistoso","Competição Oficial","Recuperação/Regenerativo","Vídeo/Teórico","Folga"]),
 ("Nível",              ["Iniciante","Intermediário","Avançado"]),
 ("Situação",           ["Planejado","Em execução","Concluído","Cancelado"]),
 ("Momento do Teste",   ["Pré-temporada","Fim Meso 1","Fim Meso 2","Fim Meso 3","Meio da Temporada","Fim da Temporada","Retorno de Lesão"]),
]

wsL = wb.active; wsL.title = "Listas"
banner(wsL, "LISTAS AUXILIARES  |  fonte das caixas de seleção",
       "Edite/complemente as opções abaixo — elas alimentam automaticamente todos os menus suspensos da planilha.", 22, GREY_T)
cab_tabela(wsL, 4, [n for n, _ in LISTAS])
for j, (_, itens) in enumerate(LISTAS, start=1):
    for i, v in enumerate(itens):
        c = wsL.cell(5 + i, j, v)
        c.font = Font(name=F, size=9); c.border = BORDER
        c.alignment = Alignment(horizontal="center")
    wsL.column_dimensions[get_column_letter(j)].width = 22
LST_F, LST_L = 5, 5 + max(len(i) for _, i in LISTAS) - 1
COL = {n: get_column_letter(j) for j, (n, _) in enumerate(LISTAS, start=1)}

def L(nome):
    """Referencia absoluta da lista na aba Listas."""
    c = COL[nome]
    return "Listas!${0}${1}:${0}${2}".format(c, LST_F, LST_L)

# coluna dinamica: destinatarios da prescricao (Equipe + atletas do cadastro)
cD = get_column_letter(len(LISTAS) + 2)
wsL.column_dimensions[cD].width = 26
h = wsL.cell(4, len(LISTAS) + 2, "Destinatário (auto)")
h.font = Font(name=F, size=9, bold=True, color=WHITE); h.fill = PatternFill("solid", fgColor=GOLD)
h.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); h.border = BORDER
wsL.cell(5, len(LISTAS) + 2, "Equipe (todos)").font = Font(name=F, size=9)
wsL.cell(5, len(LISTAS) + 2).border = BORDER
for i in range(CAD_L - CAD_F + 1):
    c = wsL.cell(6 + i, len(LISTAS) + 2, '=IF(Cadastro!$B${0}="","",Cadastro!$B${0})'.format(CAD_F + i))
    c.font = Font(name=F, size=9); c.border = BORDER
DEST_REF = "Listas!${0}$5:${0}${1}".format(cD, 5 + (CAD_L - CAD_F + 1))
wsL.freeze_panes = "A5"

ATLETAS_REF = "Cadastro!$B${}:$B${}".format(CAD_F, CAD_L)
EXERC_REF   = "Exercícios!$B${}:$B${}".format(EX_F, EX_L)
MESO_REF    = "Macrociclo!$B$19:$B$30"

# ============================================================================
# 2) INÍCIO  (instruções + legenda + índice)
# ============================================================================
wsI = wb.create_sheet("Início")
banner(wsI, "PLANILHA DE CONTROLE E PRESCRIÇÃO DE TREINAMENTO  •  VOLEIBOL",
       "Periodização Macrociclo → Mesociclo → Microciclo → Sessão  |  Controle de carga por PSE da sessão, ACWR, Monotonia/Strain e Wellness", 10, GOLD)
larguras(wsI, {"A":3,"B":22,"C":16,"D":16,"E":16,"F":16,"G":16,"H":16,"I":16,"J":18})

secao(wsI, 4, "COMO USAR — PASSO A PASSO", 10, 2)
passos = [
 ("1. Listas",      "Confira/ajuste as opções de todos os menus suspensos (posições, categorias, tipos de sessão, etc.)."),
 ("2. Cadastro",    "Cadastre os atletas. O nome cadastrado alimenta TODAS as demais abas automaticamente."),
 ("3. Macrociclo",  "Defina a temporada (datas, objetivo) e divida-a em mesociclos com volume/intensidade previstos."),
 ("4. Mesociclo",   "Escolha um mesociclo no menu e detalhe os microciclos (semanas), com carga prevista x realizada."),
 ("5. Microciclo",  "Monte a semana: sessões por dia, duração, PSE prevista e distribuição de conteúdos."),
 ("6. Exercícios",  "Biblioteca de exercícios. Já vem com 42 exercícios de voleibol — inclua os seus."),
 ("7. Prescrição",  "PRESCREVA o treino: uma linha por exercício, com séries, repetições, carga, pausa e destinatário."),
 ("8. Carga (PSE)", "REGISTRE o realizado: duração x PSE = carga em UA. ACWR, aguda e crônica são calculados sozinhos."),
 ("9. Wellness",    "Questionário diário (sono, estresse, fadiga, dor) → Índice de Hooper e classificação automática."),
 ("10. Presença",   "Chamada por sessão, com % de presença por atleta calculado ao lado."),
 ("11. Testes",     "Avaliações físicas periódicas (saltos, alcances, agilidade, sprint, flexibilidade)."),
 ("12. Atleta",     "ÁREA DO ATLETA: escolha o nome e ele vê o próprio perfil, cargas, monotonia e o treino prescrito."),
 ("13. Painel",     "PAINEL INTERATIVO: indicadores e gráficos de toda a equipe, por semana e por atleta."),
]
r = 5
for tit, desc in passos:
    c = wsI.cell(r, 2, tit); c.font = Font(name=F, size=9, bold=True, color=NAVY2)
    c.alignment = Alignment(horizontal="left", vertical="center")
    wsI.merge_cells(start_row=r, start_column=3, end_row=r, end_column=10)
    d = wsI.cell(r, 3, desc); d.font = Font(name=F, size=9)
    d.alignment = Alignment(horizontal="left", vertical="center")
    if (r % 2) == 0:
        for cc in range(2, 11):
            wsI.cell(r, cc).fill = PatternFill("solid", fgColor=LIGHT2)
    r += 1

r += 1
secao(wsI, r, "LEGENDA DE CORES", 10, 2); r += 1
legenda = [("Célula de PREENCHIMENTO (azul sobre amarelo)", GOLD_L, "0000FF"),
           ("Célula CALCULADA — não digite por cima", LIGHT, NAVY2),
           ("Cabeçalho de tabela", BLUE3, WHITE),
           ("Alerta / risco elevado", RED, RED_T),
           ("Atenção / monitorar", YELL, YELL_T),
           ("Adequado / dentro da meta", GREEN, GREEN_T)]
for txt, fill, font in legenda:
    wsI.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    c = wsI.cell(r, 2, txt)
    c.fill = PatternFill("solid", fgColor=fill); c.border = BORDER
    c.font = Font(name=F, size=9, bold=True, color=font)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    r += 1

r += 1
secao(wsI, r, "MÉTRICAS DE CONTROLE UTILIZADAS — DEFINIÇÕES E REFERÊNCIAS", 10, 2); r += 1
metricas = [
 ("Carga Interna (UA)", "Duração da sessão (min) × PSE da sessão (escala 0–10, coletada ~30 min após o treino). Método da PSE da sessão de Foster et al. (2001), J Strength Cond Res."),
 ("Carga Aguda (7 d)",  "Soma da carga em UA dos últimos 7 dias do atleta (inclui o dia atual)."),
 ("Carga Crônica (28 d)","Soma da carga dos últimos 28 dias ÷ 4 (média semanal), para ficar na mesma escala da aguda."),
 ("ACWR",               "Carga Aguda ÷ Carga Crônica. Faixas usadas na planilha: <0,80 subcarga | 0,80–1,30 zona ideal | 1,31–1,50 atenção | >1,50 risco elevado. Referência: Gabbett (2016), Br J Sports Med. Interpretar sempre junto do contexto e da clínica do atleta."),
 ("Monotonia",          "Média das cargas diárias da semana ÷ desvio-padrão dessas mesmas cargas (7 dias, dias de folga contam como zero). >2,0 indica semana monótona."),
 ("Strain",             "Carga semanal total × Monotonia. Picos de strain associam-se a maior risco de lesão/doença (Foster, 1998)."),
 ("Índice de Hooper",   "Soma de 4 itens (qualidade do sono, estresse, fadiga e dor muscular), cada um de 1 = muito bom/ausente a 7 = muito ruim/máxima. Varia de 4 a 28; QUANTO MAIOR, PIOR. Referência: Hooper & Mackinnon (1995), Sports Med."),
 ("Volume × Intensidade","No Macrociclo, valores em % representam a proporção relativa planejada para cada mesociclo (ex.: 85% = volume alto)."),
]
for nome, desc in metricas:
    c = wsI.cell(r, 2, nome); c.font = Font(name=F, size=9, bold=True, color=NAVY2)
    c.alignment = Alignment(horizontal="left", vertical="top")
    wsI.merge_cells(start_row=r, start_column=3, end_row=r, end_column=10)
    d = wsI.cell(r, 3, desc); d.font = Font(name=F, size=9)
    d.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    wsI.row_dimensions[r].height = 30
    r += 1

r += 1
wsI.merge_cells(start_row=r, start_column=2, end_row=r + 2, end_column=10)
c = wsI.cell(r, 2, "ATENÇÃO — DADOS DE EXEMPLO: a planilha vem preenchida com 12 atletas fictícios, um macrociclo modelo e "
                   "registros de carga/wellness/presença apenas para demonstrar os cálculos e os gráficos. Apague esses dados "
                   "(linhas marcadas como EXEMPLO) antes de usar com a sua equipe. As fórmulas permanecem intactas.\n"
                   "AVISO: esta planilha é uma ferramenta de organização do treino. Não substitui avaliação médica nem "
                   "fisioterápica; os indicadores de risco são apoio à decisão do técnico/preparador físico.")
c.font = Font(name=F, size=9, bold=True, color=RED_T)
c.fill = PatternFill("solid", fgColor=RED); c.border = BORDER
c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)

# ============================================================================
# 3) CADASTRO DE ATLETAS
# ============================================================================
wsC = wb.create_sheet("Cadastro")
banner(wsC, "CADASTRO DE ATLETAS", "Preencha as colunas em amarelo. As colunas cinza são calculadas. O nome digitado aqui aparece "
       "nos menus de todas as outras abas.", 22, NAVY2)
CAD_H = ["ID","Nome Completo","Data de Nasc.","Idade","Sexo","Posição","Categoria","Nº Camisa","Dominância",
         "Estatura (m)","Massa (kg)","IMC","Envergadura (cm)","Alcance em Pé (cm)","Alcance de Ataque (cm)",
         "Alcance de Bloqueio (cm)","Impulsão de Ataque (cm)","Impulsão de Bloqueio (cm)","Entrada na Equipe",
         "Status","Contato / Responsável","Observações"]
cab_tabela(wsC, 6, CAD_H)
larguras(wsC, {"A":9,"B":26,"C":13,"D":7,"E":11,"F":17,"G":10,"H":8,"I":12,"J":11,"K":11,"L":8,"M":13,
               "N":14,"O":15,"P":15,"Q":15,"R":15,"S":14,"T":14,"U":22,"V":26})
nota(wsC, 4, 2, "Campos obrigatórios: Nome, Data de Nascimento, Estatura e Massa. Alcance em Pé = altura atingida em pé com o "
     "braço dominante estendido — é a base para calcular as impulsões.", 22)

for r in range(CAD_F, CAD_L + 1):
    wsC.cell(r, 1, '=IF($B{0}="","","ATL-"&TEXT(ROW()-{1},"000"))'.format(r, CAD_F - 1))
    wsC.cell(r, 4, '=IFERROR(DATEDIF($C{0},TODAY(),"Y"),"")'.format(r))
    wsC.cell(r, 12, '=IFERROR(ROUND($K{0}/$J{0}^2,1),"")'.format(r))
    wsC.cell(r, 17, '=IF(OR($O{0}="",$N{0}=""),"",$O{0}-$N{0})'.format(r))
    wsC.cell(r, 18, '=IF(OR($P{0}="",$N{0}=""),"",$P{0}-$N{0})'.format(r))
corpo_tabela(wsC, CAD_F, CAD_L, 1, 22)
for r in range(CAD_F, CAD_L + 1):
    for c in (2, 3, 5, 6, 7, 8, 9, 10, 11, 13, 14, 15, 16, 19, 20, 21, 22):
        wsC.cell(r, c).fill = PatternFill("solid", fgColor=GOLD_L)
        wsC.cell(r, c).font = Font(name=F, size=9, color="0000FF")
    for c in (1, 4, 12, 17, 18):
        wsC.cell(r, c).font = Font(name=F, size=9, bold=True, color=NAVY2)
        wsC.cell(r, c).fill = PatternFill("solid", fgColor=LIGHT)
    wsC.cell(r, 2).alignment = Alignment(horizontal="left", vertical="center", indent=1)
    wsC.cell(r, 3).number_format = NF_DATE
    wsC.cell(r, 10).number_format = '0.00'
    wsC.cell(r, 11).number_format = '0.0'
    wsC.cell(r, 12).number_format = '0.0'
    wsC.cell(r, 19).number_format = NF_DATE
    for c in (13, 14, 15, 16, 17, 18):
        wsC.cell(r, c).number_format = '0'
dv(wsC, L("Sexo"),             "E{}:E{}".format(CAD_F, CAD_L))
dv(wsC, L("Posição"),          "F{}:F{}".format(CAD_F, CAD_L))
dv(wsC, L("Categoria"),        "G{}:G{}".format(CAD_F, CAD_L))
dv(wsC, L("Dominância"),       "I{}:I{}".format(CAD_F, CAD_L))
dv(wsC, L("Status do Atleta"), "T{}:T{}".format(CAD_F, CAD_L))
wsC.conditional_formatting.add("T{}:T{}".format(CAD_F, CAD_L),
    CellIsRule(operator="equal", formula=['"Ativo"'], fill=PatternFill("solid", fgColor=GREEN), font=Font(name=F, size=9, bold=True, color=GREEN_T)))
for txt in ('"Lesionado"', '"Departamento Médico"'):
    wsC.conditional_formatting.add("T{}:T{}".format(CAD_F, CAD_L),
        CellIsRule(operator="equal", formula=[txt], fill=PatternFill("solid", fgColor=RED), font=Font(name=F, size=9, bold=True, color=RED_T)))
wsC.freeze_panes = "C7"
wsC.auto_filter.ref = "A6:V{}".format(CAD_L)

# ---- 12 atletas de exemplo -------------------------------------------------
EXEMPLO = [
 ("Ana Beatriz Ramos",      2003, 4, 12, "Feminino","Levantador",        "Adulto",  5,"Destro",1.78,68.0,180,232,290,278),
 ("Camila Fontes Duarte",   2001, 9,  3, "Feminino","Oposto",            "Adulto", 12,"Canhoto",1.86,76.5,190,242,308,295),
 ("Larissa Mendes Prado",   2004, 1,25, "Feminino","Ponteiro (Ponta)",   "Adulto",  7,"Destro",1.81,71.0,184,236,299,286),
 ("Juliana Alves Correia",  2002,11, 8, "Feminino","Ponteiro (Ponta)",   "Adulto",  9,"Destro",1.79,69.5,182,234,296,283),
 ("Marina Tavares Lopes",   2000, 6,17, "Feminino","Central",           "Adulto",  4,"Destro",1.89,78.0,194,246,310,300),
 ("Rafaela Nunes Barros",   2003, 3,30, "Feminino","Central",           "Adulto", 15,"Destro",1.87,77.0,192,244,306,297),
 ("Isadora Pinheiro Cruz",  2005, 8,11, "Feminino","Líbero",            "Adulto",  2,"Destro",1.68,60.5,170,220,262,252),
 ("Bruna Siqueira Rocha",   2002, 2, 5, "Feminino","Ponteiro (Ponta)",   "Adulto", 11,"Destro",1.80,70.0,183,235,297,284),
 ("Gabriela Moraes Lima",   2004,12,21, "Feminino","Levantador",        "Adulto",  6,"Destro",1.76,66.0,178,230,286,275),
 ("Helena Braga Vieira",    2001, 5,14, "Feminino","Oposto",            "Adulto", 18,"Destro",1.84,74.0,188,240,303,291),
 ("Sofia Andrade Peixoto", 2006, 7, 2, "Feminino","Central","Sub-21", 3,"Destro",1.85,72.5,189,241,301,290),
 ("Letícia Carvalho Dias",  2005,10,19, "Feminino","Líbero",            "Sub-21", 1,"Canhoto",1.66,58.0,168,218,258,248),
]
for i, a in enumerate(EXEMPLO):
    r = CAD_F + i
    nome, ay, am, ad, sexo, pos, cat, cam, dom, est, mas, env, pe, atq, blq = a
    wsC.cell(r, 2, nome); wsC.cell(r, 3, date(ay, am, ad)); wsC.cell(r, 5, sexo)
    wsC.cell(r, 6, pos); wsC.cell(r, 7, cat); wsC.cell(r, 8, cam); wsC.cell(r, 9, dom)
    wsC.cell(r, 10, est); wsC.cell(r, 11, mas); wsC.cell(r, 13, env)
    wsC.cell(r, 14, pe); wsC.cell(r, 15, atq); wsC.cell(r, 16, blq)
    wsC.cell(r, 19, date(2026, 1, 5)); wsC.cell(r, 20, "Lesionado" if i == 7 else "Ativo")
    wsC.cell(r, 21, "(00) 90000-00{:02d}".format(i + 1))
    wsC.cell(r, 22, "EXEMPLO — substituir")
    wsC.cell(r, 3).number_format = NF_DATE
    wsC.cell(r, 19).number_format = NF_DATE
NOMES = [a[0] for a in EXEMPLO]

# ============================================================================
# 4) BIBLIOTECA DE EXERCÍCIOS
# ============================================================================
wsE = wb.create_sheet("Exercícios")
banner(wsE, "BIBLIOTECA DE EXERCÍCIOS", "Cada exercício cadastrado aqui fica disponível no menu suspenso da aba Prescrição, "
       "que puxa automaticamente categoria, fundamento e capacidade física.", 14, BLUE3)
EX_H = ["ID","Exercício","Categoria","Fundamento","Capacidade Física","Objetivo","Descrição / Execução",
        "Material","Nº de Atletas","Duração Sug. (min)","Intensidade Sug.","PSE Sug.","Nível","Link de Vídeo"]
cab_tabela(wsE, 6, EX_H)
larguras(wsE, {"A":8,"B":38,"C":18,"D":18,"E":22,"F":34,"G":58,"H":26,"I":11,"J":11,"K":14,"L":8,"M":14,"N":24})
for r in range(EX_F, EX_L + 1):
    wsE.cell(r, 1, '=IF($B{0}="","","EX-"&TEXT(ROW()-{1},"000"))'.format(r, EX_F - 1))
corpo_tabela(wsE, EX_F, EX_L, 1, 14)
for r in range(EX_F, EX_L + 1):
    for c in range(2, 15):
        wsE.cell(r, c).fill = PatternFill("solid", fgColor=GOLD_L)
        wsE.cell(r, c).font = Font(name=F, size=9, color="0000FF")
    for c in (2, 6, 7, 8, 14):
        wsE.cell(r, c).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    wsE.cell(r, 1).font = Font(name=F, size=9, bold=True, color=NAVY2)
    wsE.cell(r, 1).fill = PatternFill("solid", fgColor=LIGHT)
dv(wsE, L("Categoria Exerc."), "C{}:C{}".format(EX_F, EX_L))
dv(wsE, L("Fundamento"),       "D{}:D{}".format(EX_F, EX_L))
dv(wsE, L("Capacidade Física"),"E{}:E{}".format(EX_F, EX_L))
dv(wsE, L("Intensidade"),      "K{}:K{}".format(EX_F, EX_L))
dv(wsE, L("PSE (0-10)"),       "L{}:L{}".format(EX_F, EX_L))
dv(wsE, L("Nível"),            "M{}:M{}".format(EX_F, EX_L))
wsE.freeze_panes = "C7"
wsE.auto_filter.ref = "A6:N{}".format(EX_L)

T,TA,FI,PV,RC,CG = "Técnico","Tático","Físico","Preventivo/Compensatório","Recuperação","Cognitivo/Visual"
EXERCICIOS = [
("Saque flutuante de apoio",T,"Saque","Coordenação","Precisão e constância do saque flutuante","Séries de 10 saques por zona alvo (1, 5 e 6). Contato firme no centro da bola, sem rotação, braço freado no impacto.","Bolas, cones/alvos",1,12,"Moderada",4,"Iniciante",""),
("Saque viagem no fundo (jump float)",T,"Saque","Coordenação","Saque flutuante com salto e maior velocidade","Passada de 2–3 tempos, salto vertical, contato seco. 3 blocos de 8 saques com alvo no fundo da quadra.","Bolas",1,12,"Alta",5,"Intermediário",""),
("Saque potente com salto",T,"Saque","Força Explosiva/Potência","Velocidade e agressividade no saque","4 séries de 6 saques em potência máxima, alternando zona 1 e zona 5. Pausa completa entre séries.","Bolas",1,15,"Muito Alta",7,"Avançado",""),
("Recepção em duplas com bola dirigida",T,"Recepção (Passe)","Coordenação","Plataforma estável e passe ao alvo","Técnico lança bolas alternadas; atleta recebe para o alvo em zona 2/3. 3 x 15 repetições.","Bolas, alvo/cesto",2,12,"Moderada",4,"Iniciante",""),
("Recepção de saque em sistema de 2 passadores",T,"Recepção (Passe)","Coordenação","Leitura de trajetória e responsabilidade de zona","Sacador real do outro lado. Passadores em L; 20 saques por rodízio, meta de 70% de passes A/B.","Bolas, rede",4,20,"Alta",6,"Intermediário",""),
("Manchete de deslocamento lateral (leque)",T,"Recepção (Passe)","Agilidade/COD","Deslocamento e recepção fora da base","Bolas alternadas à direita e à esquerda; deslocamento em passo lateral e retorno à base. 4 x 30 s.","Bolas",1,10,"Alta",6,"Intermediário",""),
("Levantamento em suspensão contra a parede",T,"Levantamento","Coordenação","Consistência do toque e ação simétrica das mãos","3 x 30 toques contra a parede a 3 m, mantendo a bola sem rotação.","Bola, parede",1,8,"Baixa",3,"Iniciante",""),
("Levantamento em rede com alvos",T,"Levantamento","Coordenação","Precisão de bola alta, rápida e de fundo","Levantador recebe passe do técnico e distribui para alvos em P4, P2 e Pipe. 5 x 12 levantamentos.","Bolas, aros/alvos",2,18,"Moderada",5,"Intermediário",""),
("Levantamento invertido (costas)",T,"Levantamento","Coordenação","Domínio do levantamento para trás sem sinalizar","3 x 15 levantamentos de costas para P2 com o técnico corrigindo a extensão de quadril.","Bolas",2,12,"Moderada",5,"Avançado",""),
("Ataque de bola alta pela ponta (P4)",T,"Ataque","Força Explosiva/Potência","Ritmo de aproximação e ataque em bola alta","4 x 10 ataques com levantador real. Foco na passada de 3 tempos e no braço em arco.","Bolas, rede",3,18,"Alta",6,"Intermediário",""),
("Ataque de primeiro tempo (P3)",T,"Ataque","Velocidade","Sincronia central x levantador","4 x 8 ataques de bola rápida, variando o ponto de contato ao longo da rede.","Bolas, rede",3,15,"Alta",6,"Avançado",""),
("Ataque de fundo (Pipe / P6)",T,"Ataque","Força Explosiva/Potência","Ataque a partir da linha de 3 m","3 x 10 ataques de fundo saindo da recepção. Atenção à marca dos 3 m.","Bolas, rede",3,15,"Alta",6,"Avançado",""),
("Largada e explorada de bloqueio",T,"Ataque","Coordenação","Variação tática do ataque","3 x 12 alternando largada curta, exploração de bloqueio e ataque forte, comandado pelo técnico.","Bolas, rede",3,12,"Moderada",5,"Intermediário",""),
("Bloqueio individual com deslocamento lateral",T,"Bloqueio","Agilidade/COD","Deslocamento e penetração das mãos","4 x 8 deslocamentos P2→P3→P4 com salto de bloqueio em cada parada.","Rede, plataforma",1,12,"Alta",6,"Intermediário",""),
("Bloqueio duplo (fechamento ponta/central)",T,"Bloqueio","Agilidade/COD","Sincronia e fechamento do bloqueio duplo","Central desloca e fecha com o ponteiro; 4 x 10 repetições com atacante real.","Rede, bolas",4,15,"Alta",6,"Avançado",""),
("Defesa de mergulho e rolamento",T,"Defesa","Coordenação","Técnica de queda segura e recuperação de bola","3 x 10 mergulhos alternando lados, com progressão de colchonete para quadra.","Bolas, colchonete",1,12,"Alta",6,"Intermediário",""),
("Defesa em 3 zonas com técnico atacando",T,"Defesa","Agilidade/COD","Postura de defesa e leitura do atacante","Técnico ataca da plataforma para zonas 1, 6 e 5; 4 x 45 s de defesa contínua.","Bolas, plataforma",3,15,"Muito Alta",7,"Avançado",""),
("Circuito de deslocamentos específicos da posição",T,"Deslocamento","Agilidade/COD","Automatizar trajetos de saída de rede e cobertura","Circuito de cones reproduzindo o trajeto real da posição. 5 voltas com 60 s de pausa.","Cones",1,12,"Alta",6,"Intermediário",""),
("Complexo I (K1) — recepção, levantamento e ataque",TA,"Jogo/Coletivo","N/A","Eficiência do side-out","Saque adversário real; a equipe pontua se converter o side-out. Meta de 60% em 20 tentativas.","Bolas, rede",12,25,"Alta",7,"Intermediário",""),
("Complexo II (K2) — bloqueio, defesa e contra-ataque",TA,"Jogo/Coletivo","N/A","Transição defesa-ataque","Técnico ataca; equipe defende e finaliza o contra-ataque. 20 séries por rodízio.","Bolas, rede",12,25,"Alta",7,"Intermediário",""),
("Jogo 6x6 com pontuação diferenciada",TA,"Jogo/Coletivo","N/A","Aplicar o sistema em contexto real","Sets a 15 com pontos em dobro para ponto de bloqueio ou side-out no 1º toque.","Bolas, rede, placar",12,30,"Alta",7,"Intermediário",""),
("Jogo reduzido 4x4 em quadra estreita",TA,"Jogo/Coletivo","N/A","Volume de toques e leitura de jogo","3 sets a 11 pontos em meia quadra, obrigando 3 toques.","Bolas, rede",8,20,"Moderada",6,"Iniciante",""),
("Treino do sistema 5x1 — rodízios e coberturas",TA,"Jogo/Coletivo","N/A","Posicionamento em cada rodízio","Sem bola primeiro, depois com bola: 6 rodízios x 4 repetições de recepção e cobertura.","Rede, bolas",12,20,"Baixa",4,"Intermediário",""),
("Saque-recepção sob pressão de placar",TA,"Jogo/Coletivo","N/A","Desempenho técnico sob estresse competitivo","Placar simulado em 22x22; erro de saque ou recepção custa ponto. 6 rodadas.","Bolas, placar",12,20,"Muito Alta",8,"Avançado",""),
("Agachamento livre (back squat)",FI,"N/A","Força Máxima","Força máxima de membros inferiores","4 x 5 a 80–85% de 1RM, 3 min de pausa. Progressão semanal de 2,5%.","Barra, anilhas, rack",1,25,"Muito Alta",8,"Avançado",""),
("Levantamento terra romeno",FI,"N/A","Força Máxima","Força de cadeia posterior e prevenção de isquiotibiais","4 x 8 a 70% de 1RM, ênfase excêntrica de 3 s.","Barra, anilhas",1,18,"Alta",7,"Intermediário",""),
("Afundo com halteres",FI,"N/A","Força Máxima","Força unilateral e estabilidade de quadril","3 x 10 por perna com carga moderada.","Halteres",1,15,"Alta",6,"Intermediário",""),
("Power clean (levantamento olímpico)",FI,"N/A","Força Explosiva/Potência","Taxa de produção de força","5 x 3 a 70–80% de 1RM, foco em velocidade da barra.","Barra, plataforma",1,25,"Muito Alta",8,"Avançado",""),
("Agachamento com salto sob carga (jump squat)",FI,"N/A","Força Explosiva/Potência","Potência específica de salto","5 x 4 com 20–30% de 1RM, pausa completa de 2 min.","Barra leve/halteres",1,18,"Alta",7,"Avançado",""),
("Salto em profundidade (drop jump)",FI,"N/A","Pliometria","Ciclo alongamento-encurtamento rápido","5 x 5 quedas de caixa de 40 cm com salto imediato. Superfície firme e calçado adequado.","Caixa pliométrica",1,15,"Alta",7,"Avançado",""),
("Saltos consecutivos sobre barreiras",FI,"N/A","Pliometria","Rigidez muscular e reatividade","4 x 6 barreiras de 40–50 cm, contato mínimo com o solo.","Barreiras",1,12,"Alta",7,"Intermediário",""),
("Sprints de 10 e 20 m com mudança de direção",FI,"N/A","Velocidade","Aceleração e frenagem","6 sprints de 10 m e 4 de 20 m com 90 s de pausa.","Cones, cronômetro",1,15,"Muito Alta",7,"Intermediário",""),
("Circuito de agilidade T-Test",FI,"N/A","Agilidade/COD","Mudança de direção específica da quadra","4 execuções completas com 2 min de pausa; registrar o melhor tempo.","Cones, cronômetro",1,12,"Alta",6,"Intermediário",""),
("Intervalado de alta intensidade 15/15",FI,"N/A","Resistência Anaeróbia","Tolerância a esforços intermitentes","2 blocos de 8 x 15 s de corrida intensa / 15 s de trote, 3 min entre blocos.","Cronômetro",1,20,"Muito Alta",8,"Intermediário",""),
("Corrida contínua regenerativa",FI,"N/A","Resistência Aeróbia","Recuperação ativa e base aeróbia","20–25 min em intensidade leve (60–65% da FCmáx).","Cronômetro",1,25,"Baixa",3,"Iniciante",""),
("Prancha e anti-rotação (Pallof press)",FI,"N/A","Core/Estabilidade","Estabilidade de tronco para transferência de força","3 x 40 s de prancha + 3 x 12 Pallof por lado.","Elástico, colchonete",1,12,"Moderada",4,"Iniciante",""),
("Rotadores do ombro com elástico",PV,"N/A","Core/Estabilidade","Prevenção de lesão no ombro do atacante","3 x 15 rotação externa e interna, carga leve, ritmo controlado.","Elástico",1,10,"Baixa",3,"Iniciante",""),
("Nórdico de isquiotibiais (excêntrico)",PV,"N/A","Força Máxima","Prevenção de lesão de isquiotibiais","3 x 6 repetições excêntricas lentas, 2x/semana.","Colchonete, parceiro",2,10,"Alta",6,"Intermediário",""),
("Excêntrico de tornozelo e panturrilha",PV,"N/A","Força Máxima","Prevenção de entorse e tendinopatia de aquileu","3 x 12 elevações com descida lenta de 3 s no step.","Step",1,8,"Moderada",4,"Iniciante",""),
("Mobilidade de tornozelo e quadril",PV,"N/A","Mobilidade/Flexibilidade","Amplitude para agachamento e aterrissagem","6 exercícios x 8 repetições, antes da parte principal.","Colchonete, elástico",1,10,"Muito Baixa",2,"Iniciante",""),
("Alongamento e liberação miofascial",RC,"N/A","Mobilidade/Flexibilidade","Recuperação pós-sessão","15 min de rolo e alongamentos estáticos de 30 s por grupamento.","Rolo, colchonete",1,15,"Muito Baixa",2,"Iniciante",""),
("Reação visual com estímulo do técnico",CG,"Defesa","Coordenação","Tempo de reação e leitura antecipatória","Técnico sinaliza a direção no último instante; atleta desloca e defende. 5 x 30 s.","Bolas, sinalizadores",1,10,"Alta",6,"Intermediário",""),
]
for i, e in enumerate(EXERCICIOS):
    r = EX_F + i
    for j, v in enumerate(e):
        wsE.cell(r, 2 + j, v)

# ============================================================================
# 5) MACROCICLO
# ============================================================================
wsM = wb.create_sheet("Macrociclo")
banner(wsM, "MACROCICLO — PLANEJAMENTO DA TEMPORADA",
       "Defina o período total e divida-o em mesociclos. A Data de Início (C11) é a referência de TODA a numeração de semanas da planilha.", 15, NAVY)
larguras(wsM, {"A":6,"B":26,"C":18,"D":14,"E":14,"F":9,"G":10,"H":11,"I":12,"J":24,"K":24,"L":24,"M":14,"N":24,"O":13})

secao(wsM, 4, "IDENTIFICAÇÃO DO MACROCICLO", 15, 2)
IDENT = [("Temporada / Época","2026"),("Equipe / Clube","Equipe Exemplo de Voleibol"),("Categoria","Adulto Feminino"),
         ("Técnico Responsável","(preencher)"),("Preparador Físico","(preencher)"),
         ("Data de Início", date(2026,1,5)),("Data de Término", date(2026,11,29)),("Duração (semanas)", None),
         ("Objetivo Principal","Chegar ao 1º turno do estadual com alto nível técnico-tático e baixo índice de lesões"),
         ("Competição-Alvo","Campeonato Estadual Adulto — fase final em outubro/2026")]
for i, (lab, val) in enumerate(IDENT):
    r = 6 + i
    rotulo(wsM, r, 2, lab)
    if lab == "Duração (semanas)":
        c = calc(wsM, r, 3, '=IFERROR(ROUNDUP(($C$12-$C$11+1)/7,0),"")', '0" semanas"')
        wsM.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
    else:
        c = entrada(wsM, r, 3, val, NF_DATE if isinstance(val, date) else None, largura_merge=4)
    wsM.row_dimensions[r].height = 18
wsM["C11"].comment = Comment("Data de Início do Macrociclo.\nÉ a âncora do cálculo 'Semana nº' usado nas abas "
                             "Prescrição, Carga (PSE), Wellness, Presença, Atleta e Painel.\nUse preferencialmente uma "
                             "segunda-feira.", "Planilha")

secao(wsM, 17, "DIVISÃO EM MESOCICLOS", 15, 1)
MAC_H = ["Nº","Mesociclo","Período","Data Início","Data Fim","Semanas","Semana Inicial","Volume (%)","Intensidade (%)",
         "Ênfase Física","Ênfase Técnica","Ênfase Tática","Carga-Alvo Semanal (UA)","Competições no Período","Situação"]
cab_tabela(wsM, 18, MAC_H)
MESO_F, MESO_L = 19, 30
for r in range(MESO_F, MESO_L + 1):
    wsM.cell(r, 1, '=IF($B{0}="","",ROW()-{1})'.format(r, MESO_F - 1))
    wsM.cell(r, 6, '=IF(OR($D{0}="",$E{0}=""),0,ROUNDUP(($E{0}-$D{0}+1)/7,0))'.format(r))
    wsM.cell(r, 7, '=IF(OR($D{0}="",$C$11=""),0,INT(($D{0}-$C$11)/7)+1)'.format(r))
corpo_tabela(wsM, MESO_F, MESO_L, 1, 15)
for r in range(MESO_F, MESO_L + 1):
    for c in (2, 3, 4, 5, 8, 9, 10, 11, 12, 13, 14, 15):
        wsM.cell(r, c).fill = PatternFill("solid", fgColor=GOLD_L)
        wsM.cell(r, c).font = Font(name=F, size=9, color="0000FF")
    for c in (1, 6, 7):
        wsM.cell(r, c).font = Font(name=F, size=9, bold=True, color=NAVY2)
        wsM.cell(r, c).fill = PatternFill("solid", fgColor=LIGHT)
    for c in (2, 10, 11, 12, 14):
        wsM.cell(r, c).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    wsM.cell(r, 4).number_format = NF_DATE
    wsM.cell(r, 5).number_format = NF_DATE
    wsM.cell(r, 6).number_format = '0;;""'
    wsM.cell(r, 7).number_format = '0;;""'
    wsM.cell(r, 8).number_format = NF_PCT
    wsM.cell(r, 9).number_format = NF_PCT
    wsM.cell(r, 13).number_format = NF_UA
tot = MESO_L + 1
wsM.cell(tot, 2, "TOTAL / MÉDIA").font = Font(name=F, size=9, bold=True, color=WHITE)
for c in range(1, 16):
    wsM.cell(tot, c).fill = PatternFill("solid", fgColor=NAVY2); wsM.cell(tot, c).border = BORDER
    wsM.cell(tot, c).font = Font(name=F, size=9, bold=True, color=WHITE)
    wsM.cell(tot, c).alignment = Alignment(horizontal="center", vertical="center")
wsM.cell(tot, 2).alignment = Alignment(horizontal="left", vertical="center", indent=1)
wsM.cell(tot, 6, "=SUM(F{}:F{})".format(MESO_F, MESO_L)).number_format = "0"
wsM.cell(tot, 8, '=IFERROR(AVERAGE(H{}:H{}),"")'.format(MESO_F, MESO_L)).number_format = NF_PCT
wsM.cell(tot, 9, '=IFERROR(AVERAGE(I{}:I{}),"")'.format(MESO_F, MESO_L)).number_format = NF_PCT
wsM.cell(tot, 13, '=IFERROR(SUMPRODUCT($F{0}:$F{1},$M{0}:$M{1}),"")'.format(MESO_F, MESO_L)).number_format = NF_UA
nota(wsM, tot + 1, 2, "Confira: a soma de semanas dos mesociclos deve bater com a Duração (semanas) em C13. "
     "A Carga-Alvo total é a soma de (semanas × carga-alvo semanal) de cada mesociclo.", 15)
dv(wsM, L("Período do Macro"), "C{}:C{}".format(MESO_F, MESO_L))
dv(wsM, L("Situação"),         "O{}:O{}".format(MESO_F, MESO_L))
wsM.freeze_panes = "C19"

MESOS = [
 ("Meso 1 — Base Geral","Preparatório Geral",date(2026,1,5),date(2026,2,15),0.90,0.55,
  "Força máxima, base aeróbia e core","Fundamentos individuais e repetição técnica","Reconhecimento do sistema 5x1",2600,"—","Concluído"),
 ("Meso 2 — Base Específica","Preparatório Específico",date(2026,2,16),date(2026,3,29),0.85,0.70,
  "Força-potência e pliometria introdutória","Técnica sob deslocamento e sob fadiga","Complexos K1 e K2",2900,"Amistosos preparatórios","Concluído"),
 ("Meso 3 — Pré-Competitivo","Pré-Competitivo",date(2026,3,30),date(2026,5,10),0.70,0.85,
  "Potência, pliometria e velocidade","Técnica em situação de jogo","Sistemas de defesa e coberturas",2700,"Copa regional","Concluído"),
 ("Meso 4 — Competitivo I","Competitivo I",date(2026,5,11),date(2026,7,19),0.55,0.90,
  "Manutenção de força e potência","Correção pontual e saque sob pressão","Ajustes de scout adversário",2200,"Estadual — 1º turno","Em execução"),
 ("Meso 5 — Competitivo II","Competitivo II",date(2026,7,20),date(2026,10,11),0.50,0.95,
  "Manutenção e prevenção de lesões","Refinamento técnico individual","Preparação por adversário",2000,"Estadual — 2º turno e playoffs","Planejado"),
 ("Meso 6 — Transição","Transição",date(2026,10,12),date(2026,11,29),0.35,0.40,
  "Recuperação ativa e trabalho compensatório","Atividades lúdicas e técnica livre","Avaliação da temporada",900,"—","Planejado"),
]
for i, m in enumerate(MESOS):
    r = MESO_F + i
    nome, per, di, df, vol, inten, ef, et, eta, carga, comp, sit = m
    wsM.cell(r, 2, nome); wsM.cell(r, 3, per); wsM.cell(r, 4, di); wsM.cell(r, 5, df)
    wsM.cell(r, 8, vol); wsM.cell(r, 9, inten); wsM.cell(r, 10, ef); wsM.cell(r, 11, et)
    wsM.cell(r, 12, eta); wsM.cell(r, 13, carga); wsM.cell(r, 14, comp); wsM.cell(r, 15, sit)
    wsM.cell(r, 4).number_format = NF_DATE; wsM.cell(r, 5).number_format = NF_DATE
    wsM.cell(r, 8).number_format = NF_PCT;  wsM.cell(r, 9).number_format = NF_PCT
    wsM.cell(r, 13).number_format = NF_UA

ch = LineChart(); ch.title = "Dinâmica da Carga: Volume × Intensidade por Mesociclo"
ch.style = 12; ch.height = 8.5; ch.width = 22; ch.y_axis.title = "% relativo"
data = Reference(wsM, min_col=8, max_col=9, min_row=18, max_row=MESO_L)
cats = Reference(wsM, min_col=2, min_row=MESO_F, max_row=MESO_L)
ch.add_data(data, titles_from_data=True); ch.set_categories(cats)
ch.series[0].graphicalProperties.line.width = 28000
ch.series[1].graphicalProperties.line.width = 28000
wsM.add_chart(ch, "B34")

# ============================================================================
# 6) MESOCICLO
# ============================================================================
wsS = wb.create_sheet("Mesociclo")
banner(wsS, "MESOCICLO — DETALHAMENTO EM MICROCICLOS",
       "Escolha o mesociclo no menu suspenso: os dados do planejamento são puxados do Macrociclo automaticamente.", 13, NAVY2)
larguras(wsS, {"A":6,"B":24,"C":22,"D":14,"E":14,"F":40,"G":11,"H":13,"I":13,"J":14,"K":14,"L":12,"M":30})

secao(wsS, 4, "SELEÇÃO E RESUMO DO MESOCICLO", 13, 2)
rotulo(wsS, 6, 2, "Mesociclo selecionado")
sel = entrada(wsS, 6, 3, MESOS[4][0], largura_merge=2)
sel.font = Font(name=F, size=11, bold=True, color="0000FF")
dv(wsS, MESO_REF, "C6")

def idx_macro(col_letter):
    return '=IFERROR(INDEX(Macrociclo!${0}${1}:${0}${2},MATCH($C$6,Macrociclo!$B${1}:$B${2},0)),"")'.format(col_letter, MESO_F, MESO_L)

RESUMO = [("Período",("C",None)),("Data de Início",("D",NF_DATE)),("Data de Término",("E",NF_DATE)),
          ("Duração (semanas)",("F","0")),("Semana Inicial no Macro",("G","0")),("Volume Planejado",("H",NF_PCT)),
          ("Intensidade Planejada",("I",NF_PCT)),("Carga-Alvo Semanal (UA)",("M",NF_UA))]
for i, (lab, (cl, nf)) in enumerate(RESUMO):
    r = 8 + i
    rotulo(wsS, r, 2, lab)
    c = calc(wsS, r, 3, idx_macro(cl), nf)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    wsS.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
for i, (lab, cl) in enumerate([("Ênfase Física","J"),("Ênfase Técnica","K"),("Ênfase Tática","L"),("Competições","N")]):
    r = 8 + i
    rotulo(wsS, r, 6, lab)
    c = calc(wsS, r, 7, idx_macro(cl))
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    wsS.merge_cells(start_row=r, start_column=7, end_row=r, end_column=13)

secao(wsS, 17, "MICROCICLOS DO MESOCICLO — PREVISTO × REALIZADO", 13, 1)
MIC_H = ["Nº","Microciclo","Tipo de Microciclo","Data Início","Data Fim","Objetivo do Microciclo","Semana nº",
         "Sessões Previstas","Volume Previsto (min)","PSE Média Prevista","Carga Prevista (UA)","% Cumprido","Observações"]
cab_tabela(wsS, 18, MIC_H)
MIC_F, MIC_L = 19, 30
for r in range(MIC_F, MIC_L + 1):
    wsS.cell(r, 1, '=IF($B{0}="","",ROW()-{1})'.format(r, MIC_F - 1))
    wsS.cell(r, 7, '=IF(OR($D{0}="",Macrociclo!$C$11=""),0,INT(($D{0}-Macrociclo!$C$11)/7)+1)'.format(r))
    wsS.cell(r, 11, '=IF(OR($I{0}="",$J{0}=""),0,$I{0}*$J{0})'.format(r))
    wsS.cell(r, 12, '=IF($K{0}=0,0,SUMIFS(\'Carga (PSE)\'!$G${1}:$G${2},\'Carga (PSE)\'!$B${1}:$B${2},$G{0})'
                    '/MAX(1,COUNTIF(Cadastro!$T${3}:$T${4},"Ativo"))/$K{0})'.format(r, CAR_F, CAR_L, CAD_F, CAD_L))
corpo_tabela(wsS, MIC_F, MIC_L, 1, 13)
for r in range(MIC_F, MIC_L + 1):
    for c in (2, 3, 4, 5, 6, 8, 9, 10, 13):
        wsS.cell(r, c).fill = PatternFill("solid", fgColor=GOLD_L)
        wsS.cell(r, c).font = Font(name=F, size=9, color="0000FF")
    for c in (1, 7, 11, 12):
        wsS.cell(r, c).font = Font(name=F, size=9, bold=True, color=NAVY2)
        wsS.cell(r, c).fill = PatternFill("solid", fgColor=LIGHT)
    for c in (2, 6, 13):
        wsS.cell(r, c).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    wsS.cell(r, 4).number_format = NF_DATE; wsS.cell(r, 5).number_format = NF_DATE
    wsS.cell(r, 7).number_format = '0;;""'
    wsS.cell(r, 11).number_format = NF_UA;  wsS.cell(r, 12).number_format = NF_PCT
dv(wsS, L("Tipo de Microciclo"), "C{}:C{}".format(MIC_F, MIC_L))
wsS.conditional_formatting.add("L{}:L{}".format(MIC_F, MIC_L),
    CellIsRule(operator="between", formula=["0.9","1.1"], fill=PatternFill("solid", fgColor=GREEN), font=Font(name=F, size=9, bold=True, color=GREEN_T)))
wsS.conditional_formatting.add("L{}:L{}".format(MIC_F, MIC_L),
    CellIsRule(operator="greaterThan", formula=["1.1"], fill=PatternFill("solid", fgColor=YELL), font=Font(name=F, size=9, bold=True, color=YELL_T)))
wsS.conditional_formatting.add("L{}:L{}".format(MIC_F, MIC_L),
    CellIsRule(operator="between", formula=["0.0001","0.9"], fill=PatternFill("solid", fgColor=RED), font=Font(name=F, size=9, bold=True, color=RED_T)))
nota(wsS, MIC_L + 2, 2, "% Cumprido = carga média por atleta ativo realizada na semana (aba Carga (PSE)) ÷ carga prevista do microciclo. "
     "Verde = 90–110% do previsto; amarelo = acima; vermelho = abaixo.", 13)
wsS.freeze_panes = "C19"

MICROS = [("Micro 5.1","Ordinário",date(2026,8,10),date(2026,8,16),"Retomada do volume após a rodada; ênfase em side-out",7,530,6),
          ("Micro 5.2","Choque",date(2026,8,17),date(2026,8,23),"Pico de carga física da fase, mantendo o volume técnico",7,530,7),
          ("Micro 5.3","Ordinário",date(2026,8,24),date(2026,8,30),"Consolidação técnico-tática do sistema defensivo",7,530,6),
          ("Micro 5.4","Recuperativo",date(2026,8,31),date(2026,9,6),"Redução de volume; recuperação e preparação do jogo",5,380,5),
          ("Micro 5.5","Pré-Competitivo",date(2026,9,7),date(2026,9,13),"Polimento técnico e preparação por adversário",5,360,5)]
for i, m in enumerate(MICROS):
    r = MIC_F + i
    nome, tipo, di, df, obj, ses, vol, pse = m
    wsS.cell(r, 2, nome); wsS.cell(r, 3, tipo)
    if di: wsS.cell(r, 4, di); wsS.cell(r, 4).number_format = NF_DATE
    if df: wsS.cell(r, 5, df); wsS.cell(r, 5).number_format = NF_DATE
    wsS.cell(r, 6, obj); wsS.cell(r, 8, ses); wsS.cell(r, 9, vol); wsS.cell(r, 10, pse)

# ============================================================================
# 7) MICROCICLO
# ============================================================================
wsW = wb.create_sheet("Microciclo")
banner(wsW, "MICROCICLO — PLANEJAMENTO DA SEMANA",
       "Duas linhas por dia (duas sessões possíveis). As datas e os dias da semana são gerados a partir da Data de Início.", 12, BLUE3)
larguras(wsW, {"A":16,"B":12,"C":10,"D":24,"E":18,"F":11,"G":10,"H":13,"I":26,"J":26,"K":26,"L":28})

secao(wsW, 4, "IDENTIFICAÇÃO DO MICROCICLO", 12, 1)
MIC_ID = [("Microciclo","Micro 5.3"),("Semana nº no Macrociclo",34),("Data de Início (2ª feira)",date(2026,8,24)),
          ("Tipo de Microciclo","Ordinário"),("Objetivo da Semana","Consolidação técnico-tática do sistema defensivo, mantendo a carga física da fase")]
for i, (lab, val) in enumerate(MIC_ID):
    r = 5 + i
    rotulo(wsW, r, 1, lab)
    entrada(wsW, r, 2, val, NF_DATE if isinstance(val, date) else None, largura_merge=(9 if i == 4 else 2))
dv(wsW, L("Tipo de Microciclo"), "B8")
wsW["B6"].comment = Comment("Deve coincidir com a coluna 'Semana nº' da aba Mesociclo e com a numeração automática "
                            "das abas de registro.", "Planilha")

secao(wsW, 11, "SESSÕES DA SEMANA", 12, 1)
SES_H = ["Dia da Semana","Data","Turno","Tipo de Sessão","Local","Duração (min)","PSE Prevista",
         "Carga Prevista (UA)","Ênfase Técnica","Ênfase Tática","Ênfase Física / Prevenção","Observações"]
cab_tabela(wsW, 12, SES_H)
SF, SL = 13, 26
DIAS = ["Segunda-feira","Terça-feira","Quarta-feira","Quinta-feira","Sexta-feira","Sábado","Domingo"]
for r in range(SF, SL + 1):
    wsW.cell(r, 2, '=IF($B$7="","",$B$7+INT((ROW()-{})/2))'.format(SF))
    wsW.cell(r, 1, '=IF($B{0}="","",CHOOSE(WEEKDAY($B{0},2),{1}))'.format(r, ",".join('"%s"' % d for d in DIAS)))
    wsW.cell(r, 8, '=IF(OR($F{0}="",$G{0}=""),0,$F{0}*$G{0})'.format(r))
corpo_tabela(wsW, SF, SL, 1, 12)
for r in range(SF, SL + 1):
    for c in (3, 4, 5, 6, 7, 9, 10, 11, 12):
        wsW.cell(r, c).fill = PatternFill("solid", fgColor=GOLD_L)
        wsW.cell(r, c).font = Font(name=F, size=9, color="0000FF")
    for c in (1, 2, 8):
        wsW.cell(r, c).font = Font(name=F, size=9, bold=True, color=NAVY2)
        wsW.cell(r, c).fill = PatternFill("solid", fgColor=LIGHT)
    for c in (4, 9, 10, 11, 12):
        wsW.cell(r, c).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    wsW.cell(r, 2).number_format = NF_DATE
    wsW.cell(r, 8).number_format = NF_UA
    if (r - SF) % 2 == 0:
        for c in range(1, 13):
            wsW.cell(r, c).border = Border(left=thin, right=thin, top=Side(style="medium", color=BLUE3), bottom=thin)
dv(wsW, L("Turno"),          "C{}:C{}".format(SF, SL))
dv(wsW, L("Tipo de Sessão"), "D{}:D{}".format(SF, SL))
dv(wsW, L("PSE (0-10)"),     "G{}:G{}".format(SF, SL))
tr = SL + 1
for c in range(1, 13):
    wsW.cell(tr, c).fill = PatternFill("solid", fgColor=NAVY2); wsW.cell(tr, c).border = BORDER
    wsW.cell(tr, c).font = Font(name=F, size=9, bold=True, color=WHITE)
    wsW.cell(tr, c).alignment = Alignment(horizontal="center", vertical="center")
wsW.cell(tr, 1, "TOTAL DA SEMANA").alignment = Alignment(horizontal="left", vertical="center", indent=1)
wsW.cell(tr, 6, "=SUM(F{}:F{})".format(SF, SL)).number_format = NF_UA
wsW.cell(tr, 7, '=IFERROR(AVERAGE(G{}:G{}),"")'.format(SF, SL)).number_format = '0.0'
wsW.cell(tr, 8, "=SUM(H{}:H{})".format(SF, SL)).number_format = NF_UA

secao(wsW, 29, "RESUMO DA SEMANA — PREVISTO × REALIZADO", 12, 1)
kpi(wsW, 30, 1, "SESSÕES PLANEJADAS", "=COUNTA(D{}:D{})".format(SF, SL), "0", NAVY2, 2)
kpi(wsW, 30, 3, "DURAÇÃO TOTAL (min)", "=SUM(F{}:F{})".format(SF, SL), NF_UA, NAVY2, 2)
kpi(wsW, 30, 5, "CARGA PREVISTA (UA)", "=SUM(H{}:H{})".format(SF, SL), NF_UA, GOLD, 2)
kpi(wsW, 30, 7, "CARGA REALIZADA — MÉDIA/ATLETA (UA)",
    "=IFERROR(SUMIFS('Carga (PSE)'!$G${1}:$G${2},'Carga (PSE)'!$B${1}:$B${2},$B$6)"
    "/MAX(1,COUNTIF(Cadastro!$T${3}:$T${4},\"Ativo\")),0)".format(0, CAR_F, CAR_L, CAD_F, CAD_L), NF_UA, GOLD, 3)
kpi(wsW, 30, 10, "% DE CUMPRIMENTO DA CARGA",
    '=IFERROR(G31/H27,0)', NF_PCT, NAVY2, 3)

secao(wsW, 34, "DISTRIBUIÇÃO DE CONTEÚDOS POR DIA (minutos)", 12, 1)
CON_H = ["Dia","Técnico","Tático","Físico","Preventivo / Prevenção","Recuperação","Total (min)","% da Semana"]
cab_tabela(wsW, 35, CON_H)
CF, CL = 36, 42
for i in range(7):
    r = CF + i
    wsW.cell(r, 1, '=IF($A{0}="","",$A{0})'.format(SF + i * 2))
    wsW.cell(r, 7, "=SUM(B{0}:F{0})".format(r))
    wsW.cell(r, 8, '=IFERROR($G{0}/$G${1},0)'.format(r, CL + 1))
corpo_tabela(wsW, CF, CL, 1, 8)
for r in range(CF, CL + 1):
    for c in range(2, 7):
        wsW.cell(r, c).fill = PatternFill("solid", fgColor=GOLD_L)
        wsW.cell(r, c).font = Font(name=F, size=9, color="0000FF")
        wsW.cell(r, c).number_format = NF_UA
    for c in (1, 7, 8):
        wsW.cell(r, c).font = Font(name=F, size=9, bold=True, color=NAVY2)
        wsW.cell(r, c).fill = PatternFill("solid", fgColor=LIGHT)
    wsW.cell(r, 1).alignment = Alignment(horizontal="left", vertical="center", indent=1)
    wsW.cell(r, 7).number_format = NF_UA; wsW.cell(r, 8).number_format = NF_PCT
tr2 = CL + 1
for c in range(1, 9):
    wsW.cell(tr2, c).fill = PatternFill("solid", fgColor=NAVY2); wsW.cell(tr2, c).border = BORDER
    wsW.cell(tr2, c).font = Font(name=F, size=9, bold=True, color=WHITE)
    wsW.cell(tr2, c).alignment = Alignment(horizontal="center", vertical="center")
wsW.cell(tr2, 1, "TOTAL").alignment = Alignment(horizontal="left", vertical="center", indent=1)
for c in range(2, 8):
    cl_ = get_column_letter(c)
    wsW.cell(tr2, c, "=SUM({0}{1}:{0}{2})".format(cl_, CF, CL)).number_format = NF_UA
wsW.cell(tr2, 8, '=IFERROR(SUM(H{}:H{}),0)'.format(CF, CL)).number_format = NF_PCT

chW = BarChart(); chW.type = "col"; chW.grouping = "stacked"; chW.overlap = 100
chW.title = "Distribuição de Conteúdos na Semana (min)"; chW.height = 8; chW.width = 18
d = Reference(wsW, min_col=2, max_col=6, min_row=35, max_row=CL)
c_ = Reference(wsW, min_col=1, min_row=CF, max_row=CL)
chW.add_data(d, titles_from_data=True); chW.set_categories(c_)
wsW.add_chart(chW, "J35")
wsW.freeze_panes = "C13"

CONT_EX = [(60,30,45,15,10),(50,20,70,15,10),(60,40,30,15,10),(45,35,60,15,10),(70,50,20,15,15),(30,20,0,10,20),(0,0,0,0,0)]
for i, vals in enumerate(CONT_EX):
    for j, v in enumerate(vals):
        wsW.cell(CF + i, 2 + j, v)
SESS_EX = [
 (0,"Manhã","Físico (Força)","Sala de musculação",75,7,"—","—","Força máxima: agachamento e terra romeno"),
 (0,"Tarde","Técnico-Tático","Ginásio principal",90,6,"Recepção de saque e side-out","Complexo K1","Ativação e core"),
 (1,"Manhã","Físico (Potência)","Sala de musculação",60,7,"—","—","Power clean e jump squat"),
 (1,"Tarde","Técnico-Tático","Ginásio principal",90,7,"Ataque de ponta e primeiro tempo","Complexo K2","Pliometria de baixo volume"),
 (2,"Tarde","Coletivo/Jogo","Ginásio principal",100,8,"Saque sob pressão","Jogo 6x6 com pontuação diferenciada","Prevenção de ombro"),
 (3,"Manhã","Físico (Condicionamento)","Pista",50,6,"—","—","Intervalado 15/15"),
 (3,"Tarde","Técnico-Tático","Ginásio principal",90,7,"Bloqueio e defesa","Sistemas de cobertura","Mobilidade"),
 (4,"Tarde","Coletivo/Jogo","Ginásio principal",110,8,"Refinamento do side-out","Preparação por adversário","Ativação"),
 (5,"Manhã","Recuperação/Regenerativo","Piscina / sala",50,3,"Técnica livre","Vídeo do adversário","Liberação miofascial"),
 (6,"Manhã","Folga","—",0,0,"—","—","Descanso completo"),
]
for dia, turno, tipo, local, dur, pse, et, eta, ef in SESS_EX:
    base = SF + dia * 2
    r = base if wsW.cell(base, 4).value in (None, "") else base + 1
    wsW.cell(r, 3, turno); wsW.cell(r, 4, tipo); wsW.cell(r, 5, local)
    wsW.cell(r, 6, dur); wsW.cell(r, 7, pse)
    wsW.cell(r, 9, et); wsW.cell(r, 10, eta); wsW.cell(r, 11, ef)

# ============================================================================
# 8) PRESCRIÇÃO DO TREINO
# ============================================================================
wsP = wb.create_sheet("Prescrição")
banner(wsP, "PRESCRIÇÃO DO TREINO", "Uma linha por exercício prescrito. Escolha o exercício no menu (vem da aba Exercícios) e "
       "defina o destinatário: 'Equipe (todos)' ou o nome de um atleta — é assim que o treino chega à Área do Atleta.", 20, GOLD)
larguras(wsP, {"A":12,"B":8,"C":16,"D":9,"E":24,"F":18,"G":38,"H":17,"I":16,"J":20,"K":8,"L":15,"M":18,"N":9,
               "O":10,"P":9,"Q":9,"R":13,"S":26,"T":24,"U":6})
secao(wsP, 3, "RESUMO DA SESSÃO SELECIONADA", 20, 1)
rotulo(wsP, 4, 1, "Data:")
entrada(wsP, 4, 2, date(2026, 8, 31), NF_DATE, largura_merge=2)
FLT = ('$A${0}:$A${1},$B$4'.format(PRE_F, PRE_L))
kpi(wsP, 4, 5, "EXERCÍCIOS PRESCRITOS", '=COUNTIFS({})'.format(FLT), "0", NAVY2, 2)
kpi(wsP, 4, 7, "DURAÇÃO TOTAL (min)", '=SUMIFS($O${0}:$O${1},{2})'.format(PRE_F, PRE_L, FLT), NF_UA, NAVY2, 2)
kpi(wsP, 4, 9, "CARGA PREVISTA (UA)", '=SUMIFS($R${0}:$R${1},{2})'.format(PRE_F, PRE_L, FLT), NF_UA, GOLD, 2)
kpi(wsP, 4, 11, "PSE MÉDIA PREVISTA", '=IFERROR(AVERAGEIFS($P${0}:$P${1},{2}),0)'.format(PRE_F, PRE_L, FLT), '0.0;;""', GOLD, 2)
kpi(wsP, 4, 13, "MIN. TÉCNICO", '=SUMIFS($O${0}:$O${1},{2},$H${0}:$H${1},"Técnico")'.format(PRE_F, PRE_L, FLT), NF_UA, BLUE3, 2)
kpi(wsP, 4, 15, "MIN. TÁTICO", '=SUMIFS($O${0}:$O${1},{2},$H${0}:$H${1},"Tático")'.format(PRE_F, PRE_L, FLT), NF_UA, BLUE3, 2)
kpi(wsP, 4, 17, "MIN. FÍSICO", '=SUMIFS($O${0}:$O${1},{2},$H${0}:$H${1},"Físico")'.format(PRE_F, PRE_L, FLT), NF_UA, BLUE3, 2)
kpi(wsP, 4, 19, "MIN. PREVENTIVO", '=SUMIFS($O${0}:$O${1},{2},$H${0}:$H${1},"Preventivo/Compensatório")'.format(PRE_F, PRE_L, FLT), NF_UA, BLUE3, 2)
nota(wsP, 6, 1, "Digite a data em B4 para ver o resumo das sessões daquele dia. Séries × Repetições geram o Volume; "
     "Duração × PSE Prevista geram a Carga Prevista em UA (unidades arbitrárias).", 20)

PRE_H = ["Data","Semana","Sessão","Turno","Destinatário","Bloco","Exercício","Categoria","Fundamento","Capacidade Física",
         "Séries","Repetições / Tempo","Carga / Intensidade","Pausa (s)","Duração (min)","PSE Prev.","Volume",
         "Carga Prev. (UA)","Progressão / Critério","Observações","Índ."]
cab_tabela(wsP, 7, PRE_H)
for r in range(PRE_F, PRE_L + 1):
    wsP.cell(r, 2,  '=IF($A{0}="",0,INT(($A{0}-Macrociclo!$C$11)/7)+1)'.format(r))
    wsP.cell(r, 8,  '=IFERROR(INDEX(Exercícios!$C${1}:$C${2},MATCH($G{0},Exercícios!$B${1}:$B${2},0)),"")'.format(r, EX_F, EX_L))
    wsP.cell(r, 9,  '=IFERROR(INDEX(Exercícios!$D${1}:$D${2},MATCH($G{0},Exercícios!$B${1}:$B${2},0)),"")'.format(r, EX_F, EX_L))
    wsP.cell(r, 10, '=IFERROR(INDEX(Exercícios!$E${1}:$E${2},MATCH($G{0},Exercícios!$B${1}:$B${2},0)),"")'.format(r, EX_F, EX_L))
    wsP.cell(r, 17, '=IFERROR($K{0}*$L{0},"")'.format(r))
    wsP.cell(r, 18, '=IF(OR($O{0}="",$P{0}=""),0,$O{0}*$P{0})'.format(r))
    wsP.cell(r, 21, '=IF($A{0}="",0,IF(AND(OR($E{0}=Atleta!$C$4,$E{0}="Equipe (todos)"),$A{0}>=Atleta!$C$5),'
                    'MAX($U$7:U{1})+1,0))'.format(r, r - 1))
corpo_tabela(wsP, PRE_F, PRE_L, 1, 21)
for r in range(PRE_F, PRE_L + 1):
    for c in (1, 3, 4, 5, 6, 7, 11, 12, 13, 14, 15, 16, 19, 20):
        wsP.cell(r, c).fill = PatternFill("solid", fgColor=GOLD_L)
        wsP.cell(r, c).font = Font(name=F, size=9, color="0000FF")
    for c in (2, 8, 9, 10, 17, 18, 21):
        wsP.cell(r, c).font = Font(name=F, size=9, bold=True, color=NAVY2)
        wsP.cell(r, c).fill = PatternFill("solid", fgColor=LIGHT)
    for c in (5, 7, 13, 19, 20):
        wsP.cell(r, c).alignment = Alignment(horizontal="left", vertical="center", indent=1)
    wsP.cell(r, 1).number_format = NF_DATE
    wsP.cell(r, 2).number_format = '0;;""'
    wsP.cell(r, 18).number_format = NF_UA
    wsP.cell(r, 21).number_format = '0;;""'
dv(wsP, L("Turno"),          "D{}:D{}".format(PRE_F, PRE_L))
dv(wsP, DEST_REF,            "E{}:E{}".format(PRE_F, PRE_L))
dv(wsP, L("Bloco da Sessão"),"F{}:F{}".format(PRE_F, PRE_L))
dv(wsP, EXERC_REF,           "G{}:G{}".format(PRE_F, PRE_L))
dv(wsP, L("PSE (0-10)"),     "P{}:P{}".format(PRE_F, PRE_L))
wsP.column_dimensions["U"].hidden = True
wsP.freeze_panes = "C8"
wsP.auto_filter.ref = "A7:T{}".format(PRE_L)

# ============================================================================
# 9) CONTROLE DE CARGA  (PSE da sessão / ACWR)
# ============================================================================
wsG = wb.create_sheet("Carga (PSE)")
banner(wsG, "CONTROLE DE CARGA INTERNA — MÉTODO DA PSE DA SESSÃO",
       "Registre duração e PSE (0–10, coletada ~30 min após o treino). Carga, aguda, crônica, ACWR e zona de risco são automáticos.", 13, RED_T)
larguras(wsG, {"A":12,"B":8,"C":26,"D":24,"E":12,"F":9,"G":12,"H":15,"I":17,"J":10,"K":12,"L":20,"M":26})
CAR_H = ["Data","Semana","Atleta","Tipo de Sessão","Duração (min)","PSE (0-10)","Carga (UA)",
         "Carga Aguda 7 d","Carga Crônica 28 d","ACWR","Dias de Histórico","Zona / Alerta","Observações"]
cab_tabela(wsG, 7, CAR_H)
nota(wsG, 4, 1, "Carga (UA) = Duração × PSE  •  ACWR = Carga Aguda (7 d) ÷ Carga Crônica (média semanal dos 28 d).  "
     "Faixas: < 0,80 subcarga | 0,80–1,30 zona ideal | 1,31–1,50 atenção | > 1,50 risco elevado.  "
     "A Zona só é classificada após 21 dias de registros do atleta — antes disso aparece 'Histórico insuficiente'.", 13)
wsG.row_dimensions[4].height = 26
for r in range(CAR_F, CAR_L + 1):
    wsG.cell(r, 2, '=IF($A{0}="",0,INT(($A{0}-Macrociclo!$C$11)/7)+1)'.format(r))
    wsG.cell(r, 7, '=IF(OR($E{0}="",$F{0}=""),0,$E{0}*$F{0})'.format(r))
    wsG.cell(r, 8, '=IF($A{0}="",0,SUMIFS($G${1}:$G${2},$C${1}:$C${2},$C{0},$A${1}:$A${2},">="&$A{0}-6,'
                   '$A${1}:$A${2},"<="&$A{0}))'.format(r, CAR_F, CAR_L))
    wsG.cell(r, 9, '=IF($A{0}="",0,SUMIFS($G${1}:$G${2},$C${1}:$C${2},$C{0},$A${1}:$A${2},">="&$A{0}-27,'
                   '$A${1}:$A${2},"<="&$A{0})/4)'.format(r, CAR_F, CAR_L))
    wsG.cell(r, 10, '=IF(OR($A{0}="",$I{0}=0),0,$H{0}/$I{0})'.format(r))
    wsG.cell(r, 11, '=IF($A{0}="",0,$A{0}-SUMPRODUCT(MIN(($C${1}:$C${2}=$C{0})*$A${1}:$A${2}'
                    '+($C${1}:$C${2}<>$C{0})*100000))+1)'.format(r, CAR_F, CAR_L))
    wsG.cell(r, 12, '=IF($A{0}="","",IF($J{0}=0,"Sem registro anterior",IF($K{0}<21,"Histórico insuficiente",'
                    'IF($J{0}<0.8,"Subcarga",IF($J{0}<=1.3,"Zona ideal",IF($J{0}<=1.5,"Atenção","Risco elevado"))))))'.format(r))
corpo_tabela(wsG, CAR_F, CAR_L, 1, 13)
for r in range(CAR_F, CAR_L + 1):
    for c in (1, 3, 4, 5, 6, 13):
        wsG.cell(r, c).fill = PatternFill("solid", fgColor=GOLD_L)
        wsG.cell(r, c).font = Font(name=F, size=9, color="0000FF")
    for c in (2, 7, 8, 9, 10, 11, 12):
        wsG.cell(r, c).font = Font(name=F, size=9, bold=True, color=NAVY2)
        wsG.cell(r, c).fill = PatternFill("solid", fgColor=LIGHT)
    wsG.cell(r, 11).number_format = '0;;""'
    for c in (3, 4, 13):
        wsG.cell(r, c).alignment = Alignment(horizontal="left", vertical="center", indent=1)
    wsG.cell(r, 1).number_format = NF_DATE
    wsG.cell(r, 2).number_format = '0;;""'
    for c in (7, 8, 9):
        wsG.cell(r, c).number_format = NF_UA
    wsG.cell(r, 10).number_format = NF_DEC
dv(wsG, ATLETAS_REF,          "C{}:C{}".format(CAR_F, CAR_L))
dv(wsG, L("Tipo de Sessão"),  "D{}:D{}".format(CAR_F, CAR_L))
dv(wsG, L("PSE (0-10)"),      "F{}:F{}".format(CAR_F, CAR_L))
zref = "L{}:L{}".format(CAR_F, CAR_L)
for txt, fill, ft in [('"Zona ideal"', GREEN, GREEN_T), ('"Atenção"', YELL, YELL_T),
                      ('"Risco elevado"', RED, RED_T), ('"Subcarga"', "DDEBF7", NAVY2),
                      ('"Histórico insuficiente"', "EDEDED", GREY_T)]:
    wsG.conditional_formatting.add(zref, CellIsRule(operator="equal", formula=[txt],
        fill=PatternFill("solid", fgColor=fill), font=Font(name=F, size=9, bold=True, color=ft)))
wsG.conditional_formatting.add("J{}:J{}".format(CAR_F, CAR_L),
    CellIsRule(operator="greaterThan", formula=["1.5"], font=Font(name=F, size=9, bold=True, color=RED_T)))
wsG.freeze_panes = "C8"
wsG.auto_filter.ref = "A7:M{}".format(CAR_L)

# ============================================================================
# 10) WELLNESS  (Índice de Hooper)
# ============================================================================
wsWe = wb.create_sheet("Wellness")
banner(wsWe, "WELLNESS DIÁRIO — ÍNDICE DE HOOPER",
       "Cada item de 1 (muito bom / ausente) a 7 (muito ruim / máxima). Soma de 4 a 28 — QUANTO MAIOR, PIOR. Aplicar antes do treino.", 11, "6B3FA0")
larguras(wsWe, {"A":12,"B":8,"C":26,"D":16,"E":13,"F":13,"G":16,"H":14,"I":18,"J":20,"K":28})
WEL_H = ["Data","Semana","Atleta","Qualidade do Sono (1-7)","Estresse (1-7)","Fadiga (1-7)","Dor Muscular (1-7)",
         "Índice de Hooper","Classificação","Local da Dor / Queixa","Observações"]
cab_tabela(wsWe, 7, WEL_H)
nota(wsWe, 4, 1, "Classificação: 4–8 Ótimo | 9–12 Bom | 13–16 Regular | 17–20 Alerta | 21–28 Crítico. "
     "Referência: Hooper & Mackinnon (1995). Pontuações em Alerta/Crítico por 2 dias seguidos pedem conversa com o atleta e "
     "eventual ajuste da carga do dia.", 11)
wsWe.row_dimensions[4].height = 26
for r in range(WEL_F, WEL_L + 1):
    wsWe.cell(r, 2, '=IF($A{0}="",0,INT(($A{0}-Macrociclo!$C$11)/7)+1)'.format(r))
    wsWe.cell(r, 8, '=IF(COUNT($D{0}:$G{0})<4,0,SUM($D{0}:$G{0}))'.format(r))
    wsWe.cell(r, 9, '=IF($H{0}=0,"",IF($H{0}<=8,"Ótimo",IF($H{0}<=12,"Bom",IF($H{0}<=16,"Regular",'
                    'IF($H{0}<=20,"Alerta","Crítico")))))'.format(r))
corpo_tabela(wsWe, WEL_F, WEL_L, 1, 11)
for r in range(WEL_F, WEL_L + 1):
    for c in (1, 3, 4, 5, 6, 7, 10, 11):
        wsWe.cell(r, c).fill = PatternFill("solid", fgColor=GOLD_L)
        wsWe.cell(r, c).font = Font(name=F, size=9, color="0000FF")
    for c in (2, 8, 9):
        wsWe.cell(r, c).font = Font(name=F, size=9, bold=True, color=NAVY2)
        wsWe.cell(r, c).fill = PatternFill("solid", fgColor=LIGHT)
    for c in (3, 10, 11):
        wsWe.cell(r, c).alignment = Alignment(horizontal="left", vertical="center", indent=1)
    wsWe.cell(r, 1).number_format = NF_DATE
    wsWe.cell(r, 2).number_format = '0;;""'
    wsWe.cell(r, 8).number_format = '0;;""'
dv(wsWe, ATLETAS_REF,   "C{}:C{}".format(WEL_F, WEL_L))
for cl in "DEFG":
    dv(wsWe, L("Escala 1-7"), "{0}{1}:{0}{2}".format(cl, WEL_F, WEL_L))
for txt, fill, ft in [('"Ótimo"', GREEN, GREEN_T), ('"Bom"', "E2EFDA", GREEN_T),
                      ('"Regular"', YELL, YELL_T), ('"Alerta"', "FCE4D6", "C55A11"), ('"Crítico"', RED, RED_T)]:
    wsWe.conditional_formatting.add("I{}:I{}".format(WEL_F, WEL_L), CellIsRule(operator="equal", formula=[txt],
        fill=PatternFill("solid", fgColor=fill), font=Font(name=F, size=9, bold=True, color=ft)))
wsWe.freeze_panes = "C8"
wsWe.auto_filter.ref = "A7:K{}".format(WEL_L)

# ============================================================================
# 11) PRESENÇA
# ============================================================================
wsPr = wb.create_sheet("Presença")
banner(wsPr, "CONTROLE DE PRESENÇA", "Chamada por sessão à esquerda; percentual de presença por atleta calculado automaticamente à direita.", 16, "1F6F4A")
larguras(wsPr, {"A":12,"B":8,"C":24,"D":26,"E":22,"F":13,"G":30,"H":3,
                "I":26,"J":11,"K":11,"L":12,"M":10,"N":13,"O":13,"P":16})
PRS_H = ["Data","Semana","Sessão / Tipo","Atleta","Presença","Min. Participados","Justificativa"]
cab_tabela(wsPr, 7, PRS_H)
for r in range(PRS_F, PRS_L + 1):
    wsPr.cell(r, 2, '=IF($A{0}="",0,INT(($A{0}-Macrociclo!$C$11)/7)+1)'.format(r))
corpo_tabela(wsPr, PRS_F, PRS_L, 1, 7)
for r in range(PRS_F, PRS_L + 1):
    for c in (1, 3, 4, 5, 6, 7):
        wsPr.cell(r, c).fill = PatternFill("solid", fgColor=GOLD_L)
        wsPr.cell(r, c).font = Font(name=F, size=9, color="0000FF")
    wsPr.cell(r, 2).font = Font(name=F, size=9, bold=True, color=NAVY2)
    wsPr.cell(r, 2).fill = PatternFill("solid", fgColor=LIGHT)
    for c in (3, 4, 7):
        wsPr.cell(r, c).alignment = Alignment(horizontal="left", vertical="center", indent=1)
    wsPr.cell(r, 1).number_format = NF_DATE
    wsPr.cell(r, 2).number_format = '0;;""'
dv(wsPr, ATLETAS_REF,         "D{}:D{}".format(PRS_F, PRS_L))
dv(wsPr, L("Presença"),       "E{}:E{}".format(PRS_F, PRS_L))
dv(wsPr, L("Tipo de Sessão"), "C{}:C{}".format(PRS_F, PRS_L))
wsPr.conditional_formatting.add("E{}:E{}".format(PRS_F, PRS_L),
    CellIsRule(operator="equal", formula=['"Presente"'], fill=PatternFill("solid", fgColor=GREEN), font=Font(name=F, size=9, bold=True, color=GREEN_T)))
for txt in ('"Falta Não Justificada"', '"Lesionado"', '"Departamento Médico"'):
    wsPr.conditional_formatting.add("E{}:E{}".format(PRS_F, PRS_L),
        CellIsRule(operator="equal", formula=[txt], fill=PatternFill("solid", fgColor=RED), font=Font(name=F, size=9, bold=True, color=RED_T)))
wsPr.conditional_formatting.add("E{}:E{}".format(PRS_F, PRS_L),
    CellIsRule(operator="equal", formula=['"Falta Justificada"'], fill=PatternFill("solid", fgColor=YELL), font=Font(name=F, size=9, bold=True, color=YELL_T)))

secao(wsPr, 5, "RESUMO POR ATLETA", 16, 9)
cab_tabela(wsPr, 7, ["Atleta","Registros","Presenças","% Presença","Faltas","Lesão / DM","Min. Totais","Situação"], col0=9)
for i in range(CAD_L - CAD_F + 1):
    r = PRS_F + i
    a = CAD_F + i
    wsPr.cell(r, 9,  '=IF(Cadastro!$B{0}="","",Cadastro!$B{0})'.format(a))
    wsPr.cell(r, 10, '=IF($I{0}="",0,COUNTIFS($D${1}:$D${2},$I{0}))'.format(r, PRS_F, PRS_L))
    wsPr.cell(r, 11, '=IF($I{0}="",0,COUNTIFS($D${1}:$D${2},$I{0},$E${1}:$E${2},"Presente"))'.format(r, PRS_F, PRS_L))
    wsPr.cell(r, 12, '=IF($J{0}=0,0,$K{0}/$J{0})'.format(r))
    wsPr.cell(r, 13, '=IF($I{0}="",0,COUNTIFS($D${1}:$D${2},$I{0},$E${1}:$E${2},"Falta Justificada")'
                     '+COUNTIFS($D${1}:$D${2},$I{0},$E${1}:$E${2},"Falta Não Justificada"))'.format(r, PRS_F, PRS_L))
    wsPr.cell(r, 14, '=IF($I{0}="",0,COUNTIFS($D${1}:$D${2},$I{0},$E${1}:$E${2},"Lesionado")'
                     '+COUNTIFS($D${1}:$D${2},$I{0},$E${1}:$E${2},"Departamento Médico"))'.format(r, PRS_F, PRS_L))
    wsPr.cell(r, 15, '=IF($I{0}="",0,SUMIFS($F${1}:$F${2},$D${1}:$D${2},$I{0}))'.format(r, PRS_F, PRS_L))
    wsPr.cell(r, 16, '=IF($I{0}="","",IF($J{0}=0,"Sem registro",IF($L{0}>=0.9,"Excelente",'
                     'IF($L{0}>=0.75,"Regular","Baixa assiduidade"))))'.format(r))
PRS_RES_L = PRS_F + (CAD_L - CAD_F)
corpo_tabela(wsPr, PRS_F, PRS_RES_L, 9, 16)
for r in range(PRS_F, PRS_RES_L + 1):
    for c in range(9, 17):
        wsPr.cell(r, c).font = Font(name=F, size=9, bold=(c == 12), color=NAVY2)
        wsPr.cell(r, c).fill = PatternFill("solid", fgColor=LIGHT if (r - PRS_F) % 2 == 0 else LIGHT2)
    wsPr.cell(r, 9).alignment = Alignment(horizontal="left", vertical="center", indent=1)
    wsPr.cell(r, 12).number_format = NF_PCT
    for c in (10, 11, 13, 14, 15):
        wsPr.cell(r, c).number_format = '0;;""'
wsPr.conditional_formatting.add("L{}:L{}".format(PRS_F, PRS_RES_L),
    ColorScaleRule(start_type="num", start_value=0.5, start_color=RED,
                   mid_type="num", mid_value=0.8, mid_color=YELL,
                   end_type="num", end_value=1, end_color=GREEN))
wsPr.freeze_panes = "C8"

# ============================================================================
# 12) TESTES FÍSICOS
# ============================================================================
wsT = wb.create_sheet("Testes")
banner(wsT, "AVALIAÇÕES E TESTES FÍSICOS", "Aplique a mesma bateria em momentos-chave da temporada para acompanhar a evolução de cada atleta.", 18, "8A5A00")
TST_H = ["Data","Atleta","Momento","Massa (kg)","% Gordura","Squat Jump (cm)","CMJ (cm)","Salto c/ Aproximação (cm)",
         "Alcance de Ataque (cm)","Alcance de Bloqueio (cm)","Impulsão de Ataque (cm)","T-Test (s)","Sprint 10 m (s)",
         "Sentar-e-Alcançar (cm)","Yo-Yo IR1 (m)","Preensão Manual (kg)","Medicine Ball (m)","Observações"]
cab_tabela(wsT, 7, TST_H)
larguras(wsT, {"A":12,"B":26,"C":18,"D":11,"E":11,"F":13,"G":11,"H":16,"I":15,"J":15,"K":16,"L":11,"M":12,"N":15,"O":13,"P":14,"Q":13,"R":28})
nota(wsT, 4, 1, "Impulsão de Ataque é calculada automaticamente: Alcance de Ataque medido no teste − Alcance em Pé do Cadastro. "
     "Registre sempre a melhor de 3 tentativas nos testes de salto e velocidade.", 18)
for r in range(TST_F, TST_L + 1):
    wsT.cell(r, 11, '=IFERROR($I{0}-INDEX(Cadastro!$N${1}:$N${2},MATCH($B{0},Cadastro!$B${1}:$B${2},0)),"")'.format(r, CAD_F, CAD_L))
corpo_tabela(wsT, TST_F, TST_L, 1, 18)
for r in range(TST_F, TST_L + 1):
    for c in list(range(1, 11)) + list(range(12, 19)):
        wsT.cell(r, c).fill = PatternFill("solid", fgColor=GOLD_L)
        wsT.cell(r, c).font = Font(name=F, size=9, color="0000FF")
    wsT.cell(r, 11).font = Font(name=F, size=9, bold=True, color=NAVY2)
    wsT.cell(r, 11).fill = PatternFill("solid", fgColor=LIGHT)
    for c in (2, 18):
        wsT.cell(r, c).alignment = Alignment(horizontal="left", vertical="center", indent=1)
    wsT.cell(r, 1).number_format = NF_DATE
    for c in (4, 5, 6, 7, 8, 14, 16, 17):
        wsT.cell(r, c).number_format = '0.0;;""'
    for c in (12, 13):
        wsT.cell(r, c).number_format = '0.00;;""'
dv(wsT, ATLETAS_REF,          "B{}:B{}".format(TST_F, TST_L))
dv(wsT, L("Momento do Teste"),"C{}:C{}".format(TST_F, TST_L))
wsT.freeze_panes = "C8"
wsT.auto_filter.ref = "A7:R{}".format(TST_L)

# ============================================================================
# 13) ÁREA DO ATLETA
# ============================================================================
wsA = wb.create_sheet("Atleta")
banner(wsA, "ÁREA DO ATLETA", "Escolha o seu nome no menu: perfil, cargas da semana, monotonia e o treino prescrito aparecem automaticamente.", 16, GOLD)
larguras(wsA, {"A":10,"B":11,"C":13,"D":13,"E":13,"F":13,"G":13,"H":13,"I":13,"J":15,"K":10,"L":13,"M":14,"N":12,"O":13,"P":18})
rotulo(wsA, 4, 1, "ATLETA:")
selA = entrada(wsA, 4, 3, NOMES[0], largura_merge=4)
selA.font = Font(name=F, size=13, bold=True, color="0000FF")
wsA.row_dimensions[4].height = 24
dv(wsA, ATLETAS_REF, "C4")
rotulo(wsA, 5, 1, "Ver treinos a partir de:")
entrada(wsA, 5, 3, date(2026, 8, 31), NF_DATE, largura_merge=4)
wsA["C5"].comment = Comment("Digite uma data ou a fórmula  =HOJE()  (=TODAY() em inglês) para ver sempre os treinos "
                            "de hoje em diante.", "Planilha")

def idx_cad(col_letter):
    return '=IFERROR(INDEX(Cadastro!${0}${1}:${0}${2},MATCH($C$4,Cadastro!$B${1}:$B${2},0)),"")'.format(col_letter, CAD_F, CAD_L)

secao(wsA, 7, "PERFIL DO ATLETA", 16, 1)
PERFIL = [("ID","A",None),("Idade","D",'0" anos"'),("Posição","F",None),("Categoria","G",None),
          ("Nº Camisa","H",'0'),("Estatura (m)","J",'0.00'),("Massa (kg)","K",'0.0'),("IMC","L",'0.0'),
          ("Alcance de Ataque","O",NF_CM),("Impulsão de Ataque","Q",NF_CM),("Alcance de Bloqueio","P",NF_CM),
          ("Status","T",None)]
for i, (lab, cl, nf) in enumerate(PERFIL):
    col = 1 + (i // 4) * 6
    r = 8 + (i % 4)
    rotulo(wsA, r, col, lab)
    c = calc(wsA, r, col + 1, idx_cad(cl), nf)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    wsA.merge_cells(start_row=r, start_column=col + 1, end_row=r, end_column=col + 4)

CG_ = "'Carga (PSE)'"
RC_ = "${0}${1}:${0}${2}"
def cg(col): return "{}!${}${}:${}${}".format(CG_, col, CAR_F, col, CAR_L)
def we(col): return "Wellness!${0}${1}:${0}${2}".format(col, WEL_F, WEL_L)

secao(wsA, 13, "INDICADORES ATUAIS", 16, 1)
kpi(wsA, 14, 1,  "SESSÕES REGISTRADAS", '=COUNTIFS({},$C$4)'.format(cg("C")), '0', NAVY2, 2)
kpi(wsA, 14, 3,  "CARGA TOTAL (UA)", '=SUMIFS({},{},$C$4)'.format(cg("G"), cg("C")), NF_UA, NAVY2, 2)
kpi(wsA, 14, 5,  "CARGA MÉDIA SEMANAL (UA)", '=IFERROR(AVERAGEIF($J$20:$J$67,">0"),0)', NF_UA, NAVY2, 2)
kpi(wsA, 14, 7,  "ÚLTIMO REGISTRO", '=IFERROR(IF(SUMPRODUCT(MAX(({0}=$C$4)*{1}))=0,"",SUMPRODUCT(MAX(({0}=$C$4)*{1}))),"")'.format(cg("C"), cg("A")), 'DD/MM/YYYY;;""', GOLD, 2)
kpi(wsA, 14, 9,  "ACWR ATUAL", '=IF($G$15="",0,IFERROR(AVERAGEIFS({},{},$C$4,{},$G$15),0))'.format(cg("J"), cg("C"), cg("A")), NF_DEC, GOLD, 2)
kpi(wsA, 14, 11, "MONOTONIA MÉDIA", '=IFERROR(AVERAGEIF($N$20:$N$67,">0"),0)', NF_DEC, GOLD, 2)
kpi(wsA, 14, 13, "% DE PRESENÇA", '=IFERROR(INDEX(Presença!$L${0}:$L${1},MATCH($C$4,Presença!$I${0}:$I${1},0)),0)'.format(PRS_F, PRS_RES_L), NF_PCT, NAVY2, 2)
kpi(wsA, 14, 15, "HOOPER MÉDIO", '=IFERROR(AVERAGEIFS({},{},$C$4),0)'.format(we("H"), we("C")), '0.0;;""', "6B3FA0", 2)
wsA.merge_cells(start_row=16, start_column=1, end_row=16, end_column=16)
alerta = wsA.cell(16, 1, '=IF($C$4="","Selecione um atleta em C4.",'
    'IF($I$15>1.5,"ALERTA: ACWR acima de 1,50 — pico de carga aguda. Avaliar redução do volume nas próximas sessões.",'
    'IF($I$15<0.8,"ATENÇÃO: ACWR abaixo de 0,80 — carga recente baixa em relação à crônica. Progredir gradualmente.",'
    'IF($O$15>16,"ATENÇÃO: Índice de Hooper médio elevado — monitorar sono, fadiga e dor muscular.",'
    'IF($K$15>2,"ATENÇÃO: monotonia média acima de 2,0 — variar mais as cargas ao longo da semana.",'
    '"Situação dentro dos parâmetros de referência.")))))')
alerta.font = Font(name=F, size=10, bold=True, color=NAVY)
alerta.fill = PatternFill("solid", fgColor=GOLD_L)
alerta.alignment = Alignment(horizontal="left", vertical="center", indent=1)
alerta.border = BORDER
wsA.row_dimensions[16].height = 24

secao(wsA, 18, "CONTROLE SEMANAL DE CARGA — MONOTONIA E STRAIN (FOSTER)", 16, 1)
SEM_H = ["Semana","Início","Seg","Ter","Qua","Qui","Sex","Sáb","Dom","Carga Semanal (UA)","Sessões",
         "Média Diária","Desvio-Padrão","Monotonia","Strain","Classificação"]
cab_tabela(wsA, 19, SEM_H)
AW_F = 20
AW_L = AW_F + N_SEM - 1
for i in range(N_SEM):
    r = AW_F + i
    wsA.cell(r, 1, i + 1)
    wsA.cell(r, 2, '=IF(Macrociclo!$C$11="","",Macrociclo!$C$11+($A{}-1)*7)'.format(r))
    for d in range(7):
        wsA.cell(r, 3 + d, '=IF($B{0}="",0,SUMIFS({1},{2},$C$4,{3},$B{0}+{4}))'.format(r, cg("G"), cg("C"), cg("A"), d))
    wsA.cell(r, 10, "=SUM($C{0}:$I{0})".format(r))
    wsA.cell(r, 11, '=COUNTIFS({},$C$4,{},$A{})'.format(cg("C"), cg("B"), r))
    wsA.cell(r, 12, '=IF($J{0}=0,0,$J{0}/7)'.format(r))
    wsA.cell(r, 13, '=IF($J{0}=0,0,STDEV($C{0}:$I{0}))'.format(r))
    wsA.cell(r, 14, '=IF($M{0}=0,0,$L{0}/$M{0})'.format(r))
    wsA.cell(r, 15, '=$J{0}*$N{0}'.format(r))
    wsA.cell(r, 16, '=IF($J{0}=0,"",IF($N{0}>2,"Monotonia alta",IF($N{0}>=1.5,"Atenção","Adequada")))'.format(r))
corpo_tabela(wsA, AW_F, AW_L, 1, 16)
for r in range(AW_F, AW_L + 1):
    for c in range(1, 17):
        wsA.cell(r, c).font = Font(name=F, size=9, bold=(c in (10, 14, 15)), color=NAVY2)
        wsA.cell(r, c).fill = PatternFill("solid", fgColor=LIGHT if (r - AW_F) % 2 == 0 else LIGHT2)
    wsA.cell(r, 2).number_format = 'DD/MM'
    for c in list(range(3, 12)) + [15]:
        wsA.cell(r, c).number_format = NF_UA
    for c in (12, 13, 14):
        wsA.cell(r, c).number_format = NF_DEC
wsA.conditional_formatting.add("N{}:N{}".format(AW_F, AW_L),
    CellIsRule(operator="greaterThan", formula=["2"], fill=PatternFill("solid", fgColor=RED), font=Font(name=F, size=9, bold=True, color=RED_T)))
wsA.conditional_formatting.add("J{}:J{}".format(AW_F, AW_L),
    ColorScaleRule(start_type="min", start_color="FFFFFF", end_type="max", end_color="9CC3E5"))

chA = BarChart(); chA.type = "col"; chA.title = "Carga Semanal do Atleta (UA)"
chA.height = 8; chA.width = 20; chA.y_axis.title = "UA"; chA.legend = None
chA.add_data(Reference(wsA, min_col=10, min_row=19, max_row=AW_L), titles_from_data=True)
chA.set_categories(Reference(wsA, min_col=1, min_row=AW_F, max_row=AW_L))
wsA.add_chart(chA, "R19")
chA2 = LineChart(); chA2.title = "Monotonia Semanal"
chA2.height = 8; chA2.width = 20; chA2.legend = None
chA2.add_data(Reference(wsA, min_col=14, min_row=19, max_row=AW_L), titles_from_data=True)
chA2.set_categories(Reference(wsA, min_col=1, min_row=AW_F, max_row=AW_L))
wsA.add_chart(chA2, "R36")

PR_BLOCK = AW_L + 3
secao(wsA, PR_BLOCK, "MEU TREINO PRESCRITO (a partir da data informada em C5)", 16, 1)
TRE_H = ["Data","Sessão","Bloco","Exercício","Capacidade Física","Séries","Reps / Tempo","Carga / Intensidade",
         "Pausa (s)","Duração (min)","PSE Prev.","Progressão / Critério","Observações"]
cab_tabela(wsA, PR_BLOCK + 1, TRE_H)
TR_F = PR_BLOCK + 2
TR_L = TR_F + 29
SRC = ["A", "C", "F", "G", "J", "K", "L", "M", "N", "O", "P", "S", "T"]
for i in range(TR_L - TR_F + 1):
    r = TR_F + i
    for j, sc in enumerate(SRC):
        wsA.cell(r, 1 + j, '=IFERROR(INDEX(Prescrição!${0}${1}:${0}${2},MATCH({3},Prescrição!$U${1}:$U${2},0)),"")'
                 .format(sc, PRE_F, PRE_L, i + 1))
corpo_tabela(wsA, TR_F, TR_L, 1, 13)
for r in range(TR_F, TR_L + 1):
    for c in range(1, 14):
        wsA.cell(r, c).font = Font(name=F, size=9, color=NAVY2)
        wsA.cell(r, c).fill = PatternFill("solid", fgColor=LIGHT if (r - TR_F) % 2 == 0 else LIGHT2)
    wsA.cell(r, 1).number_format = NF_DATE
    for c in (4, 5, 8, 12, 13):
        wsA.cell(r, c).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    wsA.cell(r, 4).font = Font(name=F, size=9, bold=True, color=NAVY)
wsA.column_dimensions["D"].width = 34
nota(wsA, TR_L + 2, 1, "Aparecem aqui as linhas da aba Prescrição cujo Destinatário seja 'Equipe (todos)' ou o nome do atleta "
     "selecionado, com data igual ou posterior à informada em C5.", 16)
wsA.freeze_panes = "A7"

# ============================================================================
# 14) PAINEL INTERATIVO
# ============================================================================
wsD = wb.create_sheet("Painel")
banner(wsD, "PAINEL INTERATIVO DA EQUIPE", "Indicadores e gráficos alimentados por todas as abas de registro. Altere a semana de "
       "referência em C4 para atualizar os cartões.", 20, NAVY)
larguras(wsD, {"A":10,"B":11,"C":15,"D":16,"E":11,"F":12,"G":13,"H":12,"I":13,"J":3,
               "K":13,"L":13,"M":13,"N":13,"O":13,"P":13,"Q":13,"R":13,"S":26,"T":14,"U":12})
rotulo(wsD, 4, 1, "Semana de referência:")
entrada(wsD, 4, 3, '=IFERROR(MAX(1,INT((TODAY()-Macrociclo!$C$11)/7)+1),1)', '0', largura_merge=2)
wsD["C4"].comment = Comment("Vem preenchida com a semana atual do macrociclo. Digite outro número para analisar "
                            "qualquer semana da temporada.", "Planilha")

ATV = 'COUNTIF(Cadastro!$T${}:$T${},"Ativo")'.format(CAD_F, CAD_L)
kpi(wsD, 6, 1,  "ATLETAS ATIVOS", '={}'.format(ATV), '0', NAVY2, 2)
kpi(wsD, 6, 3,  "LESIONADOS / DM", '=COUNTIF(Cadastro!$T${0}:$T${1},"Lesionado")+COUNTIF(Cadastro!$T${0}:$T${1},"Departamento Médico")'.format(CAD_F, CAD_L), '0', RED_T, 2)
kpi(wsD, 6, 5,  "REGISTROS DE CARGA", '=COUNT({})'.format(cg("A")), '0', NAVY2, 2)
kpi(wsD, 6, 7,  "CARGA TOTAL DA EQUIPE (UA)", '=SUM({})'.format(cg("G")), NF_UA, NAVY2, 3)
kpi(wsD, 6, 10, "CARGA DA SEMANA (UA)", '=SUMIFS({},{},$C$4)'.format(cg("G"), cg("B")), NF_UA, GOLD, 3)
kpi(wsD, 6, 13, "CARGA MÉDIA / ATLETA NA SEMANA", '=IFERROR(J7/MAX(1,{}),0)'.format(ATV), NF_UA, GOLD, 3)
kpi(wsD, 6, 16, "ATLETAS EM RISCO (ACWR > 1,5)", '=COUNTIFS({},$C$4,{},">1.5")'.format(cg("B"), cg("J")), '0', RED_T, 2)
kpi(wsD, 6, 18, "HOOPER MÉDIO DA SEMANA", '=IFERROR(AVERAGEIFS({},{},$C$4),0)'.format(we("H"), we("B")), '0.0;;""', "6B3FA0", 3)

secao(wsD, 9, "EVOLUÇÃO SEMANAL DA EQUIPE", 9, 1)
PN_H = ["Semana","Início","Carga Total (UA)","Carga Média/Atleta","Registros","PSE Média","Hooper Médio","ACWR Médio","% Presença"]
cab_tabela(wsD, 10, PN_H)
PN_F = 11
PN_L = PN_F + N_SEM - 1
for i in range(N_SEM):
    r = PN_F + i
    wsD.cell(r, 1, i + 1)
    wsD.cell(r, 2, '=IF(Macrociclo!$C$11="","",Macrociclo!$C$11+($A{}-1)*7)'.format(r))
    wsD.cell(r, 3, '=SUMIFS({},{},$A{})'.format(cg("G"), cg("B"), r))
    wsD.cell(r, 4, '=IFERROR($C{}/MAX(1,{}),0)'.format(r, ATV))
    wsD.cell(r, 5, '=COUNTIFS({},$A{})'.format(cg("B"), r))
    wsD.cell(r, 6, '=IFERROR(AVERAGEIFS({},{},$A{}),0)'.format(cg("F"), cg("B"), r))
    wsD.cell(r, 7, '=IFERROR(AVERAGEIFS({},{},$A{}),0)'.format(we("H"), we("B"), r))
    wsD.cell(r, 8, '=IFERROR(AVERAGEIFS({},{},$A{}),0)'.format(cg("J"), cg("B"), r))
    wsD.cell(r, 9, '=IFERROR(COUNTIFS(Presença!$B${1}:$B${2},$A{0},Presença!$E${1}:$E${2},"Presente")'
                   '/MAX(1,COUNTIFS(Presença!$B${1}:$B${2},$A{0})),0)'.format(r, PRS_F, PRS_L))
corpo_tabela(wsD, PN_F, PN_L, 1, 9)
for r in range(PN_F, PN_L + 1):
    for c in range(1, 10):
        wsD.cell(r, c).font = Font(name=F, size=9, bold=(c == 3), color=NAVY2)
        wsD.cell(r, c).fill = PatternFill("solid", fgColor=LIGHT if (r - PN_F) % 2 == 0 else LIGHT2)
    wsD.cell(r, 2).number_format = 'DD/MM'
    wsD.cell(r, 3).number_format = NF_UA; wsD.cell(r, 4).number_format = NF_UA
    wsD.cell(r, 5).number_format = '0;;""'
    wsD.cell(r, 6).number_format = NF_DEC; wsD.cell(r, 7).number_format = NF_DEC
    wsD.cell(r, 8).number_format = NF_DEC; wsD.cell(r, 9).number_format = NF_PCT
wsD.conditional_formatting.add("H{}:H{}".format(PN_F, PN_L),
    CellIsRule(operator="greaterThan", formula=["1.5"], fill=PatternFill("solid", fgColor=RED), font=Font(name=F, size=9, bold=True, color=RED_T)))
wsD.conditional_formatting.add("C{}:C{}".format(PN_F, PN_L),
    ColorScaleRule(start_type="min", start_color="FFFFFF", end_type="max", end_color="8FAADC"))

c1 = BarChart(); c1.type = "col"; c1.title = "Carga Total da Equipe por Semana (UA)"
c1.height = 8.5; c1.width = 20; c1.legend = None; c1.y_axis.title = "UA"
c1.add_data(Reference(wsD, min_col=3, min_row=10, max_row=PN_L), titles_from_data=True)
c1.set_categories(Reference(wsD, min_col=1, min_row=PN_F, max_row=PN_L))
wsD.add_chart(c1, "K10")
c2 = LineChart(); c2.title = "ACWR Médio da Equipe por Semana"
c2.height = 8.5; c2.width = 20; c2.legend = None
c2.add_data(Reference(wsD, min_col=8, min_row=10, max_row=PN_L), titles_from_data=True)
c2.set_categories(Reference(wsD, min_col=1, min_row=PN_F, max_row=PN_L))
wsD.add_chart(c2, "K28")

TIPOS_ROW = 46
secao(wsD, TIPOS_ROW - 1, "CARGA POR TIPO DE SESSÃO", 13, 11)
cab_tabela(wsD, TIPOS_ROW, ["Tipo de Sessão", "Carga (UA)", "% do Total"], col0=11)
TIPOS = LISTAS[16][1]
for i, t in enumerate(TIPOS):
    r = TIPOS_ROW + 1 + i
    wsD.cell(r, 11, t)
    wsD.cell(r, 12, '=SUMIFS({},{},$K{})'.format(cg("G"), cg("D"), r))
    wsD.cell(r, 13, '=IFERROR($L{}/SUM($L${}:$L${}),0)'.format(r, TIPOS_ROW + 1, TIPOS_ROW + len(TIPOS)))
TIP_L = TIPOS_ROW + len(TIPOS)
corpo_tabela(wsD, TIPOS_ROW + 1, TIP_L, 11, 13)
for r in range(TIPOS_ROW + 1, TIP_L + 1):
    for c in range(11, 14):
        wsD.cell(r, c).font = Font(name=F, size=9, color=NAVY2)
        wsD.cell(r, c).fill = PatternFill("solid", fgColor=LIGHT if (r - TIPOS_ROW) % 2 == 1 else LIGHT2)
    wsD.cell(r, 11).alignment = Alignment(horizontal="left", vertical="center", indent=1)
    wsD.cell(r, 12).number_format = NF_UA; wsD.cell(r, 13).number_format = NF_PCT
cp = PieChart(); cp.title = "Distribuição da Carga por Tipo de Sessão"
cp.height = 9; cp.width = 13
cp.add_data(Reference(wsD, min_col=12, min_row=TIPOS_ROW, max_row=TIP_L), titles_from_data=True)
cp.set_categories(Reference(wsD, min_col=11, min_row=TIPOS_ROW + 1, max_row=TIP_L))
wsD.add_chart(cp, "O46")

RK = 62
secao(wsD, RK - 1, "RANKING E SITUAÇÃO POR ATLETA", 9, 1)
cab_tabela(wsD, RK, ["Atleta","Posição","Sessões","Carga Total (UA)","Carga Média/Sessão (UA)","ACWR Atual",
                     "% Presença","Hooper Médio","Situação"])
RK_F = RK + 1
RK_L = RK_F + (CAD_L - CAD_F)
for i in range(CAD_L - CAD_F + 1):
    r = RK_F + i
    a = CAD_F + i
    wsD.cell(r, 1, '=IF(Cadastro!$B{0}="","",Cadastro!$B{0})'.format(a))
    wsD.cell(r, 2, '=IF($A{}="","",Cadastro!$F{})'.format(r, a))
    wsD.cell(r, 3, '=IF($A{}="",0,COUNTIFS({},$A{}))'.format(r, cg("C"), r))
    wsD.cell(r, 4, '=IF($A{0}="",0,SUMIFS({1},{2},$A{0}))'.format(r, cg("G"), cg("C")))
    wsD.cell(r, 5, '=IF($C{0}=0,0,$D{0}/$C{0})'.format(r))
    wsD.cell(r, 6, '=IF($A{0}="",0,IFERROR(AVERAGEIFS({1},{2},$A{0},{3},SUMPRODUCT(MAX(({2}=$A{0})*{3}))),0))'
             .format(r, cg("J"), cg("C"), cg("A")))
    wsD.cell(r, 7, '=IF($A{0}="",0,IFERROR(INDEX(Presença!$L${1}:$L${2},MATCH($A{0},Presença!$I${1}:$I${2},0)),0))'
             .format(r, PRS_F, PRS_RES_L))
    wsD.cell(r, 8, '=IF($A{0}="",0,IFERROR(AVERAGEIFS({1},{2},$A{0}),0))'.format(r, we("H"), we("C")))
    wsD.cell(r, 9, '=IF($A{0}="","",IF(Cadastro!$T{1}="Lesionado","LESIONADO",IF($F{0}>1.5,"ALERTA — carga aguda alta",'
                   'IF($H{0}>16,"ALERTA — wellness baixo",IF(AND($G{0}>0,$G{0}<0.75),"Assiduidade baixa",'
                   'IF($C{0}=0,"Sem registros","OK"))))))'.format(r, a))
corpo_tabela(wsD, RK_F, RK_L, 1, 9)
for r in range(RK_F, RK_L + 1):
    for c in range(1, 10):
        wsD.cell(r, c).font = Font(name=F, size=9, bold=(c == 4), color=NAVY2)
        wsD.cell(r, c).fill = PatternFill("solid", fgColor=LIGHT if (r - RK_F) % 2 == 0 else LIGHT2)
    for c in (1, 2, 9):
        wsD.cell(r, c).alignment = Alignment(horizontal="left", vertical="center", indent=1)
    wsD.cell(r, 3).number_format = '0;;""'
    wsD.cell(r, 4).number_format = NF_UA; wsD.cell(r, 5).number_format = NF_UA
    wsD.cell(r, 6).number_format = NF_DEC; wsD.cell(r, 7).number_format = NF_PCT
    wsD.cell(r, 8).number_format = '0.0;;""'
wsD.conditional_formatting.add("F{}:F{}".format(RK_F, RK_L),
    CellIsRule(operator="greaterThan", formula=["1.5"], fill=PatternFill("solid", fgColor=RED), font=Font(name=F, size=9, bold=True, color=RED_T)))
wsD.conditional_formatting.add("I{}:I{}".format(RK_F, RK_L),
    CellIsRule(operator="equal", formula=['"OK"'], fill=PatternFill("solid", fgColor=GREEN), font=Font(name=F, size=9, bold=True, color=GREEN_T)))
wsD.conditional_formatting.add("I{}:I{}".format(RK_F, RK_L),
    FormulaRule(formula=['LEFT($I{},6)="ALERTA"'.format(RK_F)], fill=PatternFill("solid", fgColor=RED), font=Font(name=F, size=9, bold=True, color=RED_T)))
wsD.conditional_formatting.add("G{}:G{}".format(RK_F, RK_L),
    ColorScaleRule(start_type="num", start_value=0.5, start_color=RED, mid_type="num", mid_value=0.8,
                   mid_color=YELL, end_type="num", end_value=1, end_color=GREEN))
RK_CH = min(RK_L, RK_F + 19)
c3 = BarChart(); c3.type = "bar"; c3.title = "Carga Total por Atleta (UA)"
c3.height = 10; c3.width = 20; c3.legend = None
c3.add_data(Reference(wsD, min_col=4, min_row=RK, max_row=RK_CH), titles_from_data=True)
c3.set_categories(Reference(wsD, min_col=1, min_row=RK_F, max_row=RK_CH))
wsD.add_chart(c3, "K62")
c4 = BarChart(); c4.type = "bar"; c4.title = "% de Presença por Atleta"
c4.height = 10; c4.width = 20; c4.legend = None
c4.add_data(Reference(wsD, min_col=7, min_row=RK, max_row=RK_CH), titles_from_data=True)
c4.set_categories(Reference(wsD, min_col=1, min_row=RK_F, max_row=RK_CH))
wsD.add_chart(c4, "S62")
wsD.freeze_panes = "A11"

# ============================================================================
# 15) DADOS DE EXEMPLO
# ============================================================================
random.seed(20260903)
SEMANAS_EX = [date(2026, 8, 10), date(2026, 8, 17), date(2026, 8, 24), date(2026, 8, 31)]
SESSOES_EX = [(0, "Físico (Força)", 75, 7), (0, "Técnico-Tático", 90, 6), (1, "Físico (Potência)", 60, 7),
              (2, "Coletivo/Jogo", 100, 8), (3, "Físico (Condicionamento)", 50, 6), (3, "Técnico-Tático", 90, 7),
              (4, "Coletivo/Jogo", 110, 8)]
LESIONADA = NOMES[7]

rg, rp = CAR_F, PRS_F
for wk_i, seg in enumerate(SEMANAS_EX):
    for off, tipo, dur, pse in SESSOES_EX:
        if wk_i == 3 and off > 2:   # última semana ainda em curso
            continue
        d = seg + timedelta(days=off)
        for nome in NOMES:
            lesao = (nome == LESIONADA and wk_i >= 2)
            falta = (not lesao) and random.random() < 0.05
            if not lesao and not falta:
                p = max(1, min(10, pse + random.choice([-1, 0, 0, 0, 1])))
                du = dur + random.choice([-10, -5, 0, 0, 5])
                wsG.cell(rg, 1, d); wsG.cell(rg, 1).number_format = NF_DATE
                wsG.cell(rg, 3, nome); wsG.cell(rg, 4, tipo)
                wsG.cell(rg, 5, du); wsG.cell(rg, 6, p)
                wsG.cell(rg, 13, "EXEMPLO")
                rg += 1
            wsPr.cell(rp, 1, d); wsPr.cell(rp, 1).number_format = NF_DATE
            wsPr.cell(rp, 3, tipo); wsPr.cell(rp, 4, nome)
            wsPr.cell(rp, 5, "Lesionado" if lesao else ("Falta Justificada" if falta else "Presente"))
            wsPr.cell(rp, 6, 0 if (lesao or falta) else dur)
            wsPr.cell(rp, 7, "Entorse de tornozelo — DM" if lesao else ("Compromisso acadêmico" if falta else "EXEMPLO"))
            rp += 1

rw = WEL_F
for wk_i, seg in enumerate(SEMANAS_EX):
    for off in range(3 if wk_i == 3 else 5):
        d = seg + timedelta(days=off)
        for i, nome in enumerate(NOMES):
            base = 4 if nome == LESIONADA else 2
            wsWe.cell(rw, 1, d); wsWe.cell(rw, 1).number_format = NF_DATE
            wsWe.cell(rw, 3, nome)
            for j in range(4):
                wsWe.cell(rw, 4 + j, max(1, min(7, base + random.choice([0, 0, 1, 1, 2]))))
            wsWe.cell(rw, 11, "EXEMPLO")
            rw += 1

rt = TST_F
for momento, dt in [("Pré-temporada", date(2026, 1, 8)), ("Meio da Temporada", date(2026, 7, 22))]:
    for i, a in enumerate(EXEMPLO):
        nome, _, _, _, _, _, _, _, _, est, mas, env, pe, atq, blq = a
        g = 0 if momento == "Pré-temporada" else 1
        wsT.cell(rt, 1, dt); wsT.cell(rt, 1).number_format = NF_DATE
        wsT.cell(rt, 2, nome); wsT.cell(rt, 3, momento)
        wsT.cell(rt, 4, round(mas + g * random.uniform(-1.5, 0.5), 1))
        wsT.cell(rt, 5, round(random.uniform(17, 26) - g * 0.8, 1))
        wsT.cell(rt, 6, round(28 + (atq - pe - 50) * 0.35 + g * 1.5 + random.uniform(-2, 2), 1))
        wsT.cell(rt, 7, round(32 + (atq - pe - 50) * 0.38 + g * 1.8 + random.uniform(-2, 2), 1))
        wsT.cell(rt, 8, round(atq - pe + g * 2 + random.uniform(-2, 2), 1))
        wsT.cell(rt, 9, atq + g * 2)
        wsT.cell(rt, 10, blq + g * 2)
        wsT.cell(rt, 12, round(10.9 - g * 0.25 + random.uniform(-0.3, 0.3), 2))
        wsT.cell(rt, 13, round(1.95 - g * 0.04 + random.uniform(-0.06, 0.06), 2))
        wsT.cell(rt, 14, round(random.uniform(24, 36), 1))
        wsT.cell(rt, 15, int(random.uniform(880, 1560)))
        wsT.cell(rt, 16, round(random.uniform(28, 42), 1))
        wsT.cell(rt, 17, round(random.uniform(7.5, 11.5), 1))
        wsT.cell(rt, 18, "EXEMPLO")
        rt += 1

PRESC_EX = [
 (date(2026,8,31),"S1 — Força","Manhã","Equipe (todos)","Aquecimento","Mobilidade de tornozelo e quadril",1,"8 por exercício","Peso corporal",30,10,2,"Amplitude completa","EXEMPLO"),
 (date(2026,8,31),"S1 — Força","Manhã","Equipe (todos)","Ativação / Prevenção","Rotadores do ombro com elástico",3,15,"Elástico vermelho",45,10,3,"Progredir a resistência do elástico a cada 2 semanas","EXEMPLO"),
 (date(2026,8,31),"S1 — Força","Manhã","Equipe (todos)","Parte Principal","Agachamento livre (back squat)",4,5,"82% de 1RM",180,25,8,"+2,5% na semana seguinte se completar todas as séries","EXEMPLO"),
 (date(2026,8,31),"S1 — Força","Manhã","Equipe (todos)","Parte Principal","Levantamento terra romeno",4,8,"70% de 1RM",120,18,7,"Excêntrica de 3 s","EXEMPLO"),
 (date(2026,8,31),"S1 — Força","Manhã","Equipe (todos)","Parte Principal","Power clean (levantamento olímpico)",5,3,"75% de 1RM",180,20,8,"Interromper a série se a velocidade da barra cair","EXEMPLO"),
 (date(2026,8,31),"S1 — Força","Manhã","Equipe (todos)","Complementar","Prancha e anti-rotação (Pallof press)",3,"40 s / 12 rep","Peso corporal + elástico",45,12,4,"Aumentar 10 s por semana","EXEMPLO"),
 (date(2026,8,31),"S1 — Força","Manhã","Equipe (todos)","Volta à Calma","Alongamento e liberação miofascial",1,"30 s por grupamento","Rolo",0,15,2,"—","EXEMPLO"),
 (date(2026,8,31),"S2 — Técnico-Tático","Tarde","Equipe (todos)","Aquecimento","Recepção em duplas com bola dirigida",3,15,"Bola",30,12,4,"Meta de 80% de passes ao alvo","EXEMPLO"),
 (date(2026,8,31),"S2 — Técnico-Tático","Tarde","Equipe (todos)","Parte Principal","Recepção de saque em sistema de 2 passadores",4,20,"Saque real",60,20,6,"Meta de 70% de passes A/B","EXEMPLO"),
 (date(2026,8,31),"S2 — Técnico-Tático","Tarde","Equipe (todos)","Parte Principal","Ataque de bola alta pela ponta (P4)",4,10,"Bola alta",60,18,6,"Aumentar para 5 séries na próxima semana","EXEMPLO"),
 (date(2026,8,31),"S2 — Técnico-Tático","Tarde","Equipe (todos)","Parte Principal","Complexo I (K1) — recepção, levantamento e ataque",1,"20 tentativas","Jogo",90,25,7,"Meta de 60% de side-out","EXEMPLO"),
 (date(2026,8,31),"S2 — Técnico-Tático","Tarde","Equipe (todos)","Volta à Calma","Alongamento e liberação miofascial",1,"30 s por grupamento","Rolo",0,15,2,"—","EXEMPLO"),
 (date(2026,8,31),"S2 — Individual","Tarde","Larissa Mendes Prado","Complementar","Saque potente com salto",4,6,"Máxima",90,15,7,"Registrar acertos por série","EXEMPLO — treino individual"),
 (date(2026,9,2),"S3 — Potência","Manhã","Equipe (todos)","Aquecimento","Mobilidade de tornozelo e quadril",1,"8 por exercício","Peso corporal",30,10,2,"—","EXEMPLO"),
 (date(2026,9,2),"S3 — Potência","Manhã","Equipe (todos)","Parte Principal","Agachamento com salto sob carga (jump squat)",5,4,"25% de 1RM",120,18,7,"Manter a velocidade acima de 1,0 m/s","EXEMPLO"),
 (date(2026,9,2),"S3 — Potência","Manhã","Equipe (todos)","Parte Principal","Salto em profundidade (drop jump)",5,5,"Caixa de 40 cm",120,15,7,"Progredir a caixa só com bom controle na aterrissagem","EXEMPLO"),
 (date(2026,9,2),"S3 — Potência","Manhã","Equipe (todos)","Complementar","Nórdico de isquiotibiais (excêntrico)",3,6,"Peso corporal",90,10,6,"2x por semana ao longo do mesociclo","EXEMPLO"),
 (date(2026,9,2),"S4 — Coletivo","Tarde","Equipe (todos)","Parte Principal","Complexo II (K2) — bloqueio, defesa e contra-ataque",1,"20 séries","Jogo",60,25,7,"Avaliar eficiência do contra-ataque","EXEMPLO"),
 (date(2026,9,2),"S4 — Coletivo","Tarde","Equipe (todos)","Parte Principal","Jogo 6x6 com pontuação diferenciada",3,"sets a 15","Jogo",120,30,8,"—","EXEMPLO"),
 (date(2026,9,2),"S4 — Coletivo","Tarde","Equipe (todos)","Volta à Calma","Alongamento e liberação miofascial",1,"30 s por grupamento","Rolo",0,15,2,"—","EXEMPLO"),
]
for i, p in enumerate(PRESC_EX):
    r = PRE_F + i
    dt, ses, turno, dest, bloco, exe, ser, rep, carga, pausa, dur, pse, prog, obs = p
    wsP.cell(r, 1, dt); wsP.cell(r, 1).number_format = NF_DATE
    wsP.cell(r, 3, ses); wsP.cell(r, 4, turno); wsP.cell(r, 5, dest); wsP.cell(r, 6, bloco); wsP.cell(r, 7, exe)
    wsP.cell(r, 11, ser); wsP.cell(r, 12, rep); wsP.cell(r, 13, carga); wsP.cell(r, 14, pausa)
    wsP.cell(r, 15, dur); wsP.cell(r, 16, pse); wsP.cell(r, 19, prog); wsP.cell(r, 20, obs)

wb.move_sheet("Listas", offset=len(wb.sheetnames))
wb.active = wb.sheetnames.index("Início")
for ws in wb.worksheets:
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True

OUT = "/home/user/mdlucca/planilhas/Planilha_Voleibol_Controle_e_Prescricao_de_Treinamento.xlsx"
wb.save(OUT)
print("SALVO:", OUT)

# ---- versão em branco (mesmas fórmulas, sem os dados de exemplo) ------------
def limpar(ws, r1, r2, cols):
    for r in range(r1, r2 + 1):
        for c in cols:
            ws.cell(r, c).value = None

limpar(wsC, CAD_F, CAD_L, [2,3,5,6,7,8,9,10,11,13,14,15,16,19,20,21,22])
limpar(wsM, MESO_F, MESO_L, [2,3,4,5,8,9,10,11,12,13,14,15])
limpar(wsS, MIC_F, MIC_L, [2,3,4,5,6,8,9,10,13])
limpar(wsW, SF, SL, [3,4,5,6,7,9,10,11,12])
limpar(wsW, CF, CL, [2,3,4,5,6])
limpar(wsP, PRE_F, PRE_L, [1,3,4,5,6,7,11,12,13,14,15,16,19,20])
limpar(wsG, CAR_F, CAR_L, [1,3,4,5,6,13])
limpar(wsWe, WEL_F, WEL_L, [1,3,4,5,6,7,10,11])
limpar(wsPr, PRS_F, PRS_L, [1,3,4,5,6,7])
limpar(wsT, TST_F, TST_L, list(range(1, 11)) + list(range(12, 19)))
wsS["C6"] = None
wsW["B5"] = None; wsW["B6"] = None; wsW["B8"] = None; wsW["B9"] = None
wsC["V7"] = "← primeira linha: preencha aqui"
for lab, cel in [("Temporada / Época","C6"),("Equipe / Clube","C7"),("Categoria","C8"),
                 ("Técnico Responsável","C9"),("Preparador Físico","C10"),("Objetivo Principal","C14"),
                 ("Competição-Alvo","C15")]:
    wsM[cel] = None
OUT2 = "/home/user/mdlucca/planilhas/Planilha_Voleibol_Controle_e_Prescricao_de_Treinamento_EM_BRANCO.xlsx"
wb.save(OUT2)
print("SALVO:", OUT2)
