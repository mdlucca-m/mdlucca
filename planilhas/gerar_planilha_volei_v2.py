# -*- coding: utf-8 -*-
"""
Gerador da Planilha de Voleibol v2 - FORCA E POTENCIA
Estrutura: Macrociclo -> Mesociclo -> Microciclo -> Sessao -> Exercicio
Controles: PSE da sessao (Foster), ACWR, Monotonia/Strain, Indice de Hooper,
           presenca, testes fisicos, area do atleta e painel interativo.
"""
import random
from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule, ColorScaleRule
from openpyxl.chart import BarChart, LineChart, PieChart, ScatterChart, Reference, Series
from openpyxl.chart.marker import Marker
from openpyxl.comments import Comment

# ----------------------------------------------------------------------------
# IDENTIDADE VISUAL
# ----------------------------------------------------------------------------
F      = "Arial"
NAVY   = "13315C"   # cabecalho principal
NAVY2  = "1F4E96"   # sub-cabecalho
BLUE3  = "2E75B6"   # cabecalho de tabela
LIGHT  = "EAF1F8"   # faixa clara
LIGHT2 = "F5F8FC"
GOLD   = "E8A33D"   # acento (bola de volei)
GOLD_L = "FFF2CC"   # celula de entrada
GREEN  = "C6EFCE"; GREEN_T = "0B6B2E"
YELL   = "FFEB9C"; YELL_T  = "8A6D0B"
RED    = "FFC7CE"; RED_T   = "9C0006"
GREY_T = "7F8C9A"
WHITE  = "FFFFFF"

thin = Side(style="thin", color="BFCEDD")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

# formatos numericos (zero exibido em branco)
NF_UA   = '#,##0;-#,##0;""'
NF_DEC  = '0.00;-0.00;""'
NF_INT  = '#,##0;-#,##0;""'
NF_PCT  = '0%;-0%;""'
NF_PCT1 = '0.0%;-0.0%;""'
NF_DATE = 'DD/MM/YYYY'
NF_CM   = '0.0" cm";;""'

# ----------------------------------------------------------------------------
# LIMITES DAS TABELAS
# ----------------------------------------------------------------------------
CAD_F, CAD_L   = 8, 47      # Cadastro (40 atletas)
EX_F,  EX_L    = 7, 206     # Biblioteca de exercicios
PRE_F, PRE_L   = 8, 507     # Prescricao
CAR_F, CAR_L   = 8, 1007    # Controle de carga (PSE)
WEL_F, WEL_L   = 8, 1007    # Wellness
PRS_F, PRS_L   = 8, 1007    # Presenca
TST_F, TST_L   = 8, 407     # Testes fisicos
N_SEM          = 48         # semanas do macrociclo nos paineis

wb = Workbook()

# ----------------------------------------------------------------------------
# HELPERS DE ESTILO
# ----------------------------------------------------------------------------
def banner(ws, titulo, subtitulo, last_col, tab=NAVY):
    ws.sheet_properties.tabColor = tab
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    c = ws.cell(1, 1, titulo)
    c.font = Font(name=F, size=15, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.row_dimensions[1].height = 30
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
    c = ws.cell(2, 1, subtitulo)
    c.font = Font(name=F, size=9, color=WHITE, italic=True)
    c.fill = PatternFill("solid", fgColor=GOLD)
    c.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.row_dimensions[2].height = 16
    ws.sheet_view.showGridLines = False

def secao(ws, row, texto, last_col, col=1):
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=last_col)
    c = ws.cell(row, col, texto)
    c.font = Font(name=F, size=10, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=NAVY2)
    c.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.row_dimensions[row].height = 20

def cab_tabela(ws, row, headers, col0=1, wrap=True):
    for i, h in enumerate(headers):
        c = ws.cell(row, col0 + i, h)
        c.font = Font(name=F, size=9, bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=BLUE3)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
        c.border = BORDER
    ws.row_dimensions[row].height = 32

def rotulo(ws, row, col, texto):
    c = ws.cell(row, col, texto)
    c.font = Font(name=F, size=9, bold=True, color=NAVY)
    c.alignment = Alignment(horizontal="right", vertical="center")
    return c

def entrada(ws, row, col, valor=None, nf=None, largura_merge=None):
    c = ws.cell(row, col, valor)
    c.font = Font(name=F, size=10, color="0000FF")
    c.fill = PatternFill("solid", fgColor=GOLD_L)
    c.border = BORDER
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    if nf: c.number_format = nf
    if largura_merge:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + largura_merge - 1)
    return c

def calc(ws, row, col, formula, nf=None):
    c = ws.cell(row, col, formula)
    c.font = Font(name=F, size=10, bold=True, color=NAVY2)
    c.fill = PatternFill("solid", fgColor=LIGHT)
    c.border = BORDER
    c.alignment = Alignment(horizontal="center", vertical="center")
    if nf: c.number_format = nf
    return c

def corpo_tabela(ws, r1, r2, c1, c2, zebra=True):
    """Aplica borda, fonte e zebra ao corpo de uma tabela."""
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cel = ws.cell(r, c)
            cel.font = Font(name=F, size=9)
            cel.border = BORDER
            cel.alignment = Alignment(vertical="center", horizontal="center")
            if zebra and (r - r1) % 2 == 1:
                cel.fill = PatternFill("solid", fgColor=LIGHT2)

def larguras(ws, mapa):
    for col, w in mapa.items():
        ws.column_dimensions[col].width = w

def nota(ws, row, col, texto, last_col=None):
    c = ws.cell(row, col, texto)
    c.font = Font(name=F, size=8, italic=True, color=GREY_T)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    if last_col:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=last_col)
    return c

def dv(ws, formula1, ref, allow_blank=True):
    d = DataValidation(type="list", formula1=formula1, allow_blank=allow_blank, showErrorMessage=False)
    ws.add_data_validation(d)
    d.add(ref)
    return d

def kpi(ws, row, col, titulo, formula, nf=NF_INT, cor=NAVY2, w=3):
    """Cartao de KPI ocupando w colunas x 2 linhas."""
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + w - 1)
    t = ws.cell(row, col, titulo)
    t.font = Font(name=F, size=8, bold=True, color=WHITE)
    t.fill = PatternFill("solid", fgColor=cor)
    t.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 26
    ws.merge_cells(start_row=row + 1, start_column=col, end_row=row + 1, end_column=col + w - 1)
    v = ws.cell(row + 1, col, formula)
    v.font = Font(name=F, size=16, bold=True, color=cor)
    v.fill = PatternFill("solid", fgColor=LIGHT)
    v.alignment = Alignment(horizontal="center", vertical="center")
    v.number_format = nf
    v.border = BORDER
    ws.row_dimensions[row + 1].height = 30
    return v

# ============================================================================
# 1) LISTAS  (fonte das validacoes de dados)
# ============================================================================
LISTAS = [
 ("Posição",            ["Levantador","Oposto","Ponteiro (Ponta)","Central","Líbero","Defensor Específico"]),
 ("Categoria",          ["Sub-13","Sub-15","Sub-17","Sub-19","Sub-21","Adulto","Master"]),
 ("Sexo",               ["Masculino","Feminino"]),
 ("Dominância",         ["Destro","Canhoto","Ambidestro"]),
 ("Status do Atleta",   ["Ativo","Lesionado","Departamento Médico","Em transição","Afastado","Inativo"]),
 ("Período do Macro",   ["Preparatório Geral","Preparatório Específico","Pré-Competitivo","Competitivo I","Competitivo II","Transição"]),
 ("Tipo de Microciclo", ["Incorporação","Ordinário","Choque","Recuperativo","Pré-Competitivo","Competitivo","Polimento (Taper)"]),
 ("Bloco da Sessão",    ["Aquecimento","Ativação / Prevenção","Parte Principal","Complementar","Volta à Calma"]),
 ("Categoria Exerc.",   ["Técnico","Tático","Físico","Preventivo/Compensatório","Recuperação","Cognitivo/Visual"]),
 ("Fundamento",         ["Saque","Recepção (Passe)","Levantamento","Ataque","Bloqueio","Defesa","Deslocamento","Jogo/Coletivo","N/A"]),
 ("Capacidade Física",  ["Força Máxima","Força Explosiva/Potência","Pliometria","Velocidade","Agilidade/COD","Resistência Aeróbia",
                         "Resistência Anaeróbia","Mobilidade/Flexibilidade","Core/Estabilidade","Coordenação","N/A"]),
 ("Presença",           ["Presente","Falta Justificada","Falta Não Justificada","Lesionado","Departamento Médico","Liberado","Seleção/Convocado"]),
 ("Intensidade",        ["Muito Baixa","Baixa","Moderada","Alta","Muito Alta","Máxima"]),
 ("PSE (0-10)",         [0,1,2,3,4,5,6,7,8,9,10]),
 ("Escala 1-7",         [1,2,3,4,5,6,7]),
 ("Turno",              ["Manhã","Tarde","Noite"]),
 ("Tipo de Sessão",     ["Técnico-Tático","Físico (Força)","Físico (Potência)","Físico (Condicionamento)","Coletivo/Jogo",
                         "Amistoso","Competição Oficial","Recuperação/Regenerativo","Vídeo/Teórico","Folga"]),
 ("Nível",              ["Iniciante","Intermediário","Avançado"]),
 ("Situação",           ["Planejado","Em execução","Concluído","Cancelado"]),
 ("Momento do Teste",   ["Pré-temporada","Fim Meso 1","Fim Meso 2","Fim Meso 3","Meio da Temporada","Fim da Temporada","Retorno de Lesão"]),
 ("Perna de Impulsão",  ["Esquerda","Direita","Ambas"]),
 ("Escolaridade",       ["Ensino Fundamental incompleto","Ensino Fundamental completo","Ensino Médio incompleto",
                         "Ensino Médio completo","Superior em andamento","Superior completo","Pós-graduação"]),
 ("Situação de Estudo", ["Cursando","Concluído","Trancado","Não estuda","—"]),
 ("Turno",              ["Manhã","Tarde","Noite","Matutino","Vespertino","Noturno","Integral","EAD","—"]),
 ("Sim/Não",            ["Sim","Não"]),
 ("Classe Econômica",   ["A","B1","B2","C1","C2","D-E"]),
 ("Tipo de Moradia",    ["Própria","Alugada","Cedida","Financiada","Alojamento do clube"]),
 ("Reside com",         ["Família","Cônjuge","Sozinho","Amigos / Colegas","Alojamento do clube","República"]),
 ("Transporte",         ["A pé","Bicicleta","Transporte público","Transporte do clube","Carro próprio","Moto",
                         "Aplicativo / Carona"]),
 ("Nível Competitivo",  ["Municipal","Estadual","Superliga C","Superliga B","Superliga A","Seleção nacional",
                         "Clube no exterior","Atleta em formação"]),
 ("Tipo Sanguíneo",     ["A+","A-","B+","B-","AB+","AB-","O+","O-","Não sabe"]),
 ("Região Corporal",    ["Ombro","Cotovelo","Punho / Mão","Coluna cervical","Coluna lombar","Quadril","Coxa",
                         "Joelho","Perna","Tornozelo","Pé","—"]),
 ("Método de 1RM",      ["Direto (1RM real)","Estimado por repetições","Estimado por velocidade (VBT)",
                         "Carga máxima em treino"]),
 ("Origem do Salto",    ["Pliometria","Treino de quadra","Jogo","Musculação"]),
 ("Superfície",         ["Quadra (madeira)","Quadra (sintética)","Grama","Areia","Aquático","Colchão / Tatame"]),
 ("Ref. VBT",           ["Agachamento","Supino","Terra","—"]),
 ("Objetivo de Força",  ["Força Máxima","Força-Velocidade","Potência","Velocidade-Força","Hipertrofia","Pliometria",
                         "Preventivo","Aquecimento","Core","Velocidade"]),
 ("Bloco de Treino",    ["Acumulação","Transmutação","Realização","Descarga"]),
]

wsL = wb.active; wsL.title = "Listas"
banner(wsL, "LISTAS AUXILIARES  |  fonte das caixas de seleção",
       "Edite/complemente as opções abaixo — elas alimentam automaticamente todos os menus suspensos da planilha.", 42, GREY_T)
cab_tabela(wsL, 4, [n for n, _ in LISTAS])
for j, (_, itens) in enumerate(LISTAS, start=1):
    for i, v in enumerate(itens):
        c = wsL.cell(5 + i, j, v)
        c.font = Font(name=F, size=9); c.border = BORDER
        c.alignment = Alignment(horizontal="center")
    wsL.column_dimensions[get_column_letter(j)].width = 22
LST_F, LST_L = 5, 5 + max(len(i) for _, i in LISTAS) - 1
COL = {n: get_column_letter(j) for j, (n, _) in enumerate(LISTAS, start=1)}

def L(nome):
    """Referencia absoluta da lista na aba Listas."""
    c = COL[nome]
    return "Listas!${0}${1}:${0}${2}".format(c, LST_F, LST_L)

# coluna dinamica: destinatarios da prescricao (Equipe + atletas do cadastro)
cD = get_column_letter(len(LISTAS) + 2)
wsL.column_dimensions[cD].width = 26
h = wsL.cell(4, len(LISTAS) + 2, "Destinatário (auto)")
h.font = Font(name=F, size=9, bold=True, color=WHITE); h.fill = PatternFill("solid", fgColor=GOLD)
h.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); h.border = BORDER
wsL.cell(5, len(LISTAS) + 2, "Equipe (todos)").font = Font(name=F, size=9)
wsL.cell(5, len(LISTAS) + 2).border = BORDER
for i in range(CAD_L - CAD_F + 1):
    c = wsL.cell(6 + i, len(LISTAS) + 2, '=IF(Cadastro!$B${0}="","",Cadastro!$B${0})'.format(CAD_F + i))
    c.font = Font(name=F, size=9); c.border = BORDER
DEST_REF = "Listas!${0}$5:${0}${1}".format(cD, 5 + (CAD_L - CAD_F + 1))
wsL.freeze_panes = "A5"

ATLETAS_REF = "Cadastro!$B${}:$B${}".format(CAD_F, CAD_L)
EXERC_REF   = "Exercícios!$B${}:$B${}".format(EX_F, EX_L)
MESO_REF    = "Macrociclo!$B$19:$B$30"

# ============================================================================
# 2) INÍCIO  (instruções + legenda + índice)
# ============================================================================
wsI = wb.create_sheet("Início")
banner(wsI, "PLANILHA DE CONTROLE E PRESCRIÇÃO DE TREINAMENTO  •  VOLEIBOL",
       "Periodização Macrociclo → Mesociclo → Microciclo → Sessão  |  Controle de carga por PSE da sessão, ACWR, Monotonia/Strain e Wellness", 10, GOLD)
larguras(wsI, {"A":3,"B":22,"C":16,"D":16,"E":16,"F":16,"G":16,"H":16,"I":16,"J":18})

secao(wsI, 4, "COMO USAR — PASSO A PASSO", 10, 2)
passos = [
 ("1. Listas",           "Opções de todos os menus suspensos (posições, escolaridade, renda, tipos de sessão, superfícies…)."),
 ("2. Cadastro",         "Ficha completa do atleta: identificação, contato, SOCIOECONÔMICO, histórico esportivo e saúde."),
 ("3. Antropometria",    "Avaliações longitudinais no perfil ISAK: dobras, perímetros, diâmetros, % de gordura e somatotipo."),
 ("4. Testes",           "Bateria física: alcances e impulsões, SJ, CMJ, drop jump com RSI, sprint, agilidade e IMTP."),
 ("5. Perfil F-V-P",     "Perfil força-velocidade-potência pelo salto com carga: F0, V0, Pmax e desequilíbrio F-V."),
 ("6. Exercícios",       "Biblioteca com 89 exercícios, incluindo todos os do seu documento de Preparação Física."),
 ("7. Programa PF",      "As 8 séries do seu documento 'Preparação Física 2026/2027', na íntegra."),
 ("8. Macrociclo",       "A temporada dividida em mesociclos, com dinâmica de volume e intensidade."),
 ("9. Mesociclo",        "Microciclos do mesociclo escolhido, com carga prevista × realizada."),
 ("10. Bloco Base",      "PERIODIZAÇÃO EM BLOCOS: as 8 semanas de base (força máxima, LPO, agachamento e pliometria)."),
 ("11. Microciclo",      "A semana de quadra: sessões por dia e distribuição de conteúdos."),
 ("12. Prescrição",      "Prescrição do treino técnico-tático, por equipe ou por atleta."),
 ("13. Prescrição Força","Prescrição e controle do treino de força: %1RM, carga em kg, velocidade-alvo e tonelagem."),
 ("14. Força 1RM",       "Testes de 1RM diretos e estimados; mantém o 1RM ATUAL de cada atleta em cada exercício."),
 ("15. Carga (PSE)",     "Carga interna: duração × PSE, carga aguda, crônica, ACWR e zona de risco."),
 ("16. Saltos",          "Carga de saltos: contatos pliométricos, saltos de quadra e de jogo, com alerta de variação."),
 ("17. Wellness",        "Questionário diário e Índice de Hooper."),
 ("18. Presença",        "Chamada por sessão e percentual de assiduidade."),
 ("19. Atleta",          "ÁREA DO ATLETA: perfil, cargas, monotonia e o treino prescrito para ele."),
 ("20. Painel",          "PAINEL da equipe: indicadores e gráficos de carga, wellness e presença."),
 ("21. KPIs Força",      "PAINEL DE FORÇA: tonelagem, intensidade relativa, contatos, 1RM relativo e evolução do CMJ."),
 ("22. Evidências",      "60 referências científicas que sustentam cada escolha metodológica da planilha."),
]
r = 5
for tit, desc in passos:
    c = wsI.cell(r, 2, tit); c.font = Font(name=F, size=9, bold=True, color=NAVY2)
    c.alignment = Alignment(horizontal="left", vertical="center")
    wsI.merge_cells(start_row=r, start_column=3, end_row=r, end_column=10)
    d = wsI.cell(r, 3, desc); d.font = Font(name=F, size=9)
    d.alignment = Alignment(horizontal="left", vertical="center")
    if (r % 2) == 0:
        for cc in range(2, 11):
            wsI.cell(r, cc).fill = PatternFill("solid", fgColor=LIGHT2)
    r += 1

r += 1
secao(wsI, r, "LEGENDA DE CORES", 10, 2); r += 1
legenda = [("Célula de PREENCHIMENTO (azul sobre amarelo)", GOLD_L, "0000FF"),
           ("Célula CALCULADA — não digite por cima", LIGHT, NAVY2),
           ("Cabeçalho de tabela", BLUE3, WHITE),
           ("Alerta / risco elevado", RED, RED_T),
           ("Atenção / monitorar", YELL, YELL_T),
           ("Adequado / dentro da meta", GREEN, GREEN_T)]
for txt, fill, font in legenda:
    wsI.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    c = wsI.cell(r, 2, txt)
    c.fill = PatternFill("solid", fgColor=fill); c.border = BORDER
    c.font = Font(name=F, size=9, bold=True, color=font)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    r += 1

r += 1
secao(wsI, r, "MÉTRICAS DE CONTROLE UTILIZADAS — DEFINIÇÕES E REFERÊNCIAS", 10, 2); r += 1
metricas = [
 ("Carga Interna (UA)", "Duração da sessão (min) × PSE da sessão (escala 0–10, coletada ~30 min após o treino). Método da PSE da sessão de Foster et al. (2001), J Strength Cond Res."),
 ("Carga Aguda (7 d)",  "Soma da carga em UA dos últimos 7 dias do atleta (inclui o dia atual)."),
 ("Carga Crônica (28 d)","Soma da carga dos últimos 28 dias ÷ 4 (média semanal), para ficar na mesma escala da aguda."),
 ("ACWR",               "Carga Aguda ÷ Carga Crônica. Faixas usadas na planilha: <0,80 subcarga | 0,80–1,30 zona ideal | 1,31–1,50 atenção | >1,50 risco elevado. Referência: Gabbett (2016), Br J Sports Med. Interpretar sempre junto do contexto e da clínica do atleta."),
 ("Monotonia",          "Média das cargas diárias da semana ÷ desvio-padrão dessas mesmas cargas (7 dias, dias de folga contam como zero). >2,0 indica semana monótona."),
 ("Strain",             "Carga semanal total × Monotonia. Picos de strain associam-se a maior risco de lesão/doença (Foster, 1998)."),
 ("Índice de Hooper",   "Soma de 4 itens (qualidade do sono, estresse, fadiga e dor muscular), cada um de 1 = muito bom/ausente a 7 = muito ruim/máxima. Varia de 4 a 28; QUANTO MAIOR, PIOR. Referência: Hooper & Mackinnon (1995), Sports Med."),
 ("Volume × Intensidade","No Macrociclo, valores em % representam a proporção relativa planejada para cada mesociclo (ex.: 85% = volume alto)."),
 ("Tonelagem", "Séries × repetições × carga usada, em kg. É o volume-carga do treino de força; some por sessão, semana e bloco."),
 ("Intensidade média relativa", "Média dos %1RM prescritos, ponderada pelo número de repetições de cada exercício. Mostra o quanto a semana foi 'pesada' de verdade."),
 ("Perda de velocidade", "Queda percentual da velocidade da barra dentro da série. Limiares ≤ 25% favorecem força; acima de 20–25% favorecem hipertrofia (Hickmott et al., 2022)."),
 ("F0, V0 e Pmax", "Força teórica máxima, velocidade teórica máxima e potência máxima estimadas por regressão linear a partir de saltos com cargas progressivas (Morin & Samozino, 2016)."),
 ("FVimb", "Razão entre a inclinação real do perfil força-velocidade e a inclinação ótima teórica. Abaixo de 90% indica déficit de força; acima de 110%, déficit de velocidade."),
 ("RSI", "Reactive Strength Index: altura do drop jump ÷ tempo de contato com o solo. Mede a qualidade do ciclo alongamento-encurtamento."),
 ("Contatos pliométricos", "Número de aterrissagens por sessão e por semana. É a dose que a meta-análise de Sáez de Villarreal et al. (2009) mostrou determinar o ganho de impulsão."),
 ("Somatotipo", "Endomorfia, mesomorfia e ectomorfia pelo método Heath-Carter, a partir de dobras, perímetros, diâmetros, estatura e massa."),
]
for nome, desc in metricas:
    c = wsI.cell(r, 2, nome); c.font = Font(name=F, size=9, bold=True, color=NAVY2)
    c.alignment = Alignment(horizontal="left", vertical="top")
    wsI.merge_cells(start_row=r, start_column=3, end_row=r, end_column=10)
    d = wsI.cell(r, 3, desc); d.font = Font(name=F, size=9)
    d.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    wsI.row_dimensions[r].height = 30
    r += 1

r += 1
wsI.merge_cells(start_row=r, start_column=2, end_row=r + 2, end_column=10)
c = wsI.cell(r, 2, "ATENÇÃO — DADOS DE EXEMPLO: a planilha vem preenchida com 12 atletas fictícios, um macrociclo modelo e "
                   "registros de carga/wellness/presença apenas para demonstrar os cálculos e os gráficos. Apague esses dados "
                   "(linhas marcadas como EXEMPLO) antes de usar com a sua equipe. As fórmulas permanecem intactas.\n"
                   "AVISO: esta planilha é uma ferramenta de organização do treino. Não substitui avaliação médica nem "
                   "fisioterápica; os indicadores de risco são apoio à decisão do técnico/preparador físico.")
c.font = Font(name=F, size=9, bold=True, color=RED_T)
c.fill = PatternFill("solid", fgColor=RED); c.border = BORDER
c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)


# ============================================================================
# 3) CADASTRO DE ATLETAS  (identificação + socioeconômico + histórico + saúde)
#    Sem qualquer campo financeiro: nenhuma renda, salário ou valor de bolsa.
# ============================================================================
wsC = wb.create_sheet("Cadastro")
banner(wsC, "CADASTRO DE ATLETAS — FICHA COMPLETA",
       "Uma linha por atleta. Blocos: Identificação · Antropometria (puxada da aba Antropometria) · Contato · "
       "Socioeconômico · Histórico Esportivo · Saúde · Situação.", 65, NAVY2)
nota(wsC, 4, 2, "LGPD — esta aba armazena dados pessoais e de saúde (dados sensíveis). Colete somente o necessário, "
     "com consentimento por escrito do atleta (ou do responsável, se menor), e restrinja o acesso ao arquivo. "
     "Por decisão do usuário, NENHUM dado financeiro do atleta é registrado aqui.", 40)

GRUPOS = [("IDENTIFICAÇÃO", 1, 12, NAVY),
          ("ANTROPOMETRIA (puxada da aba Antropometria)", 13, 17, "1F6F4A"),
          ("CONTATO", 18, 25, BLUE3),
          ("SOCIOECONÔMICO", 26, 44, GOLD),
          ("HISTÓRICO ESPORTIVO", 45, 50, "6B3FA0"),
          ("SAÚDE", 51, 62, RED_T),
          ("SITUAÇÃO", 63, 65, GREY_T)]
for txt, c1, c2, cor in GRUPOS:
    wsC.merge_cells(start_row=6, start_column=c1, end_row=6, end_column=c2)
    c = wsC.cell(6, c1, txt)
    c.font = Font(name=F, size=9, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=cor)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = BORDER
wsC.row_dimensions[6].height = 18

CAD_H = [
 # IDENTIFICAÇÃO (1-12)
 "ID","Nome Completo","Nome na Equipe","Data de Nasc.","Idade","Sexo","Nacionalidade","Naturalidade (Cidade/UF)",
 "Documento (RG/CPF)","Nº Camisa","Categoria","Posição",
 # ANTROPOMETRIA automática (13-17)
 "Dominância de Mão","Perna de Impulsão","Estatura (cm)","Massa (kg)","IMC",
 # CONTATO (18-25)
 "Telefone","E-mail","Endereço","Cidade / UF","CEP","Contato de Emergência","Telefone de Emergência","Parentesco",
 # SOCIOECONÔMICO (26-44) — sem campos financeiros
 "Escolaridade","Situação de Estudo","Turno de Estudo","Instituição de Ensino","Trabalha?","Horas de Trabalho / Semana",
 "Ocupação","Nº de Pessoas no Domicílio","Classe Econômica (Critério Brasil)","Tipo de Moradia","Reside com",
 "Transporte para o Treino","Tempo de Deslocamento (min)","Recebe Bolsa / Auxílio?","Benefício Social",
 "Plano de Saúde","Acesso à Internet em Casa","Refeições por Dia","Acompanhamento Nutricional",
 # HISTÓRICO ESPORTIVO (45-50)
 "Idade de Início no Vôlei","Anos de Prática","Anos de Treino de Força","Nível Competitivo Máximo",
 "Clubes Anteriores","Seleções / Convocações",
 # SAÚDE (51-62)
 "Tipo Sanguíneo","Alergias","Medicamentos em Uso","Cirurgias Prévias","Lesões Prévias (região)",
 "Nº de Lesões (12 meses)","Queixa / Dor Atual","Data do Atestado Médico","Validade do Atestado",
 "Situação do Atestado","PAR-Q Respondido?","Usa Óculos / Lentes?",
 # SITUAÇÃO (63-65)
 "Data de Entrada na Equipe","Status","Observações",
 # auxiliar (66)
 "últ. aval."]
cab_tabela(wsC, 7, CAD_H)
larguras(wsC, {"A":9,"B":26,"C":16,"D":13,"E":7,"F":11,"G":13,"H":19,"I":16,"J":8,"K":10,"L":18,
               "M":14,"N":15,"O":12,"P":11,"Q":8,
               "R":15,"S":26,"T":30,"U":18,"V":12,"W":24,"X":18,"Y":13,
               "Z":22,"AA":18,"AB":14,"AC":24,"AD":11,"AE":15,"AF":20,"AG":13,"AH":20,"AI":16,"AJ":18,
               "AK":22,"AL":15,"AM":16,"AN":14,"AO":14,"AP":16,"AQ":12,"AR":18,
               "AS":15,"AT":12,"AU":15,"AV":22,"AW":26,"AX":22,
               "AY":11,"AZ":20,"BA":22,"BB":22,"BC":24,"BD":13,"BE":22,"BF":14,"BG":14,"BH":16,"BI":14,"BJ":14,
               "BK":15,"BL":14,"BM":26,"BN":11})

ANT = "Antropometria"
for r in range(CAD_F, CAD_L + 1):
    wsC.cell(r, 1,  '=IF($B{0}="","","ATL-"&TEXT(ROW()-{1},"000"))'.format(r, CAD_F - 1))
    wsC.cell(r, 5,  '=IFERROR(DATEDIF($D{0},TODAY(),"Y"),"")'.format(r))
    # última avaliação antropométrica do atleta (coluna auxiliar BN = 66)
    wsC.cell(r, 66, '=IF($B{0}="",0,IFERROR(SUMPRODUCT(MAX(({1}!$B$8:$B$407=$B{0})*{1}!$A$8:$A$407)),0))'.format(r, ANT))
    for col, src in ((15, "F"), (16, "E")):     # estatura, massa
        wsC.cell(r, col, '=IF(OR($B{0}="",$BN{0}=0),"",IFERROR(AVERAGEIFS({1}!${2}$8:${2}$407,'
                         '{1}!$B$8:$B$407,$B{0},{1}!$A$8:$A$407,$BN{0}),""))'.format(r, ANT, src))
    wsC.cell(r, 17, '=IFERROR(ROUND($P{0}/($O{0}/100)^2,1),"")'.format(r))
    wsC.cell(r, 46, '=IFERROR($E{0}-$AS{0},"")'.format(r))                        # anos de prática
    wsC.cell(r, 60, '=IF($BG{0}="","",IF($BG{0}<TODAY(),"VENCIDO",'               # situação do atestado
                    'IF($BG{0}<=TODAY()+30,"Vence em 30 dias","Válido")))'.format(r))
corpo_tabela(wsC, CAD_F, CAD_L, 1, 66)
CALCULADAS = (1, 5, 15, 16, 17, 46, 60, 66)
for r in range(CAD_F, CAD_L + 1):
    for c in range(1, 67):
        if c in CALCULADAS:
            wsC.cell(r, c).font = Font(name=F, size=9, bold=True, color=NAVY2)
            wsC.cell(r, c).fill = PatternFill("solid", fgColor=LIGHT)
        else:
            wsC.cell(r, c).fill = PatternFill("solid", fgColor=GOLD_L)
            wsC.cell(r, c).font = Font(name=F, size=9, color="0000FF")
    for c in (2, 8, 19, 20, 21, 29, 32, 48, 49, 50, 52, 53, 54, 55, 57, 65):
        wsC.cell(r, c).alignment = Alignment(horizontal="left", vertical="center", indent=1)
    for c in (4, 58, 59, 63):
        wsC.cell(r, c).number_format = NF_DATE
    wsC.cell(r, 15).number_format = '0.0;;""'
    wsC.cell(r, 16).number_format = '0.0;;""'
    wsC.cell(r, 17).number_format = '0.0;;""'
    wsC.cell(r, 66).number_format = NF_DATE
wsC.column_dimensions["BN"].hidden = True

DVS = [("Sexo","F"),("Categoria","K"),("Posição","L"),("Dominância","M"),("Perna de Impulsão","N"),
       ("Escolaridade","Z"),("Situação de Estudo","AA"),("Turno","AB"),("Sim/Não","AD"),
       ("Classe Econômica","AH"),("Tipo de Moradia","AI"),("Reside com","AJ"),("Transporte","AK"),
       ("Sim/Não","AM"),("Sim/Não","AN"),("Sim/Não","AO"),("Sim/Não","AP"),("Sim/Não","AR"),
       ("Nível Competitivo","AV"),("Tipo Sanguíneo","AY"),("Região Corporal","BC"),("Região Corporal","BE"),
       ("Sim/Não","BI"),("Sim/Não","BJ"),("Status do Atleta","BL")]
for lista, col in DVS:
    dv(wsC, L(lista), "{0}{1}:{0}{2}".format(col, CAD_F, CAD_L))
wsC.conditional_formatting.add("BL{}:BL{}".format(CAD_F, CAD_L),
    CellIsRule(operator="equal", formula=['"Ativo"'], fill=PatternFill("solid", fgColor=GREEN),
               font=Font(name=F, size=9, bold=True, color=GREEN_T)))
for txt in ('"Lesionado"', '"Departamento Médico"'):
    wsC.conditional_formatting.add("BL{}:BL{}".format(CAD_F, CAD_L),
        CellIsRule(operator="equal", formula=[txt], fill=PatternFill("solid", fgColor=RED),
                   font=Font(name=F, size=9, bold=True, color=RED_T)))
wsC.conditional_formatting.add("BH{}:BH{}".format(CAD_F, CAD_L),
    CellIsRule(operator="equal", formula=['"VENCIDO"'], fill=PatternFill("solid", fgColor=RED),
               font=Font(name=F, size=9, bold=True, color=RED_T)))
wsC.conditional_formatting.add("BH{}:BH{}".format(CAD_F, CAD_L),
    CellIsRule(operator="equal", formula=['"Vence em 30 dias"'], fill=PatternFill("solid", fgColor=YELL),
               font=Font(name=F, size=9, bold=True, color=YELL_T)))
wsC.conditional_formatting.add("BH{}:BH{}".format(CAD_F, CAD_L),
    CellIsRule(operator="equal", formula=['"Válido"'], fill=PatternFill("solid", fgColor=GREEN),
               font=Font(name=F, size=9, bold=True, color=GREEN_T)))
wsC.freeze_panes = "C8"
wsC.auto_filter.ref = "A7:BM{}".format(CAD_L)

# ---- 12 atletas de exemplo (ELASE Voleibol Masculino Adulto) ---------------
EXEMPLO = [
 ("Rafael Monteiro Alves","Rafa",1998,3,14,"Levantador",5,"Destro","Esquerda",196,
  "Superior em andamento","Cursando","Noturno","Não",0,"Atleta profissional",3,"B1","Própria","Família",
  "Carro próprio",25,"Sim","Não","Sim","Sim",6,"Sim",11,4,"Seleção estadual"),
 ("Diego Salgado Ferraz","Diego",1996,8,2,"Oposto",12,"Canhoto","Esquerda",202,
  "Superior completo","Concluído","—","Não",0,"Atleta profissional",2,"B1","Alugada","Cônjuge",
  "Carro próprio",20,"Sim","Não","Sim","Sim",6,"Sim",12,6,"Seleção nacional Sub-21"),
 ("Lucas Prado Bittencourt","Lucas",2000,1,27,"Ponteiro (Ponta)",7,"Destro","Esquerda",198,
  "Superior em andamento","Cursando","Matutino","Não",0,"Atleta profissional",4,"B2","Própria","Família",
  "Transporte do clube",35,"Sim","Não","Sim","Sim",5,"Sim",13,3,"Seleção estadual"),
 ("Bruno Rezende Camargo","Bruninho",1999,11,5,"Ponteiro (Ponta)",9,"Destro","Esquerda",197,
  "Superior em andamento","Cursando","Noturno","Sim",20,"Estagiário",4,"C1","Alugada","Família",
  "Transporte público",55,"Sim","Sim","Não","Sim",4,"Não",12,3,"—"),
 ("Thiago Nogueira Vasques","Thiago",1995,5,19,"Central",4,"Destro","Direita",205,
  "Superior completo","Concluído","—","Não",0,"Atleta profissional",3,"A","Própria","Família",
  "Carro próprio",18,"Sim","Não","Sim","Sim",6,"Sim",10,8,"Seleção nacional adulta"),
 ("Matheus Caldeira Lins","Matheus",2001,9,8,"Central",15,"Destro","Esquerda",203,
  "Superior em andamento","Cursando","Noturno","Não",0,"Atleta profissional",5,"C1","Cedida","Alojamento do clube",
  "A pé",10,"Sim","Sim","Sim","Sim",5,"Sim",14,2,"—"),
 ("Felipe Andrade Rocha","Felipão",2003,2,11,"Líbero",2,"Destro","Direita",184,
  "Superior em andamento","Cursando","Matutino","Não",0,"Atleta profissional",4,"C2","Alugada","Alojamento do clube",
  "Transporte do clube",30,"Sim","Sim","Não","Sim",5,"Sim",12,2,"Seleção estadual Sub-19"),
 ("Gustavo Peixoto Maia","Gu",1997,7,23,"Ponteiro (Ponta)",11,"Destro","Esquerda",199,
  "Superior completo","Concluído","—","Não",0,"Atleta profissional",2,"B1","Própria","Cônjuge",
  "Carro próprio",22,"Sim","Não","Sim","Sim",6,"Sim",11,7,"—"),
 ("Vinícius Barreto Duarte","Vini",2002,12,3,"Levantador",6,"Destro","Esquerda",192,
  "Superior em andamento","Cursando","Noturno","Não",0,"Atleta profissional",5,"C2","Cedida","Alojamento do clube",
  "A pé",8,"Sim","Sim","Não","Sim",5,"Sim",13,2,"—"),
 ("André Luiz Sampaio","Dedé",1994,4,30,"Oposto",18,"Destro","Esquerda",201,
  "Superior completo","Concluído","—","Não",0,"Atleta profissional",4,"A","Própria","Família",
  "Carro próprio",15,"Sim","Não","Sim","Sim",6,"Sim",10,9,"Seleção nacional adulta"),
 ("Pedro Henrique Coutinho","PH",2004,6,17,"Central",3,"Destro","Direita",204,
  "Ensino Médio completo","Concluído","—","Não",0,"Atleta em formação",6,"D-E","Cedida","Alojamento do clube",
  "A pé",5,"Sim","Sim","Não","Sim",5,"Sim",13,1,"Seleção estadual Sub-21"),
 ("Caio Fernandes Bastos","Caio",2001,10,9,"Líbero",1,"Canhoto","Direita",186,
  "Superior em andamento","Cursando","Matutino","Sim",12,"Auxiliar administrativo",3,"C1","Alugada","Família",
  "Transporte público",45,"Não","Não","Sim","Sim",4,"Não",12,3,"—"),
]
LESOES_EX = [("—",0,"—"),("Entorse de tornozelo D",1,"—"),("—",0,"—"),
             ("Tendinopatia patelar",2,"Dor anterior no joelho D (leve)"),("Lombalgia",1,"—"),
             ("—",0,"—"),("Entorse de tornozelo E",1,"—"),("Tendinopatia de ombro D",1,"—"),
             ("—",0,"—"),("Cirurgia de ombro D (2021)",0,"—"),("—",0,"—"),("Entorse de dedo",1,"—")]
for i, a in enumerate(EXEMPLO):
    r = CAD_F + i
    (nome, apel, ay, am, ad, pos, cam, dom, perna, est, esc, sit, turno, trab, horas, ocup,
     pes, classe, mor, reside, transp, desloc, bolsa, benef, plano, net, refs, nutri,
     ini, forca, selec) = a
    V = {2:nome, 3:apel, 4:date(ay,am,ad), 6:"Masculino", 7:"Brasileira", 8:"—", 9:"—", 10:cam,
         11:"Adulto", 12:pos, 13:dom, 14:perna,
         18:"(00) 90000-00{:02d}".format(i+1), 19:"atleta{}@elase.com.br".format(i+1), 20:"—",
         21:"—", 22:"—", 23:"Responsável / familiar", 24:"(00) 90000-10{:02d}".format(i+1), 25:"—",
         26:esc, 27:sit, 28:turno, 29:"—", 30:trab, 31:horas, 32:ocup, 33:pes,
         34:classe, 35:mor, 36:reside, 37:transp, 38:desloc, 39:bolsa, 40:benef,
         41:plano, 42:net, 43:refs, 44:nutri, 45:ini, 47:forca,
         48:"Superliga B" if i % 3 else "Superliga A", 49:"—", 50:selec,
         51:"O+", 52:"—", 53:"—", 54:LESOES_EX[i][0] if "Cirurgia" in LESOES_EX[i][0] else "—",
         55:LESOES_EX[i][0], 56:LESOES_EX[i][1], 57:LESOES_EX[i][2],
         58:date(2026,1,12), 59:date(2027,1,12), 61:"Sim", 62:"Não" if i % 4 else "Sim",
         63:date(2026,1,5), 64:"Lesionado" if i == 3 else "Ativo", 65:"EXEMPLO — substituir"}
    for c, v in V.items():
        wsC.cell(r, c, v)
        if isinstance(v, date):
            wsC.cell(r, c).number_format = NF_DATE
NOMES = [a[0] for a in EXEMPLO]
EST_EX = {a[0]: a[9] for a in EXEMPLO}
POS_EX = {a[0]: a[5] for a in EXEMPLO}


# ============================================================================
# 4) BIBLIOTECA DE EXERCÍCIOS
# ============================================================================
wsE = wb.create_sheet("Exercícios")
banner(wsE, "BIBLIOTECA DE EXERCÍCIOS", "Cada exercício cadastrado aqui fica disponível no menu suspenso da aba Prescrição, "
       "que puxa automaticamente categoria, fundamento e capacidade física.", 14, BLUE3)
EX_H = ["ID","Exercício","Categoria","Fundamento","Capacidade Física","Objetivo","Descrição / Execução",
        "Material","Nº de Atletas","Duração Sug. (min)","Intensidade Sug.","PSE Sug.","Nível","Link de Vídeo"]
cab_tabela(wsE, 6, EX_H)
larguras(wsE, {"A":8,"B":38,"C":18,"D":18,"E":22,"F":34,"G":58,"H":26,"I":11,"J":11,"K":14,"L":8,"M":14,"N":24})
for r in range(EX_F, EX_L + 1):
    wsE.cell(r, 1, '=IF($B{0}="","","EX-"&TEXT(ROW()-{1},"000"))'.format(r, EX_F - 1))
corpo_tabela(wsE, EX_F, EX_L, 1, 14)
for r in range(EX_F, EX_L + 1):
    for c in range(2, 15):
        wsE.cell(r, c).fill = PatternFill("solid", fgColor=GOLD_L)
        wsE.cell(r, c).font = Font(name=F, size=9, color="0000FF")
    for c in (2, 6, 7, 8, 14):
        wsE.cell(r, c).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    wsE.cell(r, 1).font = Font(name=F, size=9, bold=True, color=NAVY2)
    wsE.cell(r, 1).fill = PatternFill("solid", fgColor=LIGHT)
dv(wsE, L("Categoria Exerc."), "C{}:C{}".format(EX_F, EX_L))
dv(wsE, L("Fundamento"),       "D{}:D{}".format(EX_F, EX_L))
dv(wsE, L("Capacidade Física"),"E{}:E{}".format(EX_F, EX_L))
dv(wsE, L("Intensidade"),      "K{}:K{}".format(EX_F, EX_L))
dv(wsE, L("PSE (0-10)"),       "L{}:L{}".format(EX_F, EX_L))
dv(wsE, L("Nível"),            "M{}:M{}".format(EX_F, EX_L))
wsE.freeze_panes = "C7"
wsE.auto_filter.ref = "A6:N{}".format(EX_L)

T,TA,FI,PV,RC,CG = "Técnico","Tático","Físico","Preventivo/Compensatório","Recuperação","Cognitivo/Visual"
EXERCICIOS = [
("Saque flutuante de apoio",T,"Saque","Coordenação","Precisão e constância do saque flutuante","Séries de 10 saques por zona alvo (1, 5 e 6). Contato firme no centro da bola, sem rotação, braço freado no impacto.","Bolas, cones/alvos",1,12,"Moderada",4,"Iniciante",""),
("Saque viagem no fundo (jump float)",T,"Saque","Coordenação","Saque flutuante com salto e maior velocidade","Passada de 2–3 tempos, salto vertical, contato seco. 3 blocos de 8 saques com alvo no fundo da quadra.","Bolas",1,12,"Alta",5,"Intermediário",""),
("Saque potente com salto",T,"Saque","Força Explosiva/Potência","Velocidade e agressividade no saque","4 séries de 6 saques em potência máxima, alternando zona 1 e zona 5. Pausa completa entre séries.","Bolas",1,15,"Muito Alta",7,"Avançado",""),
("Recepção em duplas com bola dirigida",T,"Recepção (Passe)","Coordenação","Plataforma estável e passe ao alvo","Técnico lança bolas alternadas; atleta recebe para o alvo em zona 2/3. 3 x 15 repetições.","Bolas, alvo/cesto",2,12,"Moderada",4,"Iniciante",""),
("Recepção de saque em sistema de 2 passadores",T,"Recepção (Passe)","Coordenação","Leitura de trajetória e responsabilidade de zona","Sacador real do outro lado. Passadores em L; 20 saques por rodízio, meta de 70% de passes A/B.","Bolas, rede",4,20,"Alta",6,"Intermediário",""),
("Manchete de deslocamento lateral (leque)",T,"Recepção (Passe)","Agilidade/COD","Deslocamento e recepção fora da base","Bolas alternadas à direita e à esquerda; deslocamento em passo lateral e retorno à base. 4 x 30 s.","Bolas",1,10,"Alta",6,"Intermediário",""),
("Levantamento em suspensão contra a parede",T,"Levantamento","Coordenação","Consistência do toque e ação simétrica das mãos","3 x 30 toques contra a parede a 3 m, mantendo a bola sem rotação.","Bola, parede",1,8,"Baixa",3,"Iniciante",""),
("Levantamento em rede com alvos",T,"Levantamento","Coordenação","Precisão de bola alta, rápida e de fundo","Levantador recebe passe do técnico e distribui para alvos em P4, P2 e Pipe. 5 x 12 levantamentos.","Bolas, aros/alvos",2,18,"Moderada",5,"Intermediário",""),
("Levantamento invertido (costas)",T,"Levantamento","Coordenação","Domínio do levantamento para trás sem sinalizar","3 x 15 levantamentos de costas para P2 com o técnico corrigindo a extensão de quadril.","Bolas",2,12,"Moderada",5,"Avançado",""),
("Ataque de bola alta pela ponta (P4)",T,"Ataque","Força Explosiva/Potência","Ritmo de aproximação e ataque em bola alta","4 x 10 ataques com levantador real. Foco na passada de 3 tempos e no braço em arco.","Bolas, rede",3,18,"Alta",6,"Intermediário",""),
("Ataque de primeiro tempo (P3)",T,"Ataque","Velocidade","Sincronia central x levantador","4 x 8 ataques de bola rápida, variando o ponto de contato ao longo da rede.","Bolas, rede",3,15,"Alta",6,"Avançado",""),
("Ataque de fundo (Pipe / P6)",T,"Ataque","Força Explosiva/Potência","Ataque a partir da linha de 3 m","3 x 10 ataques de fundo saindo da recepção. Atenção à marca dos 3 m.","Bolas, rede",3,15,"Alta",6,"Avançado",""),
("Largada e explorada de bloqueio",T,"Ataque","Coordenação","Variação tática do ataque","3 x 12 alternando largada curta, exploração de bloqueio e ataque forte, comandado pelo técnico.","Bolas, rede",3,12,"Moderada",5,"Intermediário",""),
("Bloqueio individual com deslocamento lateral",T,"Bloqueio","Agilidade/COD","Deslocamento e penetração das mãos","4 x 8 deslocamentos P2→P3→P4 com salto de bloqueio em cada parada.","Rede, plataforma",1,12,"Alta",6,"Intermediário",""),
("Bloqueio duplo (fechamento ponta/central)",T,"Bloqueio","Agilidade/COD","Sincronia e fechamento do bloqueio duplo","Central desloca e fecha com o ponteiro; 4 x 10 repetições com atacante real.","Rede, bolas",4,15,"Alta",6,"Avançado",""),
("Defesa de mergulho e rolamento",T,"Defesa","Coordenação","Técnica de queda segura e recuperação de bola","3 x 10 mergulhos alternando lados, com progressão de colchonete para quadra.","Bolas, colchonete",1,12,"Alta",6,"Intermediário",""),
("Defesa em 3 zonas com técnico atacando",T,"Defesa","Agilidade/COD","Postura de defesa e leitura do atacante","Técnico ataca da plataforma para zonas 1, 6 e 5; 4 x 45 s de defesa contínua.","Bolas, plataforma",3,15,"Muito Alta",7,"Avançado",""),
("Circuito de deslocamentos específicos da posição",T,"Deslocamento","Agilidade/COD","Automatizar trajetos de saída de rede e cobertura","Circuito de cones reproduzindo o trajeto real da posição. 5 voltas com 60 s de pausa.","Cones",1,12,"Alta",6,"Intermediário",""),
("Complexo I (K1) — recepção, levantamento e ataque",TA,"Jogo/Coletivo","N/A","Eficiência do side-out","Saque adversário real; a equipe pontua se converter o side-out. Meta de 60% em 20 tentativas.","Bolas, rede",12,25,"Alta",7,"Intermediário",""),
("Complexo II (K2) — bloqueio, defesa e contra-ataque",TA,"Jogo/Coletivo","N/A","Transição defesa-ataque","Técnico ataca; equipe defende e finaliza o contra-ataque. 20 séries por rodízio.","Bolas, rede",12,25,"Alta",7,"Intermediário",""),
("Jogo 6x6 com pontuação diferenciada",TA,"Jogo/Coletivo","N/A","Aplicar o sistema em contexto real","Sets a 15 com pontos em dobro para ponto de bloqueio ou side-out no 1º toque.","Bolas, rede, placar",12,30,"Alta",7,"Intermediário",""),
("Jogo reduzido 4x4 em quadra estreita",TA,"Jogo/Coletivo","N/A","Volume de toques e leitura de jogo","3 sets a 11 pontos em meia quadra, obrigando 3 toques.","Bolas, rede",8,20,"Moderada",6,"Iniciante",""),
("Treino do sistema 5x1 — rodízios e coberturas",TA,"Jogo/Coletivo","N/A","Posicionamento em cada rodízio","Sem bola primeiro, depois com bola: 6 rodízios x 4 repetições de recepção e cobertura.","Rede, bolas",12,20,"Baixa",4,"Intermediário",""),
("Saque-recepção sob pressão de placar",TA,"Jogo/Coletivo","N/A","Desempenho técnico sob estresse competitivo","Placar simulado em 22x22; erro de saque ou recepção custa ponto. 6 rodadas.","Bolas, placar",12,20,"Muito Alta",8,"Avançado",""),
("Agachamento livre (back squat)",FI,"N/A","Força Máxima","Força máxima de membros inferiores","4 x 5 a 80–85% de 1RM, 3 min de pausa. Progressão semanal de 2,5%.","Barra, anilhas, rack",1,25,"Muito Alta",8,"Avançado",""),
("Levantamento terra romeno",FI,"N/A","Força Máxima","Força de cadeia posterior e prevenção de isquiotibiais","4 x 8 a 70% de 1RM, ênfase excêntrica de 3 s.","Barra, anilhas",1,18,"Alta",7,"Intermediário",""),
("Afundo com halteres",FI,"N/A","Força Máxima","Força unilateral e estabilidade de quadril","3 x 10 por perna com carga moderada.","Halteres",1,15,"Alta",6,"Intermediário",""),
("Power clean (levantamento olímpico)",FI,"N/A","Força Explosiva/Potência","Taxa de produção de força","5 x 3 a 70–80% de 1RM, foco em velocidade da barra.","Barra, plataforma",1,25,"Muito Alta",8,"Avançado",""),
("Agachamento com salto sob carga (jump squat)",FI,"N/A","Força Explosiva/Potência","Potência específica de salto","5 x 4 com 20–30% de 1RM, pausa completa de 2 min.","Barra leve/halteres",1,18,"Alta",7,"Avançado",""),
("Salto em profundidade (drop jump)",FI,"N/A","Pliometria","Ciclo alongamento-encurtamento rápido","5 x 5 quedas de caixa de 40 cm com salto imediato. Superfície firme e calçado adequado.","Caixa pliométrica",1,15,"Alta",7,"Avançado",""),
("Saltos consecutivos sobre barreiras",FI,"N/A","Pliometria","Rigidez muscular e reatividade","4 x 6 barreiras de 40–50 cm, contato mínimo com o solo.","Barreiras",1,12,"Alta",7,"Intermediário",""),
("Sprints de 10 e 20 m com mudança de direção",FI,"N/A","Velocidade","Aceleração e frenagem","6 sprints de 10 m e 4 de 20 m com 90 s de pausa.","Cones, cronômetro",1,15,"Muito Alta",7,"Intermediário",""),
("Circuito de agilidade T-Test",FI,"N/A","Agilidade/COD","Mudança de direção específica da quadra","4 execuções completas com 2 min de pausa; registrar o melhor tempo.","Cones, cronômetro",1,12,"Alta",6,"Intermediário",""),
("Intervalado de alta intensidade 15/15",FI,"N/A","Resistência Anaeróbia","Tolerância a esforços intermitentes","2 blocos de 8 x 15 s de corrida intensa / 15 s de trote, 3 min entre blocos.","Cronômetro",1,20,"Muito Alta",8,"Intermediário",""),
("Corrida contínua regenerativa",FI,"N/A","Resistência Aeróbia","Recuperação ativa e base aeróbia","20–25 min em intensidade leve (60–65% da FCmáx).","Cronômetro",1,25,"Baixa",3,"Iniciante",""),
("Prancha e anti-rotação (Pallof press)",FI,"N/A","Core/Estabilidade","Estabilidade de tronco para transferência de força","3 x 40 s de prancha + 3 x 12 Pallof por lado.","Elástico, colchonete",1,12,"Moderada",4,"Iniciante",""),
("Rotadores do ombro com elástico",PV,"N/A","Core/Estabilidade","Prevenção de lesão no ombro do atacante","3 x 15 rotação externa e interna, carga leve, ritmo controlado.","Elástico",1,10,"Baixa",3,"Iniciante",""),
("Nórdico de isquiotibiais (excêntrico)",PV,"N/A","Força Máxima","Prevenção de lesão de isquiotibiais","3 x 6 repetições excêntricas lentas, 2x/semana.","Colchonete, parceiro",2,10,"Alta",6,"Intermediário",""),
("Excêntrico de tornozelo e panturrilha",PV,"N/A","Força Máxima","Prevenção de entorse e tendinopatia de aquileu","3 x 12 elevações com descida lenta de 3 s no step.","Step",1,8,"Moderada",4,"Iniciante",""),
("Mobilidade de tornozelo e quadril",PV,"N/A","Mobilidade/Flexibilidade","Amplitude para agachamento e aterrissagem","6 exercícios x 8 repetições, antes da parte principal.","Colchonete, elástico",1,10,"Muito Baixa",2,"Iniciante",""),
("Alongamento e liberação miofascial",RC,"N/A","Mobilidade/Flexibilidade","Recuperação pós-sessão","15 min de rolo e alongamentos estáticos de 30 s por grupamento.","Rolo, colchonete",1,15,"Muito Baixa",2,"Iniciante",""),
("Reação visual com estímulo do técnico",CG,"Defesa","Coordenação","Tempo de reação e leitura antecipatória","Técnico sinaliza a direção no último instante; atleta desloca e defende. 5 x 30 s.","Bolas, sinalizadores",1,10,"Alta",6,"Intermediário",""),
]
for i, e in enumerate(EXERCICIOS):
    r = EX_F + i
    for j, v in enumerate(e):
        wsE.cell(r, 2 + j, v)

# ============================================================================
# 5) MACROCICLO
# ============================================================================
wsM = wb.create_sheet("Macrociclo")
banner(wsM, "MACROCICLO — PLANEJAMENTO DA TEMPORADA",
       "Defina o período total e divida-o em mesociclos. A Data de Início (C11) é a referência de TODA a numeração de semanas da planilha.", 15, NAVY)
larguras(wsM, {"A":6,"B":26,"C":18,"D":14,"E":14,"F":9,"G":10,"H":11,"I":12,"J":24,"K":24,"L":24,"M":14,"N":24,"O":13})

secao(wsM, 4, "IDENTIFICAÇÃO DO MACROCICLO", 15, 2)
IDENT = [("Temporada / Época","2026"),("Equipe / Clube","Equipe Exemplo de Voleibol"),("Categoria","Adulto Feminino"),
         ("Técnico Responsável","(preencher)"),("Preparador Físico","(preencher)"),
         ("Data de Início", date(2026,1,5)),("Data de Término", date(2026,11,29)),("Duração (semanas)", None),
         ("Objetivo Principal","Chegar ao 1º turno do estadual com alto nível técnico-tático e baixo índice de lesões"),
         ("Competição-Alvo","Campeonato Estadual Adulto — fase final em outubro/2026")]
for i, (lab, val) in enumerate(IDENT):
    r = 6 + i
    rotulo(wsM, r, 2, lab)
    if lab == "Duração (semanas)":
        c = calc(wsM, r, 3, '=IFERROR(ROUNDUP(($C$12-$C$11+1)/7,0),"")', '0" semanas"')
        wsM.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
    else:
        c = entrada(wsM, r, 3, val, NF_DATE if isinstance(val, date) else None, largura_merge=4)
    wsM.row_dimensions[r].height = 18
wsM["C11"].comment = Comment("Data de Início do Macrociclo.\nÉ a âncora do cálculo 'Semana nº' usado nas abas "
                             "Prescrição, Carga (PSE), Wellness, Presença, Atleta e Painel.\nUse preferencialmente uma "
                             "segunda-feira.", "Planilha")

secao(wsM, 17, "DIVISÃO EM MESOCICLOS", 15, 1)
MAC_H = ["Nº","Mesociclo","Período","Data Início","Data Fim","Semanas","Semana Inicial","Volume (%)","Intensidade (%)",
         "Ênfase Física","Ênfase Técnica","Ênfase Tática","Carga-Alvo Semanal (UA)","Competições no Período","Situação"]
cab_tabela(wsM, 18, MAC_H)
MESO_F, MESO_L = 19, 30
for r in range(MESO_F, MESO_L + 1):
    wsM.cell(r, 1, '=IF($B{0}="","",ROW()-{1})'.format(r, MESO_F - 1))
    wsM.cell(r, 6, '=IF(OR($D{0}="",$E{0}=""),0,ROUNDUP(($E{0}-$D{0}+1)/7,0))'.format(r))
    wsM.cell(r, 7, '=IF(OR($D{0}="",$C$11=""),0,INT(($D{0}-$C$11)/7)+1)'.format(r))
corpo_tabela(wsM, MESO_F, MESO_L, 1, 15)
for r in range(MESO_F, MESO_L + 1):
    for c in (2, 3, 4, 5, 8, 9, 10, 11, 12, 13, 14, 15):
        wsM.cell(r, c).fill = PatternFill("solid", fgColor=GOLD_L)
        wsM.cell(r, c).font = Font(name=F, size=9, color="0000FF")
    for c in (1, 6, 7):
        wsM.cell(r, c).font = Font(name=F, size=9, bold=True, color=NAVY2)
        wsM.cell(r, c).fill = PatternFill("solid", fgColor=LIGHT)
    for c in (2, 10, 11, 12, 14):
        wsM.cell(r, c).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    wsM.cell(r, 4).number_format = NF_DATE
    wsM.cell(r, 5).number_format = NF_DATE
    wsM.cell(r, 6).number_format = '0;;""'
    wsM.cell(r, 7).number_format = '0;;""'
    wsM.cell(r, 8).number_format = NF_PCT
    wsM.cell(r, 9).number_format = NF_PCT
    wsM.cell(r, 13).number_format = NF_UA
tot = MESO_L + 1
wsM.cell(tot, 2, "TOTAL / MÉDIA").font = Font(name=F, size=9, bold=True, color=WHITE)
for c in range(1, 16):
    wsM.cell(tot, c).fill = PatternFill("solid", fgColor=NAVY2); wsM.cell(tot, c).border = BORDER
    wsM.cell(tot, c).font = Font(name=F, size=9, bold=True, color=WHITE)
    wsM.cell(tot, c).alignment = Alignment(horizontal="center", vertical="center")
wsM.cell(tot, 2).alignment = Alignment(horizontal="left", vertical="center", indent=1)
wsM.cell(tot, 6, "=SUM(F{}:F{})".format(MESO_F, MESO_L)).number_format = "0"
wsM.cell(tot, 8, '=IFERROR(AVERAGE(H{}:H{}),"")'.format(MESO_F, MESO_L)).number_format = NF_PCT
wsM.cell(tot, 9, '=IFERROR(AVERAGE(I{}:I{}),"")'.format(MESO_F, MESO_L)).number_format = NF_PCT
wsM.cell(tot, 13, '=IFERROR(SUMPRODUCT($F{0}:$F{1},$M{0}:$M{1}),"")'.format(MESO_F, MESO_L)).number_format = NF_UA
nota(wsM, tot + 1, 2, "Confira: a soma de semanas dos mesociclos deve bater com a Duração (semanas) em C13. "
     "A Carga-Alvo total é a soma de (semanas × carga-alvo semanal) de cada mesociclo.", 15)
dv(wsM, L("Período do Macro"), "C{}:C{}".format(MESO_F, MESO_L))
dv(wsM, L("Situação"),         "O{}:O{}".format(MESO_F, MESO_L))
wsM.freeze_panes = "C19"

MESOS = [
 ("Meso 1 — Base Geral","Preparatório Geral",date(2026,1,5),date(2026,2,15),0.90,0.55,
  "Força máxima, base aeróbia e core","Fundamentos individuais e repetição técnica","Reconhecimento do sistema 5x1",2600,"—","Concluído"),
 ("Meso 2 — Base Específica","Preparatório Específico",date(2026,2,16),date(2026,3,29),0.85,0.70,
  "Força-potência e pliometria introdutória","Técnica sob deslocamento e sob fadiga","Complexos K1 e K2",2900,"Amistosos preparatórios","Concluído"),
 ("Meso 3 — Pré-Competitivo","Pré-Competitivo",date(2026,3,30),date(2026,5,10),0.70,0.85,
  "Potência, pliometria e velocidade","Técnica em situação de jogo","Sistemas de defesa e coberturas",2700,"Copa regional","Concluído"),
 ("Meso 4 — Competitivo I","Competitivo I",date(2026,5,11),date(2026,7,19),0.55,0.90,
  "Manutenção de força e potência","Correção pontual e saque sob pressão","Ajustes de scout adversário",2200,"Estadual — 1º turno","Em execução"),
 ("Meso 5 — Competitivo II","Competitivo II",date(2026,7,20),date(2026,10,11),0.50,0.95,
  "Manutenção e prevenção de lesões","Refinamento técnico individual","Preparação por adversário",2000,"Estadual — 2º turno e playoffs","Planejado"),
 ("Meso 6 — Transição","Transição",date(2026,10,12),date(2026,11,29),0.35,0.40,
  "Recuperação ativa e trabalho compensatório","Atividades lúdicas e técnica livre","Avaliação da temporada",900,"—","Planejado"),
]
for i, m in enumerate(MESOS):
    r = MESO_F + i
    nome, per, di, df, vol, inten, ef, et, eta, carga, comp, sit = m
    wsM.cell(r, 2, nome); wsM.cell(r, 3, per); wsM.cell(r, 4, di); wsM.cell(r, 5, df)
    wsM.cell(r, 8, vol); wsM.cell(r, 9, inten); wsM.cell(r, 10, ef); wsM.cell(r, 11, et)
    wsM.cell(r, 12, eta); wsM.cell(r, 13, carga); wsM.cell(r, 14, comp); wsM.cell(r, 15, sit)
    wsM.cell(r, 4).number_format = NF_DATE; wsM.cell(r, 5).number_format = NF_DATE
    wsM.cell(r, 8).number_format = NF_PCT;  wsM.cell(r, 9).number_format = NF_PCT
    wsM.cell(r, 13).number_format = NF_UA

ch = LineChart(); ch.title = "Dinâmica da Carga: Volume × Intensidade por Mesociclo"
ch.style = 12; ch.height = 8.5; ch.width = 22; ch.y_axis.title = "% relativo"
data = Reference(wsM, min_col=8, max_col=9, min_row=18, max_row=MESO_L)
cats = Reference(wsM, min_col=2, min_row=MESO_F, max_row=MESO_L)
ch.add_data(data, titles_from_data=True); ch.set_categories(cats)
ch.series[0].graphicalProperties.line.width = 28000
ch.series[1].graphicalProperties.line.width = 28000
wsM.add_chart(ch, "B34")

# ============================================================================
# 6) MESOCICLO
# ============================================================================
wsS = wb.create_sheet("Mesociclo")
banner(wsS, "MESOCICLO — DETALHAMENTO EM MICROCICLOS",
       "Escolha o mesociclo no menu suspenso: os dados do planejamento são puxados do Macrociclo automaticamente.", 13, NAVY2)
larguras(wsS, {"A":6,"B":24,"C":22,"D":14,"E":14,"F":40,"G":11,"H":13,"I":13,"J":14,"K":14,"L":12,"M":30})

secao(wsS, 4, "SELEÇÃO E RESUMO DO MESOCICLO", 13, 2)
rotulo(wsS, 6, 2, "Mesociclo selecionado")
sel = entrada(wsS, 6, 3, MESOS[4][0], largura_merge=2)
sel.font = Font(name=F, size=11, bold=True, color="0000FF")
dv(wsS, MESO_REF, "C6")

def idx_macro(col_letter):
    return '=IFERROR(INDEX(Macrociclo!${0}${1}:${0}${2},MATCH($C$6,Macrociclo!$B${1}:$B${2},0)),"")'.format(col_letter, MESO_F, MESO_L)

RESUMO = [("Período",("C",None)),("Data de Início",("D",NF_DATE)),("Data de Término",("E",NF_DATE)),
          ("Duração (semanas)",("F","0")),("Semana Inicial no Macro",("G","0")),("Volume Planejado",("H",NF_PCT)),
          ("Intensidade Planejada",("I",NF_PCT)),("Carga-Alvo Semanal (UA)",("M",NF_UA))]
for i, (lab, (cl, nf)) in enumerate(RESUMO):
    r = 8 + i
    rotulo(wsS, r, 2, lab)
    c = calc(wsS, r, 3, idx_macro(cl), nf)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    wsS.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
for i, (lab, cl) in enumerate([("Ênfase Física","J"),("Ênfase Técnica","K"),("Ênfase Tática","L"),("Competições","N")]):
    r = 8 + i
    rotulo(wsS, r, 6, lab)
    c = calc(wsS, r, 7, idx_macro(cl))
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    wsS.merge_cells(start_row=r, start_column=7, end_row=r, end_column=13)

secao(wsS, 17, "MICROCICLOS DO MESOCICLO — PREVISTO × REALIZADO", 13, 1)
MIC_H = ["Nº","Microciclo","Tipo de Microciclo","Data Início","Data Fim","Objetivo do Microciclo","Semana nº",
         "Sessões Previstas","Volume Previsto (min)","PSE Média Prevista","Carga Prevista (UA)","% Cumprido","Observações"]
cab_tabela(wsS, 18, MIC_H)
MIC_F, MIC_L = 19, 30
for r in range(MIC_F, MIC_L + 1):
    wsS.cell(r, 1, '=IF($B{0}="","",ROW()-{1})'.format(r, MIC_F - 1))
    wsS.cell(r, 7, '=IF(OR($D{0}="",Macrociclo!$C$11=""),0,INT(($D{0}-Macrociclo!$C$11)/7)+1)'.format(r))
    wsS.cell(r, 11, '=IF(OR($I{0}="",$J{0}=""),0,$I{0}*$J{0})'.format(r))
    wsS.cell(r, 12, '=IF($K{0}=0,0,SUMIFS(\'Carga (PSE)\'!$G${1}:$G${2},\'Carga (PSE)\'!$B${1}:$B${2},$G{0})'
                    '/MAX(1,COUNTIF(Cadastro!$BL${3}:$BL${4},"Ativo"))/$K{0})'.format(r, CAR_F, CAR_L, CAD_F, CAD_L))
corpo_tabela(wsS, MIC_F, MIC_L, 1, 13)
for r in range(MIC_F, MIC_L + 1):
    for c in (2, 3, 4, 5, 6, 8, 9, 10, 13):
        wsS.cell(r, c).fill = PatternFill("solid", fgColor=GOLD_L)
        wsS.cell(r, c).font = Font(name=F, size=9, color="0000FF")
    for c in (1, 7, 11, 12):
        wsS.cell(r, c).font = Font(name=F, size=9, bold=True, color=NAVY2)
        wsS.cell(r, c).fill = PatternFill("solid", fgColor=LIGHT)
    for c in (2, 6, 13):
        wsS.cell(r, c).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    wsS.cell(r, 4).number_format = NF_DATE; wsS.cell(r, 5).number_format = NF_DATE
    wsS.cell(r, 7).number_format = '0;;""'
    wsS.cell(r, 11).number_format = NF_UA;  wsS.cell(r, 12).number_format = NF_PCT
dv(wsS, L("Tipo de Microciclo"), "C{}:C{}".format(MIC_F, MIC_L))
wsS.conditional_formatting.add("L{}:L{}".format(MIC_F, MIC_L),
    CellIsRule(operator="between", formula=["0.9","1.1"], fill=PatternFill("solid", fgColor=GREEN), font=Font(name=F, size=9, bold=True, color=GREEN_T)))
wsS.conditional_formatting.add("L{}:L{}".format(MIC_F, MIC_L),
    CellIsRule(operator="greaterThan", formula=["1.1"], fill=PatternFill("solid", fgColor=YELL), font=Font(name=F, size=9, bold=True, color=YELL_T)))
wsS.conditional_formatting.add("L{}:L{}".format(MIC_F, MIC_L),
    CellIsRule(operator="between", formula=["0.0001","0.9"], fill=PatternFill("solid", fgColor=RED), font=Font(name=F, size=9, bold=True, color=RED_T)))
nota(wsS, MIC_L + 2, 2, "% Cumprido = carga média por atleta ativo realizada na semana (aba Carga (PSE)) ÷ carga prevista do microciclo. "
     "Verde = 90–110% do previsto; amarelo = acima; vermelho = abaixo.", 13)
wsS.freeze_panes = "C19"

MICROS = [("Micro 5.1","Ordinário",date(2026,8,10),date(2026,8,16),"Retomada do volume após a rodada; ênfase em side-out",7,530,6),
          ("Micro 5.2","Choque",date(2026,8,17),date(2026,8,23),"Pico de carga física da fase, mantendo o volume técnico",7,530,7),
          ("Micro 5.3","Ordinário",date(2026,8,24),date(2026,8,30),"Consolidação técnico-tática do sistema defensivo",7,530,6),
          ("Micro 5.4","Recuperativo",date(2026,8,31),date(2026,9,6),"Redução de volume; recuperação e preparação do jogo",5,380,5),
          ("Micro 5.5","Pré-Competitivo",date(2026,9,7),date(2026,9,13),"Polimento técnico e preparação por adversário",5,360,5)]
for i, m in enumerate(MICROS):
    r = MIC_F + i
    nome, tipo, di, df, obj, ses, vol, pse = m
    wsS.cell(r, 2, nome); wsS.cell(r, 3, tipo)
    if di: wsS.cell(r, 4, di); wsS.cell(r, 4).number_format = NF_DATE
    if df: wsS.cell(r, 5, df); wsS.cell(r, 5).number_format = NF_DATE
    wsS.cell(r, 6, obj); wsS.cell(r, 8, ses); wsS.cell(r, 9, vol); wsS.cell(r, 10, pse)

# ============================================================================
# 7) MICROCICLO
# ============================================================================
wsW = wb.create_sheet("Microciclo")
banner(wsW, "MICROCICLO — PLANEJAMENTO DA SEMANA",
       "Duas linhas por dia (duas sessões possíveis). As datas e os dias da semana são gerados a partir da Data de Início.", 12, BLUE3)
larguras(wsW, {"A":16,"B":12,"C":10,"D":24,"E":18,"F":11,"G":10,"H":13,"I":26,"J":26,"K":26,"L":28})

secao(wsW, 4, "IDENTIFICAÇÃO DO MICROCICLO", 12, 1)
MIC_ID = [("Microciclo","Micro 5.3"),("Semana nº no Macrociclo",34),("Data de Início (2ª feira)",date(2026,8,24)),
          ("Tipo de Microciclo","Ordinário"),("Objetivo da Semana","Consolidação técnico-tática do sistema defensivo, mantendo a carga física da fase")]
for i, (lab, val) in enumerate(MIC_ID):
    r = 5 + i
    rotulo(wsW, r, 1, lab)
    entrada(wsW, r, 2, val, NF_DATE if isinstance(val, date) else None, largura_merge=(9 if i == 4 else 2))
dv(wsW, L("Tipo de Microciclo"), "B8")
wsW["B6"].comment = Comment("Deve coincidir com a coluna 'Semana nº' da aba Mesociclo e com a numeração automática "
                            "das abas de registro.", "Planilha")

secao(wsW, 11, "SESSÕES DA SEMANA", 12, 1)
SES_H = ["Dia da Semana","Data","Turno","Tipo de Sessão","Local","Duração (min)","PSE Prevista",
         "Carga Prevista (UA)","Ênfase Técnica","Ênfase Tática","Ênfase Física / Prevenção","Observações"]
cab_tabela(wsW, 12, SES_H)
SF, SL = 13, 26
DIAS = ["Segunda-feira","Terça-feira","Quarta-feira","Quinta-feira","Sexta-feira","Sábado","Domingo"]
for r in range(SF, SL + 1):
    wsW.cell(r, 2, '=IF($B$7="","",$B$7+INT((ROW()-{})/2))'.format(SF))
    wsW.cell(r, 1, '=IF($B{0}="","",CHOOSE(WEEKDAY($B{0},2),{1}))'.format(r, ",".join('"%s"' % d for d in DIAS)))
    wsW.cell(r, 8, '=IF(OR($F{0}="",$G{0}=""),0,$F{0}*$G{0})'.format(r))
corpo_tabela(wsW, SF, SL, 1, 12)
for r in range(SF, SL + 1):
    for c in (3, 4, 5, 6, 7, 9, 10, 11, 12):
        wsW.cell(r, c).fill = PatternFill("solid", fgColor=GOLD_L)
        wsW.cell(r, c).font = Font(name=F, size=9, color="0000FF")
    for c in (1, 2, 8):
        wsW.cell(r, c).font = Font(name=F, size=9, bold=True, color=NAVY2)
        wsW.cell(r, c).fill = PatternFill("solid", fgColor=LIGHT)
    for c in (4, 9, 10, 11, 12):
        wsW.cell(r, c).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    wsW.cell(r, 2).number_format = NF_DATE
    wsW.cell(r, 8).number_format = NF_UA
    if (r - SF) % 2 == 0:
        for c in range(1, 13):
            wsW.cell(r, c).border = Border(left=thin, right=thin, top=Side(style="medium", color=BLUE3), bottom=thin)
dv(wsW, L("Turno"),          "C{}:C{}".format(SF, SL))
dv(wsW, L("Tipo de Sessão"), "D{}:D{}".format(SF, SL))
dv(wsW, L("PSE (0-10)"),     "G{}:G{}".format(SF, SL))
tr = SL + 1
for c in range(1, 13):
    wsW.cell(tr, c).fill = PatternFill("solid", fgColor=NAVY2); wsW.cell(tr, c).border = BORDER
    wsW.cell(tr, c).font = Font(name=F, size=9, bold=True, color=WHITE)
    wsW.cell(tr, c).alignment = Alignment(horizontal="center", vertical="center")
wsW.cell(tr, 1, "TOTAL DA SEMANA").alignment = Alignment(horizontal="left", vertical="center", indent=1)
wsW.cell(tr, 6, "=SUM(F{}:F{})".format(SF, SL)).number_format = NF_UA
wsW.cell(tr, 7, '=IFERROR(AVERAGE(G{}:G{}),"")'.format(SF, SL)).number_format = '0.0'
wsW.cell(tr, 8, "=SUM(H{}:H{})".format(SF, SL)).number_format = NF_UA

secao(wsW, 29, "RESUMO DA SEMANA — PREVISTO × REALIZADO", 12, 1)
kpi(wsW, 30, 1, "SESSÕES PLANEJADAS", "=COUNTA(D{}:D{})".format(SF, SL), "0", NAVY2, 2)
kpi(wsW, 30, 3, "DURAÇÃO TOTAL (min)", "=SUM(F{}:F{})".format(SF, SL), NF_UA, NAVY2, 2)
kpi(wsW, 30, 5, "CARGA PREVISTA (UA)", "=SUM(H{}:H{})".format(SF, SL), NF_UA, GOLD, 2)
kpi(wsW, 30, 7, "CARGA REALIZADA — MÉDIA/ATLETA (UA)",
    "=IFERROR(SUMIFS('Carga (PSE)'!$G${1}:$G${2},'Carga (PSE)'!$B${1}:$B${2},$B$6)"
    "/MAX(1,COUNTIF(Cadastro!$BL${3}:$BL${4},\"Ativo\")),0)".format(0, CAR_F, CAR_L, CAD_F, CAD_L), NF_UA, GOLD, 3)
kpi(wsW, 30, 10, "% DE CUMPRIMENTO DA CARGA",
    '=IFERROR(G31/H27,0)', NF_PCT, NAVY2, 3)

secao(wsW, 34, "DISTRIBUIÇÃO DE CONTEÚDOS POR DIA (minutos)", 12, 1)
CON_H = ["Dia","Técnico","Tático","Físico","Preventivo / Prevenção","Recuperação","Total (min)","% da Semana"]
cab_tabela(wsW, 35, CON_H)
CF, CL = 36, 42
for i in range(7):
    r = CF + i
    wsW.cell(r, 1, '=IF($A{0}="","",$A{0})'.format(SF + i * 2))
    wsW.cell(r, 7, "=SUM(B{0}:F{0})".format(r))
    wsW.cell(r, 8, '=IFERROR($G{0}/$G${1},0)'.format(r, CL + 1))
corpo_tabela(wsW, CF, CL, 1, 8)
for r in range(CF, CL + 1):
    for c in range(2, 7):
        wsW.cell(r, c).fill = PatternFill("solid", fgColor=GOLD_L)
        wsW.cell(r, c).font = Font(name=F, size=9, color="0000FF")
        wsW.cell(r, c).number_format = NF_UA
    for c in (1, 7, 8):
        wsW.cell(r, c).font = Font(name=F, size=9, bold=True, color=NAVY2)
        wsW.cell(r, c).fill = PatternFill("solid", fgColor=LIGHT)
    wsW.cell(r, 1).alignment = Alignment(horizontal="left", vertical="center", indent=1)
    wsW.cell(r, 7).number_format = NF_UA; wsW.cell(r, 8).number_format = NF_PCT
tr2 = CL + 1
for c in range(1, 9):
    wsW.cell(tr2, c).fill = PatternFill("solid", fgColor=NAVY2); wsW.cell(tr2, c).border = BORDER
    wsW.cell(tr2, c).font = Font(name=F, size=9, bold=True, color=WHITE)
    wsW.cell(tr2, c).alignment = Alignment(horizontal="center", vertical="center")
wsW.cell(tr2, 1, "TOTAL").alignment = Alignment(horizontal="left", vertical="center", indent=1)
for c in range(2, 8):
    cl_ = get_column_letter(c)
    wsW.cell(tr2, c, "=SUM({0}{1}:{0}{2})".format(cl_, CF, CL)).number_format = NF_UA
wsW.cell(tr2, 8, '=IFERROR(SUM(H{}:H{}),0)'.format(CF, CL)).number_format = NF_PCT

chW = BarChart(); chW.type = "col"; chW.grouping = "stacked"; chW.overlap = 100
chW.title = "Distribuição de Conteúdos na Semana (min)"; chW.height = 8; chW.width = 18
d = Reference(wsW, min_col=2, max_col=6, min_row=35, max_row=CL)
c_ = Reference(wsW, min_col=1, min_row=CF, max_row=CL)
chW.add_data(d, titles_from_data=True); chW.set_categories(c_)
wsW.add_chart(chW, "J35")
wsW.freeze_panes = "C13"

CONT_EX = [(60,30,45,15,10),(50,20,70,15,10),(60,40,30,15,10),(45,35,60,15,10),(70,50,20,15,15),(30,20,0,10,20),(0,0,0,0,0)]
for i, vals in enumerate(CONT_EX):
    for j, v in enumerate(vals):
        wsW.cell(CF + i, 2 + j, v)
SESS_EX = [
 (0,"Manhã","Físico (Força)","Sala de musculação",75,7,"—","—","Força máxima: agachamento e terra romeno"),
 (0,"Tarde","Técnico-Tático","Ginásio principal",90,6,"Recepção de saque e side-out","Complexo K1","Ativação e core"),
 (1,"Manhã","Físico (Potência)","Sala de musculação",60,7,"—","—","Power clean e jump squat"),
 (1,"Tarde","Técnico-Tático","Ginásio principal",90,7,"Ataque de ponta e primeiro tempo","Complexo K2","Pliometria de baixo volume"),
 (2,"Tarde","Coletivo/Jogo","Ginásio principal",100,8,"Saque sob pressão","Jogo 6x6 com pontuação diferenciada","Prevenção de ombro"),
 (3,"Manhã","Físico (Condicionamento)","Pista",50,6,"—","—","Intervalado 15/15"),
 (3,"Tarde","Técnico-Tático","Ginásio principal",90,7,"Bloqueio e defesa","Sistemas de cobertura","Mobilidade"),
 (4,"Tarde","Coletivo/Jogo","Ginásio principal",110,8,"Refinamento do side-out","Preparação por adversário","Ativação"),
 (5,"Manhã","Recuperação/Regenerativo","Piscina / sala",50,3,"Técnica livre","Vídeo do adversário","Liberação miofascial"),
 (6,"Manhã","Folga","—",0,0,"—","—","Descanso completo"),
]
for dia, turno, tipo, local, dur, pse, et, eta, ef in SESS_EX:
    base = SF + dia * 2
    r = base if wsW.cell(base, 4).value in (None, "") else base + 1
    wsW.cell(r, 3, turno); wsW.cell(r, 4, tipo); wsW.cell(r, 5, local)
    wsW.cell(r, 6, dur); wsW.cell(r, 7, pse)
    wsW.cell(r, 9, et); wsW.cell(r, 10, eta); wsW.cell(r, 11, ef)

# ============================================================================
# 8) PRESCRIÇÃO DO TREINO
# ============================================================================
wsP = wb.create_sheet("Prescrição")
banner(wsP, "PRESCRIÇÃO DO TREINO", "Uma linha por exercício prescrito. Escolha o exercício no menu (vem da aba Exercícios) e "
       "defina o destinatário: 'Equipe (todos)' ou o nome de um atleta — é assim que o treino chega à Área do Atleta.", 20, GOLD)
larguras(wsP, {"A":12,"B":8,"C":16,"D":9,"E":24,"F":18,"G":38,"H":17,"I":16,"J":20,"K":8,"L":15,"M":18,"N":9,
               "O":10,"P":9,"Q":9,"R":13,"S":26,"T":24,"U":6})
secao(wsP, 3, "RESUMO DA SESSÃO SELECIONADA", 20, 1)
rotulo(wsP, 4, 1, "Data:")
entrada(wsP, 4, 2, date(2026, 8, 31), NF_DATE, largura_merge=2)
FLT = ('$A${0}:$A${1},$B$4'.format(PRE_F, PRE_L))
kpi(wsP, 4, 5, "EXERCÍCIOS PRESCRITOS", '=COUNTIFS({})'.format(FLT), "0", NAVY2, 2)
kpi(wsP, 4, 7, "DURAÇÃO TOTAL (min)", '=SUMIFS($O${0}:$O${1},{2})'.format(PRE_F, PRE_L, FLT), NF_UA, NAVY2, 2)
kpi(wsP, 4, 9, "CARGA PREVISTA (UA)", '=SUMIFS($R${0}:$R${1},{2})'.format(PRE_F, PRE_L, FLT), NF_UA, GOLD, 2)
kpi(wsP, 4, 11, "PSE MÉDIA PREVISTA", '=IFERROR(AVERAGEIFS($P${0}:$P${1},{2}),0)'.format(PRE_F, PRE_L, FLT), '0.0;;""', GOLD, 2)
kpi(wsP, 4, 13, "MIN. TÉCNICO", '=SUMIFS($O${0}:$O${1},{2},$H${0}:$H${1},"Técnico")'.format(PRE_F, PRE_L, FLT), NF_UA, BLUE3, 2)
kpi(wsP, 4, 15, "MIN. TÁTICO", '=SUMIFS($O${0}:$O${1},{2},$H${0}:$H${1},"Tático")'.format(PRE_F, PRE_L, FLT), NF_UA, BLUE3, 2)
kpi(wsP, 4, 17, "MIN. FÍSICO", '=SUMIFS($O${0}:$O${1},{2},$H${0}:$H${1},"Físico")'.format(PRE_F, PRE_L, FLT), NF_UA, BLUE3, 2)
kpi(wsP, 4, 19, "MIN. PREVENTIVO", '=SUMIFS($O${0}:$O${1},{2},$H${0}:$H${1},"Preventivo/Compensatório")'.format(PRE_F, PRE_L, FLT), NF_UA, BLUE3, 2)
nota(wsP, 6, 1, "Digite a data em B4 para ver o resumo das sessões daquele dia. Séries × Repetições geram o Volume; "
     "Duração × PSE Prevista geram a Carga Prevista em UA (unidades arbitrárias).", 20)

PRE_H = ["Data","Semana","Sessão","Turno","Destinatário","Bloco","Exercício","Categoria","Fundamento","Capacidade Física",
         "Séries","Repetições / Tempo","Carga / Intensidade","Pausa (s)","Duração (min)","PSE Prev.","Volume",
         "Carga Prev. (UA)","Progressão / Critério","Observações","Índ."]
cab_tabela(wsP, 7, PRE_H)
for r in range(PRE_F, PRE_L + 1):
    wsP.cell(r, 2,  '=IF($A{0}="",0,INT(($A{0}-Macrociclo!$C$11)/7)+1)'.format(r))
    wsP.cell(r, 8,  '=IFERROR(INDEX(Exercícios!$C${1}:$C${2},MATCH($G{0},Exercícios!$B${1}:$B${2},0)),"")'.format(r, EX_F, EX_L))
    wsP.cell(r, 9,  '=IFERROR(INDEX(Exercícios!$D${1}:$D${2},MATCH($G{0},Exercícios!$B${1}:$B${2},0)),"")'.format(r, EX_F, EX_L))
    wsP.cell(r, 10, '=IFERROR(INDEX(Exercícios!$E${1}:$E${2},MATCH($G{0},Exercícios!$B${1}:$B${2},0)),"")'.format(r, EX_F, EX_L))
    wsP.cell(r, 17, '=IFERROR($K{0}*$L{0},"")'.format(r))
    wsP.cell(r, 18, '=IF(OR($O{0}="",$P{0}=""),0,$O{0}*$P{0})'.format(r))
    wsP.cell(r, 21, '=IF($A{0}="",0,IF(AND(OR($E{0}=Atleta!$C$4,$E{0}="Equipe (todos)"),$A{0}>=Atleta!$C$5),'
                    'MAX($U$7:U{1})+1,0))'.format(r, r - 1))
corpo_tabela(wsP, PRE_F, PRE_L, 1, 21)
for r in range(PRE_F, PRE_L + 1):
    for c in (1, 3, 4, 5, 6, 7, 11, 12, 13, 14, 15, 16, 19, 20):
        wsP.cell(r, c).fill = PatternFill("solid", fgColor=GOLD_L)
        wsP.cell(r, c).font = Font(name=F, size=9, color="0000FF")
    for c in (2, 8, 9, 10, 17, 18, 21):
        wsP.cell(r, c).font = Font(name=F, size=9, bold=True, color=NAVY2)
        wsP.cell(r, c).fill = PatternFill("solid", fgColor=LIGHT)
    for c in (5, 7, 13, 19, 20):
        wsP.cell(r, c).alignment = Alignment(horizontal="left", vertical="center", indent=1)
    wsP.cell(r, 1).number_format = NF_DATE
    wsP.cell(r, 2).number_format = '0;;""'
    wsP.cell(r, 18).number_format = NF_UA
    wsP.cell(r, 21).number_format = '0;;""'
dv(wsP, L("Turno"),          "D{}:D{}".format(PRE_F, PRE_L))
dv(wsP, DEST_REF,            "E{}:E{}".format(PRE_F, PRE_L))
dv(wsP, L("Bloco da Sessão"),"F{}:F{}".format(PRE_F, PRE_L))
dv(wsP, EXERC_REF,           "G{}:G{}".format(PRE_F, PRE_L))
dv(wsP, L("PSE (0-10)"),     "P{}:P{}".format(PRE_F, PRE_L))
wsP.column_dimensions["U"].hidden = True
wsP.freeze_panes = "C8"
wsP.auto_filter.ref = "A7:T{}".format(PRE_L)

# ============================================================================
# 9) CONTROLE DE CARGA  (PSE da sessão / ACWR)
# ============================================================================
wsG = wb.create_sheet("Carga (PSE)")
banner(wsG, "CONTROLE DE CARGA INTERNA — MÉTODO DA PSE DA SESSÃO",
       "Registre duração e PSE (0–10, coletada ~30 min após o treino). Carga, aguda, crônica, ACWR e zona de risco são automáticos.", 13, RED_T)
larguras(wsG, {"A":12,"B":8,"C":26,"D":24,"E":12,"F":9,"G":12,"H":15,"I":17,"J":10,"K":12,"L":20,"M":26})
CAR_H = ["Data","Semana","Atleta","Tipo de Sessão","Duração (min)","PSE (0-10)","Carga (UA)",
         "Carga Aguda 7 d","Carga Crônica 28 d","ACWR","Dias de Histórico","Zona / Alerta","Observações"]
cab_tabela(wsG, 7, CAR_H)
nota(wsG, 4, 1, "Carga (UA) = Duração × PSE  •  ACWR = Carga Aguda (7 d) ÷ Carga Crônica (média semanal dos 28 d).  "
     "Faixas: < 0,80 subcarga | 0,80–1,30 zona ideal | 1,31–1,50 atenção | > 1,50 risco elevado.  "
     "A Zona só é classificada após 21 dias de registros do atleta — antes disso aparece 'Histórico insuficiente'.", 13)
wsG.row_dimensions[4].height = 26
for r in range(CAR_F, CAR_L + 1):
    wsG.cell(r, 2, '=IF($A{0}="",0,INT(($A{0}-Macrociclo!$C$11)/7)+1)'.format(r))
    wsG.cell(r, 7, '=IF(OR($E{0}="",$F{0}=""),0,$E{0}*$F{0})'.format(r))
    wsG.cell(r, 8, '=IF($A{0}="",0,SUMIFS($G${1}:$G${2},$C${1}:$C${2},$C{0},$A${1}:$A${2},">="&$A{0}-6,'
                   '$A${1}:$A${2},"<="&$A{0}))'.format(r, CAR_F, CAR_L))
    wsG.cell(r, 9, '=IF($A{0}="",0,SUMIFS($G${1}:$G${2},$C${1}:$C${2},$C{0},$A${1}:$A${2},">="&$A{0}-27,'
                   '$A${1}:$A${2},"<="&$A{0})/4)'.format(r, CAR_F, CAR_L))
    wsG.cell(r, 10, '=IF(OR($A{0}="",$I{0}=0),0,$H{0}/$I{0})'.format(r))
    wsG.cell(r, 11, '=IF($A{0}="",0,$A{0}-SUMPRODUCT(MIN(($C${1}:$C${2}=$C{0})*$A${1}:$A${2}'
                    '+($C${1}:$C${2}<>$C{0})*100000))+1)'.format(r, CAR_F, CAR_L))
    wsG.cell(r, 12, '=IF($A{0}="","",IF($J{0}=0,"Sem registro anterior",IF($K{0}<21,"Histórico insuficiente",'
                    'IF($J{0}<0.8,"Subcarga",IF($J{0}<=1.3,"Zona ideal",IF($J{0}<=1.5,"Atenção","Risco elevado"))))))'.format(r))
corpo_tabela(wsG, CAR_F, CAR_L, 1, 13)
for r in range(CAR_F, CAR_L + 1):
    for c in (1, 3, 4, 5, 6, 13):
        wsG.cell(r, c).fill = PatternFill("solid", fgColor=GOLD_L)
        wsG.cell(r, c).font = Font(name=F, size=9, color="0000FF")
    for c in (2, 7, 8, 9, 10, 11, 12):
        wsG.cell(r, c).font = Font(name=F, size=9, bold=True, color=NAVY2)
        wsG.cell(r, c).fill = PatternFill("solid", fgColor=LIGHT)
    wsG.cell(r, 11).number_format = '0;;""'
    for c in (3, 4, 13):
        wsG.cell(r, c).alignment = Alignment(horizontal="left", vertical="center", indent=1)
    wsG.cell(r, 1).number_format = NF_DATE
    wsG.cell(r, 2).number_format = '0;;""'
    for c in (7, 8, 9):
        wsG.cell(r, c).number_format = NF_UA
    wsG.cell(r, 10).number_format = NF_DEC
dv(wsG, ATLETAS_REF,          "C{}:C{}".format(CAR_F, CAR_L))
dv(wsG, L("Tipo de Sessão"),  "D{}:D{}".format(CAR_F, CAR_L))
dv(wsG, L("PSE (0-10)"),      "F{}:F{}".format(CAR_F, CAR_L))
zref = "L{}:L{}".format(CAR_F, CAR_L)
for txt, fill, ft in [('"Zona ideal"', GREEN, GREEN_T), ('"Atenção"', YELL, YELL_T),
                      ('"Risco elevado"', RED, RED_T), ('"Subcarga"', "DDEBF7", NAVY2),
                      ('"Histórico insuficiente"', "EDEDED", GREY_T)]:
    wsG.conditional_formatting.add(zref, CellIsRule(operator="equal", formula=[txt],
        fill=PatternFill("solid", fgColor=fill), font=Font(name=F, size=9, bold=True, color=ft)))
wsG.conditional_formatting.add("J{}:J{}".format(CAR_F, CAR_L),
    CellIsRule(operator="greaterThan", formula=["1.5"], font=Font(name=F, size=9, bold=True, color=RED_T)))
wsG.freeze_panes = "C8"
wsG.auto_filter.ref = "A7:M{}".format(CAR_L)

# ============================================================================
# 10) WELLNESS  (Índice de Hooper)
# ============================================================================
wsWe = wb.create_sheet("Wellness")
banner(wsWe, "WELLNESS DIÁRIO — ÍNDICE DE HOOPER",
       "Cada item de 1 (muito bom / ausente) a 7 (muito ruim / máxima). Soma de 4 a 28 — QUANTO MAIOR, PIOR. Aplicar antes do treino.", 11, "6B3FA0")
larguras(wsWe, {"A":12,"B":8,"C":26,"D":16,"E":13,"F":13,"G":16,"H":14,"I":18,"J":20,"K":28})
WEL_H = ["Data","Semana","Atleta","Qualidade do Sono (1-7)","Estresse (1-7)","Fadiga (1-7)","Dor Muscular (1-7)",
         "Índice de Hooper","Classificação","Local da Dor / Queixa","Observações"]
cab_tabela(wsWe, 7, WEL_H)
nota(wsWe, 4, 1, "Classificação: 4–8 Ótimo | 9–12 Bom | 13–16 Regular | 17–20 Alerta | 21–28 Crítico. "
     "Referência: Hooper & Mackinnon (1995). Pontuações em Alerta/Crítico por 2 dias seguidos pedem conversa com o atleta e "
     "eventual ajuste da carga do dia.", 11)
wsWe.row_dimensions[4].height = 26
for r in range(WEL_F, WEL_L + 1):
    wsWe.cell(r, 2, '=IF($A{0}="",0,INT(($A{0}-Macrociclo!$C$11)/7)+1)'.format(r))
    wsWe.cell(r, 8, '=IF(COUNT($D{0}:$G{0})<4,0,SUM($D{0}:$G{0}))'.format(r))
    wsWe.cell(r, 9, '=IF($H{0}=0,"",IF($H{0}<=8,"Ótimo",IF($H{0}<=12,"Bom",IF($H{0}<=16,"Regular",'
                    'IF($H{0}<=20,"Alerta","Crítico")))))'.format(r))
corpo_tabela(wsWe, WEL_F, WEL_L, 1, 11)
for r in range(WEL_F, WEL_L + 1):
    for c in (1, 3, 4, 5, 6, 7, 10, 11):
        wsWe.cell(r, c).fill = PatternFill("solid", fgColor=GOLD_L)
        wsWe.cell(r, c).font = Font(name=F, size=9, color="0000FF")
    for c in (2, 8, 9):
        wsWe.cell(r, c).font = Font(name=F, size=9, bold=True, color=NAVY2)
        wsWe.cell(r, c).fill = PatternFill("solid", fgColor=LIGHT)
    for c in (3, 10, 11):
        wsWe.cell(r, c).alignment = Alignment(horizontal="left", vertical="center", indent=1)
    wsWe.cell(r, 1).number_format = NF_DATE
    wsWe.cell(r, 2).number_format = '0;;""'
    wsWe.cell(r, 8).number_format = '0;;""'
dv(wsWe, ATLETAS_REF,   "C{}:C{}".format(WEL_F, WEL_L))
for cl in "DEFG":
    dv(wsWe, L("Escala 1-7"), "{0}{1}:{0}{2}".format(cl, WEL_F, WEL_L))
for txt, fill, ft in [('"Ótimo"', GREEN, GREEN_T), ('"Bom"', "E2EFDA", GREEN_T),
                      ('"Regular"', YELL, YELL_T), ('"Alerta"', "FCE4D6", "C55A11"), ('"Crítico"', RED, RED_T)]:
    wsWe.conditional_formatting.add("I{}:I{}".format(WEL_F, WEL_L), CellIsRule(operator="equal", formula=[txt],
        fill=PatternFill("solid", fgColor=fill), font=Font(name=F, size=9, bold=True, color=ft)))
wsWe.freeze_panes = "C8"
wsWe.auto_filter.ref = "A7:K{}".format(WEL_L)

# ============================================================================
# 11) PRESENÇA
# ============================================================================
wsPr = wb.create_sheet("Presença")
banner(wsPr, "CONTROLE DE PRESENÇA", "Chamada por sessão à esquerda; percentual de presença por atleta calculado automaticamente à direita.", 16, "1F6F4A")
larguras(wsPr, {"A":12,"B":8,"C":24,"D":26,"E":22,"F":13,"G":30,"H":3,
                "I":26,"J":11,"K":11,"L":12,"M":10,"N":13,"O":13,"P":16})
PRS_H = ["Data","Semana","Sessão / Tipo","Atleta","Presença","Min. Participados","Justificativa"]
cab_tabela(wsPr, 7, PRS_H)
for r in range(PRS_F, PRS_L + 1):
    wsPr.cell(r, 2, '=IF($A{0}="",0,INT(($A{0}-Macrociclo!$C$11)/7)+1)'.format(r))
corpo_tabela(wsPr, PRS_F, PRS_L, 1, 7)
for r in range(PRS_F, PRS_L + 1):
    for c in (1, 3, 4, 5, 6, 7):
        wsPr.cell(r, c).fill = PatternFill("solid", fgColor=GOLD_L)
        wsPr.cell(r, c).font = Font(name=F, size=9, color="0000FF")
    wsPr.cell(r, 2).font = Font(name=F, size=9, bold=True, color=NAVY2)
    wsPr.cell(r, 2).fill = PatternFill("solid", fgColor=LIGHT)
    for c in (3, 4, 7):
        wsPr.cell(r, c).alignment = Alignment(horizontal="left", vertical="center", indent=1)
    wsPr.cell(r, 1).number_format = NF_DATE
    wsPr.cell(r, 2).number_format = '0;;""'
dv(wsPr, ATLETAS_REF,         "D{}:D{}".format(PRS_F, PRS_L))
dv(wsPr, L("Presença"),       "E{}:E{}".format(PRS_F, PRS_L))
dv(wsPr, L("Tipo de Sessão"), "C{}:C{}".format(PRS_F, PRS_L))
wsPr.conditional_formatting.add("E{}:E{}".format(PRS_F, PRS_L),
    CellIsRule(operator="equal", formula=['"Presente"'], fill=PatternFill("solid", fgColor=GREEN), font=Font(name=F, size=9, bold=True, color=GREEN_T)))
for txt in ('"Falta Não Justificada"', '"Lesionado"', '"Departamento Médico"'):
    wsPr.conditional_formatting.add("E{}:E{}".format(PRS_F, PRS_L),
        CellIsRule(operator="equal", formula=[txt], fill=PatternFill("solid", fgColor=RED), font=Font(name=F, size=9, bold=True, color=RED_T)))
wsPr.conditional_formatting.add("E{}:E{}".format(PRS_F, PRS_L),
    CellIsRule(operator="equal", formula=['"Falta Justificada"'], fill=PatternFill("solid", fgColor=YELL), font=Font(name=F, size=9, bold=True, color=YELL_T)))

secao(wsPr, 5, "RESUMO POR ATLETA", 16, 9)
cab_tabela(wsPr, 7, ["Atleta","Registros","Presenças","% Presença","Faltas","Lesão / DM","Min. Totais","Situação"], col0=9)
for i in range(CAD_L - CAD_F + 1):
    r = PRS_F + i
    a = CAD_F + i
    wsPr.cell(r, 9,  '=IF(Cadastro!$B{0}="","",Cadastro!$B{0})'.format(a))
    wsPr.cell(r, 10, '=IF($I{0}="",0,COUNTIFS($D${1}:$D${2},$I{0}))'.format(r, PRS_F, PRS_L))
    wsPr.cell(r, 11, '=IF($I{0}="",0,COUNTIFS($D${1}:$D${2},$I{0},$E${1}:$E${2},"Presente"))'.format(r, PRS_F, PRS_L))
    wsPr.cell(r, 12, '=IF($J{0}=0,0,$K{0}/$J{0})'.format(r))
    wsPr.cell(r, 13, '=IF($I{0}="",0,COUNTIFS($D${1}:$D${2},$I{0},$E${1}:$E${2},"Falta Justificada")'
                     '+COUNTIFS($D${1}:$D${2},$I{0},$E${1}:$E${2},"Falta Não Justificada"))'.format(r, PRS_F, PRS_L))
    wsPr.cell(r, 14, '=IF($I{0}="",0,COUNTIFS($D${1}:$D${2},$I{0},$E${1}:$E${2},"Lesionado")'
                     '+COUNTIFS($D${1}:$D${2},$I{0},$E${1}:$E${2},"Departamento Médico"))'.format(r, PRS_F, PRS_L))
    wsPr.cell(r, 15, '=IF($I{0}="",0,SUMIFS($F${1}:$F${2},$D${1}:$D${2},$I{0}))'.format(r, PRS_F, PRS_L))
    wsPr.cell(r, 16, '=IF($I{0}="","",IF($J{0}=0,"Sem registro",IF($L{0}>=0.9,"Excelente",'
                     'IF($L{0}>=0.75,"Regular","Baixa assiduidade"))))'.format(r))
PRS_RES_L = PRS_F + (CAD_L - CAD_F)
corpo_tabela(wsPr, PRS_F, PRS_RES_L, 9, 16)
for r in range(PRS_F, PRS_RES_L + 1):
    for c in range(9, 17):
        wsPr.cell(r, c).font = Font(name=F, size=9, bold=(c == 12), color=NAVY2)
        wsPr.cell(r, c).fill = PatternFill("solid", fgColor=LIGHT if (r - PRS_F) % 2 == 0 else LIGHT2)
    wsPr.cell(r, 9).alignment = Alignment(horizontal="left", vertical="center", indent=1)
    wsPr.cell(r, 12).number_format = NF_PCT
    for c in (10, 11, 13, 14, 15):
        wsPr.cell(r, c).number_format = '0;;""'
wsPr.conditional_formatting.add("L{}:L{}".format(PRS_F, PRS_RES_L),
    ColorScaleRule(start_type="num", start_value=0.5, start_color=RED,
                   mid_type="num", mid_value=0.8, mid_color=YELL,
                   end_type="num", end_value=1, end_color=GREEN))
wsPr.freeze_panes = "C8"

# ============================================================================
# 12) TESTES FÍSICOS  (bateria geral + específica do voleibol)
# ============================================================================
wsT = wb.create_sheet("Testes")
banner(wsT, "AVALIAÇÕES E TESTES FÍSICOS", "Bateria longitudinal. Impulsões, RSI, índice elástico e potência de pico "
       "são calculados automaticamente. Registre a melhor de 3 tentativas nos testes de salto, sprint e arremesso.", 29, "8A5A00")
TST_H = ["Data","Atleta","Momento","Avaliador","Massa (kg)",
         "Alcance em Pé (cm)","Alcance de Ataque (cm)","Alcance de Bloqueio (cm)",
         "Impulsão de Ataque (cm)","Impulsão de Bloqueio (cm)",
         "Squat Jump (cm)","CMJ (cm)","CMJ c/ Braços (cm)","Drop Jump 40 cm (cm)","Tempo de Contato DJ (s)",
         "RSI (m/s)","Índice Elástico (%)","Salto c/ Aproximação (cm)","Potência de Pico — Sayers (W)",
         "Sprint 5 m (s)","Sprint 10 m (s)","T-Test (s)","Sheppard Agility (s)","Arremesso Medicine Ball (m)",
         "Sentar-e-Alcançar (cm)","Yo-Yo IR1 (m)","IMTP Pico de Força (N)","Força Relativa IMTP (N/kg)","Observações"]
GRUPOS_T = [("IDENTIFICAÇÃO", 1, 5, NAVY), ("ALCANCES E IMPULSÕES (específico do voleibol)", 6, 10, "1F6F4A"),
            ("SALTOS E POTÊNCIA", 11, 19, GOLD), ("VELOCIDADE, AGILIDADE E OUTROS", 20, 26, BLUE3),
            ("FORÇA ISOMÉTRICA", 27, 28, RED_T), ("", 29, 29, GREY_T)]
for txt, c1, c2, cor in GRUPOS_T:
    wsT.merge_cells(start_row=6, start_column=c1, end_row=6, end_column=c2)
    c = wsT.cell(6, c1, txt)
    c.font = Font(name=F, size=9, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=cor)
    c.alignment = Alignment(horizontal="center", vertical="center"); c.border = BORDER
wsT.row_dimensions[6].height = 18
cab_tabela(wsT, 7, TST_H)
larguras(wsT, {"A":12,"B":24,"C":18,"D":18,"E":11,"F":14,"G":16,"H":16,"I":15,"J":15,
               "K":13,"L":11,"M":14,"N":15,"O":16,"P":11,"Q":14,"R":16,"S":18,
               "T":12,"U":12,"V":11,"W":16,"X":18,"Y":15,"Z":13,"AA":15,"AB":16,"AC":28})
nota(wsT, 4, 1, "RSI (Reactive Strength Index) = altura do drop jump (m) ÷ tempo de contato (s).  •  "
     "Índice Elástico = (CMJ − Squat Jump) ÷ Squat Jump — quanto o atleta aproveita do ciclo alongamento-encurtamento.  •  "
     "Potência de Pico pela equação de Sayers et al. (1999): P = 60,7 × CMJ(cm) + 45,3 × massa(kg) − 2055.", 29)
wsT.row_dimensions[4].height = 26
for r in range(TST_F, TST_L + 1):
    wsT.cell(r,  9, '=IF(OR($G{0}="",$F{0}=""),"",$G{0}-$F{0})'.format(r))
    wsT.cell(r, 10, '=IF(OR($H{0}="",$F{0}=""),"",$H{0}-$F{0})'.format(r))
    wsT.cell(r, 16, '=IF(OR($N{0}="",$O{0}="",$O{0}=0),"",($N{0}/100)/$O{0})'.format(r))
    wsT.cell(r, 17, '=IF(OR($K{0}="",$L{0}="",$K{0}=0),"",($L{0}-$K{0})/$K{0})'.format(r))
    wsT.cell(r, 19, '=IF(OR($L{0}="",$E{0}=""),"",60.7*$L{0}+45.3*$E{0}-2055)'.format(r))
    wsT.cell(r, 28, '=IF(OR($AA{0}="",$E{0}="",$E{0}=0),"",$AA{0}/$E{0})'.format(r))
corpo_tabela(wsT, TST_F, TST_L, 1, 29)
CALC_T = (9, 10, 16, 17, 19, 28)
for r in range(TST_F, TST_L + 1):
    for c in range(1, 30):
        if c in CALC_T:
            wsT.cell(r, c).font = Font(name=F, size=9, bold=True, color=NAVY2)
            wsT.cell(r, c).fill = PatternFill("solid", fgColor=LIGHT)
        else:
            wsT.cell(r, c).fill = PatternFill("solid", fgColor=GOLD_L)
            wsT.cell(r, c).font = Font(name=F, size=9, color="0000FF")
    for c in (2, 4, 29):
        wsT.cell(r, c).alignment = Alignment(horizontal="left", vertical="center", indent=1)
    wsT.cell(r, 1).number_format = NF_DATE
    for c in (5, 11, 12, 13, 14, 18, 24, 25):
        wsT.cell(r, c).number_format = '0.0;;""'
    for c in (6, 7, 8, 9, 10, 26):
        wsT.cell(r, c).number_format = '0;;""'
    for c in (15,):
        wsT.cell(r, c).number_format = '0.000;;""'
    for c in (16, 20, 21, 22, 23, 28):
        wsT.cell(r, c).number_format = '0.00;;""'
    wsT.cell(r, 17).number_format = NF_PCT1
    wsT.cell(r, 19).number_format = NF_UA
    wsT.cell(r, 27).number_format = NF_UA
dv(wsT, ATLETAS_REF,          "B{}:B{}".format(TST_F, TST_L))
dv(wsT, L("Momento do Teste"),"C{}:C{}".format(TST_F, TST_L))
wsT.conditional_formatting.add("I{}:I{}".format(TST_F, TST_L),
    ColorScaleRule(start_type="min", start_color="FFFFFF", end_type="max", end_color="A9D08E"))
wsT.conditional_formatting.add("L{}:L{}".format(TST_F, TST_L),
    ColorScaleRule(start_type="min", start_color="FFFFFF", end_type="max", end_color="9CC3E5"))
wsT.freeze_panes = "C8"
wsT.auto_filter.ref = "A7:AC{}".format(TST_L)

# ============================================================================
# 13) ÁREA DO ATLETA
# ============================================================================
wsA = wb.create_sheet("Atleta")
banner(wsA, "ÁREA DO ATLETA", "Escolha o seu nome no menu: perfil, cargas da semana, monotonia e o treino prescrito aparecem automaticamente.", 16, GOLD)
larguras(wsA, {"A":10,"B":11,"C":13,"D":13,"E":13,"F":13,"G":13,"H":13,"I":13,"J":15,"K":10,"L":13,"M":14,"N":12,"O":13,"P":18})
rotulo(wsA, 4, 1, "ATLETA:")
selA = entrada(wsA, 4, 3, NOMES[0], largura_merge=4)
selA.font = Font(name=F, size=13, bold=True, color="0000FF")
wsA.row_dimensions[4].height = 24
dv(wsA, ATLETAS_REF, "C4")
rotulo(wsA, 5, 1, "Ver treinos a partir de:")
entrada(wsA, 5, 3, date(2026, 8, 31), NF_DATE, largura_merge=4)
wsA["C5"].comment = Comment("Digite uma data ou a fórmula  =HOJE()  (=TODAY() em inglês) para ver sempre os treinos "
                            "de hoje em diante.", "Planilha")

def idx_cad(col_letter):
    return '=IFERROR(INDEX(Cadastro!${0}${1}:${0}${2},MATCH($C$4,Cadastro!$B${1}:$B${2},0)),"")'.format(col_letter, CAD_F, CAD_L)

secao(wsA, 7, "PERFIL DO ATLETA", 16, 1)
PERFIL = [("ID","A",None),("Idade","E",'0" anos"'),("Posição","L",None),("Categoria","K",None),
          ("Nº Camisa","J",'0'),("Estatura (cm)","O",'0.0'),("Massa (kg)","P",'0.0'),("IMC","Q",'0.0'),
          ("Dominância","M",None),("Perna de Impulsão","N",None),("Anos de Prática","AT",'0" anos"'),
          ("Status","BL",None)]
for i, (lab, cl, nf) in enumerate(PERFIL):
    col = 1 + (i // 4) * 6
    r = 8 + (i % 4)
    rotulo(wsA, r, col, lab)
    c = calc(wsA, r, col + 1, idx_cad(cl), nf)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    wsA.merge_cells(start_row=r, start_column=col + 1, end_row=r, end_column=col + 4)

CG_ = "'Carga (PSE)'"
RC_ = "${0}${1}:${0}${2}"
def cg(col): return "{}!${}${}:${}${}".format(CG_, col, CAR_F, col, CAR_L)
def we(col): return "Wellness!${0}${1}:${0}${2}".format(col, WEL_F, WEL_L)

secao(wsA, 13, "INDICADORES ATUAIS", 16, 1)
kpi(wsA, 14, 1,  "SESSÕES REGISTRADAS", '=COUNTIFS({},$C$4)'.format(cg("C")), '0', NAVY2, 2)
kpi(wsA, 14, 3,  "CARGA TOTAL (UA)", '=SUMIFS({},{},$C$4)'.format(cg("G"), cg("C")), NF_UA, NAVY2, 2)
kpi(wsA, 14, 5,  "CARGA MÉDIA SEMANAL (UA)", '=IFERROR(AVERAGEIF($J$20:$J$67,">0"),0)', NF_UA, NAVY2, 2)
kpi(wsA, 14, 7,  "ÚLTIMO REGISTRO", '=IFERROR(IF(SUMPRODUCT(MAX(({0}=$C$4)*{1}))=0,"",SUMPRODUCT(MAX(({0}=$C$4)*{1}))),"")'.format(cg("C"), cg("A")), 'DD/MM/YYYY;;""', GOLD, 2)
kpi(wsA, 14, 9,  "ACWR ATUAL", '=IF($G$15="",0,IFERROR(AVERAGEIFS({},{},$C$4,{},$G$15),0))'.format(cg("J"), cg("C"), cg("A")), NF_DEC, GOLD, 2)
kpi(wsA, 14, 11, "MONOTONIA MÉDIA", '=IFERROR(AVERAGEIF($N$20:$N$67,">0"),0)', NF_DEC, GOLD, 2)
kpi(wsA, 14, 13, "% DE PRESENÇA", '=IFERROR(INDEX(Presença!$L${0}:$L${1},MATCH($C$4,Presença!$I${0}:$I${1},0)),0)'.format(PRS_F, PRS_RES_L), NF_PCT, NAVY2, 2)
kpi(wsA, 14, 15, "HOOPER MÉDIO", '=IFERROR(AVERAGEIFS({},{},$C$4),0)'.format(we("H"), we("C")), '0.0;;""', "6B3FA0", 2)
wsA.merge_cells(start_row=16, start_column=1, end_row=16, end_column=16)
alerta = wsA.cell(16, 1, '=IF($C$4="","Selecione um atleta em C4.",'
    'IF($I$15>1.5,"ALERTA: ACWR acima de 1,50 — pico de carga aguda. Avaliar redução do volume nas próximas sessões.",'
    'IF($I$15<0.8,"ATENÇÃO: ACWR abaixo de 0,80 — carga recente baixa em relação à crônica. Progredir gradualmente.",'
    'IF($O$15>16,"ATENÇÃO: Índice de Hooper médio elevado — monitorar sono, fadiga e dor muscular.",'
    'IF($K$15>2,"ATENÇÃO: monotonia média acima de 2,0 — variar mais as cargas ao longo da semana.",'
    '"Situação dentro dos parâmetros de referência.")))))')
alerta.font = Font(name=F, size=10, bold=True, color=NAVY)
alerta.fill = PatternFill("solid", fgColor=GOLD_L)
alerta.alignment = Alignment(horizontal="left", vertical="center", indent=1)
alerta.border = BORDER
wsA.row_dimensions[16].height = 24

secao(wsA, 18, "CONTROLE SEMANAL DE CARGA — MONOTONIA E STRAIN (FOSTER)", 16, 1)
SEM_H = ["Semana","Início","Seg","Ter","Qua","Qui","Sex","Sáb","Dom","Carga Semanal (UA)","Sessões",
         "Média Diária","Desvio-Padrão","Monotonia","Strain","Classificação"]
cab_tabela(wsA, 19, SEM_H)
AW_F = 20
AW_L = AW_F + N_SEM - 1
for i in range(N_SEM):
    r = AW_F + i
    wsA.cell(r, 1, i + 1)
    wsA.cell(r, 2, '=IF(Macrociclo!$C$11="","",Macrociclo!$C$11+($A{}-1)*7)'.format(r))
    for d in range(7):
        wsA.cell(r, 3 + d, '=IF($B{0}="",0,SUMIFS({1},{2},$C$4,{3},$B{0}+{4}))'.format(r, cg("G"), cg("C"), cg("A"), d))
    wsA.cell(r, 10, "=SUM($C{0}:$I{0})".format(r))
    wsA.cell(r, 11, '=COUNTIFS({},$C$4,{},$A{})'.format(cg("C"), cg("B"), r))
    wsA.cell(r, 12, '=IF($J{0}=0,0,$J{0}/7)'.format(r))
    wsA.cell(r, 13, '=IF($J{0}=0,0,STDEV($C{0}:$I{0}))'.format(r))
    wsA.cell(r, 14, '=IF($M{0}=0,0,$L{0}/$M{0})'.format(r))
    wsA.cell(r, 15, '=$J{0}*$N{0}'.format(r))
    wsA.cell(r, 16, '=IF($J{0}=0,"",IF($N{0}>2,"Monotonia alta",IF($N{0}>=1.5,"Atenção","Adequada")))'.format(r))
corpo_tabela(wsA, AW_F, AW_L, 1, 16)
for r in range(AW_F, AW_L + 1):
    for c in range(1, 17):
        wsA.cell(r, c).font = Font(name=F, size=9, bold=(c in (10, 14, 15)), color=NAVY2)
        wsA.cell(r, c).fill = PatternFill("solid", fgColor=LIGHT if (r - AW_F) % 2 == 0 else LIGHT2)
    wsA.cell(r, 2).number_format = 'DD/MM'
    for c in list(range(3, 12)) + [15]:
        wsA.cell(r, c).number_format = NF_UA
    for c in (12, 13, 14):
        wsA.cell(r, c).number_format = NF_DEC
wsA.conditional_formatting.add("N{}:N{}".format(AW_F, AW_L),
    CellIsRule(operator="greaterThan", formula=["2"], fill=PatternFill("solid", fgColor=RED), font=Font(name=F, size=9, bold=True, color=RED_T)))
wsA.conditional_formatting.add("J{}:J{}".format(AW_F, AW_L),
    ColorScaleRule(start_type="min", start_color="FFFFFF", end_type="max", end_color="9CC3E5"))

chA = BarChart(); chA.type = "col"; chA.title = "Carga Semanal do Atleta (UA)"
chA.height = 8; chA.width = 20; chA.y_axis.title = "UA"; chA.legend = None
chA.add_data(Reference(wsA, min_col=10, min_row=19, max_row=AW_L), titles_from_data=True)
chA.set_categories(Reference(wsA, min_col=1, min_row=AW_F, max_row=AW_L))
wsA.add_chart(chA, "R19")
chA2 = LineChart(); chA2.title = "Monotonia Semanal"
chA2.height = 8; chA2.width = 20; chA2.legend = None
chA2.add_data(Reference(wsA, min_col=14, min_row=19, max_row=AW_L), titles_from_data=True)
chA2.set_categories(Reference(wsA, min_col=1, min_row=AW_F, max_row=AW_L))
wsA.add_chart(chA2, "R36")

PR_BLOCK = AW_L + 3
secao(wsA, PR_BLOCK, "MEU TREINO PRESCRITO (a partir da data informada em C5)", 16, 1)
TRE_H = ["Data","Sessão","Bloco","Exercício","Capacidade Física","Séries","Reps / Tempo","Carga / Intensidade",
         "Pausa (s)","Duração (min)","PSE Prev.","Progressão / Critério","Observações"]
cab_tabela(wsA, PR_BLOCK + 1, TRE_H)
TR_F = PR_BLOCK + 2
TR_L = TR_F + 29
SRC = ["A", "C", "F", "G", "J", "K", "L", "M", "N", "O", "P", "S", "T"]
for i in range(TR_L - TR_F + 1):
    r = TR_F + i
    for j, sc in enumerate(SRC):
        wsA.cell(r, 1 + j, '=IFERROR(INDEX(Prescrição!${0}${1}:${0}${2},MATCH({3},Prescrição!$U${1}:$U${2},0)),"")'
                 .format(sc, PRE_F, PRE_L, i + 1))
corpo_tabela(wsA, TR_F, TR_L, 1, 13)
for r in range(TR_F, TR_L + 1):
    for c in range(1, 14):
        wsA.cell(r, c).font = Font(name=F, size=9, color=NAVY2)
        wsA.cell(r, c).fill = PatternFill("solid", fgColor=LIGHT if (r - TR_F) % 2 == 0 else LIGHT2)
    wsA.cell(r, 1).number_format = NF_DATE
    for c in (4, 5, 8, 12, 13):
        wsA.cell(r, c).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    wsA.cell(r, 4).font = Font(name=F, size=9, bold=True, color=NAVY)
wsA.column_dimensions["D"].width = 34
nota(wsA, TR_L + 2, 1, "Aparecem aqui as linhas da aba Prescrição cujo Destinatário seja 'Equipe (todos)' ou o nome do atleta "
     "selecionado, com data igual ou posterior à informada em C5.", 16)
wsA.freeze_panes = "A7"

# ============================================================================
# 14) PAINEL INTERATIVO
# ============================================================================
wsD = wb.create_sheet("Painel")
banner(wsD, "PAINEL INTERATIVO DA EQUIPE", "Indicadores e gráficos alimentados por todas as abas de registro. Altere a semana de "
       "referência em C4 para atualizar os cartões.", 20, NAVY)
larguras(wsD, {"A":10,"B":11,"C":15,"D":16,"E":11,"F":12,"G":13,"H":12,"I":13,"J":3,
               "K":13,"L":13,"M":13,"N":13,"O":13,"P":13,"Q":13,"R":13,"S":26,"T":14,"U":12})
rotulo(wsD, 4, 1, "Semana de referência:")
entrada(wsD, 4, 3, '=IFERROR(MAX(1,INT((TODAY()-Macrociclo!$C$11)/7)+1),1)', '0', largura_merge=2)
wsD["C4"].comment = Comment("Vem preenchida com a semana atual do macrociclo. Digite outro número para analisar "
                            "qualquer semana da temporada.", "Planilha")

ATV = 'COUNTIF(Cadastro!$BL${}:$BL${},"Ativo")'.format(CAD_F, CAD_L)
kpi(wsD, 6, 1,  "ATLETAS ATIVOS", '={}'.format(ATV), '0', NAVY2, 2)
kpi(wsD, 6, 3,  "LESIONADOS / DM", '=COUNTIF(Cadastro!$BL${0}:$BL${1},"Lesionado")+COUNTIF(Cadastro!$BL${0}:$BL${1},"Departamento Médico")'.format(CAD_F, CAD_L), '0', RED_T, 2)
kpi(wsD, 6, 5,  "REGISTROS DE CARGA", '=COUNT({})'.format(cg("A")), '0', NAVY2, 2)
kpi(wsD, 6, 7,  "CARGA TOTAL DA EQUIPE (UA)", '=SUM({})'.format(cg("G")), NF_UA, NAVY2, 3)
kpi(wsD, 6, 10, "CARGA DA SEMANA (UA)", '=SUMIFS({},{},$C$4)'.format(cg("G"), cg("B")), NF_UA, GOLD, 3)
kpi(wsD, 6, 13, "CARGA MÉDIA / ATLETA NA SEMANA", '=IFERROR(J7/MAX(1,{}),0)'.format(ATV), NF_UA, GOLD, 3)
kpi(wsD, 6, 16, "ATLETAS EM RISCO (ACWR > 1,5)", '=COUNTIFS({},$C$4,{},">1.5")'.format(cg("B"), cg("J")), '0', RED_T, 2)
kpi(wsD, 6, 18, "HOOPER MÉDIO DA SEMANA", '=IFERROR(AVERAGEIFS({},{},$C$4),0)'.format(we("H"), we("B")), '0.0;;""', "6B3FA0", 3)

secao(wsD, 9, "EVOLUÇÃO SEMANAL DA EQUIPE", 9, 1)
PN_H = ["Semana","Início","Carga Total (UA)","Carga Média/Atleta","Registros","PSE Média","Hooper Médio","ACWR Médio","% Presença"]
cab_tabela(wsD, 10, PN_H)
PN_F = 11
PN_L = PN_F + N_SEM - 1
for i in range(N_SEM):
    r = PN_F + i
    wsD.cell(r, 1, i + 1)
    wsD.cell(r, 2, '=IF(Macrociclo!$C$11="","",Macrociclo!$C$11+($A{}-1)*7)'.format(r))
    wsD.cell(r, 3, '=SUMIFS({},{},$A{})'.format(cg("G"), cg("B"), r))
    wsD.cell(r, 4, '=IFERROR($C{}/MAX(1,{}),0)'.format(r, ATV))
    wsD.cell(r, 5, '=COUNTIFS({},$A{})'.format(cg("B"), r))
    wsD.cell(r, 6, '=IFERROR(AVERAGEIFS({},{},$A{}),0)'.format(cg("F"), cg("B"), r))
    wsD.cell(r, 7, '=IFERROR(AVERAGEIFS({},{},$A{}),0)'.format(we("H"), we("B"), r))
    wsD.cell(r, 8, '=IFERROR(AVERAGEIFS({},{},$A{}),0)'.format(cg("J"), cg("B"), r))
    wsD.cell(r, 9, '=IFERROR(COUNTIFS(Presença!$B${1}:$B${2},$A{0},Presença!$E${1}:$E${2},"Presente")'
                   '/MAX(1,COUNTIFS(Presença!$B${1}:$B${2},$A{0})),0)'.format(r, PRS_F, PRS_L))
corpo_tabela(wsD, PN_F, PN_L, 1, 9)
for r in range(PN_F, PN_L + 1):
    for c in range(1, 10):
        wsD.cell(r, c).font = Font(name=F, size=9, bold=(c == 3), color=NAVY2)
        wsD.cell(r, c).fill = PatternFill("solid", fgColor=LIGHT if (r - PN_F) % 2 == 0 else LIGHT2)
    wsD.cell(r, 2).number_format = 'DD/MM'
    wsD.cell(r, 3).number_format = NF_UA; wsD.cell(r, 4).number_format = NF_UA
    wsD.cell(r, 5).number_format = '0;;""'
    wsD.cell(r, 6).number_format = NF_DEC; wsD.cell(r, 7).number_format = NF_DEC
    wsD.cell(r, 8).number_format = NF_DEC; wsD.cell(r, 9).number_format = NF_PCT
wsD.conditional_formatting.add("H{}:H{}".format(PN_F, PN_L),
    CellIsRule(operator="greaterThan", formula=["1.5"], fill=PatternFill("solid", fgColor=RED), font=Font(name=F, size=9, bold=True, color=RED_T)))
wsD.conditional_formatting.add("C{}:C{}".format(PN_F, PN_L),
    ColorScaleRule(start_type="min", start_color="FFFFFF", end_type="max", end_color="8FAADC"))

c1 = BarChart(); c1.type = "col"; c1.title = "Carga Total da Equipe por Semana (UA)"
c1.height = 8.5; c1.width = 20; c1.legend = None; c1.y_axis.title = "UA"
c1.add_data(Reference(wsD, min_col=3, min_row=10, max_row=PN_L), titles_from_data=True)
c1.set_categories(Reference(wsD, min_col=1, min_row=PN_F, max_row=PN_L))
wsD.add_chart(c1, "K10")
c2 = LineChart(); c2.title = "ACWR Médio da Equipe por Semana"
c2.height = 8.5; c2.width = 20; c2.legend = None
c2.add_data(Reference(wsD, min_col=8, min_row=10, max_row=PN_L), titles_from_data=True)
c2.set_categories(Reference(wsD, min_col=1, min_row=PN_F, max_row=PN_L))
wsD.add_chart(c2, "K28")

TIPOS_ROW = 46
secao(wsD, TIPOS_ROW - 1, "CARGA POR TIPO DE SESSÃO", 13, 11)
cab_tabela(wsD, TIPOS_ROW, ["Tipo de Sessão", "Carga (UA)", "% do Total"], col0=11)
TIPOS = LISTAS[16][1]
for i, t in enumerate(TIPOS):
    r = TIPOS_ROW + 1 + i
    wsD.cell(r, 11, t)
    wsD.cell(r, 12, '=SUMIFS({},{},$K{})'.format(cg("G"), cg("D"), r))
    wsD.cell(r, 13, '=IFERROR($L{}/SUM($L${}:$L${}),0)'.format(r, TIPOS_ROW + 1, TIPOS_ROW + len(TIPOS)))
TIP_L = TIPOS_ROW + len(TIPOS)
corpo_tabela(wsD, TIPOS_ROW + 1, TIP_L, 11, 13)
for r in range(TIPOS_ROW + 1, TIP_L + 1):
    for c in range(11, 14):
        wsD.cell(r, c).font = Font(name=F, size=9, color=NAVY2)
        wsD.cell(r, c).fill = PatternFill("solid", fgColor=LIGHT if (r - TIPOS_ROW) % 2 == 1 else LIGHT2)
    wsD.cell(r, 11).alignment = Alignment(horizontal="left", vertical="center", indent=1)
    wsD.cell(r, 12).number_format = NF_UA; wsD.cell(r, 13).number_format = NF_PCT
cp = PieChart(); cp.title = "Distribuição da Carga por Tipo de Sessão"
cp.height = 9; cp.width = 13
cp.add_data(Reference(wsD, min_col=12, min_row=TIPOS_ROW, max_row=TIP_L), titles_from_data=True)
cp.set_categories(Reference(wsD, min_col=11, min_row=TIPOS_ROW + 1, max_row=TIP_L))
wsD.add_chart(cp, "O46")

RK = 62
secao(wsD, RK - 1, "RANKING E SITUAÇÃO POR ATLETA", 9, 1)
cab_tabela(wsD, RK, ["Atleta","Posição","Sessões","Carga Total (UA)","Carga Média/Sessão (UA)","ACWR Atual",
                     "% Presença","Hooper Médio","Situação"])
RK_F = RK + 1
RK_L = RK_F + (CAD_L - CAD_F)
for i in range(CAD_L - CAD_F + 1):
    r = RK_F + i
    a = CAD_F + i
    wsD.cell(r, 1, '=IF(Cadastro!$B{0}="","",Cadastro!$B{0})'.format(a))
    wsD.cell(r, 2, '=IF($A{}="","",Cadastro!$L{})'.format(r, a))
    wsD.cell(r, 3, '=IF($A{}="",0,COUNTIFS({},$A{}))'.format(r, cg("C"), r))
    wsD.cell(r, 4, '=IF($A{0}="",0,SUMIFS({1},{2},$A{0}))'.format(r, cg("G"), cg("C")))
    wsD.cell(r, 5, '=IF($C{0}=0,0,$D{0}/$C{0})'.format(r))
    wsD.cell(r, 6, '=IF($A{0}="",0,IFERROR(AVERAGEIFS({1},{2},$A{0},{3},SUMPRODUCT(MAX(({2}=$A{0})*{3}))),0))'
             .format(r, cg("J"), cg("C"), cg("A")))
    wsD.cell(r, 7, '=IF($A{0}="",0,IFERROR(INDEX(Presença!$L${1}:$L${2},MATCH($A{0},Presença!$I${1}:$I${2},0)),0))'
             .format(r, PRS_F, PRS_RES_L))
    wsD.cell(r, 8, '=IF($A{0}="",0,IFERROR(AVERAGEIFS({1},{2},$A{0}),0))'.format(r, we("H"), we("C")))
    wsD.cell(r, 9, '=IF($A{0}="","",IF(Cadastro!$BL{1}="Lesionado","LESIONADO",IF($F{0}>1.5,"ALERTA — carga aguda alta",'
                   'IF($H{0}>16,"ALERTA — wellness baixo",IF(AND($G{0}>0,$G{0}<0.75),"Assiduidade baixa",'
                   'IF($C{0}=0,"Sem registros","OK"))))))'.format(r, a))
corpo_tabela(wsD, RK_F, RK_L, 1, 9)
for r in range(RK_F, RK_L + 1):
    for c in range(1, 10):
        wsD.cell(r, c).font = Font(name=F, size=9, bold=(c == 4), color=NAVY2)
        wsD.cell(r, c).fill = PatternFill("solid", fgColor=LIGHT if (r - RK_F) % 2 == 0 else LIGHT2)
    for c in (1, 2, 9):
        wsD.cell(r, c).alignment = Alignment(horizontal="left", vertical="center", indent=1)
    wsD.cell(r, 3).number_format = '0;;""'
    wsD.cell(r, 4).number_format = NF_UA; wsD.cell(r, 5).number_format = NF_UA
    wsD.cell(r, 6).number_format = NF_DEC; wsD.cell(r, 7).number_format = NF_PCT
    wsD.cell(r, 8).number_format = '0.0;;""'
wsD.conditional_formatting.add("F{}:F{}".format(RK_F, RK_L),
    CellIsRule(operator="greaterThan", formula=["1.5"], fill=PatternFill("solid", fgColor=RED), font=Font(name=F, size=9, bold=True, color=RED_T)))
wsD.conditional_formatting.add("I{}:I{}".format(RK_F, RK_L),
    CellIsRule(operator="equal", formula=['"OK"'], fill=PatternFill("solid", fgColor=GREEN), font=Font(name=F, size=9, bold=True, color=GREEN_T)))
wsD.conditional_formatting.add("I{}:I{}".format(RK_F, RK_L),
    FormulaRule(formula=['LEFT($I{},6)="ALERTA"'.format(RK_F)], fill=PatternFill("solid", fgColor=RED), font=Font(name=F, size=9, bold=True, color=RED_T)))
wsD.conditional_formatting.add("G{}:G{}".format(RK_F, RK_L),
    ColorScaleRule(start_type="num", start_value=0.5, start_color=RED, mid_type="num", mid_value=0.8,
                   mid_color=YELL, end_type="num", end_value=1, end_color=GREEN))
RK_CH = min(RK_L, RK_F + 19)
c3 = BarChart(); c3.type = "bar"; c3.title = "Carga Total por Atleta (UA)"
c3.height = 10; c3.width = 20; c3.legend = None
c3.add_data(Reference(wsD, min_col=4, min_row=RK, max_row=RK_CH), titles_from_data=True)
c3.set_categories(Reference(wsD, min_col=1, min_row=RK_F, max_row=RK_CH))
wsD.add_chart(c3, "K62")
c4 = BarChart(); c4.type = "bar"; c4.title = "% de Presença por Atleta"
c4.height = 10; c4.width = 20; c4.legend = None
c4.add_data(Reference(wsD, min_col=7, min_row=RK, max_row=RK_CH), titles_from_data=True)
c4.set_categories(Reference(wsD, min_col=1, min_row=RK_F, max_row=RK_CH))
wsD.add_chart(c4, "S62")
wsD.freeze_panes = "A11"


# ============================================================================
# 16) ANTROPOMETRIA  (perfil restrito ISAK + composição corporal + somatotipo)
# ============================================================================
wsAn = wb.create_sheet("Antropometria")
ANT_F, ANT_L = 8, 407
banner(wsAn, "ANTROPOMETRIA E COMPOSIÇÃO CORPORAL",
       "Avaliações longitudinais. Preencha as medidas em amarelo; densidade, % de gordura, massa magra, índices e "
       "somatotipo são calculados. Sexo e data de nascimento são lidos do Cadastro.", 48, "1F6F4A")
ANT_H = ["Data","Atleta","Momento","Avaliador","Massa (kg)","Estatura (cm)","Estatura Sentado (cm)","Envergadura (cm)",
         # dobras 9-18
         "Tríceps","Subescapular","Bíceps","Peitoral","Axilar Média","Suprailíaca","Supraespinhal","Abdominal",
         "Coxa Medial","Perna Medial",
         # perímetros 19-28
         "Braço Relaxado","Braço Contraído","Antebraço","Tórax","Cintura","Abdômen","Quadril","Coxa (1/3)",
         "Perna","Punho",
         # diâmetros 29-32
         "Biacromial","Biileocristal","Úmero","Fêmur",
         # calculados 33-48
         "Idade na Avaliação","Σ 4 Dobras (mm)","Σ 7 Dobras JP (mm)","Σ 8 Dobras (mm)","Densidade Corporal",
         "% de Gordura","Massa Gorda (kg)","Massa Magra (kg)","IMC","Índice Córmico","Cintura / Estatura",
         "AMB (cm²)","Endomorfia","Mesomorfia","Ectomorfia","Somatotipo Dominante"]
GRUPOS_A = [("IDENTIFICAÇÃO E MEDIDAS BÁSICAS", 1, 8, NAVY),
            ("DOBRAS CUTÂNEAS (mm)", 9, 18, GOLD),
            ("PERÍMETROS (cm)", 19, 28, BLUE3),
            ("DIÂMETROS ÓSSEOS (cm)", 29, 32, "6B3FA0"),
            ("CALCULADOS", 33, 48, "1F6F4A")]
for txt, c1, c2, cor in GRUPOS_A:
    wsAn.merge_cells(start_row=6, start_column=c1, end_row=6, end_column=c2)
    c = wsAn.cell(6, c1, txt)
    c.font = Font(name=F, size=9, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=cor)
    c.alignment = Alignment(horizontal="center", vertical="center"); c.border = BORDER
wsAn.row_dimensions[6].height = 18
cab_tabela(wsAn, 7, ANT_H)
larguras(wsAn, dict([("A",12),("B",24),("C",18),("D",16),("E",11),("F",12),("G",15),("H",15)] +
                    [(get_column_letter(i), 11) for i in range(9, 33)] +
                    [("AG",15),("AH",13),("AI",15),("AJ",13),("AK",14),("AL",12),("AM",13),("AN",13),
                     ("AO",9),("AP",13),("AQ",14),("AR",11),("AS",12),("AT",12),("AU",12),("AV",20)]))
nota(wsAn, 4, 1, "Protocolo: perfil restrito ISAK. Dobras em mm, perímetros e diâmetros em cm.  •  "
     "% de gordura pela equação de 7 dobras de Jackson & Pollock (1978 masculino / 1980 feminino) com conversão de "
     "Siri (1961) — a planilha escolhe a equação pelo sexo cadastrado.  •  Somatotipo pelo método Heath-Carter.  •  "
     "Σ 7 Dobras JP = peitoral + axilar média + tríceps + subescapular + abdominal + suprailíaca + coxa medial.", 48)
wsAn.row_dimensions[4].height = 34

CADR = "Cadastro!$B${}:$B${}".format(CAD_F, CAD_L)
def cadcol(letra):
    return "Cadastro!${0}${1}:${0}${2}".format(letra, CAD_F, CAD_L)

for r in range(ANT_F, ANT_L + 1):
    # idade na data da avaliação
    wsAn.cell(r, 33, '=IFERROR(DATEDIF(INDEX({1},MATCH($B{0},{2},0)),$A{0},"Y"),"")'.format(r, cadcol("D"), CADR))
    wsAn.cell(r, 34, '=IF(COUNT($I{0},$K{0},$J{0},$N{0})<4,"",$I{0}+$K{0}+$J{0}+$N{0})'.format(r))
    wsAn.cell(r, 35, '=IF(COUNT($L{0},$M{0},$I{0},$J{0},$P{0},$N{0},$Q{0})<7,"",'
                     '$L{0}+$M{0}+$I{0}+$J{0}+$P{0}+$N{0}+$Q{0})'.format(r))
    wsAn.cell(r, 36, '=IF(COUNT($I{0},$J{0},$K{0},$O{0},$P{0},$Q{0},$R{0},$N{0})<8,"",'
                     '$I{0}+$J{0}+$K{0}+$O{0}+$P{0}+$Q{0}+$R{0}+$N{0})'.format(r))
    # densidade corporal — Jackson & Pollock 7 dobras, por sexo
    wsAn.cell(r, 37, '=IF(OR($AI{0}="",$AG{0}=""),"",IF(IFERROR(INDEX({1},MATCH($B{0},{2},0)),"")="Feminino",'
                     '1.097-0.00046971*$AI{0}+0.00000056*$AI{0}^2-0.00012828*$AG{0},'
                     '1.112-0.00043499*$AI{0}+0.00000055*$AI{0}^2-0.00028826*$AG{0}))'
                     .format(r, cadcol("F"), CADR))
    wsAn.cell(r, 38, '=IF(OR($AK{0}="",$AK{0}=0),"",495/$AK{0}-450)'.format(r))
    wsAn.cell(r, 39, '=IF(OR($AL{0}="",$E{0}=""),"",$E{0}*$AL{0}/100)'.format(r))
    wsAn.cell(r, 40, '=IF(OR($AM{0}="",$E{0}=""),"",$E{0}-$AM{0})'.format(r))
    wsAn.cell(r, 41, '=IFERROR($E{0}/($F{0}/100)^2,"")'.format(r))
    wsAn.cell(r, 42, '=IFERROR($G{0}/$F{0}*100,"")'.format(r))
    wsAn.cell(r, 43, '=IFERROR($W{0}/$F{0},"")'.format(r))
    # área muscular do braço
    wsAn.cell(r, 44, '=IF(OR($S{0}="",$I{0}=""),"",($S{0}-PI()*$I{0}/10)^2/(4*PI())'
                     '-IF(IFERROR(INDEX({1},MATCH($B{0},{2},0)),"")="Feminino",6.5,10))'
                     .format(r, cadcol("F"), CADR))
    # Heath-Carter
    wsAn.cell(r, 45, '=IF(OR($I{0}="",$J{0}="",$O{0}="",$F{0}=""),"",'
                     '-0.7182+0.1451*(($I{0}+$J{0}+$O{0})*(170.18/$F{0}))'
                     '-0.00068*(($I{0}+$J{0}+$O{0})*(170.18/$F{0}))^2'
                     '+0.0000014*(($I{0}+$J{0}+$O{0})*(170.18/$F{0}))^3)'.format(r))
    wsAn.cell(r, 46, '=IF(OR($AE{0}="",$AF{0}="",$T{0}="",$AA{0}="",$F{0}=""),"",'
                     '0.858*$AE{0}+0.601*$AF{0}+0.188*($T{0}-$I{0}/10)+0.161*($AA{0}-$R{0}/10)'
                     '-0.131*$F{0}+4.5)'.format(r))
    wsAn.cell(r, 47, '=IF(OR($E{0}="",$F{0}=""),"",IF($F{0}/$E{0}^(1/3)>=40.75,0.732*($F{0}/$E{0}^(1/3))-28.58,'
                     'IF($F{0}/$E{0}^(1/3)>=38.25,0.463*($F{0}/$E{0}^(1/3))-17.63,0.1)))'.format(r))
    wsAn.cell(r, 48, '=IF(OR($AS{0}="",$AT{0}="",$AU{0}=""),"",'
                     'IF(AND($AT{0}>=$AS{0},$AT{0}>=$AU{0}),"Mesomorfo dominante",'
                     'IF(AND($AU{0}>=$AS{0},$AU{0}>=$AT{0}),"Ectomorfo dominante","Endomorfo dominante")))'.format(r))
corpo_tabela(wsAn, ANT_F, ANT_L, 1, 48)
for r in range(ANT_F, ANT_L + 1):
    for c in range(1, 49):
        if c >= 33:
            wsAn.cell(r, c).font = Font(name=F, size=9, bold=(c in (38, 40, 48)), color=NAVY2)
            wsAn.cell(r, c).fill = PatternFill("solid", fgColor=LIGHT)
        else:
            wsAn.cell(r, c).fill = PatternFill("solid", fgColor=GOLD_L)
            wsAn.cell(r, c).font = Font(name=F, size=9, color="0000FF")
    for c in (2, 4, 48):
        wsAn.cell(r, c).alignment = Alignment(horizontal="left", vertical="center", indent=1)
    wsAn.cell(r, 1).number_format = NF_DATE
    for c in range(5, 33):
        wsAn.cell(r, c).number_format = '0.0;;""'
    for c in (33, 34, 35, 36):
        wsAn.cell(r, c).number_format = '0;;""'
    wsAn.cell(r, 37).number_format = '0.0000;;""'
    for c in (38, 39, 40, 41, 42, 44, 45, 46, 47):
        wsAn.cell(r, c).number_format = '0.0;;""'
    wsAn.cell(r, 43).number_format = '0.00;;""'
dv(wsAn, ATLETAS_REF,           "B{}:B{}".format(ANT_F, ANT_L))
dv(wsAn, L("Momento do Teste"), "C{}:C{}".format(ANT_F, ANT_L))
wsAn.conditional_formatting.add("AL{}:AL{}".format(ANT_F, ANT_L),
    ColorScaleRule(start_type="num", start_value=6, start_color=GREEN, mid_type="num", mid_value=14,
                   mid_color=YELL, end_type="num", end_value=22, end_color=RED))
wsAn.freeze_panes = "C8"
wsAn.auto_filter.ref = "A7:AV{}".format(ANT_L)


# ============================================================================
# 17) PROGRAMA PF — documento "Preparação Física 2026/2027" do usuário
# ============================================================================
wsPF = wb.create_sheet("Programa PF")
banner(wsPF, "PROGRAMA DE PREPARAÇÃO FÍSICA 2026/2027 — SÉRIES 1 A 8",
       "Reprodução fiel do documento original (Adulto e Sub-19, fases Básica 1.1, 1.2, 2.1 e 3.1). As colunas "
       "%1RM, Carga e Pausa são sugestões editáveis; o restante é o conteúdo do documento.", 14, GOLD)
larguras(wsPF, {"A":8,"B":13,"C":11,"D":13,"E":12,"F":9,"G":22,"H":7,"I":46,"J":9,"K":16,"L":11,"M":11,"N":26})
PF_H = ["Série","Temporada","Categoria","Fase","Objetivo","Bloco","Esquema (séries/reps)","Ordem","Exercício",
        "Nº de Séries","Repetições","%1RM Sug.","Carga (kg)","Observações"]
cab_tabela(wsPF, 6, PF_H)

# (série, temporada, categoria, fase, objetivo, bloco, esquema, [exercícios])
PROGRAMA = [
 (1,"2026/2027","Adulto","Básica 1.1","Força","Core","3x25",
  ["Abdominal reto com braços esticados","Dorsal perdigueiro"]),
 (1,"2026/2027","Adulto","Básica 1.1","Força","Braço","8/8/6/6",
  ["Armação de braço","Supino deitado com halteres","Puxada alta"]),
 (1,"2026/2027","Adulto","Básica 1.1","Força","Perna","8/8/6/6",
  ["Agachamento profundo","Stiff com barra","Afundo lateral sem passada"]),
 (2,"2027","Adulto","Básica 1.1","Força","Core","3x25",
  ["Abdominal cruzado esticado","Dorsal reto"]),
 (2,"2027","Adulto","Básica 1.1","Força","Braço","8/8/6/6",
  ["Remada unilateral na polia","Bíceps / Tríceps","Pullover deitado com haltere"]),
 (2,"2027","Adulto","Básica 1.1","Força","Perna","8/8/6/6",
  ["Afundo frontal sem passada","Flexão de joelho unilateral","Agachamento sumô no minitramp"]),
 (3,"2026","Sub-19","Fase 1.2","Força 1","Braço","carga alta – 8/8/6",
  ["Supino sentado","Remada serrote","Pullover na polia"]),
 (3,"2026","Sub-19","Fase 1.2","Força 1","Perna","carga alta – 8/8/6",
  ["Agachamento profundo","Stiff","Afundo lateral com tornozeleira"]),
 (3,"2026","Sub-19","Fase 1.2","Força 1","Core","3x25",
  ["Abdominal infra","Dorsal perdigueiro"]),
 (4,"2026","Sub-19","Fase 1.2","Força 1","Braço","carga alta – 8/8/6",
  ["Tríceps / Bíceps","Puxada alta","Cross-over"]),
 (4,"2026","Sub-19","Fase 1.2","Força 1","Perna","carga alta – 8/8/6",
  ["Elevação de calcanhares","Afundo frontal sem passada com tornozeleira","Flexão de joelhos (unilateral)"]),
 (4,"2026","Sub-19","Fase 1.2","Força 1","Core","3x25",
  ["Abdominal reto","Dorsal reto"]),
 (5,"2026","Sub-19","Fase 2.1","Potência","Braço","carga alta menos – 10/8/6",
  ["Supino sentado","Remada unilateral na polia","Pullover na polia"]),
 (5,"2026","Sub-19","Fase 2.1","Potência","Perna","carga alta menos – 10/8/6",
  ["Agachamento","Stiff","Salto no caixote"]),
 (5,"2026","Sub-19","Fase 2.1","Potência","Core","3x20",
  ["Abdominal cruzado com anilha"]),
 (6,"2026","Sub-19","Fase 2.1","Potência","Braço","carga alta menos – 10/8/6",
  ["Tríceps / Bíceps","Puxada alta","Arremesso de bola"]),
 (6,"2026","Sub-19","Fase 2.1","Potência","Perna","carga alta menos – 10/8/6",
  ["Elevação de calcanhares","Salto unilateral com sobrepeso","Flexão de joelhos"]),
 (6,"2026","Sub-19","Fase 2.1","Potência","Core","3x20",
  ["Abdominal remador com anilha"]),
 (7,"2026","Sub-19","Fase 3.1","Força 3","Braço","carga alta – 8/8/6/6",
  ["Tríceps","Remada unilateral na polia (perna contrária à frente)","Pullover com haltere no banco"]),
 (7,"2026","Sub-19","Fase 3.1","Força 3","Perna","carga alta – 8/8/6/6",
  ["Agachamento profundo","Stiff","Afundo lateral com sobrepeso"]),
 (7,"2026","Sub-19","Fase 3.1","Força 3","Core","3x25",
  ["Abdominal cruzado com anilha de 20"]),
 (8,"2026","Sub-19","Fase 3.1","Força 3","Braço","carga alta – 8/8/6/6",
  ["Bíceps","Puxada alta","Cross-over supino"]),
 (8,"2026","Sub-19","Fase 3.1","Força 3","Perna","carga alta – 8/8/6/6",
  ["Elevação de calcanhares","Afundo frontal com sobrepeso (sem passada)","Flexão de joelhos"]),
 (8,"2026","Sub-19","Fase 3.1","Força 3","Core","3x25",
  ["Abdominal reto canivete"]),
]
PCT_OBJ = {"Força": 0.75, "Força 1": 0.80, "Potência": 0.65, "Força 3": 0.85}
PAUSA_BLOCO = {"Core": 45, "Braço": 120, "Perna": 150}
PF_F = 7
r = PF_F
COR_SERIE = {1: LIGHT, 2: LIGHT2}
for serie, temp, cat, fase, obj, bloco, esq, exs in PROGRAMA:
    if "x" in esq:
        nser = int(esq.split("x")[0])
        reps = esq.split("x")[1]
    else:
        partes_esq = esq.split("–")[-1].strip()
        nser = len(partes_esq.split("/"))
        reps = partes_esq
    for i, ex in enumerate(exs, start=1):
        wsPF.cell(r, 1, serie); wsPF.cell(r, 2, temp); wsPF.cell(r, 3, cat)
        wsPF.cell(r, 4, fase); wsPF.cell(r, 5, obj); wsPF.cell(r, 6, bloco)
        wsPF.cell(r, 7, esq); wsPF.cell(r, 8, i); wsPF.cell(r, 9, ex)
        wsPF.cell(r, 10, nser); wsPF.cell(r, 11, reps)
        if bloco != "Core":
            wsPF.cell(r, 12, PCT_OBJ[obj])
        wsPF.cell(r, 14, "Do documento original")
        r += 1
PF_L = r - 1
corpo_tabela(wsPF, PF_F, PF_L, 1, 14)
for rr in range(PF_F, PF_L + 1):
    serie = wsPF.cell(rr, 1).value
    fill = PatternFill("solid", fgColor=LIGHT if serie % 2 else LIGHT2)
    for c in range(1, 15):
        wsPF.cell(rr, c).fill = fill
        wsPF.cell(rr, c).font = Font(name=F, size=9, color=NAVY2)
    for c in (12, 13):
        wsPF.cell(rr, c).fill = PatternFill("solid", fgColor=GOLD_L)
        wsPF.cell(rr, c).font = Font(name=F, size=9, color="0000FF")
    wsPF.cell(rr, 9).alignment = Alignment(horizontal="left", vertical="center", indent=1)
    wsPF.cell(rr, 9).font = Font(name=F, size=9, bold=True, color=NAVY)
    wsPF.cell(rr, 12).number_format = NF_PCT
    wsPF.cell(rr, 13).number_format = '0.0;;""'
    wsPF.cell(rr, 14).alignment = Alignment(horizontal="left", vertical="center", indent=1)
    wsPF.cell(rr, 14).font = Font(name=F, size=8, italic=True, color=GREY_T)
secao(wsPF, PF_L + 2, "COMO ESTE PROGRAMA SE ENCAIXA NA PERIODIZAÇÃO EM BLOCOS (aba Bloco Base)", 14, 1)
MAPA = [
 ("Séries 1 e 2 — Básica 1.1 (Adulto), 8/8/6/6",
  "Bloco de ACUMULAÇÃO — semanas 1 a 4. Volume alto, intensidade moderada (70–80% 1RM). É aqui que entram a técnica "
  "de LPO com carga leve e a pliometria de baixa intensidade."),
 ("Séries 3 e 4 — Fase 1.2 'Força 1', carga alta 8/8/6",
  "Fim da ACUMULAÇÃO e início da TRANSMUTAÇÃO — semanas 4 a 6. Intensidade sobe para 80–88% 1RM no agachamento."),
 ("Séries 5 e 6 — Fase 2.1 'Potência', carga alta menos 10/8/6 (já com salto no caixote e arremesso de bola)",
  "Bloco de TRANSMUTAÇÃO — semanas 6 e 7, incluindo o microciclo de CHOQUE. É o momento do jump squat com 30% 1RM, "
  "do drop jump e do power clean em intensidade alta."),
 ("Séries 7 e 8 — Fase 3.1 'Força 3', carga alta 8/8/6/6",
  "Bloco de REALIZAÇÃO — semana 8 e o mesociclo seguinte. Volume baixo, intensidade ≥ 90% 1RM, pliometria de alta "
  "qualidade e pouco volume."),
]
rr = PF_L + 3
for tit, desc in MAPA:
    c = wsPF.cell(rr, 1, tit)
    wsPF.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=7)
    c.font = Font(name=F, size=9, bold=True, color=NAVY2)
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    wsPF.merge_cells(start_row=rr, start_column=8, end_row=rr, end_column=14)
    d = wsPF.cell(rr, 8, desc)
    d.font = Font(name=F, size=9)
    d.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    for c2 in range(1, 15):
        wsPF.cell(rr, c2).border = BORDER
        if (rr % 2) == 0:
            wsPF.cell(rr, c2).fill = PatternFill("solid", fgColor=LIGHT2)
    wsPF.row_dimensions[rr].height = 34
    rr += 1
wsPF.freeze_panes = "A7"
wsPF.auto_filter.ref = "A6:N{}".format(PF_L)
PF_EXERCICIOS = sorted({ex for *_, exs in PROGRAMA for ex in exs})


# ============================================================================
# 18) BLOCO BASE — 8 semanas de periodização em blocos (força máxima e potência)
# ============================================================================
wsB = wb.create_sheet("Bloco Base")
banner(wsB, "PERIODIZAÇÃO EM BLOCOS — 2 MESES DE PERÍODO DE BASE",
       "Acumulação → Transmutação → Realização, com microciclos de CHOQUE nas semanas 3 e 7 e descarga na semana 4. "
       "Foco em força máxima, força explosiva com movimentos de LPO, agachamento e pliometria.", 18, "9C0006")
larguras(wsB, {"A":8,"B":13,"C":15,"D":17,"E":34,"F":10,"G":11,"H":15,"I":11,"J":26,"K":14,"L":11,
               "M":13,"N":18,"O":11,"P":12,"Q":12,"R":40})

secao(wsB, 4, "PARÂMETROS DO BLOCO", 18, 1)
BL_ID = [("Início do bloco (2ª feira)", date(2026, 9, 7)), ("Sessões de força por semana", 3),
         ("Sessões de quadra por semana", 5), ("Atletas no bloco", 12)]
for i, (lab, val) in enumerate(BL_ID):
    rotulo(wsB, 5 + i, 1, lab)
    entrada(wsB, 5 + i, 3, val, NF_DATE if isinstance(val, date) else '0', largura_merge=2)
wsB.merge_cells(start_row=5, start_column=6, end_row=8, end_column=18)
obj = wsB.cell(5, 6,
    "OBJETIVO DO BLOCO — elevar a força máxima de membros inferiores e a taxa de produção de força, criando a base "
    "sobre a qual a potência será expressa nos blocos competitivos.\n"
    "LÓGICA — a periodização em blocos concentra poucas capacidades por vez, em vez de desenvolver muitas ao mesmo "
    "tempo (Issurin, 2010; Stone et al., 2021). Estudos com esportes coletivos mostram vantagem dos blocos sobre a "
    "periodização tradicional em força, potência e salto (Rønnestad et al., 2018; Manchado et al., 2017).\n"
    "CHOQUE — as semanas 3 e 7 concentram a carga deliberadamente para induzir overreaching funcional; espera-se "
    "queda transitória do CMJ nessas semanas, com supercompensação após a descarga (Micke et al., 2026). MONITORE o "
    "CMJ e o wellness nessas semanas: queda de CMJ acima de 10% por mais de 5 dias exige reduzir a carga.")
obj.font = Font(name=F, size=9, color=NAVY)
obj.fill = PatternFill("solid", fgColor=LIGHT)
obj.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
obj.border = BORDER

secao(wsB, 10, "MATRIZ SEMANA A SEMANA", 18, 1)
BL_H = ["Semana","Data Início","Bloco","Microciclo","Ênfase Principal","Sessões de Força","Séries MMII/Sessão",
        "Agachamento (sér.×reps)","%1RM Agach.","Exercício de LPO","LPO (sér.×reps)","%1RM LPO",
        "Pliometria (contatos/sem)","Intensidade Pliométrica","Índice de Volume","Índice de Intensidade",
        "ACWR Alvo","O que monitorar nesta semana"]
cab_tabela(wsB, 11, BL_H)
BL_F = 12
BLOCO_8 = [
 (1,"Acumulação","Incorporação","Adaptação anatômica + técnica de LPO",3,13,"4 × 8",0.65,
  "Power clean a partir do joelho (técnica)","5 × 3",0.55,60,"Baixa (solo, CMJ e saltos horizontais)",0.75,0.62,1.00,
  "Qualidade técnica do power clean e da aterrissagem. Sem busca de carga."),
 (2,"Acumulação","Ordinário","Força máxima de base",3,15,"4 × 6",0.75,
  "Clean pull","4 × 4",0.70,90,"Baixa a média (saltos sobre barreiras 40 cm)",0.88,0.72,1.10,
  "Velocidade média no agachamento a 75% ≥ 0,62 m/s. PSE das sessões de força entre 6 e 7."),
 (3,"Acumulação","CHOQUE","Pico de volume da fase de acumulação",4,18,"5 × 5",0.80,
  "Power clean","5 × 3",0.75,120,"Média (barreiras 50 cm + salto no caixote)",1.00,0.78,1.30,
  "SEMANA DE CHOQUE — espere queda de CMJ e piora do Hooper. Monitore diariamente e não some carga de quadra."),
 (4,"Descarga","Recuperativo","Assimilação e supercompensação",2,8,"3 × 5",0.70,
  "Técnica leve (hang high pull)","3 × 3",0.60,45,"Baixa (qualidade, volume reduzido)",0.45,0.66,0.75,
  "Retorno do CMJ ao valor da semana 1 ou acima. Se não voltar, prolongue a descarga."),
 (5,"Transmutação","Ordinário","Força máxima",3,13,"5 × 4",0.85,
  "Power clean","5 × 2",0.82,90,"Média (drop jump 40 cm)",0.75,0.85,1.10,
  "Reteste indireto do 1RM pela velocidade. Ganho esperado de 3 a 6% no agachamento."),
 (6,"Transmutação","Ordinário","Força-velocidade",3,12,"4 × 3",0.88,
  "Clean pull (carga supramáxima)","4 × 3",0.95,110,"Média a alta (drop jump 40–50 cm + unilaterais)",0.70,0.90,1.15,
  "Introduzir jump squat a 30% 1RM. Velocidade média do jump squat ≥ 1,0 m/s."),
 (7,"Transmutação","CHOQUE","Sobrecarga concentrada de potência",4,15,"4 × 3 (cluster 2+2)",0.90,
  "Power clean","6 × 2",0.88,140,"Alta (drop jump 50 cm + saltos com sobrecarga)",0.85,0.93,1.45,
  "SEGUNDA SEMANA DE CHOQUE. ACWR chega a ~1,45 de propósito; reduza o volume de saltos de quadra para compensar."),
 (8,"Realização","Polimento","Expressão de potência",3,9,"3 × 2",0.92,
  "Power clean","4 × 1–2",0.90,70,"Alta qualidade, volume baixo (CMJ e DJ máximos)",0.42,0.95,0.85,
  "Reteste de CMJ, salto de ataque e 1RM. Espera-se o pico do bloco no fim desta semana."),
]
for i, w in enumerate(BLOCO_8):
    r = BL_F + i
    (sem, bloco, micro, enf, ses, ser, agach, pagach, lpo, slpo, plpo, plio, iplio, vol, inten, acwr, mon) = w
    wsB.cell(r, 1, sem)
    wsB.cell(r, 2, '=IF($C$5="","",$C$5+($A{}-1)*7)'.format(r))
    for c, v in ((3,bloco),(4,micro),(5,enf),(6,ses),(7,ser),(8,agach),(9,pagach),(10,lpo),(11,slpo),
                 (12,plpo),(13,plio),(14,iplio),(15,vol),(16,inten),(17,acwr),(18,mon)):
        wsB.cell(r, c, v)
corpo_tabela(wsB, BL_F, BL_F + 7, 1, 18)
for i in range(8):
    r = BL_F + i
    choque = wsB.cell(r, 4).value == "CHOQUE"
    desc = wsB.cell(r, 3).value == "Descarga"
    base = RED if choque else ("DDEBF7" if desc else (LIGHT if i % 2 == 0 else LIGHT2))
    for c in range(1, 19):
        wsB.cell(r, c).fill = PatternFill("solid", fgColor=base)
        wsB.cell(r, c).font = Font(name=F, size=9, bold=(c in (1, 3, 4)),
                                   color=RED_T if choque else NAVY2)
    for c in (5, 10, 14, 18):
        wsB.cell(r, c).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    wsB.cell(r, 2).number_format = NF_DATE
    for c in (9, 12, 15, 16):
        wsB.cell(r, c).number_format = NF_PCT
    wsB.cell(r, 17).number_format = NF_DEC
    wsB.row_dimensions[r].height = 34
tot = BL_F + 8
for c in range(1, 19):
    wsB.cell(tot, c).fill = PatternFill("solid", fgColor=NAVY2); wsB.cell(tot, c).border = BORDER
    wsB.cell(tot, c).font = Font(name=F, size=9, bold=True, color=WHITE)
    wsB.cell(tot, c).alignment = Alignment(horizontal="center", vertical="center")
wsB.cell(tot, 1, "TOTAL").alignment = Alignment(horizontal="left", vertical="center", indent=1)
wsB.cell(tot, 6, "=SUM(F{}:F{})".format(BL_F, BL_F + 7)).number_format = "0"
wsB.cell(tot, 13, "=SUM(M{}:M{})".format(BL_F, BL_F + 7)).number_format = NF_UA
wsB.cell(tot, 15, "=AVERAGE(O{}:O{})".format(BL_F, BL_F + 7)).number_format = NF_PCT
wsB.cell(tot, 16, "=AVERAGE(P{}:P{})".format(BL_F, BL_F + 7)).number_format = NF_PCT

chB = LineChart(); chB.title = "Dinâmica do Bloco: Volume × Intensidade (índices relativos)"
chB.height = 8; chB.width = 19; chB.y_axis.title = "% relativo"
chB.add_data(Reference(wsB, min_col=15, max_col=16, min_row=11, max_row=BL_F + 7), titles_from_data=True)
chB.set_categories(Reference(wsB, min_col=1, min_row=BL_F, max_row=BL_F + 7))
for s in chB.series:
    s.graphicalProperties.line.width = 28000
wsB.add_chart(chB, "A23")
chB2 = BarChart(); chB2.type = "col"; chB2.title = "Contatos Pliométricos por Semana"
chB2.height = 8; chB2.width = 19; chB2.legend = None; chB2.y_axis.title = "contatos"
chB2.add_data(Reference(wsB, min_col=13, min_row=11, max_row=BL_F + 7), titles_from_data=True)
chB2.set_categories(Reference(wsB, min_col=1, min_row=BL_F, max_row=BL_F + 7))
wsB.add_chart(chB2, "J23")

secao(wsB, 40, "SESSÕES-MODELO DA SEMANA (A, B e C) — exercícios do seu programa + LPO e pliometria", 18, 1)
cab_tabela(wsB, 41, ["Sessão","Dia","Ordem","Exercício","Objetivo","Ref. VBT","Observação de execução"])
SESSOES_BLOCO = [
 ("A — Força MMII + LPO","Segunda", [
   ("Mobilidade de tornozelo e quadril","Aquecimento","—","6 exercícios × 8 rep, antes da barra"),
   ("Power clean a partir do joelho","Força explosiva (LPO)","—","Interromper a série se a velocidade da barra cair visivelmente"),
   ("Agachamento profundo","Força máxima","Agachamento","Exercício-âncora do bloco; controlar pela velocidade média"),
   ("Stiff com barra","Força máxima (cadeia posterior)","Terra","Excêntrica de 3 s nas semanas 1 a 4"),
   ("Afundo lateral sem passada","Força unilateral","—","Do seu programa (Séries 1, 3 e 7)"),
   ("Nórdico de isquiotibiais (excêntrico)","Preventivo","—","3 × 6, 2x por semana durante todo o bloco"),
   ("Abdominal reto com braços esticados + Dorsal perdigueiro","Core","—","3 × 25, do seu programa (Série 1)")]),
 ("B — Potência, pliometria e MMSS","Quarta", [
   ("Rotadores do ombro com elástico","Preventivo (ativação)","—","3 × 15, antes de qualquer trabalho de ombro"),
   ("Agachamento com salto sob carga (jump squat)","Potência","Agachamento","30% 1RM a partir da semana 6; pausa completa"),
   ("Salto no caixote / Drop jump","Pliometria","—","Do seu programa (Série 5); altura conforme a matriz semanal"),
   ("Saltos consecutivos sobre barreiras","Pliometria","—","Contato mínimo com o solo; contar os contatos"),
   ("Supino deitado com halteres","Força MMSS","Supino","Do seu programa (Série 1)"),
   ("Remada unilateral na polia","Força MMSS","—","Do seu programa (Séries 2, 5 e 7)"),
   ("Arremesso de bola","Potência MMSS","—","Do seu programa (Série 6)"),
   ("Abdominal cruzado esticado + Dorsal reto","Core","—","3 × 25, do seu programa (Série 2)")]),
 ("C — Força-velocidade e unilateral","Sexta", [
   ("Mobilidade de tornozelo e quadril","Aquecimento","—","—"),
   ("Snatch pull / Hang high pull","Força explosiva (LPO)","—","Puxadas sem fase de recepção: mesmo estímulo de tripla extensão, técnica mais simples"),
   ("Agachamento sumô no minitramp","Força-velocidade","Agachamento","Do seu programa (Série 2)"),
   ("Elevação de calcanhares","Preventivo (tornozelo)","—","Do seu programa (Séries 4, 6 e 8); descida lenta de 3 s"),
   ("Afundo frontal sem passada com sobrepeso","Força unilateral","—","Do seu programa (Séries 2, 4 e 8)"),
   ("Salto unilateral com sobrepeso","Pliometria unilateral","—","Do seu programa (Série 6)"),
   ("Pullover deitado com haltere","Força MMSS","—","Do seu programa (Série 2)"),
   ("Sprints de 10 e 20 m com mudança de direção","Velocidade","—","Só nas semanas sem choque"),
   ("Abdominal remador com anilha","Core","—","3 × 20, do seu programa (Série 6)")]),
]
r = 42
for nome, dia, itens in SESSOES_BLOCO:
    ini = r
    for i, (ex, objx, vbt, obs) in enumerate(itens, start=1):
        wsB.cell(r, 1, nome if i == 1 else "")
        wsB.cell(r, 2, dia if i == 1 else "")
        wsB.cell(r, 3, i); wsB.cell(r, 4, ex); wsB.cell(r, 5, objx); wsB.cell(r, 6, vbt); wsB.cell(r, 7, obs)
        r += 1
    wsB.merge_cells(start_row=ini, start_column=1, end_row=r - 1, end_column=1)
    wsB.merge_cells(start_row=ini, start_column=2, end_row=r - 1, end_column=2)
SESS_L = r - 1
corpo_tabela(wsB, 42, SESS_L, 1, 7)
for rr in range(42, SESS_L + 1):
    for c in range(1, 8):
        wsB.cell(rr, c).font = Font(name=F, size=9, color=NAVY2)
        wsB.cell(rr, c).fill = PatternFill("solid", fgColor=LIGHT if (rr % 2) else LIGHT2)
    for c in (1, 4, 5, 7):
        wsB.cell(rr, c).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    wsB.cell(rr, 4).font = Font(name=F, size=9, bold=True, color=NAVY)
    wsB.cell(rr, 1).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
wsB.column_dimensions["D"].width = 46
wsB.column_dimensions["E"].width = 28
wsB.column_dimensions["G"].width = 62
nota(wsB, SESS_L + 2, 1, "As sessões A, B e C já estão prescritas semana a semana na aba 'Prescrição Força', com "
     "séries, repetições, %1RM, carga calculada e velocidade-alvo para o atleta de referência.", 18)
wsB.freeze_panes = "A12"


# ============================================================================
# 19) FORÇA 1RM — testes diretos e estimados + matriz de 1RM atual
# ============================================================================
wsR = wb.create_sheet("Força 1RM")
RM_F, RM_L = 8, 407
banner(wsR, "CONTROLE DE FORÇA MÁXIMA (1RM)",
       "Registre testes diretos ou séries até a falha técnica: a planilha estima o 1RM por Epley e Brzycki. A matriz "
       "à direita mantém o 1RM ATUAL de cada atleta e alimenta o cálculo de carga da aba Prescrição Força.", 26, "9C0006")
RM_H = ["Data","Semana","Atleta","Exercício","Método","Carga (kg)","Repetições","1RM Epley (kg)","1RM Brzycki (kg)",
        "1RM Adotado (kg)","Massa Corporal (kg)","Força Relativa (×massa)","Velocidade Média (m/s)","Observações"]
cab_tabela(wsR, 7, RM_H)
larguras(wsR, {"A":12,"B":8,"C":24,"D":30,"E":24,"F":11,"G":11,"H":13,"I":14,"J":14,"K":16,"L":15,"M":16,"N":24,
               "O":3,"P":24,"Q":14,"R":14,"S":13,"T":14,"U":13,"V":15,"W":13,"X":12,"Y":13,"Z":15})
nota(wsR, 4, 1, "Epley: 1RM = carga × (1 + reps/30).  •  Brzycki: 1RM = carga ÷ (1,0278 − 0,0278 × reps).  "
     "Ambas perdem precisão acima de ~10 repetições — use 3 a 6 repetições.  "
     "A estimativa do 1RM pela velocidade da barra tende a SUPERESTIMAR a força real (Greig et al., 2023; "
     "LeMense et al., 2024): sempre que possível, teste o 1RM diretamente.", 14)
wsR.row_dimensions[4].height = 30
for r in range(RM_F, RM_L + 1):
    wsR.cell(r, 2, '=IF($A{0}="",0,INT(($A{0}-Macrociclo!$C$11)/7)+1)'.format(r))
    wsR.cell(r, 8, '=IF(OR($F{0}="",$G{0}=""),"",ROUND($F{0}*(1+$G{0}/30),1))'.format(r))
    wsR.cell(r, 9, '=IF(OR($F{0}="",$G{0}="",$G{0}>36),"",ROUND($F{0}/(1.0278-0.0278*$G{0}),1))'.format(r))
    wsR.cell(r, 10, '=IF($E{0}="Direto (1RM real)",$F{0},IF(OR($H{0}="",$I{0}=""),"",'
                    'ROUND(AVERAGE($H{0},$I{0}),1)))'.format(r))
    wsR.cell(r, 11, '=IF($C{0}="","",IFERROR(INDEX(Cadastro!$P${1}:$P${2},MATCH($C{0},Cadastro!$B${1}:$B${2},0)),""))'
             .format(r, CAD_F, CAD_L))
    wsR.cell(r, 12, '=IFERROR($J{0}/$K{0},"")'.format(r))
corpo_tabela(wsR, RM_F, RM_L, 1, 14)
for r in range(RM_F, RM_L + 1):
    for c in (1, 3, 4, 5, 6, 7, 13, 14):
        wsR.cell(r, c).fill = PatternFill("solid", fgColor=GOLD_L)
        wsR.cell(r, c).font = Font(name=F, size=9, color="0000FF")
    for c in (2, 8, 9, 10, 11, 12):
        wsR.cell(r, c).font = Font(name=F, size=9, bold=(c == 10), color=NAVY2)
        wsR.cell(r, c).fill = PatternFill("solid", fgColor=LIGHT)
    for c in (3, 4, 5, 14):
        wsR.cell(r, c).alignment = Alignment(horizontal="left", vertical="center", indent=1)
    wsR.cell(r, 1).number_format = NF_DATE
    wsR.cell(r, 2).number_format = '0;;""'
    for c in (6, 8, 9, 10, 11):
        wsR.cell(r, c).number_format = '0.0;;""'
    wsR.cell(r, 12).number_format = '0.00;;""'
    wsR.cell(r, 13).number_format = '0.00;;""'
EX_1RM = ["Agachamento profundo","Agachamento frontal","Stiff com barra","Levantamento terra",
          "Supino sentado","Supino deitado com halteres","Remada serrote","Power clean","Clean pull"]
dv(wsR, ATLETAS_REF,       "C{}:C{}".format(RM_F, RM_L))
dv(wsR, L("Método de 1RM"),"E{}:E{}".format(RM_F, RM_L))
dv(wsR, "'Força 1RM'!$Q$7:$Y$7", "D{}:D{}".format(RM_F, RM_L))

secao(wsR, 5, "1RM ATUAL POR ATLETA (última avaliação registrada, em kg)", 25, 16)
cab_tabela(wsR, 7, ["Atleta"] + EX_1RM, col0=16)
for i in range(CAD_L - CAD_F + 1):
    r = RM_F + i
    a = CAD_F + i
    wsR.cell(r, 16, '=IF(Cadastro!$B{0}="","",Cadastro!$B{0})'.format(a))
    for j in range(9):
        cl = get_column_letter(17 + j)
        wsR.cell(r, 17 + j,
          '=IF($P{0}="",0,IFERROR(AVERAGEIFS($J${1}:$J${2},$C${1}:$C${2},$P{0},$D${1}:$D${2},{3}$7,'
          '$A${1}:$A${2},SUMPRODUCT(MAX(($C${1}:$C${2}=$P{0})*($D${1}:$D${2}={3}$7)*$A${1}:$A${2}))),0))'
          .format(r, RM_F, RM_L, cl))
RM_MAT_L = RM_F + (CAD_L - CAD_F)
corpo_tabela(wsR, RM_F, RM_MAT_L, 16, 25)
for r in range(RM_F, RM_MAT_L + 1):
    for c in range(16, 26):
        wsR.cell(r, c).font = Font(name=F, size=9, bold=(c > 16), color=NAVY2)
        wsR.cell(r, c).fill = PatternFill("solid", fgColor=LIGHT if (r - RM_F) % 2 == 0 else LIGHT2)
        wsR.cell(r, c).number_format = '0.0;;""'
    wsR.cell(r, 16).alignment = Alignment(horizontal="left", vertical="center", indent=1)
    wsR.cell(r, 16).number_format = "General"
wsR.conditional_formatting.add("Q{}:Q{}".format(RM_F, RM_MAT_L),
    ColorScaleRule(start_type="min", start_color="FFFFFF", end_type="max", end_color="A9D08E"))
wsR.freeze_panes = "C8"


# ============================================================================
# 20) PRESCRIÇÃO FORÇA — programa das 8 semanas, %1RM, carga e velocidade-alvo
# ============================================================================
wsPQ = wb.create_sheet("Prescrição Força")
PRF_F, PRF_L = 8, 407
banner(wsPQ, "PRESCRIÇÃO E CONTROLE DO TREINO DE FORÇA E POTÊNCIA",
       "As 8 semanas do bloco de base já vêm prescritas. A carga em kg sai de %1RM × 1RM atual do atleta de "
       "referência; a velocidade-alvo sai da tabela carga-velocidade (editável) à direita.", 28, "9C0006")
larguras(wsPQ, {"A":12,"B":9,"C":15,"D":10,"E":22,"F":7,"G":44,"H":18,"I":13,"J":8,"K":9,"L":9,"M":12,"N":14,
                "O":12,"P":13,"Q":15,"R":15,"S":15,"T":16,"U":7,"V":8,"W":10,"X":10,"Y":13,"Z":26,"AA":24,
                "AB":26,"AC":3,"AD":9,"AE":15,"AF":13,"AG":13})
secao(wsPQ, 3, "PARÂMETROS E RESUMO DA SEMANA", 28, 1)
rotulo(wsPQ, 4, 1, "Atleta de referência:")
selPQ = entrada(wsPQ, 4, 3, NOMES[0], largura_merge=4)
selPQ.font = Font(name=F, size=11, bold=True, color="0000FF")
dv(wsPQ, ATLETAS_REF, "C4")
rotulo(wsPQ, 5, 1, "Semana do bloco (1 a 8):")
entrada(wsPQ, 5, 3, 1, '0', largura_merge=2)
FW = '$B${0}:$B${1},$C$5'.format(PRF_F, PRF_L)
kpi(wsPQ, 4, 8,  "EXERCÍCIOS NA SEMANA", '=COUNTIFS({})'.format(FW), '0', NAVY2, 3)
kpi(wsPQ, 4, 11, "SÉRIES TOTAIS", '=SUMIFS($J${0}:$J${1},{2})'.format(PRF_F, PRF_L, FW), '0', NAVY2, 3)
kpi(wsPQ, 4, 14, "REPETIÇÕES TOTAIS", '=SUMIFS($X${0}:$X${1},{2})'.format(PRF_F, PRF_L, FW), NF_UA, NAVY2, 3)
kpi(wsPQ, 4, 17, "TONELAGEM DA SEMANA (kg)", '=SUMIFS($Y${0}:$Y${1},{2})'.format(PRF_F, PRF_L, FW), NF_UA, GOLD, 3)
kpi(wsPQ, 4, 20, "INTENSIDADE MÉDIA RELATIVA",
    '=IFERROR(SUMPRODUCT(($B${0}:$B${1}=$C$5)*$X${0}:$X${1}*$L${0}:$L${1})'
    '/SUMPRODUCT(($B${0}:$B${1}=$C$5)*$X${0}:$X${1}*($L${0}:$L${1}<>"")),0)'.format(PRF_F, PRF_L), NF_PCT, GOLD, 3)
kpi(wsPQ, 4, 23, "CONTATOS PLIOMÉTRICOS",
    '=SUMIFS($X${0}:$X${1},{2},$H${0}:$H${1},"Pliometria")'.format(PRF_F, PRF_L, FW), NF_UA, "6B3FA0", 3)
nota(wsPQ, 6, 1, "Preencha em amarelo o que foi EXECUTADO (carga usada, velocidade obtida, perda de velocidade, RIR e PSE). "
     "A tonelagem usa a carga usada quando informada e, na falta dela, a carga calculada.", 28)

PRF_H = ["Data","Semana do Bloco","Bloco","Sessão","Destinatário","Ordem","Exercício","Objetivo","Ref. VBT",
         "Séries","Reps","%1RM","1RM Atual (kg)","Carga Calculada (kg)","Carga Usada (kg)","Velocidade Alvo (m/s)",
         "Velocidade Obtida (m/s)","Perda de Vel. Limite (%)","Perda de Vel. Obtida (%)","Situação VBT",
         "RIR","PSE","Pausa (s)","Reps Totais","Tonelagem (kg)","Zona de Intensidade","Observações","Ref. 1RM"]
cab_tabela(wsPQ, 7, PRF_H)
F1 = "'Força 1RM'"
for r in range(PRF_F, PRF_L + 1):
    wsPQ.cell(r, 2,  '=IF($A{0}="",0,INT(($A{0}-\'Bloco Base\'!$C$5)/7)+1)'.format(r))
    wsPQ.cell(r, 13, '=IF($AB{0}="","",IFERROR(INDEX({1}!$Q${2}:$Y${3},'
                     'MATCH(IF($E{0}="Equipe (todos)",$C$4,$E{0}),{1}!$P${2}:$P${3},0),'
                     'MATCH($AB{0},{1}!$Q$7:$Y$7,0)),""))'.format(r, F1, RM_F, RM_MAT_L))
    wsPQ.cell(r, 14, '=IFERROR(IF(OR($L{0}="",$M{0}="",$M{0}=0),0,ROUND($L{0}*$M{0}*2,0)/2),0)'.format(r))
    wsPQ.cell(r, 16, '=IFERROR(IF(OR($L{0}="",$I{0}="—",$I{0}=""),"",'
                     'IF($I{0}="Agachamento",INDEX($AE$8:$AE$20,MATCH($L{0},$AD$8:$AD$20,1)),'
                     'IF($I{0}="Supino",INDEX($AF$8:$AF$20,MATCH($L{0},$AD$8:$AD$20,1)),'
                     'IF($I{0}="Terra",INDEX($AG$8:$AG$20,MATCH($L{0},$AD$8:$AD$20,1)),"")))),"")'.format(r))
    wsPQ.cell(r, 20, '=IF(OR($R{0}="",$S{0}=""),"",IF($S{0}>$R{0},"Acima do limite","Dentro do limite"))'.format(r))
    wsPQ.cell(r, 24, '=IFERROR(IF(OR($J{0}="",$K{0}=""),0,$J{0}*$K{0}),0)'.format(r))
    wsPQ.cell(r, 25, '=IFERROR(IF($X{0}=0,0,$X{0}*IF($O{0}="",$N{0},$O{0})),0)'.format(r))
    wsPQ.cell(r, 26, '=IF($L{0}="","",IF($L{0}<0.5,"Balística / Velocidade (<50%)",'
                     'IF($L{0}<0.7,"Velocidade-Força (50–69%)",IF($L{0}<0.8,"Força-Velocidade (70–79%)",'
                     'IF($L{0}<0.9,"Força Máxima (80–89%)","Força Máxima Alta (≥90%)")))))'.format(r))
corpo_tabela(wsPQ, PRF_F, PRF_L, 1, 28)
CALC_PQ = (2, 13, 14, 16, 20, 24, 25, 26)
for r in range(PRF_F, PRF_L + 1):
    for c in range(1, 29):
        if c in CALC_PQ:
            wsPQ.cell(r, c).font = Font(name=F, size=9, bold=(c in (14, 16, 25)), color=NAVY2)
            wsPQ.cell(r, c).fill = PatternFill("solid", fgColor=LIGHT)
        else:
            wsPQ.cell(r, c).fill = PatternFill("solid", fgColor=GOLD_L)
            wsPQ.cell(r, c).font = Font(name=F, size=9, color="0000FF")
    for c in (5, 7, 8, 26, 27, 28):
        wsPQ.cell(r, c).alignment = Alignment(horizontal="left", vertical="center", indent=1)
    wsPQ.cell(r, 7).font = Font(name=F, size=9, bold=True, color=NAVY)
    wsPQ.cell(r, 1).number_format = NF_DATE
    wsPQ.cell(r, 2).number_format = '0;;""'
    wsPQ.cell(r, 12).number_format = NF_PCT
    for c in (13, 14, 15):
        wsPQ.cell(r, c).number_format = '0.0;;""'
    for c in (16, 17):
        wsPQ.cell(r, c).number_format = '0.00;;""'
    for c in (18, 19):
        wsPQ.cell(r, c).number_format = NF_PCT
    for c in (24, 25):
        wsPQ.cell(r, c).number_format = NF_UA
dv(wsPQ, DEST_REF,               "E{}:E{}".format(PRF_F, PRF_L))
dv(wsPQ, L("Objetivo de Força"), "H{}:H{}".format(PRF_F, PRF_L))
dv(wsPQ, L("Ref. VBT"),          "I{}:I{}".format(PRF_F, PRF_L))
dv(wsPQ, L("Bloco de Treino"),   "C{}:C{}".format(PRF_F, PRF_L))
dv(wsPQ, L("PSE (0-10)"),        "V{}:V{}".format(PRF_F, PRF_L))
dv(wsPQ, "{}!$Q$7:$Y$7".format(F1), "AB{}:AB{}".format(PRF_F, PRF_L))
wsPQ.conditional_formatting.add("T{}:T{}".format(PRF_F, PRF_L),
    CellIsRule(operator="equal", formula=['"Acima do limite"'], fill=PatternFill("solid", fgColor=RED),
               font=Font(name=F, size=9, bold=True, color=RED_T)))
wsPQ.conditional_formatting.add("T{}:T{}".format(PRF_F, PRF_L),
    CellIsRule(operator="equal", formula=['"Dentro do limite"'], fill=PatternFill("solid", fgColor=GREEN),
               font=Font(name=F, size=9, bold=True, color=GREEN_T)))
wsPQ.freeze_panes = "C8"
wsPQ.auto_filter.ref = "A7:AB{}".format(PRF_L)

# --- tabela de referência carga-velocidade (editável) -----------------------
secao(wsPQ, 6, "TABELA CARGA-VELOCIDADE (editável)", 33, 30)
cab_tabela(wsPQ, 7, ["%1RM", "Agachamento — MPV (m/s)", "Supino — MPV (m/s)", "Terra — MV (m/s)"], col0=30)
LV = [(0.40,1.09,0.94,1.05),(0.45,1.02,0.86,0.98),(0.50,0.96,0.79,0.90),(0.55,0.89,0.72,0.83),
      (0.60,0.83,0.65,0.76),(0.65,0.75,0.58,0.69),(0.70,0.68,0.51,0.62),(0.75,0.62,0.44,0.55),
      (0.80,0.55,0.37,0.48),(0.85,0.47,0.31,0.41),(0.90,0.40,0.24,0.35),(0.95,0.36,0.20,0.29),
      (1.00,0.32,0.17,0.24)]
for i, row in enumerate(LV):
    r = 8 + i
    for j, v in enumerate(row):
        c = wsPQ.cell(r, 30 + j, v)
        c.font = Font(name=F, size=9, color="0000FF")
        c.fill = PatternFill("solid", fgColor=GOLD_L)
        c.border = BORDER
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.number_format = NF_PCT if j == 0 else '0.00'
nt = wsPQ.cell(22, 30, "Valores APROXIMADOS de referência compilados da literatura de treinamento baseado em "
                       "velocidade (velocidade média propulsiva no agachamento completo e no supino; velocidade média "
                       "no terra). A relação carga-velocidade é específica de exercício, equipamento e atleta: "
                       "levante o SEU perfil individual e substitua estes números. Use a velocidade para conferir a "
                       "intensidade da série, não para substituir o teste de 1RM.")
wsPQ.merge_cells(start_row=22, start_column=30, end_row=27, end_column=33)
nt.font = Font(name=F, size=8, italic=True, color=GREY_T)
nt.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
nt.fill = PatternFill("solid", fgColor=LIGHT2); nt.border = BORDER

# --- geração das 8 semanas --------------------------------------------------
BLOCO_INI = date(2026, 9, 7)
PROG = {
 "AGACH":  [(4,8,.65),(4,6,.75),(5,5,.80),(3,5,.70),(5,4,.85),(4,3,.88),(4,3,.90),(3,2,.92)],
 "AGACH2": [(4,8,.50),(4,6,.60),(4,5,.65),(3,5,.55),(4,4,.70),(4,4,.72),(4,3,.75),(3,3,.75)],
 "LPO":    [(5,3,.55),(4,4,.70),(5,3,.75),(3,3,.60),(5,2,.82),(4,3,.95),(6,2,.88),(4,2,.90)],
 "LPO2":   [(4,3,.50),(4,3,.65),(4,3,.70),(3,3,.55),(4,3,.78),(4,3,.85),(5,2,.85),(3,2,.85)],
 "STIFF":  [(4,8,.60),(4,8,.65),(4,6,.70),(3,6,.60),(4,6,.72),(4,5,.75),(4,5,.78),(3,4,.75)],
 "MMSS":   [(4,10,.65),(4,8,.70),(4,8,.75),(3,8,.65),(4,6,.80),(4,6,.82),(4,5,.85),(3,5,.85)],
 "UNILAT": [(3,10,None),(3,10,None),(3,8,None),(2,8,None),(3,8,None),(3,8,None),(3,6,None),(2,6,None)],
 "JUMPSQ": [(0,0,None),(4,4,.20),(4,4,.25),(3,3,.20),(5,4,.30),(5,4,.30),(5,3,.35),(4,3,.30)],
 "PREV":   [(3,10,None),(3,10,None),(3,8,None),(2,10,None),(3,8,None),(3,8,None),(3,6,None),(2,8,None)],
 "MB":     [(3,8,None)] * 8,
 "SPRINT": [(4,1,None),(6,1,None),(0,0,None),(4,1,None),(6,1,None),(6,1,None),(0,0,None),(4,1,None)],
 "CORE":   [(3,25,None),(3,25,None),(3,25,None),(3,20,None),(3,20,None),(3,20,None),(3,20,None),(3,15,None)],
 "AQUEC":  [(1,8,None)] * 8,
}
PLIO_SEM = [60, 90, 120, 45, 90, 110, 140, 70]
for slot, share in (("PLIO1", .40), ("PLIO2", .35), ("PLIO3", .25)):
    PROG[slot] = [(max(1, round(PLIO_SEM[w] * share / 5)), 5, None) for w in range(8)]
PAUSA = {"AGACH":180,"AGACH2":150,"LPO":180,"LPO2":150,"STIFF":120,"MMSS":120,"UNILAT":90,"JUMPSQ":180,
         "PLIO1":120,"PLIO2":120,"PLIO3":120,"PREV":60,"MB":90,"SPRINT":120,"CORE":45,"AQUEC":0}
OBJ = {"AGACH":"Força Máxima","AGACH2":"Força-Velocidade","LPO":"Força-Velocidade","LPO2":"Força-Velocidade",
       "STIFF":"Força Máxima","MMSS":"Força Máxima","UNILAT":"Força Máxima","JUMPSQ":"Potência",
       "PLIO1":"Pliometria","PLIO2":"Pliometria","PLIO3":"Pliometria","PREV":"Preventivo","MB":"Potência",
       "SPRINT":"Velocidade","CORE":"Core","AQUEC":"Aquecimento"}
VL = {"AGACH":(.15,.10),"AGACH2":(.15,.10),"LPO":(.10,.10),"LPO2":(.10,.10),"STIFF":(.20,.15),
      "MMSS":(.20,.15),"JUMPSQ":(.10,.10)}
SESSOES_PRF = [
 ("A", 0, [("Mobilidade de tornozelo e quadril","AQUEC","—",""),
           ("Power clean a partir do joelho","LPO","—","Power clean"),
           ("Agachamento profundo","AGACH","Agachamento","Agachamento profundo"),
           ("Stiff com barra","STIFF","Terra","Stiff com barra"),
           ("Afundo lateral sem passada","UNILAT","—",""),
           ("Nórdico de isquiotibiais (excêntrico)","PREV","—",""),
           ("Abdominal reto com braços esticados + Dorsal perdigueiro","CORE","—","")]),
 ("B", 2, [("Rotadores do ombro com elástico","PREV","—",""),
           ("Agachamento com salto sob carga (jump squat)","JUMPSQ","Agachamento","Agachamento profundo"),
           ("Salto no caixote / Drop jump","PLIO1","—",""),
           ("Saltos consecutivos sobre barreiras","PLIO2","—",""),
           ("Supino deitado com halteres","MMSS","Supino","Supino deitado com halteres"),
           ("Remada unilateral na polia","MMSS","—",""),
           ("Arremesso de bola","MB","—",""),
           ("Abdominal cruzado esticado + Dorsal reto","CORE","—","")]),
 ("C", 4, [("Mobilidade de tornozelo e quadril","AQUEC","—",""),
           ("Snatch pull / Hang high pull","LPO2","—","Clean pull"),
           ("Agachamento sumô no minitramp","AGACH2","Agachamento","Agachamento profundo"),
           ("Elevação de calcanhares","PREV","—",""),
           ("Afundo frontal sem passada com sobrepeso","UNILAT","—",""),
           ("Salto unilateral com sobrepeso","PLIO3","—",""),
           ("Pullover deitado com haltere","MMSS","—",""),
           ("Sprints de 10 e 20 m com mudança de direção","SPRINT","—",""),
           ("Abdominal remador com anilha","CORE","—","")]),
]
BLOCO_NOME = ["Acumulação","Acumulação","Acumulação","Descarga",
              "Transmutação","Transmutação","Transmutação","Realização"]
r = PRF_F
for w in range(8):
    for sess, off, itens in SESSOES_PRF:
        d = BLOCO_INI + timedelta(days=w * 7 + off)
        ordem = 0
        for ex, fam, vbt, ref1rm in itens:
            ser, rep, pct = PROG[fam][w]
            if ser == 0:
                continue
            ordem += 1
            wsPQ.cell(r, 1, d); wsPQ.cell(r, 1).number_format = NF_DATE
            wsPQ.cell(r, 3, BLOCO_NOME[w]); wsPQ.cell(r, 4, sess)
            wsPQ.cell(r, 5, "Equipe (todos)"); wsPQ.cell(r, 6, ordem)
            wsPQ.cell(r, 7, ex); wsPQ.cell(r, 8, OBJ[fam]); wsPQ.cell(r, 9, vbt)
            wsPQ.cell(r, 10, ser); wsPQ.cell(r, 11, rep)
            if pct is not None:
                wsPQ.cell(r, 12, pct)
            if fam in VL:
                wsPQ.cell(r, 18, VL[fam][0 if w < 4 else 1])
            wsPQ.cell(r, 23, PAUSA[fam])
            if ref1rm:
                wsPQ.cell(r, 28, ref1rm)
            r += 1
PRF_USADO = r - 1


# ============================================================================
# 21) PERFIL FORÇA-VELOCIDADE-POTÊNCIA  (método de Samozino no salto com carga)
# ============================================================================
wsFV = wb.create_sheet("Perfil F-V-P")
banner(wsFV, "PERFIL FORÇA–VELOCIDADE–POTÊNCIA (SALTO COM CARGA)",
       "Protocolo: squat jump com 0, 20, 40, 60 e 80 kg (ou % da massa corporal). A planilha calcula F0, V0, Pmax e "
       "a inclinação do perfil por regressão linear, pelo método de Samozino/Morin.", 14, "6B3FA0")
larguras(wsFV, {"A":18,"B":16,"C":16,"D":16,"E":16,"F":16,"G":16,"H":16,"I":16,"J":22,"K":20,"L":52,"M":14,"N":14})
secao(wsFV, 3, "PARÂMETROS DO TESTE", 14, 1)
rotulo(wsFV, 4, 1, "Atleta:")
selFV = entrada(wsFV, 4, 3, NOMES[0], largura_merge=4)
selFV.font = Font(name=F, size=12, bold=True, color="0000FF")
dv(wsFV, ATLETAS_REF, "C4")
rotulo(wsFV, 5, 1, "Massa corporal (kg):")
calc(wsFV, 5, 3, '=IFERROR(INDEX(Cadastro!$P${0}:$P${1},MATCH($C$4,Cadastro!$B${0}:$B${1},0)),"")'.format(CAD_F, CAD_L), '0.0')
rotulo(wsFV, 6, 1, "Distância de push-off hPO (m):")
entrada(wsFV, 6, 3, 0.40, '0.00')
wsFV["C6"].comment = Comment("Distância percorrida pelo centro de massa na fase propulsiva: da posição inicial do "
                             "squat jump (≈90° de joelho) até a extensão completa.\nMeça a diferença de altura do "
                             "trocânter maior entre as duas posições. Faixa usual: 0,32 a 0,48 m.\n"
                             "hPO errado desloca todo o perfil — meça, não estime.", "Planilha")
rotulo(wsFV, 7, 1, "SFvopt (N·s/m/kg):")
entrada(wsFV, 7, 3, None, '0.00')
wsFV["C7"].comment = Comment("Inclinação ÓTIMA teórica do perfil força-velocidade do atleta.\nÉ calculada pelo "
                             "método de Samozino/Jiménez-Reyes e sai pronta de softwares como o My Jump ou do "
                             "relatório da sua plataforma de força.\nDeixe em branco se não tiver: F0, V0, Pmax e "
                             "a inclinação real continuam sendo calculados.", "Planilha")

secao(wsFV, 10, "SALTOS COM CARGA PROGRESSIVA", 14, 1)
cab_tabela(wsFV, 11, ["Carga Adicional (kg)","Altura do Salto (cm)","Massa Total (kg)","Força Média (N)",
                      "Força Relativa (N/kg)","Velocidade Média (m/s)","Potência Média (W)","Potência Relativa (W/kg)"])
FV_F, FV_L = 12, 16
for i in range(5):
    r = FV_F + i
    wsFV.cell(r, 1, [0, 20, 40, 60, 80][i])
    wsFV.cell(r, 3, '=IFERROR($C$5+$A{0},"")'.format(r))
    wsFV.cell(r, 4, '=IFERROR($C{0}*9.81*(($B{0}/100)/$C$6+1),"")'.format(r))
    wsFV.cell(r, 5, '=IFERROR($D{0}/$C$5,"")'.format(r))
    wsFV.cell(r, 6, '=IFERROR(SQRT(9.81*($B{0}/100)/2),"")'.format(r))
    wsFV.cell(r, 7, '=IFERROR($D{0}*$F{0},"")'.format(r))
    wsFV.cell(r, 8, '=IFERROR($G{0}/$C$5,"")'.format(r))
corpo_tabela(wsFV, FV_F, FV_L, 1, 8)
for r in range(FV_F, FV_L + 1):
    for c in (1, 2):
        wsFV.cell(r, c).fill = PatternFill("solid", fgColor=GOLD_L)
        wsFV.cell(r, c).font = Font(name=F, size=10, color="0000FF")
        wsFV.cell(r, c).number_format = '0.0'
    for c in range(3, 9):
        wsFV.cell(r, c).font = Font(name=F, size=10, bold=True, color=NAVY2)
        wsFV.cell(r, c).fill = PatternFill("solid", fgColor=LIGHT)
        wsFV.cell(r, c).number_format = '0.00' if c in (5, 6, 8) else '0.0'
for i, h in enumerate([48.0, 39.0, 32.0, 26.0, 21.0]):
    wsFV.cell(FV_F + i, 2, h)

secao(wsFV, 18, "RESULTADOS DO PERFIL", 14, 1)
RES = [("F0 — Força teórica máxima (N)", '=IFERROR(INTERCEPT($D${0}:$D${1},$F${0}:$F${1}),"")'.format(FV_F, FV_L), '0'),
       ("F0 relativa (N/kg)", '=IFERROR($C$19/$C$5,"")', '0.00'),
       ("V0 — Velocidade teórica máxima (m/s)", '=IFERROR(-$C$19/$C$22,"")', '0.00'),
       ("Sfv — Inclinação do perfil (N·s/m)", '=IFERROR(SLOPE($D${0}:$D${1},$F${0}:$F${1}),"")'.format(FV_F, FV_L), '0.0'),
       ("Sfv relativa (N·s/m/kg)", '=IFERROR($C$22/$C$5,"")', '0.00'),
       ("Pmax — Potência máxima (W)", '=IFERROR($C$19*$C$21/4,"")', '0'),
       ("Pmax relativa (W/kg)", '=IFERROR($C$24/$C$5,"")', '0.00'),
       ("FVimb — Desequilíbrio F-V (%)", '=IF($C$7="","",IFERROR(ABS($C$23)/ABS($C$7),""))', NF_PCT),
       ("R² da regressão", '=IFERROR(RSQ($D${0}:$D${1},$F${0}:$F${1}),"")'.format(FV_F, FV_L), '0.000')]
for i, (lab, f, nf) in enumerate(RES):
    r = 19 + i
    rotulo(wsFV, r, 1, lab)
    c = calc(wsFV, r, 3, f, nf)
    c.font = Font(name=F, size=11, bold=True, color=NAVY2)
nota(wsFV, 17, 1, "F0 e V0 são EXTRAPOLAÇÕES da reta força-velocidade: dependem muito da distância de push-off (C6) "
     "e da amplitude de cargas testada. Use pelo menos 5 cargas cobrindo uma faixa ampla e confira o R² (linha 27): "
     "abaixo de 0,95, refaça o teste.", 14)
rotulo(wsFV, 28, 1, "PERFIL:")
p = calc(wsFV, 28, 3, '=IF($C$26="","Informe o SFvopt em C7 para classificar",'
                      'IF($C$26<0.9,"Déficit de FORÇA",IF($C$26<=1.1,"Equilibrado","Déficit de VELOCIDADE")))')
p.font = Font(name=F, size=12, bold=True, color=NAVY)
wsFV.merge_cells(start_row=28, start_column=3, end_row=28, end_column=6)
wsFV.merge_cells(start_row=29, start_column=1, end_row=32, end_column=14)
rec = wsFV.cell(29, 1,
  '=IF($C$26="","Sem o SFvopt não é possível estimar o desequilíbrio. Ainda assim, F0, V0 e Pmax servem para '
  'acompanhar a evolução do atleta ao longo dos blocos e para comparar atletas do mesmo elenco na tabela abaixo.",'
  'IF($C$26<0.9,"DÉFICIT DE FORÇA — priorize cargas altas: agachamento e stiff a 80–90% 1RM, LPO (power clean e '
  'clean pull) a 80–90%, 70 a 100% do volume de força na região pesada do perfil. Reduza o volume balístico.",'
  'IF($C$26<=1.1,"PERFIL EQUILIBRADO — mantenha a distribuição mista e busque elevar a Pmax como um todo: alterne '
  'blocos de força máxima (85–90% 1RM) e blocos balísticos (jump squat 30% 1RM + pliometria).",'
  '"DÉFICIT DE VELOCIDADE — priorize o lado rápido do perfil: jump squat a 0–40% 1RM, saltos sem carga, pliometria '
  'reativa (drop jump) e LPO com ênfase em velocidade da barra. Mantenha a força máxima com volume mínimo.")))')
rec.font = Font(name=F, size=10, bold=True, color=NAVY)
rec.fill = PatternFill("solid", fgColor=GOLD_L)
rec.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
rec.border = BORDER

chFV = ScatterChart(); chFV.title = "Relação Força–Velocidade"; chFV.style = 13
chFV.x_axis.title = "Velocidade média (m/s)"; chFV.y_axis.title = "Força média (N)"
chFV.height = 9; chFV.width = 13; chFV.legend = None
s = Series(Reference(wsFV, min_col=4, min_row=FV_F, max_row=FV_L),
           Reference(wsFV, min_col=6, min_row=FV_F, max_row=FV_L), title="F-v")
s.marker = Marker(symbol="circle", size=8); s.graphicalProperties.line.noFill = True
chFV.series.append(s)
wsFV.add_chart(chFV, "H10")
chPV = ScatterChart(); chPV.title = "Relação Potência–Velocidade"; chPV.style = 13
chPV.x_axis.title = "Velocidade média (m/s)"; chPV.y_axis.title = "Potência média (W)"
chPV.height = 9; chPV.width = 13; chPV.legend = None
s2 = Series(Reference(wsFV, min_col=7, min_row=FV_F, max_row=FV_L),
            Reference(wsFV, min_col=6, min_row=FV_F, max_row=FV_L), title="P-v")
s2.marker = Marker(symbol="triangle", size=8); s2.graphicalProperties.line.noFill = True
chPV.series.append(s2)
wsFV.add_chart(chPV, "H29")

secao(wsFV, 34, "PERFIL F-V DE TODO O ELENCO", 12, 1)
cab_tabela(wsFV, 35, ["Atleta","Massa (kg)","F0 (N)","V0 (m/s)","F0 rel (N/kg)","Pmax (W)","Pmax rel (W/kg)",
                      "Sfv rel (N·s/m/kg)","SFvopt","FVimb (%)","Perfil","Prioridade de Treino"])
FVE_F = 36
FVE_L = FVE_F + (CAD_L - CAD_F)
for i in range(CAD_L - CAD_F + 1):
    r = FVE_F + i
    a = CAD_F + i
    wsFV.cell(r, 1, '=IF(Cadastro!$B{0}="","",Cadastro!$B{0})'.format(a))
    wsFV.cell(r, 2, '=IF($A{0}="","",IFERROR(Cadastro!$P{1},""))'.format(r, a))
    wsFV.cell(r, 5, '=IFERROR($C{0}/$B{0},"")'.format(r))
    wsFV.cell(r, 6, '=IFERROR($C{0}*$D{0}/4,"")'.format(r))
    wsFV.cell(r, 7, '=IFERROR($F{0}/$B{0},"")'.format(r))
    wsFV.cell(r, 8, '=IFERROR(-$E{0}/$D{0},"")'.format(r))
    wsFV.cell(r, 10, '=IF(OR($I{0}="",$H{0}=""),"",IFERROR(ABS($H{0})/ABS($I{0}),""))'.format(r))
    wsFV.cell(r, 11, '=IF($J{0}="","",IF($J{0}<0.9,"Déficit de FORÇA",'
                     'IF($J{0}<=1.1,"Equilibrado","Déficit de VELOCIDADE")))'.format(r))
    wsFV.cell(r, 12, '=IF($K{0}="","",IF($K{0}="Déficit de FORÇA","Cargas altas: 80–90% 1RM + LPO pesado",'
                     'IF($K{0}="Equilibrado","Misto: alternar força máxima e balístico",'
                     '"Balístico: 0–40% 1RM, jump squat e pliometria reativa")))'.format(r))
corpo_tabela(wsFV, FVE_F, FVE_L, 1, 12)
for r in range(FVE_F, FVE_L + 1):
    for c in range(1, 13):
        if c in (3, 4, 9):
            wsFV.cell(r, c).fill = PatternFill("solid", fgColor=GOLD_L)
            wsFV.cell(r, c).font = Font(name=F, size=9, color="0000FF")
        else:
            wsFV.cell(r, c).font = Font(name=F, size=9, bold=(c == 11), color=NAVY2)
            wsFV.cell(r, c).fill = PatternFill("solid", fgColor=LIGHT if (r - FVE_F) % 2 == 0 else LIGHT2)
    for c in (1, 11, 12):
        wsFV.cell(r, c).alignment = Alignment(horizontal="left", vertical="center", indent=1)
    for c in (2, 3, 6):
        wsFV.cell(r, c).number_format = '0.0;;""'
    for c in (4, 5, 7, 8, 9):
        wsFV.cell(r, c).number_format = '0.00;;""'
    wsFV.cell(r, 10).number_format = NF_PCT
for txt, fill, ft in [('"Equilibrado"', GREEN, GREEN_T), ('"Déficit de FORÇA"', "FCE4D6", "C55A11"),
                      ('"Déficit de VELOCIDADE"', "DDEBF7", NAVY2)]:
    wsFV.conditional_formatting.add("K{}:K{}".format(FVE_F, FVE_L),
        CellIsRule(operator="equal", formula=[txt], fill=PatternFill("solid", fgColor=fill),
                   font=Font(name=F, size=9, bold=True, color=ft)))
chEQ = ScatterChart(); chEQ.title = "Elenco: Força relativa (F0/kg) × Velocidade (V0)"
chEQ.x_axis.title = "F0 relativa (N/kg)"; chEQ.y_axis.title = "V0 (m/s)"
chEQ.height = 10; chEQ.width = 16; chEQ.legend = None
sE = Series(Reference(wsFV, min_col=4, min_row=FVE_F, max_row=FVE_L),
            Reference(wsFV, min_col=5, min_row=FVE_F, max_row=FVE_L), title="Atletas")
sE.marker = Marker(symbol="diamond", size=9); sE.graphicalProperties.line.noFill = True
chEQ.series.append(sE)
wsFV.add_chart(chEQ, "N35")
wsFV.merge_cells(start_row=FVE_L + 2, start_column=1, end_row=FVE_L + 6, end_column=12)
cav = wsFV.cell(FVE_L + 2, 1,
  "LEIA ANTES DE USAR — o treino individualizado pelo desequilíbrio F-V mostrou bons resultados em alguns estudos "
  "(Jiménez-Reyes et al., 2017 e 2019) e ganhos de F0, V0 e altura de salto em meta-análises recentes, mas ensaios "
  "controlados independentes não reproduziram a superioridade sobre um programa bem feito não individualizado "
  "(Lindberg et al., 2021; Solberg et al., 2025), e há questionamento de que parte do efeito seja aprendizado da "
  "tarefa e não adaptação neuromuscular (Bobbert et al., 2024). Use o perfil como MAIS UMA informação para decidir "
  "a ênfase do bloco — não como regra única de prescrição. A base continua sendo elevar a força máxima e a Pmax.")
cav.font = Font(name=F, size=9, color=RED_T)
cav.fill = PatternFill("solid", fgColor="FDECEA"); cav.border = BORDER
cav.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
wsFV.freeze_panes = "A11"


# ============================================================================
# 22) SALTOS — carga de saltos (pliometria + quadra + jogo)
# ============================================================================
wsS2 = wb.create_sheet("Saltos")
SLT_F, SLT_L = 8, 807
banner(wsS2, "CARGA DE SALTOS — PLIOMETRIA, QUADRA E JOGO",
       "O salto é a carga externa que mais explica lesão por sobrecarga no voleibol. Registre TODOS os saltos: "
       "contatos pliométricos da musculação, saltos de treino de quadra e saltos de jogo.", 13, "6B3FA0")
larguras(wsS2, {"A":12,"B":8,"C":24,"D":20,"E":34,"F":16,"G":14,"H":20,"I":15,"J":17,"K":18,"L":16,"M":26,"N":3,
                "O":24,"P":15,"Q":14,"R":13,"S":15,"T":18,"U":11,"V":22})
SLT_H = ["Data","Semana","Atleta","Origem","Exercício / Contexto","Nº de Saltos ou Contatos","Intensidade",
         "Superfície","Altura da Caixa (cm)","Altura Média do Salto (cm)","% de Saltos > 80% do Máx",
         "Qualidade Técnica (1-5)","Observações"]
cab_tabela(wsS2, 7, SLT_H)
rotulo(wsS2, 4, 15, "Semana de referência:")
entrada(wsS2, 4, 18, '=IFERROR(MAX(1,INT((TODAY()-Macrociclo!$C$11)/7)+1),1)', '0')
nota(wsS2, 4, 1, "Referências de dose: programas pliométricos com mais de 10 semanas, mais de 20 sessões e mais de 50 "
     "saltos por sessão, combinando tipos de salto, maximizam o ganho de impulsão (Sáez de Villarreal et al., 2009). "
     "Em temporada, equipes universitárias masculinas chegam a ~466 saltos semanais no mesociclo competitivo "
     "(Lin et al., 2024). O alerta aqui é o SALTO SEMANAL e a variação brusca em relação à semana anterior.", 13)
wsS2.row_dimensions[4].height = 30
for r in range(SLT_F, SLT_L + 1):
    wsS2.cell(r, 2, '=IF($A{0}="",0,INT(($A{0}-Macrociclo!$C$11)/7)+1)'.format(r))
corpo_tabela(wsS2, SLT_F, SLT_L, 1, 13)
for r in range(SLT_F, SLT_L + 1):
    for c in (1, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13):
        wsS2.cell(r, c).fill = PatternFill("solid", fgColor=GOLD_L)
        wsS2.cell(r, c).font = Font(name=F, size=9, color="0000FF")
    wsS2.cell(r, 2).font = Font(name=F, size=9, bold=True, color=NAVY2)
    wsS2.cell(r, 2).fill = PatternFill("solid", fgColor=LIGHT)
    for c in (3, 4, 5, 13):
        wsS2.cell(r, c).alignment = Alignment(horizontal="left", vertical="center", indent=1)
    wsS2.cell(r, 1).number_format = NF_DATE
    wsS2.cell(r, 2).number_format = '0;;""'
    for c in (6, 9, 10):
        wsS2.cell(r, c).number_format = '0;;""'
    wsS2.cell(r, 11).number_format = NF_PCT
dv(wsS2, ATLETAS_REF,          "C{}:C{}".format(SLT_F, SLT_L))
dv(wsS2, L("Origem do Salto"), "D{}:D{}".format(SLT_F, SLT_L))
dv(wsS2, L("Intensidade"),     "G{}:G{}".format(SLT_F, SLT_L))
dv(wsS2, L("Superfície"),      "H{}:H{}".format(SLT_F, SLT_L))

secao(wsS2, 5, "RESUMO SEMANAL POR ATLETA", 22, 15)
cab_tabela(wsS2, 7, ["Atleta","Pliometria","Quadra","Jogo","Total da Semana","Semana Anterior","Variação",
                     "Situação"], col0=15)
SLT_RF = SLT_F
SLT_RL = SLT_F + (CAD_L - CAD_F)
for i in range(CAD_L - CAD_F + 1):
    r = SLT_RF + i
    a = CAD_F + i
    wsS2.cell(r, 15, '=IF(Cadastro!$B{0}="","",Cadastro!$B{0})'.format(a))
    for j, orig in enumerate(["Pliometria", "Treino de quadra", "Jogo"]):
        wsS2.cell(r, 16 + j, '=IF($O{0}="",0,SUMIFS($F${1}:$F${2},$C${1}:$C${2},$O{0},'
                             '$B${1}:$B${2},$R$4,$D${1}:$D${2},"{3}"))'.format(r, SLT_F, SLT_L, orig))
    wsS2.cell(r, 19, '=SUM($P{0}:$R{0})+IF($O{0}="",0,SUMIFS($F${1}:$F${2},$C${1}:$C${2},$O{0},'
                     '$B${1}:$B${2},$R$4,$D${1}:$D${2},"Musculação"))'.format(r, SLT_F, SLT_L))
    wsS2.cell(r, 20, '=IF($O{0}="",0,SUMIFS($F${1}:$F${2},$C${1}:$C${2},$O{0},$B${1}:$B${2},$R$4-1))'
             .format(r, SLT_F, SLT_L))
    wsS2.cell(r, 21, '=IF($T{0}=0,0,$S{0}/$T{0}-1)'.format(r))
    wsS2.cell(r, 22, '=IF($O{0}="","",IF($S{0}=0,"Sem registro",IF($U{0}>0.3,"ALERTA — aumento acima de 30%",'
                     'IF($S{0}>600,"Volume semanal muito alto",IF($S{0}>=350,"Volume alto",'
                     'IF($S{0}>=150,"Volume adequado","Volume baixo"))))))'.format(r))
corpo_tabela(wsS2, SLT_RF, SLT_RL, 15, 22)
for r in range(SLT_RF, SLT_RL + 1):
    for c in range(15, 23):
        wsS2.cell(r, c).font = Font(name=F, size=9, bold=(c == 19), color=NAVY2)
        wsS2.cell(r, c).fill = PatternFill("solid", fgColor=LIGHT if (r - SLT_RF) % 2 == 0 else LIGHT2)
    for c in (15, 22):
        wsS2.cell(r, c).alignment = Alignment(horizontal="left", vertical="center", indent=1)
    for c in range(16, 21):
        wsS2.cell(r, c).number_format = NF_UA
    wsS2.cell(r, 21).number_format = '+0%;-0%;""'
for txt, fill, ft in [('"Volume adequado"', GREEN, GREEN_T), ('"Volume alto"', YELL, YELL_T),
                      ('"Volume semanal muito alto"', RED, RED_T),
                      ('"ALERTA — aumento acima de 30%"', RED, RED_T), ('"Volume baixo"', "DDEBF7", NAVY2)]:
    wsS2.conditional_formatting.add("V{}:V{}".format(SLT_RF, SLT_RL),
        CellIsRule(operator="equal", formula=[txt], fill=PatternFill("solid", fgColor=fill),
                   font=Font(name=F, size=9, bold=True, color=ft)))
chS = BarChart(); chS.type = "col"; chS.grouping = "stacked"; chS.overlap = 100
chS.title = "Saltos da Semana por Atleta (origem)"; chS.height = 9; chS.width = 20
chS.add_data(Reference(wsS2, min_col=16, max_col=18, min_row=7, max_row=min(SLT_RL, SLT_RF + 19)),
             titles_from_data=True)
chS.set_categories(Reference(wsS2, min_col=15, min_row=SLT_RF, max_row=min(SLT_RL, SLT_RF + 19)))
wsS2.add_chart(chS, "X7")
wsS2.freeze_panes = "C8"
wsS2.auto_filter.ref = "A7:M{}".format(SLT_L)


# ============================================================================
# 23) KPIs FORÇA — painel de acompanhamento do bloco de força e potência
# ============================================================================
wsK = wb.create_sheet("KPIs Força")
banner(wsK, "KPIs DE FORÇA E POTÊNCIA — ACOMPANHAMENTO DO BLOCO",
       "Indicadores de carga (tonelagem, intensidade relativa, contatos pliométricos) e de evolução (1RM relativo, "
       "CMJ, potência de pico). Altere a semana em C4 para atualizar os cartões.", 20, "9C0006")
larguras(wsK, {"A":22,"B":16,"C":13,"D":15,"E":16,"F":18,"G":18,"H":16,"I":16,"J":3,
               "K":13,"L":13,"M":13,"N":13,"O":13,"P":13,"Q":13,"R":13,"S":13,"T":13})
PQ = "'Prescrição Força'"
rotulo(wsK, 4, 1, "Semana do bloco (1 a 8):")
entrada(wsK, 4, 3, 1, '0', largura_merge=2)
FK = "{0}!$B${1}:$B${2},$C$4".format(PQ, PRF_F, PRF_L)
kpi(wsK, 6, 1,  "TONELAGEM DA SEMANA (kg)", '=SUMIFS({0}!$Y${1}:$Y${2},{3})'.format(PQ, PRF_F, PRF_L, FK), NF_UA, "9C0006", 3)
kpi(wsK, 6, 4,  "REPETIÇÕES TOTAIS", '=SUMIFS({0}!$X${1}:$X${2},{3})'.format(PQ, PRF_F, PRF_L, FK), NF_UA, NAVY2, 3)
kpi(wsK, 6, 7,  "INTENSIDADE MÉDIA RELATIVA",
    '=IFERROR(SUMPRODUCT(({0}!$B${1}:$B${2}=$C$4)*{0}!$X${1}:$X${2}*{0}!$L${1}:$L${2})'
    '/SUMPRODUCT(({0}!$B${1}:$B${2}=$C$4)*{0}!$X${1}:$X${2}*({0}!$L${1}:$L${2}<>"")),0)'.format(PQ, PRF_F, PRF_L),
    NF_PCT, GOLD, 3)
kpi(wsK, 6, 10, "CONTATOS PLIOMÉTRICOS",
    '=SUMIFS({0}!$X${1}:$X${2},{3},{0}!$H${1}:$H${2},"Pliometria")'.format(PQ, PRF_F, PRF_L, FK), NF_UA, "6B3FA0", 3)
kpi(wsK, 6, 13, "1RM MÉDIO — AGACHAMENTO (kg)",
    '=IFERROR(AVERAGEIF(\'Força 1RM\'!$Q${0}:$Q${1},">0"),0)'.format(RM_F, RM_MAT_L), '0.0;;""', NAVY2, 3)
kpi(wsK, 6, 16, "FORÇA RELATIVA MÉDIA (× massa)", '=IFERROR(AVERAGEIF($C$31:$C$70,">0"),0)', NF_DEC, NAVY2, 3)
kpi(wsK, 6, 19, "Δ CMJ MÉDIO DO BLOCO", '=IFERROR(AVERAGEIF($F$31:$F$70,"<>0"),0)', '+0.0%;-0.0%;""', GOLD, 2)

secao(wsK, 9, "CARGA DE FORÇA POR SEMANA DO BLOCO", 9, 1)
cab_tabela(wsK, 10, ["Semana","Bloco","Séries","Repetições","Tonelagem (kg)","Intensidade Média Relativa",
                     "Contatos Pliométricos","Exercícios Prescritos","Tonelagem por Exercício"])
KW_F, KW_L = 11, 18
for i in range(8):
    r = KW_F + i
    wsK.cell(r, 1, i + 1)
    wsK.cell(r, 2, '=IFERROR(INDEX(\'Bloco Base\'!$C$12:$C$19,MATCH($A{},\'Bloco Base\'!$A$12:$A$19,0)),"")'.format(r))
    wsK.cell(r, 3, '=SUMIFS({0}!$J${1}:$J${2},{0}!$B${1}:$B${2},$A{3})'.format(PQ, PRF_F, PRF_L, r))
    wsK.cell(r, 4, '=SUMIFS({0}!$X${1}:$X${2},{0}!$B${1}:$B${2},$A{3})'.format(PQ, PRF_F, PRF_L, r))
    wsK.cell(r, 5, '=SUMIFS({0}!$Y${1}:$Y${2},{0}!$B${1}:$B${2},$A{3})'.format(PQ, PRF_F, PRF_L, r))
    wsK.cell(r, 6, '=IFERROR(SUMPRODUCT(({0}!$B${1}:$B${2}=$A{3})*{0}!$X${1}:$X${2}*{0}!$L${1}:$L${2})'
                   '/SUMPRODUCT(({0}!$B${1}:$B${2}=$A{3})*{0}!$X${1}:$X${2}*({0}!$L${1}:$L${2}<>"")),0)'
             .format(PQ, PRF_F, PRF_L, r))
    wsK.cell(r, 7, '=SUMIFS({0}!$X${1}:$X${2},{0}!$B${1}:$B${2},$A{3},{0}!$H${1}:$H${2},"Pliometria")'
             .format(PQ, PRF_F, PRF_L, r))
    wsK.cell(r, 8, '=COUNTIFS({0}!$B${1}:$B${2},$A{3})'.format(PQ, PRF_F, PRF_L, r))
    wsK.cell(r, 9, '=IFERROR($E{0}/$H{0},0)'.format(r))
corpo_tabela(wsK, KW_F, KW_L, 1, 9)
for r in range(KW_F, KW_L + 1):
    for c in range(1, 10):
        wsK.cell(r, c).font = Font(name=F, size=9, bold=(c in (1, 5)), color=NAVY2)
        wsK.cell(r, c).fill = PatternFill("solid", fgColor=LIGHT if (r - KW_F) % 2 == 0 else LIGHT2)
    wsK.cell(r, 2).alignment = Alignment(horizontal="left", vertical="center", indent=1)
    for c in (3, 4, 5, 7, 8, 9):
        wsK.cell(r, c).number_format = NF_UA
    wsK.cell(r, 6).number_format = NF_PCT
wsK.conditional_formatting.add("E{}:E{}".format(KW_F, KW_L),
    ColorScaleRule(start_type="min", start_color="FFFFFF", end_type="max", end_color="F4B183"))

cB = BarChart(); cB.type = "col"; cB.title = "Tonelagem × Intensidade Média Relativa por Semana"
cB.height = 9; cB.width = 20; cB.y_axis.title = "Tonelagem (kg)"; cB.y_axis.majorGridlines = None
cB.add_data(Reference(wsK, min_col=5, min_row=10, max_row=KW_L), titles_from_data=True)
cB.set_categories(Reference(wsK, min_col=1, min_row=KW_F, max_row=KW_L))
cL = LineChart()
cL.add_data(Reference(wsK, min_col=6, min_row=10, max_row=KW_L), titles_from_data=True)
cL.y_axis.axId = 200; cL.y_axis.title = "% de 1RM"; cL.y_axis.crosses = "max"
cL.series[0].graphicalProperties.line.width = 30000
cB += cL
wsK.add_chart(cB, "K10")
cP = BarChart(); cP.type = "col"; cP.title = "Contatos Pliométricos por Semana"
cP.height = 9; cP.width = 20; cP.legend = None; cP.y_axis.title = "contatos"
cP.add_data(Reference(wsK, min_col=7, min_row=10, max_row=KW_L), titles_from_data=True)
cP.set_categories(Reference(wsK, min_col=1, min_row=KW_F, max_row=KW_L))
wsK.add_chart(cP, "K29")

secao(wsK, 20, "DISTRIBUIÇÃO DE REPETIÇÕES POR ZONA DE INTENSIDADE (bloco completo)", 9, 1)
cab_tabela(wsK, 21, ["Zona de Intensidade","Repetições","% do Total","Séries"])
ZONAS = ["Balística / Velocidade (<50%)","Velocidade-Força (50–69%)","Força-Velocidade (70–79%)",
         "Força Máxima (80–89%)","Força Máxima Alta (≥90%)"]
KZ_F = 22
for i, z in enumerate(ZONAS):
    r = KZ_F + i
    wsK.cell(r, 1, z)
    wsK.cell(r, 2, '=SUMIFS({0}!$X${1}:$X${2},{0}!$Z${1}:$Z${2},$A{3})'.format(PQ, PRF_F, PRF_L, r))
    wsK.cell(r, 3, '=IFERROR($B{0}/SUM($B${1}:$B${2}),0)'.format(r, KZ_F, KZ_F + 4))
    wsK.cell(r, 4, '=SUMIFS({0}!$J${1}:$J${2},{0}!$Z${1}:$Z${2},$A{3})'.format(PQ, PRF_F, PRF_L, r))
KZ_L = KZ_F + 4
corpo_tabela(wsK, KZ_F, KZ_L, 1, 4)
for r in range(KZ_F, KZ_L + 1):
    for c in range(1, 5):
        wsK.cell(r, c).font = Font(name=F, size=9, color=NAVY2)
        wsK.cell(r, c).fill = PatternFill("solid", fgColor=LIGHT if (r - KZ_F) % 2 == 0 else LIGHT2)
    wsK.cell(r, 1).alignment = Alignment(horizontal="left", vertical="center", indent=1)
    wsK.cell(r, 2).number_format = NF_UA; wsK.cell(r, 4).number_format = NF_UA
    wsK.cell(r, 3).number_format = NF_PCT
cZ = BarChart(); cZ.type = "bar"; cZ.title = "Repetições por Zona de Intensidade"
cZ.height = 7; cZ.width = 20; cZ.legend = None
cZ.add_data(Reference(wsK, min_col=2, min_row=21, max_row=KZ_L), titles_from_data=True)
cZ.set_categories(Reference(wsK, min_col=1, min_row=KZ_F, max_row=KZ_L))
wsK.add_chart(cZ, "K48")

secao(wsK, 29, "EVOLUÇÃO POR ATLETA", 9, 1)
cab_tabela(wsK, 30, ["Atleta","1RM Agachamento (kg)","Força Relativa (× massa)","CMJ Inicial (cm)","CMJ Atual (cm)",
                     "Δ CMJ","Potência de Pico Atual (W)","Perfil F-V","Prioridade"])
KA_F = 31
KA_L = KA_F + (CAD_L - CAD_F)
TS = "Testes"
def _prim(r):
    return ('SUMPRODUCT(MIN((Testes!$B${1}:$B${2}=$A{0})*Testes!$A${1}:$A${2}'
            '+(Testes!$B${1}:$B${2}<>$A{0})*100000))').format(r, TST_F, TST_L)
def _ult(r):
    return 'SUMPRODUCT(MAX((Testes!$B${1}:$B${2}=$A{0})*Testes!$A${1}:$A${2}))'.format(r, TST_F, TST_L)
def _teste(r, col, quando):
    return ('=IF($A{0}="",0,IFERROR(AVERAGEIFS(Testes!${3}${1}:${3}${2},Testes!$B${1}:$B${2},$A{0},'
            'Testes!$A${1}:$A${2},{4}),0))').format(r, TST_F, TST_L, col, quando)
for i in range(CAD_L - CAD_F + 1):
    r = KA_F + i
    a = CAD_F + i
    wsK.cell(r, 1, '=IF(Cadastro!$B{0}="","",Cadastro!$B{0})'.format(a))
    wsK.cell(r, 2, '=IF($A{0}="",0,IFERROR(INDEX(\'Força 1RM\'!$Q${1}:$Q${2},'
                   'MATCH($A{0},\'Força 1RM\'!$P${1}:$P${2},0)),0))'.format(r, RM_F, RM_MAT_L))
    wsK.cell(r, 3, '=IF($A{0}="",0,IFERROR($B{0}/INDEX(Cadastro!$P${1}:$P${2},'
                   'MATCH($A{0},Cadastro!$B${1}:$B${2},0)),0))'.format(r, CAD_F, CAD_L))
    wsK.cell(r, 4, _teste(r, "L", _prim(r)))
    wsK.cell(r, 5, _teste(r, "L", _ult(r)))
    wsK.cell(r, 6, '=IF(OR($D{0}=0,$E{0}=0),0,$E{0}/$D{0}-1)'.format(r))
    wsK.cell(r, 7, _teste(r, "S", _ult(r)))
    wsK.cell(r, 8, '=IF($A{0}="","",IFERROR(INDEX(\'Perfil F-V-P\'!$K${1}:$K${2},'
                   'MATCH($A{0},\'Perfil F-V-P\'!$A${1}:$A${2},0)),""))'.format(r, FVE_F, FVE_L))
    wsK.cell(r, 9, '=IF($A{0}="","",IFERROR(INDEX(\'Perfil F-V-P\'!$L${1}:$L${2},'
                   'MATCH($A{0},\'Perfil F-V-P\'!$A${1}:$A${2},0)),""))'.format(r, FVE_F, FVE_L))
corpo_tabela(wsK, KA_F, KA_L, 1, 9)
for r in range(KA_F, KA_L + 1):
    for c in range(1, 10):
        wsK.cell(r, c).font = Font(name=F, size=9, bold=(c in (3, 6)), color=NAVY2)
        wsK.cell(r, c).fill = PatternFill("solid", fgColor=LIGHT if (r - KA_F) % 2 == 0 else LIGHT2)
    for c in (1, 8, 9):
        wsK.cell(r, c).alignment = Alignment(horizontal="left", vertical="center", indent=1)
    for c in (2, 4, 5):
        wsK.cell(r, c).number_format = '0.0;;""'
    wsK.cell(r, 3).number_format = NF_DEC
    wsK.cell(r, 6).number_format = '+0.0%;-0.0%;""'
    wsK.cell(r, 7).number_format = NF_UA
wsK.conditional_formatting.add("F{}:F{}".format(KA_F, KA_L),
    CellIsRule(operator="greaterThan", formula=["0.03"], fill=PatternFill("solid", fgColor=GREEN),
               font=Font(name=F, size=9, bold=True, color=GREEN_T)))
wsK.conditional_formatting.add("F{}:F{}".format(KA_F, KA_L),
    CellIsRule(operator="lessThan", formula=["-0.05"], fill=PatternFill("solid", fgColor=RED),
               font=Font(name=F, size=9, bold=True, color=RED_T)))
wsK.conditional_formatting.add("C{}:C{}".format(KA_F, KA_L),
    ColorScaleRule(start_type="min", start_color="FFFFFF", end_type="max", end_color="A9D08E"))
KA_CH = min(KA_L, KA_F + 19)
cC = BarChart(); cC.type = "col"; cC.title = "CMJ: primeiro teste × teste mais recente (cm)"
cC.height = 9; cC.width = 20
cC.add_data(Reference(wsK, min_col=4, max_col=5, min_row=30, max_row=KA_CH), titles_from_data=True)
cC.set_categories(Reference(wsK, min_col=1, min_row=KA_F, max_row=KA_CH))
wsK.add_chart(cC, "K67")
cFR = BarChart(); cFR.type = "bar"; cFR.title = "Força Relativa no Agachamento (× massa corporal)"
cFR.height = 10; cFR.width = 20; cFR.legend = None
cFR.add_data(Reference(wsK, min_col=3, min_row=30, max_row=KA_CH), titles_from_data=True)
cFR.set_categories(Reference(wsK, min_col=1, min_row=KA_F, max_row=KA_CH))
wsK.add_chart(cFR, "K86")
wsK.freeze_panes = "A11"


# ============================================================================
# 24) EVIDÊNCIAS CIENTÍFICAS
# ============================================================================
wsEv = wb.create_sheet("Evidências")
banner(wsEv, "BASE DE EVIDÊNCIAS — PREPARAÇÃO FÍSICA PARA VOLEIBOL",
       "Referências que sustentam as escolhas metodológicas desta planilha: periodização em blocos, LPO, pliometria, "
       "treino baseado em velocidade, perfil força-velocidade, antropometria e controle de carga.", 8, "1F6F4A")
larguras(wsEv, {"A":5,"B":30,"C":66,"D":30,"E":26,"F":72,"G":30,"H":54})
cab_tabela(wsEv, 6, ["Nº","Autores (ano)","Título","Periódico","Tipo de Estudo","Achado Principal",
                     "Onde é aplicado na planilha","Link"])
CS = "https://consensus.app/papers/details/"
EVID = [
("PERIODIZAÇÃO EM BLOCOS",),
("Issurin (2010)","New Horizons for the Methodology and Physiology of Training Periodization","Sports Medicine","Revisão",
 "Base conceitual da periodização em blocos: mesociclos com cargas altamente concentradas em poucas capacidades, em vez do desenvolvimento simultâneo de muitas.",
 "Aba Bloco Base — estrutura acumulação/transmutação/realização", CS+"f382ecd540d75e349364da666ea4c096/"),
("Issurin (2016)","Benefits and Limitations of Block Periodized Training Approaches to Athletes' Preparation","Sports Medicine","Revisão",
 "O modelo em blocos MULTIALVO superou a preparação tradicional em 28 estudos com esportes coletivos, de resistência e de força; o modelo unidirecional concentrado só serve a disciplinas de uma única capacidade.",
 "Bloco Base — escolha do modelo multialvo, e não do concentrado puro", CS+"c45056fdee8e538689e4b3e3cedc15f7/"),
("Stone et al. (2021)","Periodization and Block Periodization in Sports: Emphasis on Strength-Power Training","J Strength Cond Res","Narrativa",
 "Separa PERIODIZAÇÃO (fases e prazos) de PROGRAMAÇÃO (exercícios, volume, intensidade); esportes coletivos precisam da variante de blocos com múltiplos alvos.",
 "Macrociclo, Mesociclo e Bloco Base", CS+"f256bd9422b65ada8b27fd28db364b55/"),
("Rønnestad et al. (2018)","Block periodization of strength and endurance training is superior to traditional periodization in ice hockey players","Scand J Med Sci Sports","ECR",
 "Com volume e intensidade IGUAIS, o grupo em blocos melhorou mais o torque de extensores de joelho e o VO2máx que o grupo tradicional.",
 "Bloco Base — justificativa do modelo", CS+"db1410c0c5ca5209811bb4e58eb7ec1b/"),
("Manchado et al. (2017)","Effects of Two Different Training Periodization Models on Elite Female Team Handball Players","J Strength Cond Res","Longitudinal",
 "Blocos superaram a periodização tradicional em salto agachado (+5,97%), CMJ (+8,76%), preensão manual, 1RM de supino e sprints de 10 e 20 m.",
 "Bloco Base — transferência para esporte coletivo de quadra", CS+"d8c5245dada2544285c48a147bbc308f/"),
("Bartolomei et al. (2014)","A Comparison of Traditional and Block Periodized Strength Training Programs in Trained Athletes","J Strength Cond Res","ECR",
 "Em 15 semanas com volume igual, o grupo em blocos ampliou mais a área sob a curva força-potência de MMSS; sem diferença em MMII.",
 "Bloco Base — expectativa realista de efeito", CS+"51ebe06543295477a6e3b87ed2fffb56/"),
("Painter et al. (2012)","Strength gains: block versus daily undulating periodization weight training among track and field athletes","Int J Sports Physiol Perform","ECR",
 "Sem diferença estatística entre blocos e ondulatório diário, mas os blocos foram mais EFICIENTES: mais ganho por unidade de volume-carga.",
 "Prescrição Força — controle de tonelagem", CS+"37aba2c3463d5db18d9682ba4f77c696/"),
("Gavanda et al. (2019)","The Effect of Block Versus Daily Undulating Periodization on Strength and Performance in Adolescent Football Players","Int J Sports Physiol Perform","ECR",
 "Em 12 semanas com adolescentes, blocos e ondulatório diário produziram ganhos equivalentes de massa muscular, força e potência.",
 "Bloco Base — ressalva para categorias de base (Sub-19)", CS+"8fc30754f83059c691319023497a547a/"),
("Micke et al. (2026)","Block periodization vs. traditional periodization in high-intensity functional training","Front Physiol","Crossover randomizado",
 "Blocos produziram maior ganho de força máxima; houve queda AGUDA do CMJ na semana de carga concentrada, confirmando overreaching funcional.",
 "Semanas 3 e 7 (CHOQUE) — o que esperar e o que monitorar", CS+"cb319ced4f525fccb29ad766c357d07a/"),
("FORÇA E POTÊNCIA NO VOLEIBOL",),
("Berriel et al. (2022)","Does Complex Training Enhance Vertical Jump Performance and Muscle Power in Elite Male Volleyball Players?","Int J Sports Physiol Perform","ECR",
 "Em 4 semanas, salto e potência melhoraram tanto com treino de saltos quanto com treino complexo; acrescentar estímulo pesado NÃO trouxe ganho adicional em atletas de elite.",
 "Prescrição Força — não sobrecarregar o elenco de elite com estímulo pesado desnecessário", CS+"342916b5711d52e1a81ed4f1f87bdecb/"),
("Rong et al. (2024)","Effects of Cluster vs. Traditional Sets Complex Training on Trained Male Volleyball Players","J Sports Sci Med","ECR",
 "Em 6 semanas, séries em CLUSTER produziram melhores CMJ, salto de ataque, T-test e potência de pico, além de menor cortisol de repouso, que séries tradicionais.",
 "Semana 7 do Bloco Base — agachamento em cluster 2+2", CS+"cc1f0acd730e55469151f4ddd60ad2d8/"),
("Cin et al. (2021)","Cluster Resistance Training Results Higher Improvements on Sprint, Agility, Strength and Vertical Jump in Professional Volleyball Players","Turkiye Klinikleri J Sports Sci","ECR",
 "Cluster superou o treino tradicional em 1RM de agachamento, terra, supino e pullover, sprint de 20 m, T-test e salto vertical.",
 "Prescrição Força — configuração de séries", CS+"5d6956f74040559da080d9e04b1cb3cc/"),
("Moussi et al. (2025)","Effects of two periodization models (linear vs. nonlinear) in young adult male volleyball players","J Bodyw Mov Ther","ECR",
 "6 semanas: ambos eficazes. O não linear foi melhor para salto vertical, salto de bloqueio e SJ; o linear, melhor para sprint de 10 m.",
 "Microciclo — variação de intensidade dentro da semana", CS+"9c60d35e1632588d8f6e0bf8707415fd/"),
("Wang et al. (2022)","Effect of Leg Half-Squat Training With Blood Flow Restriction Under Different External Loads in Volleyball Players","Dose-Response","ECR",
 "Carga alta (70% 1RM) foi o que melhorou o salto; carga baixa isolada aumentou pouco a força e não melhorou o salto.",
 "Prescrição Força — justificativa das faixas de 80-90% 1RM", CS+"9f2b4f05a98c52f2827718b3afa7bd36/"),
("Marques et al. (2009)","Physical Fitness Qualities of Professional Volleyball Players: Determination of Positional Differences","J Strength Cond Res","Transversal",
 "Centrais e opostos são os mais altos, pesados e fortes no supino; levantadores têm o pior desempenho no agachamento paralelo.",
 "Cadastro e Prescrição Força — individualização por posição", CS+"7f91b9f3f4aa502db2c5986d5aff1136/"),
("PLIOMETRIA",),
("Sáez de Villarreal et al. (2009)","Determining Variables of Plyometric Training for Improving Vertical Jump Height Performance","J Strength Cond Res","Meta-análise (56 estudos)",
 "Mais de 10 semanas, mais de 20 sessões, programas de alta intensidade com mais de 50 saltos por sessão e a COMBINAÇÃO de tipos de salto (SJ + CMJ + drop jump) maximizam o ganho. Peso adicional não trouxe benefício extra.",
 "Bloco Base e Saltos — dose de contatos por semana", CS+"e60a777fc6a654d29ca1c61c7d29da22/"),
("Sáez-Sáez de Villarreal et al. (2010)","Does plyometric training improve strength performance? A meta-analysis","J Sci Med Sport","Meta-análise",
 "Combinar pliometria COM treino de força supera usar qualquer uma delas isoladamente; mais de 40 saltos por sessão em alta intensidade otimiza o ganho de força.",
 "Sessões A e B — força e pliometria na mesma semana", CS+"049130fbe9d05f56a4810dde0ff51ef8/"),
("Iranpour et al. (2025)","The effects of plyometric training with speed and weight overloads on volleyball players","PLOS One","ECR",
 "Pliometria com sobrecarga de velocidade E de peso foi superior à pliometria simples em salto de ataque, altura de salto e taxa de produção de força.",
 "Sessão B — salto unilateral com sobrepeso e jump squat", CS+"07f5b6752bba54198af3c2d0e5cd9d87/"),
("Yu et al. (2025)","The influence of training surface on the effectiveness of plyometric training in volleyball players","Scientific Reports","ECR",
 "Pliometria na AREIA e na ÁGUA gerou mais ganho de CMJ que no solo; a aquática também superou o solo em sprint, mudança de direção e força máxima.",
 "Aba Saltos — coluna Superfície", CS+"75bc1970db385dd2a7fc798a2d496428/"),
("Atıcı et al. (2025)","Effects of plyometric and agility-based training in adolescent male volleyball players","Ped Phys Cult Sports","Experimental controlado",
 "8 semanas de pliometria: +14,1% no salto vertical, +13,25% na potência de pico e +3,07% no sprint. Agilidade melhora COD, não potência.",
 "Bloco Base — expectativa de ganho em 8 semanas", CS+"97766fe403ed5d0db00747ff1e0c3b71/"),
("Zhou et al. (2024)","Meta-analysis of the effect of plyometric training on youth basketball players","Front Physiol","Meta-análise (24 estudos)",
 "Pliometria de BAIXA frequência (1-2x/semana), ALTO volume (>150 saltos/semana) e de tipo MISTO melhorou salto, sprint, COD e equilíbrio; alta frequência com baixo volume só melhorou o salto.",
 "Saltos — faixas de referência de volume semanal", CS+"127de9c932de5022bf7207226214e479/"),
("LEVANTAMENTO DE PESO OLÍMPICO (LPO) E DERIVADOS",),
("Suchomel et al. (2015)","Weightlifting Pulling Derivatives: Rationale for Implementation and Application","Sports Medicine","Revisão",
 "Clean pull, snatch pull, hang high pull, jump shrug e mid-thigh pull dão estímulo igual ou melhor que os levantamentos completos, com técnica mais simples — ideal para quem não é levantador de peso.",
 "Sessões A e C — escolha de puxadas em vez do clean completo", CS+"e55b8645d8505f2bb685fbf171e41da5/"),
("Suchomel et al. (2017)","Force-Time-Curve Comparison Between Weight-Lifting Derivatives","Int J Sports Physiol Perform","Transversal",
 "O jump shrug produziu maior força relativa, impulso relativo e taxa de produção de força que o hang power clean e o hang high pull.",
 "Biblioteca de Exercícios — seleção de derivados", CS+"832a4d9da8035145b9b54feac5caa891/"),
("Suchomel et al. (2017)","Power-Time Curve Comparison between Weightlifting Derivatives","J Sports Sci Med","Transversal",
 "Cargas de 30-45% 1RM maximizam a potência no jump shrug e no hang high pull; 65-80% 1RM é a faixa ótima para o hang power clean.",
 "Prescrição Força — %1RM prescrito para o LPO", CS+"d4fd5a5594c45196bd8036e06e06ddd5/"),
("Suchomel et al. (2020)","Training With Weightlifting Derivatives: The Effects of Force and Velocity Overload Stimuli","J Strength Cond Res","ECR (10 semanas)",
 "Sobrecarga específica de força e de velocidade nas puxadas produziu os maiores ganhos em força isométrica relativa, sprint e mudança de direção.",
 "Semanas 6 e 7 — clean pull com carga supramáxima (95%)", CS+"bf6021458f755ca3a8d88965905119ed/"),
("Meechan et al. (2025)","The Effect of Load on Subphase Analysis During the Hang Pull","J Strength Cond Res","Transversal",
 "40% 1RM maximiza a velocidade de propulsão e 140% 1RM maximiza a força: o hang pull cabe tanto no mesociclo de força máxima quanto no de força-velocidade.",
 "Semana 6 — justificativa da carga supramáxima", CS+"45131b3cbdca5cfdb31431fdb5b7b004/"),
("James et al. (2022)","Rate of Force Development Adaptations After Weightlifting-Style Training: The Influence of Power Clean Ability","J Strength Cond Res","ECR",
 "Atletas FORTES no power clean melhoram a RFD em cargas leves; atletas FRACOS deslocam-se para o lado da força do perfil. Quem é fraco precisa primeiro de força máxima.",
 "Perfil F-V-P — decisão de ênfase por atleta", CS+"df935f1fd927508eb7134ed0c1a65ea2/"),
("Mehls et al. (2022)","An Examination of Loading Profiles for Youth Athletes Performing the Hang Power Clean","Mont J Sports Sci Med","Transversal",
 "Em jovens, a potência é máxima a ~70% 1RM, mas a velocidade da barra cai em cargas mais leves que em adultos: precisam de mais força antes de treinar pesado.",
 "Categoria Sub-19 — cautela com carga alta", CS+"011fd93e694d529a90e28eae7334b9c4/"),
("TREINAMENTO BASEADO EM VELOCIDADE (VBT)",),
("Weakley et al. (2020)","Velocity-Based Training: From Theory to Application","Strength Cond J","Revisão aplicada",
 "Como montar perfis carga-velocidade, dar feedback objetivo, usar limiares de perda de velocidade e integrar o VBT aos modelos de periodização.",
 "Prescrição Força — colunas de velocidade", CS+"a39766be63eb5ac7a0b76b6229d01c5f/"),
("García Ramos (2023)","Resistance Training Intensity Prescription Methods Based on Lifting Velocity Monitoring","Int J Sports Med","Revisão",
 "Três formas de prescrever %1RM pela velocidade — zonas, relação generalizada e relação INDIVIDUALIZADA — e os fatores que afetam a precisão de cada uma.",
 "Tabela carga-velocidade da aba Prescrição Força", CS+"7f88e021c22957a980dea51d84dd960a/"),
("Balsalobre-Fernández et al. (2021)","The Implementation of Velocity-Based Training Paradigm for Team Sports","Sports","Revisão",
 "Como aplicar VBT em elencos grandes, combinando métricas de velocidade com escalas subjetivas e estimando o 1RM diariamente.",
 "Prescrição Força — uso prático com o elenco todo", CS+"bd88451c5a7955fa8fa376e203a6aa0e/"),
("Hickmott et al. (2022)","The Effect of Load and Volume Autoregulation on Muscular Strength and Hypertrophy","Sports Med Open","Meta-análise",
 "Perda de velocidade ≤ 25% favorece FORÇA (menos fadiga aguda, mais adaptação crônica); perda > 20-25% favorece HIPERTROFIA por acumular mais volume.",
 "Coluna Perda de Velocidade Limite (10-15% no bloco de força)", CS+"0e6bb3b98d875b1392222cf931190cb8/"),
("Jiménez-Reyes et al. (2021)","Differences between adjusted vs. non-adjusted loads in velocity-based training","PeerJ","ECR (8 semanas)",
 "Sem ajuste diário pela velocidade, os atletas treinaram ~15% 1RM mais leve que o programado e não chegaram à intensidade alvo.",
 "Prescrição Força — por que registrar a velocidade obtida", CS+"c8cd652b51485409ae4e353f96d0cecd/"),
("Greig et al. (2023)","The Predictive Validity of Individualised Load-Velocity Relationships for Predicting 1RM","Sports Medicine","Revisão sistemática com meta-análise de dados individuais",
 "A estimativa do 1RM pela velocidade SUPERESTIMA a força real em ~3,7% (SEE ~9,8%). Use teste direto sempre que possível; a velocidade serve para acompanhar tendências.",
 "Força 1RM — ressalva sobre estimativas", CS+"94ecaddd8a3257c1ac5ba343f1a358be/"),
("LeMense et al. (2024)","Validity of Using the Load-Velocity Relationship to Estimate 1RM in the Back Squat","J Strength Cond Res","Revisão sistemática com meta-análise",
 "O método do limiar de velocidade mínima superestima o 1RM no agachamento livre e não é opção confiável de substituição do teste.",
 "Força 1RM — ressalva sobre estimativas", CS+"9f572c3e92ea5983aec88ccbe3db6da9/"),
("Morán-Navarro et al. (2020)","Load-velocity relationship of the deadlift exercise","Eur J Sport Sci","Transversal",
 "Velocidade média no 1RM do terra ≈ 0,24 m/s, consistente entre atletas de força diferente; a potência é máxima a ~60% 1RM.",
 "Tabela carga-velocidade — coluna Terra", CS+"2699a62b8e4350bd918acff60c192346/"),
("PERFIL FORÇA-VELOCIDADE-POTÊNCIA",),
("Morin & Samozino (2016)","Interpreting Power-Force-Velocity Profiles for Individualized and Specific Training","Int J Sports Physiol Perform","Revisão aplicada",
 "Método de campo para calcular F0, V0, Pmax e o desequilíbrio força-velocidade a partir da altura do salto com cargas progressivas e da distância de push-off.",
 "Perfil F-V-P — fórmulas usadas na planilha", CS+"0b47720f42de500d95c89cf488fa6bc6/"),
("Jiménez-Reyes et al. (2017)","Effectiveness of an Individualized Training Based on Force-Velocity Profiling during Jumping","Front Physiol","ECR",
 "Treino individualizado pelo desequilíbrio F-V melhorou o salto (+7 a +14%) mais que um programa igual para todos.",
 "Perfil F-V-P — lógica da recomendação por perfil", CS+"ac9d7b3fb607592887e952929efe8b2e/"),
("Jiménez-Reyes et al. (2019)","Optimized training for jumping performance using the force-velocity imbalance","PLoS ONE","Longitudinal",
 "O tempo necessário para corrigir o perfil é proporcional ao desequilíbrio inicial (12,6 ± 4,6 semanas em média); os ganhos se mantiveram após 3 semanas sem treino específico.",
 "Perfil F-V-P — duração realista da intervenção", CS+"07f17089e4935553836bfe8b6c443414/"),
("Li et al. (2026)","FV profile-based individualized vs. non-individualized strength training: systematic review and meta-analysis","BMC Sports Sci Med Rehabil","Meta-análise",
 "Vantagem grande do treino individualizado para F0, V0, desequilíbrio F-V e altura de salto; SEM efeito sobre Pmax e sprint.",
 "Perfil F-V-P — o que esperar de fato", CS+"bdbc8fc9675a54d29cca3b5eabf701bb/"),
("Lindberg et al. (2021)","Should we individualize training based on force-velocity profiling to improve physical performance?","Scand J Med Sci Sports","ECR (10 semanas)",
 "RESSALVA: não houve diferença entre treinar a favor, contra ou independentemente do perfil ótimo em 40 atletas de esportes coletivos.",
 "Perfil F-V-P — aviso na própria aba", CS+"5eda6a3b427f55098d20848a20b32336/"),
("Solberg et al. (2025)","Force-velocity profile based training to improve vertical jump performance: systematic review and meta-analysis","Scientific Reports","Meta-análise",
 "RESSALVA: os ganhos de salto com treino otimizado foram comparáveis aos do treino não otimizado; permanece incerto se o método é superior.",
 "Perfil F-V-P — aviso na própria aba", CS+"92d673eefa9b57158c87d11c3ec5fe71/"),
("Bobbert et al. (2024)","Is the Force-Velocity Profile for Free Jumping a Sound Basis for Individualized Jump Training Prescriptions?","Med Sci Sports Exerc","Simulação musculoesquelética",
 "RESSALVA: mudanças no perfil podem refletir aprendizado da tarefa (habilidade de saltar com e sem carga) e não adaptação neuromuscular.",
 "Perfil F-V-P — aviso na própria aba", CS+"6ab00db8b08e56fd86ec01c0c1b4054a/"),
("CARGA DE SALTOS E MONITORAMENTO",),
("Skazalski et al. (2018)","A valid and reliable method to measure jump-specific training and competition load in elite volleyball players","Scand J Med Sci Sports","Validação",
 "Dispositivo inercial contou 99,3% dos 3.637 saltos de treinos e jogos, mas superestimou a altura em ~5,5 cm: serve para CONTAR saltos, não para medir salto máximo.",
 "Aba Saltos — como registrar o jump count", CS+"4113ba2b7f9254a78670457ae8f99ac2/"),
("Charlton et al. (2017)","A simple method for quantifying jump loads in volleyball athletes","J Sci Med Sport","Validação",
 "Propõe um índice de carga externa a partir do produto entre número de saltos e energia cinética média.",
 "Aba Saltos — conceito de carga de salto", CS+"9895fbb37b4f5a628d174d7e58c624c4/"),
("Lin et al. (2024)","Quantifying internal and external training loads in collegiate male volleyball players","BMC Sports Sci Med Rehabil","Longitudinal (29 semanas)",
 "Carga interna correlaciona com o número de saltos (ρ = 0,477); MUITOS saltos associam-se a MENOR altura média e a menos saltos acima de 80% do máximo — sinal de fadiga.",
 "Saltos e Carga (PSE) — leitura conjunta", CS+"e05450c06ad55ab29a85378baaaf6adb/"),
("Taylor et al. (2022)","Quantifying External Load and Injury Occurrence in Women's Collegiate Volleyball Players","J Strength Cond Res","Longitudinal",
 "Atletas que se lesionaram tiveram MAIOR VARIABILIDADE da carga (CV 54% vs 41%): a instabilidade da carga importa tanto quanto o volume absoluto.",
 "Saltos — alerta de variação acima de 30% entre semanas", CS+"49131de7b79c5ffa86f91d25bcf2469a/"),
("Wang et al. (2025)","Comparison of external load and specific activities of starters vs. non-starters in men's professional volleyball","J Men's Health","Transversal",
 "Titulares acumulam mais player load, saltos totais e saltos de alta intensidade; centrais lideram os saltos de alta intensidade e os esforços repetidos.",
 "Saltos — controle individual por posição e por minutagem", CS+"8c773e2fbe795c968e9ea5185529f375/"),
("Villarejo-García et al. (2023)","Use, Validity and Reliability of Inertial Movement Units in Volleyball","Sensors","Revisão sistemática",
 "As IMUs têm boa validade para CONTAR saltos; a confiabilidade entre medidas de ALTURA ainda é limitada e contraditória.",
 "Saltos — o que confiar no dado do dispositivo", CS+"9be281c8d4d25bfaa2a0d31a09c1c0f8/"),
("ANTROPOMETRIA E PERFIL DO VOLEIBOLISTA",),
("Sheppard et al. (2009)","An Analysis of Playing Positions in Elite Men's Volleyball","J Strength Cond Res","Transversal (142 atletas)",
 "Centrais executam muito mais saltos de bloqueio e de ataque que levantadores e ponteiros; a seleção adulta supera a de base no CMJ e no salto de ataque RELATIVOS.",
 "Cadastro e Testes — metas por posição", CS+"acb87a06dc0b57e7896f92d25162767f/"),
("Palao et al. (2014)","Anthropometric, Physical, and Age Differences by the Player Position and the Performance Level in Volleyball","J Hum Kinet","Transversal (2.906 atletas)",
 "Normas de estatura, massa, alcance de ataque e alcance de bloqueio por posição em Jogos Olímpicos e Mundiais entre 2000 e 2012.",
 "Testes — referências de alcance por posição", CS+"bed94db8efef54d0b856b82b62dadcaf/"),
("Toselli & Campa (2018)","Anthropometry and Functional Movement Patterns in Elite Male Volleyball Players of Different Competitive Levels","J Strength Cond Res","Transversal",
 "O que separou as duas divisões foram medidas NÃO modificáveis (largura de úmero, estatura) e MODIFICÁVEIS (perímetro de braço contraído e área muscular do braço).",
 "Antropometria — perímetros e área muscular do braço", CS+"5d2c1a9f86b35fd89bf83100980361d1/"),
("Giannopoulos et al. (2017)","Somatotype, Level of Competition, and Performance in Attack in Elite Male Volleyball","J Hum Kinet","Transversal",
 "Atletas da divisão A1 são mais altos, pesados, musculosos e menos endomórficos que os da A2; centrais e opostos são endomorfo-ectomorfos.",
 "Antropometria — somatotipo Heath-Carter", CS+"8eb017eaf75f51948d230ec23aa8d1dd/"),
("De la Rosa et al. (2025)","Positional Profiling of Anthropometric, Baropodometric, and Grip Strength Traits in Male Volleyball Players","J Funct Morphol Kinesiol","Transversal (92 Sub-23)",
 "Centrais, ponteiros e opostos superam líberos e levantadores na maioria das medidas de membro superior; largura da mão e área muscular do braço predizem a preensão manual.",
 "Antropometria — perfil por posição", CS+"7dcedc0290ee5117bc3e007990be1d87/"),
("MÉTODOS CLÁSSICOS DE CONTROLE (já usados na v1 da planilha)",),
("Foster et al. (2001)","A new approach to monitoring exercise training","J Strength Cond Res","Método",
 "Carga interna = PSE da sessão (0-10) × duração em minutos; base também da monotonia e do strain.","Aba Carga (PSE)",""),
("Gabbett (2016)","The training-injury prevention paradox: should athletes be training smarter and harder?","Br J Sports Med","Revisão",
 "Razão entre carga aguda (7 dias) e crônica (28 dias) como ferramenta de gestão de risco; interpretar sempre com o contexto clínico.","Aba Carga (PSE) — ACWR",""),
("Hooper & Mackinnon (1995)","Monitoring overtraining in athletes","Sports Medicine","Revisão",
 "Índice de bem-estar de 4 itens (sono, estresse, fadiga e dor muscular), de 1 a 7 cada.","Aba Wellness",""),
("Sayers et al. (1999)","Cross-validation of three jump power equations","Med Sci Sports Exerc","Validação",
 "Equação de potência de pico a partir da altura do CMJ e da massa corporal.","Aba Testes — Potência de Pico",""),
("Jackson & Pollock (1978)","Generalized equations for predicting body density of men","Br J Nutr","Validação",
 "Equação de densidade corporal por 7 dobras cutâneas, convertida em % de gordura pela equação de Siri (1961).","Aba Antropometria",""),
("Carter & Heath (1990)","Somatotyping: Development and Applications","Cambridge University Press","Livro / método",
 "Método antropométrico de somatotipo (endomorfia, mesomorfia e ectomorfia).","Aba Antropometria — somatotipo",""),
]
EV_F = 7
r = EV_F
n = 0
for item in EVID:
    if len(item) == 1:
        wsEv.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        c = wsEv.cell(r, 1, item[0])
        c.font = Font(name=F, size=10, bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=NAVY2)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c.border = BORDER
        wsEv.row_dimensions[r].height = 20
        r += 1
        continue
    n += 1
    aut, tit, rev, tipo, ach, apl, url = item
    for c_, v in ((1, n), (2, aut), (3, tit), (4, rev), (5, tipo), (6, ach), (7, apl)):
        cc = wsEv.cell(r, c_, v)
        cc.font = Font(name=F, size=9, bold=(c_ == 2), color=NAVY2)
        cc.alignment = Alignment(horizontal="left" if c_ > 1 else "center", vertical="top", wrap_text=True, indent=1)
        cc.border = BORDER
        cc.fill = PatternFill("solid", fgColor=LIGHT if n % 2 else LIGHT2)
    lk = wsEv.cell(r, 8, url if url else "—")
    lk.border = BORDER
    lk.fill = PatternFill("solid", fgColor=LIGHT if n % 2 else LIGHT2)
    lk.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    if url:
        lk.hyperlink = url
        lk.font = Font(name=F, size=8, color="0563C1", underline="single")
    else:
        lk.font = Font(name=F, size=8, italic=True, color=GREY_T)
    wsEv.row_dimensions[r].height = 46
    r += 1
EV_L = r - 1
nota(wsEv, EV_L + 2, 2, "Os links levam ao registro do artigo. Referências sem link são obras clássicas de método, "
     "citadas pela fonte original.", 8)
wsEv.freeze_panes = "A7"
wsEv.auto_filter.ref = "A6:H{}".format(EV_L)


# ============================================================================
# 25) EXERCÍCIOS ADICIONAIS (do documento + LPO + pliometria)
# ============================================================================
EXTRA_EX = [
 ("Power clean a partir do joelho",FI,"N/A","Força Explosiva/Potência","Tripla extensão explosiva com técnica simplificada","Barra na altura do joelho, extensão explosiva de quadril, joelho e tornozelo, recepção em meio agachamento.","Barra, plataforma, anilhas emborrachadas",1,20,"Muito Alta",8,"Intermediário",""),
 ("Clean pull",FI,"N/A","Força Explosiva/Potência","Segunda puxada do arranco/arremesso sem recepção","Puxada completa sem receber a barra; permite cargas de 90 a 140% do 1RM do power clean.","Barra, plataforma",1,20,"Muito Alta",8,"Intermediário",""),
 ("Snatch pull / Hang high pull",FI,"N/A","Força Explosiva/Potência","Estímulo de tripla extensão com pegada aberta","Puxada alta a partir da suspensão; carga de 40% 1RM maximiza velocidade, 140% maximiza força.","Barra, plataforma",1,18,"Alta",7,"Intermediário",""),
 ("Jump shrug",FI,"N/A","Força Explosiva/Potência","Maior força relativa e taxa de produção de força entre os derivados de LPO","Encolhimento de ombros com salto, sem recepção. Cargas leves (30–45% 1RM do power clean).","Barra, plataforma",1,15,"Alta",7,"Iniciante",""),
 ("Salto no caixote",FI,"N/A","Pliometria","Potência concêntrica com baixa demanda de aterrissagem","Salto sobre caixote com aterrissagem amortecida em cima; descer andando.","Caixa pliométrica",1,12,"Alta",6,"Iniciante",""),
 ("Salto unilateral com sobrepeso",FI,"N/A","Pliometria","Pliometria unilateral com sobrecarga","Saltos alternados em uma perna com colete ou halteres leves; controle da aterrissagem.","Colete lastrado / halteres",1,12,"Alta",7,"Avançado",""),
 ("Arremesso de bola",FI,"N/A","Força Explosiva/Potência","Potência de membros superiores e tronco","Arremesso de medicine ball acima da cabeça ou em rotação; máxima velocidade de saída.","Medicine ball",2,10,"Alta",6,"Iniciante",""),
 ("Armação de braço",T,"Ataque","Coordenação","Padrão do braço de ataque com resistência leve","Simulação do movimento de ataque com elástico ou halter leve, foco na trajetória do cotovelo.","Elástico / halter leve",1,10,"Baixa",3,"Iniciante",""),
 ("Supino deitado com halteres",FI,"N/A","Força Máxima","Força horizontal de empurrar","Amplitude completa, halteres permitem trajetória mais livre para o ombro do atacante.","Halteres, banco",1,15,"Alta",7,"Intermediário",""),
 ("Supino sentado",FI,"N/A","Força Máxima","Força de empurrar em ângulo inclinado","Do programa original (Séries 3 e 5).","Máquina / halteres",1,15,"Alta",7,"Intermediário",""),
 ("Puxada alta",FI,"N/A","Força Máxima","Força de puxar vertical","Do programa original (Séries 1, 4, 6 e 8).","Polia alta",1,15,"Alta",7,"Iniciante",""),
 ("Remada unilateral na polia",FI,"N/A","Força Máxima","Força de puxar horizontal unilateral","Do programa original (Séries 2, 5 e 7).","Polia baixa",1,15,"Alta",6,"Iniciante",""),
 ("Remada serrote",FI,"N/A","Força Máxima","Força de puxar horizontal com apoio","Do programa original (Série 3).","Haltere, banco",1,15,"Alta",6,"Iniciante",""),
 ("Pullover deitado com haltere",FI,"N/A","Força Máxima","Força de latíssimo e mobilidade de ombro","Do programa original (Série 2).","Haltere, banco",1,12,"Moderada",6,"Intermediário",""),
 ("Pullover na polia",FI,"N/A","Força Máxima","Força de latíssimo com tensão constante","Do programa original (Séries 3 e 5).","Polia alta",1,12,"Moderada",6,"Iniciante",""),
 ("Pullover com haltere no banco",FI,"N/A","Força Máxima","Variação do pullover com maior amplitude","Do programa original (Série 7).","Haltere, banco",1,12,"Moderada",6,"Intermediário",""),
 ("Cross-over",FI,"N/A","Força Máxima","Adução horizontal de ombro","Do programa original (Série 4).","Polia dupla",1,12,"Moderada",5,"Iniciante",""),
 ("Cross-over supino",FI,"N/A","Força Máxima","Adução horizontal deitado","Do programa original (Série 8).","Polia dupla, banco",1,12,"Moderada",5,"Intermediário",""),
 ("Bíceps / Tríceps",FI,"N/A","Força Máxima","Trabalho complementar de braço","Do programa original (Séries 2, 4 e 6).","Halteres / polia",1,12,"Moderada",5,"Iniciante",""),
 ("Tríceps",FI,"N/A","Força Máxima","Extensores de cotovelo","Do programa original (Série 7).","Polia / halter",1,10,"Moderada",5,"Iniciante",""),
 ("Bíceps",FI,"N/A","Força Máxima","Flexores de cotovelo","Do programa original (Série 8).","Halteres / barra",1,10,"Moderada",5,"Iniciante",""),
 ("Agachamento profundo",FI,"N/A","Força Máxima","Força máxima de MMII em amplitude completa","Exercício-âncora do bloco de base. Do programa original (Séries 1, 3 e 7).","Barra, rack",1,25,"Muito Alta",8,"Intermediário",""),
 ("Agachamento",FI,"N/A","Força Máxima","Força máxima de MMII","Do programa original (Série 5).","Barra, rack",1,25,"Alta",7,"Iniciante",""),
 ("Agachamento frontal",FI,"N/A","Força Máxima","Força de MMII com maior demanda de tronco","Complemento ao agachamento profundo; útil como exercício de teste.","Barra, rack",1,20,"Alta",7,"Intermediário",""),
 ("Agachamento sumô no minitramp",FI,"N/A","Força Explosiva/Potência","Agachamento com base ampla e superfície instável/elástica","Do programa original (Série 2).","Minitramp, halteres",1,15,"Moderada",6,"Intermediário",""),
 ("Stiff com barra",FI,"N/A","Força Máxima","Cadeia posterior e prevenção de isquiotibiais","Do programa original (Séries 1, 3, 5 e 7).","Barra, anilhas",1,18,"Alta",7,"Intermediário",""),
 ("Stiff",FI,"N/A","Força Máxima","Cadeia posterior","Do programa original (Séries 3, 5 e 7).","Barra / halteres",1,18,"Alta",7,"Iniciante",""),
 ("Afundo lateral sem passada",FI,"N/A","Força Máxima","Força unilateral no plano frontal","Do programa original (Série 1).","Halteres",1,12,"Moderada",6,"Iniciante",""),
 ("Afundo lateral com tornozeleira",FI,"N/A","Força Máxima","Força unilateral com sobrecarga distal","Do programa original (Série 3).","Tornozeleira",1,12,"Moderada",5,"Iniciante",""),
 ("Afundo lateral com sobrepeso",FI,"N/A","Força Máxima","Força unilateral com carga externa","Do programa original (Série 7).","Colete / halteres",1,12,"Alta",6,"Intermediário",""),
 ("Afundo frontal sem passada",FI,"N/A","Força Máxima","Força unilateral no plano sagital","Do programa original (Série 2).","Halteres",1,12,"Moderada",6,"Iniciante",""),
 ("Afundo frontal sem passada com tornozeleira",FI,"N/A","Força Máxima","Força unilateral com sobrecarga distal","Do programa original (Série 4).","Tornozeleira",1,12,"Moderada",5,"Iniciante",""),
 ("Afundo frontal com sobrepeso (sem passada)",FI,"N/A","Força Máxima","Força unilateral com carga externa","Do programa original (Série 8).","Colete / halteres",1,12,"Alta",6,"Intermediário",""),
 ("Flexão de joelho unilateral",FI,"N/A","Força Máxima","Isquiotibiais unilateral","Do programa original (Séries 2 e 4).","Mesa flexora",1,12,"Moderada",6,"Iniciante",""),
 ("Flexão de joelhos",FI,"N/A","Força Máxima","Isquiotibiais bilateral","Do programa original (Séries 6 e 8).","Mesa flexora",1,12,"Moderada",6,"Iniciante",""),
 ("Elevação de calcanhares",PV,"N/A","Força Máxima","Tríceps sural e prevenção de tornozelo","Do programa original (Séries 4, 6 e 8). Descida lenta de 3 s.","Step, halteres",1,10,"Moderada",5,"Iniciante",""),
 ("Abdominal reto com braços esticados",FI,"N/A","Core/Estabilidade","Flexão de tronco","Do programa original (Série 1).","Colchonete",1,8,"Moderada",4,"Iniciante",""),
 ("Abdominal cruzado esticado",FI,"N/A","Core/Estabilidade","Flexão com rotação","Do programa original (Série 2).","Colchonete",1,8,"Moderada",4,"Iniciante",""),
 ("Abdominal infra",FI,"N/A","Core/Estabilidade","Porção inferior do reto abdominal","Do programa original (Série 3).","Colchonete / barra fixa",1,8,"Moderada",4,"Iniciante",""),
 ("Abdominal reto",FI,"N/A","Core/Estabilidade","Flexão de tronco","Do programa original (Série 4).","Colchonete",1,8,"Moderada",4,"Iniciante",""),
 ("Abdominal cruzado com anilha",FI,"N/A","Core/Estabilidade","Flexão com rotação e sobrecarga","Do programa original (Séries 5 e 7).","Anilha, colchonete",1,8,"Alta",5,"Intermediário",""),
 ("Abdominal cruzado com anilha de 20",FI,"N/A","Core/Estabilidade","Flexão com rotação e carga alta","Do programa original (Série 7).","Anilha de 20 kg",1,8,"Alta",6,"Avançado",""),
 ("Abdominal remador com anilha",FI,"N/A","Core/Estabilidade","Flexão simultânea de tronco e quadril","Do programa original (Série 6).","Anilha, colchonete",1,8,"Alta",5,"Intermediário",""),
 ("Abdominal reto canivete",FI,"N/A","Core/Estabilidade","Flexão simultânea de tronco e quadril","Do programa original (Série 8).","Colchonete",1,8,"Alta",5,"Intermediário",""),
 ("Dorsal perdigueiro",PV,"N/A","Core/Estabilidade","Estabilidade lombo-pélvica em quatro apoios","Do programa original (Séries 1 e 3).","Colchonete",1,8,"Baixa",3,"Iniciante",""),
 ("Dorsal reto",PV,"N/A","Core/Estabilidade","Extensores de tronco","Do programa original (Séries 2 e 4).","Banco romano / colchonete",1,8,"Baixa",3,"Iniciante",""),
 ("Levantamento terra",FI,"N/A","Força Máxima","Força máxima de cadeia posterior","Exercício de teste de 1RM; velocidade média no 1RM ≈ 0,24 m/s.","Barra, anilhas",1,25,"Muito Alta",8,"Avançado",""),
]
rex = EX_F + len(EXERCICIOS)
for e in EXTRA_EX:
    for j, v in enumerate(e):
        wsE.cell(rex, 2 + j, v)
    rex += 1

# ============================================================================
# 26) DADOS DE EXEMPLO
# ============================================================================
random.seed(20260903)
MASSA_EX = {"Rafael Monteiro Alves":88.0,"Diego Salgado Ferraz":95.0,"Lucas Prado Bittencourt":89.0,
            "Bruno Rezende Camargo":87.0,"Thiago Nogueira Vasques":99.0,"Matheus Caldeira Lins":95.0,
            "Felipe Andrade Rocha":78.0,"Gustavo Peixoto Maia":91.0,"Vinícius Barreto Duarte":84.0,
            "André Luiz Sampaio":97.0,"Pedro Henrique Coutinho":93.0,"Caio Fernandes Bastos":79.0}

# ---- Antropometria ---------------------------------------------------------
ra = ANT_F
for momento, dt in [("Pré-temporada", date(2026, 1, 12)), ("Meio da Temporada", date(2026, 7, 22))]:
    g = 0 if momento == "Pré-temporada" else 1
    for nome in NOMES:
        est = EST_EX[nome] * 1.0
        m = MASSA_EX[nome] + (0 if g == 0 else round(random.uniform(-1.5, 1.0), 1))
        vals = {1: dt, 2: nome, 3: momento, 4: "Comissão técnica",
                5: round(m, 1), 6: float(est), 7: round(est * 0.522, 1), 8: round(est + random.uniform(2, 7), 1),
                9: round(random.uniform(7, 11) - g * 0.6, 1), 10: round(random.uniform(9, 13) - g * 0.5, 1),
                11: round(random.uniform(3, 6), 1), 12: round(random.uniform(5, 9) - g * 0.4, 1),
                13: round(random.uniform(6, 10) - g * 0.5, 1), 14: round(random.uniform(8, 14) - g * 0.9, 1),
                15: round(random.uniform(6, 10) - g * 0.5, 1), 16: round(random.uniform(10, 18) - g * 1.2, 1),
                17: round(random.uniform(9, 15) - g * 0.7, 1), 18: round(random.uniform(6, 10) - g * 0.4, 1),
                19: round(random.uniform(31, 35) + g * 0.4, 1), 20: round(random.uniform(33, 38) + g * 0.5, 1),
                21: round(random.uniform(27, 30), 1), 22: round(random.uniform(96, 105) + g * 0.6, 1),
                23: round(random.uniform(78, 86) - g * 0.8, 1), 24: round(random.uniform(80, 88) - g * 0.8, 1),
                25: round(random.uniform(95, 102), 1), 26: round(random.uniform(55, 62) + g * 0.5, 1),
                27: round(random.uniform(37, 41) + g * 0.3, 1), 28: round(random.uniform(17, 19), 1),
                29: round(random.uniform(41, 45), 1), 30: round(random.uniform(27, 30), 1),
                31: round(random.uniform(7.0, 7.8), 1), 32: round(random.uniform(9.5, 10.5), 1)}
        for c, v in vals.items():
            wsAn.cell(ra, c, v)
        wsAn.cell(ra, 1).number_format = NF_DATE
        ra += 1

# ---- Testes ----------------------------------------------------------------
rt = TST_F
ALC_PE = {n: round(EST_EX[n] * 1.29) for n in NOMES}
CMJ_BASE = {n: round(random.uniform(45, 58), 1) for n in NOMES}
for momento, dt in [("Pré-temporada", date(2026, 1, 14)), ("Meio da Temporada", date(2026, 7, 24))]:
    g = 0 if momento == "Pré-temporada" else 1
    for nome in NOMES:
        cmj = round(CMJ_BASE[nome] + g * random.uniform(1.4, 3.4), 1)
        sj = round(cmj - random.uniform(3, 5), 1)
        dj = round(cmj - random.uniform(1, 3), 1)
        tc = round(random.uniform(0.20, 0.26), 3)
        pe = ALC_PE[nome]
        vals = {1: dt, 2: nome, 3: momento, 4: "Preparação física",
                5: round(MASSA_EX[nome] + (0 if g == 0 else random.uniform(-1.5, 1.0)), 1),
                6: pe, 7: pe + round(random.uniform(68, 82) + g * 2), 8: pe + round(random.uniform(55, 70) + g * 2),
                11: sj, 12: cmj, 13: round(cmj + random.uniform(6, 9), 1), 14: dj, 15: tc,
                18: round(cmj + random.uniform(14, 20), 1),
                20: round(random.uniform(1.05, 1.20) - g * 0.02, 2), 21: round(random.uniform(1.72, 1.90) - g * 0.03, 2),
                22: round(random.uniform(9.2, 10.2) - g * 0.15, 2), 23: round(random.uniform(5.5, 6.2) - g * 0.08, 2),
                24: round(random.uniform(12, 16) + g * 0.5, 1), 25: round(random.uniform(22, 34), 1),
                26: int(random.uniform(1200, 2000) + g * 120), 27: int(random.uniform(2800, 4200) + g * 150),
                29: "EXEMPLO"}
        for c, v in vals.items():
            wsT.cell(rt, c, v)
        wsT.cell(rt, 1).number_format = NF_DATE
        rt += 1

# ---- Força 1RM -------------------------------------------------------------
FATOR_1RM = {"Agachamento profundo":1.70,"Agachamento frontal":1.35,"Stiff com barra":1.45,
             "Levantamento terra":2.00,"Supino sentado":1.05,"Supino deitado com halteres":0.80,
             "Remada serrote":0.55,"Power clean":1.05,"Clean pull":1.30}
rr = RM_F
for dt, g in [(date(2026, 1, 15), 0), (date(2026, 7, 25), 1)]:
    for nome in NOMES:
        for ex, fat in FATOR_1RM.items():
            carga = round(MASSA_EX[nome] * fat * random.uniform(0.92, 1.08) * (1 + g * 0.05) / 2.5) * 2.5
            wsR.cell(rr, 1, dt); wsR.cell(rr, 1).number_format = NF_DATE
            wsR.cell(rr, 3, nome); wsR.cell(rr, 4, ex)
            wsR.cell(rr, 5, "Direto (1RM real)"); wsR.cell(rr, 6, carga); wsR.cell(rr, 7, 1)
            wsR.cell(rr, 14, "EXEMPLO")
            rr += 1

# ---- Perfil F-V-P do elenco ------------------------------------------------
for i, nome in enumerate(NOMES):
    r = FVE_F + i
    massa = MASSA_EX[nome]
    f0 = round(massa * random.uniform(30, 37))
    v0 = round(random.uniform(3.0, 4.2), 2)
    sfv_rel = -(f0 / massa) / v0
    wsFV.cell(r, 3, f0); wsFV.cell(r, 4, v0)
    wsFV.cell(r, 9, round(sfv_rel * random.uniform(0.82, 1.18), 2))

# ---- Carga (PSE), Presença, Wellness, Saltos -------------------------------
SEMANAS_EX = [date(2026, 8, 10), date(2026, 8, 17), date(2026, 8, 24), date(2026, 8, 31)]
SESSOES_EX = [(0, "Físico (Força)", 75, 7), (0, "Técnico-Tático", 90, 6), (1, "Físico (Potência)", 60, 7),
              (2, "Coletivo/Jogo", 100, 8), (3, "Físico (Condicionamento)", 50, 6), (3, "Técnico-Tático", 90, 7),
              (4, "Coletivo/Jogo", 110, 8)]
LESIONADO = NOMES[3]
rg, rp = CAR_F, PRS_F
for wk_i, seg in enumerate(SEMANAS_EX):
    for off, tipo, dur, pse in SESSOES_EX:
        if wk_i == 3 and off > 2:
            continue
        d = seg + timedelta(days=off)
        for nome in NOMES:
            lesao = (nome == LESIONADO and wk_i >= 2)
            falta = (not lesao) and random.random() < 0.05
            if not lesao and not falta:
                p = max(1, min(10, pse + random.choice([-1, 0, 0, 0, 1])))
                du = dur + random.choice([-10, -5, 0, 0, 5])
                wsG.cell(rg, 1, d); wsG.cell(rg, 1).number_format = NF_DATE
                wsG.cell(rg, 3, nome); wsG.cell(rg, 4, tipo)
                wsG.cell(rg, 5, du); wsG.cell(rg, 6, p); wsG.cell(rg, 13, "EXEMPLO")
                rg += 1
            wsPr.cell(rp, 1, d); wsPr.cell(rp, 1).number_format = NF_DATE
            wsPr.cell(rp, 3, tipo); wsPr.cell(rp, 4, nome)
            wsPr.cell(rp, 5, "Lesionado" if lesao else ("Falta Justificada" if falta else "Presente"))
            wsPr.cell(rp, 6, 0 if (lesao or falta) else dur)
            wsPr.cell(rp, 7, "Tendinopatia patelar — DM" if lesao else
                             ("Compromisso acadêmico" if falta else "EXEMPLO"))
            rp += 1
rw = WEL_F
for wk_i, seg in enumerate(SEMANAS_EX):
    for off in range(3 if wk_i == 3 else 5):
        d = seg + timedelta(days=off)
        for nome in NOMES:
            base = 4 if nome == LESIONADO else 2
            wsWe.cell(rw, 1, d); wsWe.cell(rw, 1).number_format = NF_DATE
            wsWe.cell(rw, 3, nome)
            for j in range(4):
                wsWe.cell(rw, 4 + j, max(1, min(7, base + random.choice([0, 0, 1, 1, 2]))))
            wsWe.cell(rw, 11, "EXEMPLO")
            rw += 1
rs = SLT_F
for wk_i, seg in enumerate(SEMANAS_EX):
    for nome in NOMES:
        if nome == LESIONADO and wk_i >= 2:
            continue
        for orig, ctx, n, inten, sup in [
                ("Pliometria", "Sessão B — drop jump e barreiras", random.randint(80, 130), "Alta", "Quadra (madeira)"),
                ("Treino de quadra", "Complexos K1 e K2", random.randint(190, 320), "Média", "Quadra (madeira)"),
                ("Jogo", "Rodada do estadual", random.choice([0, 70, 85]), "Muito Alta", "Quadra (madeira)")]:
            if n == 0:
                continue
            wsS2.cell(rs, 1, seg + timedelta(days=random.randint(0, 4)))
            wsS2.cell(rs, 1).number_format = NF_DATE
            wsS2.cell(rs, 3, nome); wsS2.cell(rs, 4, orig); wsS2.cell(rs, 5, ctx)
            wsS2.cell(rs, 6, n); wsS2.cell(rs, 7, inten); wsS2.cell(rs, 8, sup)
            wsS2.cell(rs, 10, random.randint(52, 74)); wsS2.cell(rs, 12, random.randint(3, 5))
            wsS2.cell(rs, 13, "EXEMPLO")
            rs += 1

# ---- Prescrição técnica de exemplo ----------------------------------------
PRESC_EX = [
 (date(2026,8,31),"S1 — Força","Manhã","Equipe (todos)","Aquecimento","Mobilidade de tornozelo e quadril",1,"8 por exercício","Peso corporal",30,10,2,"Amplitude completa","EXEMPLO"),
 (date(2026,8,31),"S1 — Força","Manhã","Equipe (todos)","Ativação / Prevenção","Rotadores do ombro com elástico",3,15,"Elástico vermelho",45,10,3,"Progredir a resistência a cada 2 semanas","EXEMPLO"),
 (date(2026,8,31),"S1 — Força","Manhã","Equipe (todos)","Parte Principal","Agachamento profundo",4,6,"75% de 1RM",180,25,8,"+2,5% se completar todas as séries","EXEMPLO"),
 (date(2026,8,31),"S1 — Força","Manhã","Equipe (todos)","Parte Principal","Stiff com barra",4,8,"65% de 1RM",120,18,7,"Excêntrica de 3 s","EXEMPLO"),
 (date(2026,8,31),"S1 — Força","Manhã","Equipe (todos)","Parte Principal","Power clean a partir do joelho",4,4,"70% de 1RM",180,20,8,"Parar a série se a barra desacelerar","EXEMPLO"),
 (date(2026,8,31),"S1 — Força","Manhã","Equipe (todos)","Complementar","Nórdico de isquiotibiais (excêntrico)",3,6,"Peso corporal",90,10,6,"2x por semana no bloco","EXEMPLO"),
 (date(2026,8,31),"S1 — Força","Manhã","Equipe (todos)","Volta à Calma","Alongamento e liberação miofascial",1,"30 s por grupamento","Rolo",0,15,2,"—","EXEMPLO"),
 (date(2026,8,31),"S2 — Técnico-Tático","Tarde","Equipe (todos)","Aquecimento","Recepção em duplas com bola dirigida",3,15,"Bola",30,12,4,"Meta de 80% ao alvo","EXEMPLO"),
 (date(2026,8,31),"S2 — Técnico-Tático","Tarde","Equipe (todos)","Parte Principal","Recepção de saque em sistema de 2 passadores",4,20,"Saque real",60,20,6,"Meta de 70% de passes A/B","EXEMPLO"),
 (date(2026,8,31),"S2 — Técnico-Tático","Tarde","Equipe (todos)","Parte Principal","Ataque de bola alta pela ponta (P4)",4,10,"Bola alta",60,18,6,"5 séries na semana seguinte","EXEMPLO"),
 (date(2026,8,31),"S2 — Técnico-Tático","Tarde","Equipe (todos)","Parte Principal","Complexo I (K1) — recepção, levantamento e ataque",1,"20 tentativas","Jogo",90,25,7,"Meta de 60% de side-out","EXEMPLO"),
 (date(2026,8,31),"S2 — Técnico-Tático","Tarde",NOMES[2],"Complementar","Saque potente com salto",4,6,"Máxima",90,15,7,"Registrar acertos por série","EXEMPLO — treino individual"),
 (date(2026,9,2),"S3 — Potência","Manhã","Equipe (todos)","Parte Principal","Agachamento com salto sob carga (jump squat)",5,4,"30% de 1RM",120,18,7,"Velocidade acima de 1,0 m/s","EXEMPLO"),
 (date(2026,9,2),"S3 — Potência","Manhã","Equipe (todos)","Parte Principal","Salto em profundidade (drop jump)",5,5,"Caixa de 40 cm",120,15,7,"Progredir só com boa aterrissagem","EXEMPLO"),
 (date(2026,9,2),"S4 — Coletivo","Tarde","Equipe (todos)","Parte Principal","Complexo II (K2) — bloqueio, defesa e contra-ataque",1,"20 séries","Jogo",60,25,7,"Avaliar eficiência do contra-ataque","EXEMPLO"),
 (date(2026,9,2),"S4 — Coletivo","Tarde","Equipe (todos)","Parte Principal","Jogo 6x6 com pontuação diferenciada",3,"sets a 15","Jogo",120,30,8,"—","EXEMPLO"),
]
for i, pex in enumerate(PRESC_EX):
    r = PRE_F + i
    dt, ses, turno, dest, bloco, exe, ser, rep, carga, pausa, dur, pse, prog, obs = pex
    wsP.cell(r, 1, dt); wsP.cell(r, 1).number_format = NF_DATE
    wsP.cell(r, 3, ses); wsP.cell(r, 4, turno); wsP.cell(r, 5, dest); wsP.cell(r, 6, bloco); wsP.cell(r, 7, exe)
    wsP.cell(r, 11, ser); wsP.cell(r, 12, rep); wsP.cell(r, 13, carga); wsP.cell(r, 14, pausa)
    wsP.cell(r, 15, dur); wsP.cell(r, 16, pse); wsP.cell(r, 19, prog); wsP.cell(r, 20, obs)

# ============================================================================
# 27) ORDEM DAS ABAS E GRAVAÇÃO
# ============================================================================
ORDEM = ["Início","Cadastro","Antropometria","Testes","Perfil F-V-P","Exercícios","Programa PF","Macrociclo",
         "Mesociclo","Bloco Base","Microciclo","Prescrição","Prescrição Força","Força 1RM","Carga (PSE)","Saltos",
         "Wellness","Presença","Atleta","Painel","KPIs Força","Evidências","Listas"]
for i, nome in enumerate(ORDEM):
    if nome in wb.sheetnames:
        wb.move_sheet(nome, offset=i - wb.sheetnames.index(nome))
wb.active = wb.sheetnames.index("Início")
for ws in wb.worksheets:
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True

BASE = "/home/user/mdlucca/planilhas/"
OUT = BASE + "Planilha_Voleibol_Forca_e_Potencia_v2.xlsx"
wb.save(OUT)
print("SALVO:", OUT)

# ---- versão em branco ------------------------------------------------------
def limpar(ws, r1, r2, cols):
    for r in range(r1, r2 + 1):
        for c in cols:
            ws.cell(r, c).value = None

limpar(wsC, CAD_F, CAD_L, [c for c in range(2, 66) if c not in (5, 15, 16, 17, 46, 60)])
limpar(wsAn, ANT_F, ANT_L, list(range(1, 33)))
limpar(wsT, TST_F, TST_L, [c for c in range(1, 30) if c not in (9, 10, 16, 17, 19, 28)])
limpar(wsR, RM_F, RM_L, [1, 3, 4, 5, 6, 7, 13, 14])
limpar(wsFV, FVE_F, FVE_L, [3, 4, 9])
limpar(wsM, MESO_F, MESO_L, [2, 3, 4, 5, 8, 9, 10, 11, 12, 13, 14, 15])
limpar(wsS, MIC_F, MIC_L, [2, 3, 4, 5, 6, 8, 9, 10, 13])
limpar(wsW, SF, SL, [3, 4, 5, 6, 7, 9, 10, 11, 12])
limpar(wsW, CF, CL, [2, 3, 4, 5, 6])
limpar(wsP, PRE_F, PRE_L, [1, 3, 4, 5, 6, 7, 11, 12, 13, 14, 15, 16, 19, 20])
limpar(wsG, CAR_F, CAR_L, [1, 3, 4, 5, 6, 13])
limpar(wsWe, WEL_F, WEL_L, [1, 3, 4, 5, 6, 7, 10, 11])
limpar(wsPr, PRS_F, PRS_L, [1, 3, 4, 5, 6, 7])
limpar(wsS2, SLT_F, SLT_L, [1, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13])
wsS["C6"] = None
wsW["B5"] = None; wsW["B6"] = None; wsW["B8"] = None; wsW["B9"] = None
for cel in ("C6", "C7", "C8", "C9", "C10", "C14", "C15"):
    wsM[cel] = None
OUT2 = BASE + "Planilha_Voleibol_Forca_e_Potencia_v2_EM_BRANCO.xlsx"
wb.save(OUT2)
print("SALVO:", OUT2)
