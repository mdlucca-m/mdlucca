#!/usr/bin/env python3
"""Testes do painel analítico: rota, exportação e a página.

    python3 -m unittest tests.test_panorama -v

O perigo aqui é o painel bonito que mente. Um gráfico que recebe o campo
com o nome errado não dá erro — ele desenha vazio, e ninguém percebe que
a rede temática está sem uma única linha. Foi o que aconteceu.
"""
from __future__ import annotations

import io
import json
import re
import sys
import tempfile
import threading
import unittest
import urllib.error
import urllib.request
from http.server import ThreadingHTTPServer
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "scripts"))

from lape import api, auth, ingest_excel, variaveis  # noqa: E402
from lape.db import Database  # noqa: E402

TEMPLATES = ROOT / "scripts" / "lape" / "templates"


class BasePanorama(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.tmp = tempfile.TemporaryDirectory()
        cls.db_path = Path(cls.tmp.name) / "t.sqlite"
        db = Database(cls.db_path)
        db.migrate()
        ingest_excel.ingest_articles(db, [
            {"title": "Treinamento resistido e ansiedade em fibromialgia",
             "authors": "Ana Souza; Beto Lima", "status": "Publicado",
             "year_published": 2024, "journal": "Pain", "doi": "10.1000/a.1"},
            {"title": "Dropout no treinamento resistido", "authors": "Ana Souza",
             "status": "Submetido", "first_submission_on": "2025-02-01"},
            {"title": "Um título sem variável nenhuma do vocabulário",
             "authors": "Beto Lima", "status": "Em produção", "started_on": "2023-05-01"},
        ])
        auth.create_account(db, "Ana Souza", "ana@udesc.br", "senhaforte123", role="admin")
        auth.create_account(db, "Curioso Silva", "curioso@udesc.br", "senhaforte123",
                            role="leitura")
        variaveis.instalar(db)
        variaveis.marcar_artigos(db)
        db.close()
        cls.publico = api.PUBLIC_DASHBOARD
        api.PUBLIC_DASHBOARD = False
        api.Handler.db_path = cls.db_path
        api.Handler.log_message = lambda *a, **k: None
        cls.server = ThreadingHTTPServer(("127.0.0.1", 0), api.Handler)
        cls.port = cls.server.server_address[1]
        threading.Thread(target=cls.server.serve_forever, daemon=True).start()

    @classmethod
    def tearDownClass(cls):
        cls.server.shutdown()
        cls.server.server_close()
        api.PUBLIC_DASHBOARD = cls.publico
        cls.tmp.cleanup()

    def entrar(self, login):
        pedido = urllib.request.Request(
            f"http://127.0.0.1:{self.port}/api/auth/login",
            data=json.dumps({"login": login, "senha": "senhaforte123"}).encode(),
            headers={"Content-Type": "application/json"}, method="POST")
        with urllib.request.urlopen(pedido, timeout=30) as r:
            return r.headers.get("Set-Cookie", "").split(";")[0]

    def buscar(self, caminho, cookie=None, cru=False):
        pedido = urllib.request.Request(f"http://127.0.0.1:{self.port}{caminho}")
        if cookie:
            pedido.add_header("Cookie", cookie)
        try:
            with urllib.request.urlopen(pedido, timeout=30) as r:
                bruto = r.read()
                return r.status, r.headers, (bruto if cru else json.loads(bruto or b"{}"))
        except urllib.error.HTTPError as exc:
            return exc.code, exc.headers, {}


class TestRotaDoPanorama(BasePanorama):
    def setUp(self):
        self.ana = self.entrar("ana@udesc.br")
        self.curioso = self.entrar("curioso@udesc.br")

    def test_sem_entrar_nao_ve(self):
        self.assertEqual(self.buscar("/api/panorama")[0], 401)

    def test_quem_so_le_ve(self):
        # o panorama da produção é do laboratório inteiro
        self.assertEqual(self.buscar("/api/panorama", self.curioso)[0], 200)

    def test_traz_tudo_o_que_a_tela_precisa(self):
        _, _, dados = self.buscar("/api/panorama", self.ana)
        for chave in ("panorama", "sintese", "lacunas", "artigos", "linhas",
                      "laboratorio", "vocabulario", "dendrograma"):
            self.assertIn(chave, dados)
        self.assertEqual(len(dados["artigos"]), 3)

    def test_a_arvore_dos_assuntos_vem_junto(self):
        # a rota inteira cai se um `set` sobrar dentro de um nó -- não só
        # o gráfico do dendrograma
        _, _, dados = self.buscar("/api/panorama", self.ana)
        arvore = dados["dendrograma"]
        self.assertIn("raiz", arvore)
        self.assertIn("altura_maxima", arvore)

    def test_cada_artigo_chega_com_as_suas_variaveis(self):
        _, _, dados = self.buscar("/api/panorama", self.ana)
        com_var = [a for a in dados["artigos"] if a["variaveis"]]
        self.assertEqual(len(com_var), 2)
        codigos = {v["code"] for v in com_var[0]["variaveis"]}
        self.assertTrue(codigos)

    def test_cada_variavel_diz_se_e_principal(self):
        # sem esse campo a tela não tem como destacar nada: todo selo
        # sairia igual, e "o artigo é sobre" viraria "o artigo cita"
        _, _, dados = self.buscar("/api/panorama", self.ana)
        for artigo in dados["artigos"]:
            for v in artigo["variaveis"]:
                with self.subTest(code=v["code"]):
                    self.assertIn(v["principal"], (0, 1))
                    self.assertIn("onde", v)

    def test_as_abas_novas_chegam_com_o_painel(self):
        # cada uma dessas abas precisaria de uma viagem própria se não
        # viesse aqui, e a tela abriria vazia por um instante
        _, _, dados = self.buscar("/api/panorama", self.ana)
        for chave in ("incidencia", "prevalencia", "triangulacao", "projetos",
                      "raio_x"):
            with self.subTest(chave=chave):
                self.assertIn(chave, dados)

    def test_a_janela_pode_ser_apertada_pela_consulta(self):
        _, _, dados = self.buscar("/api/panorama?desde=2023&ate=2025", self.ana)
        self.assertEqual(dados["panorama"]["janela"]["anos"], [2023, 2024, 2025])

    def test_a_sintese_nao_afirma_tendencia_sem_serie(self):
        # três artigos em três anos diferentes, mas cada variável aparece
        # em poucos: nada aqui pode virar "subindo" com confiança
        _, _, dados = self.buscar("/api/panorama", self.ana)
        for v in dados["panorama"]["variaveis"]:
            if not v["confiavel"]:
                self.assertEqual(v["tendencia"], "sem série suficiente")
                self.assertIsNotNone(v["porque"])


class TestMapaMundiServido(BasePanorama):
    """O contorno do mundo é servido à parte da página, e guardado."""

    def test_o_contorno_sai_sem_login(self):
        # é geografia, não é dado do laboratório
        status, cab, corpo = self.buscar("/api/geo/mundo.json", cru=True)
        self.assertEqual(status, 200)
        self.assertIn("application/json", cab.get("Content-Type", ""))
        self.assertGreater(len(json.loads(corpo)["paises"]), 120)

    def test_o_contorno_e_guardado_pelo_navegador(self):
        # 70 KB que nunca mudam; sem cache, viajam a cada visita à aba
        _, cab, _ = self.buscar("/api/geo/mundo.json", cru=True)
        self.assertIn("max-age", cab.get("Cache-Control", ""))

    def test_o_contorno_nao_e_reserializado(self):
        # a regra era "o tipo é application/json, então serialize": um
        # JSON já pronto em disco só passava mentindo o tipo, ou inchado
        # com indentação a cada visita
        _, _, corpo = self.buscar("/api/geo/mundo.json", cru=True)
        self.assertNotIn(b'\n  ', corpo[:400])


class TestProducaoDasBases(BasePanorama):
    """O botão que traz a produção sem ninguém abrir um terminal."""

    def setUp(self):
        self.ana = self.entrar("ana@udesc.br")
        self.curioso = self.entrar("curioso@udesc.br")

    def test_a_lista_de_quem_trazer_esta_na_rota(self):
        _, _, dados = self.buscar("/api/producao", self.ana)
        nomes = {p["nome"] for p in dados["pesquisadores"]}
        self.assertIn("Alexandro Andrade", nomes)
        self.assertIn("Guilherme Torres Vilarino", nomes)

    def test_o_link_do_lattes_vem_pronto(self):
        # o selo na tela vira o caminho para conferir se é a pessoa certa
        _, _, dados = self.buscar("/api/producao", self.ana)
        andrade = next(p for p in dados["pesquisadores"]
                       if p["nome"] == "Alexandro Andrade")
        self.assertEqual(andrade["link_lattes"],
                         "http://lattes.cnpq.br/5577164706111568")

    def test_o_painel_ja_chega_com_ela(self):
        # sem isso a aba precisaria de uma segunda viagem só para o botão
        _, _, dados = self.buscar("/api/panorama", self.ana)
        self.assertIn("producao", dados)
        self.assertIn("usuario", dados)

    def test_quem_so_le_nao_ve_botao_que_grava(self):
        # o botão que devolve 403 é pior que botão nenhum
        _, _, dados = self.buscar("/api/panorama", self.curioso)
        self.assertEqual(dados["usuario"]["papel"], "leitura")

    def test_importar_e_da_coordenacao(self):
        pedido = urllib.request.Request(
            f"http://127.0.0.1:{self.port}/api/producao/importar",
            data=b"{}", headers={"Content-Type": "application/json",
                                 "Cookie": self.curioso}, method="POST")
        with self.assertRaises(urllib.error.HTTPError) as caso:
            urllib.request.urlopen(pedido, timeout=30)
        self.assertEqual(caso.exception.code, 403)


class TestExportacaoDaExtracao(BasePanorama):
    def setUp(self):
        self.ana = self.entrar("ana@udesc.br")

    def test_o_csv_abre_no_excel_em_portugues(self):
        status, cab, corpo = self.buscar("/api/panorama/extracao.csv", self.ana, cru=True)
        self.assertEqual(status, 200)
        self.assertIn("text/csv", cab.get("Content-Type", ""))
        texto = corpo.decode("utf-8")
        self.assertTrue(texto.startswith("﻿"))
        cabecalho = texto.splitlines()[0].lstrip("﻿")
        self.assertIn("Variáveis", cabecalho)
        self.assertIn("Nº de variáveis", cabecalho)

    def test_o_csv_separa_principal_de_secundaria(self):
        # quem filtra a planilha por assunto quer os artigos que SÃO
        # sobre o assunto, não os que o mencionam de passagem
        _, _, corpo = self.buscar("/api/panorama/extracao.csv", self.ana, cru=True)
        linhas = corpo.decode("utf-8").lstrip("\ufeff").splitlines()
        cabecalho = linhas[0].split(";")
        self.assertIn("Variáveis principais", cabecalho)
        self.assertIn("Variáveis secundárias", cabecalho)
        coluna = cabecalho.index("Variáveis principais")
        principais = [l.split(";")[coluna] for l in linhas[1:]]
        self.assertTrue(any(p.strip('"') for p in principais),
                        "nenhum artigo saiu com variável principal")

    def test_o_csv_traz_as_variaveis_escritas(self):
        _, _, corpo = self.buscar("/api/panorama/extracao.csv", self.ana, cru=True)
        self.assertIn("Fibromialgia", corpo.decode("utf-8"))

    def test_o_xlsx_tem_as_duas_abas(self):
        # a segunda não é recorte por comodidade: é a tabela que responde
        # "como estes assuntos se relacionam", que a primeira não responde
        status, cab, corpo = self.buscar("/api/panorama/extracao.xlsx", self.ana, cru=True)
        self.assertEqual(status, 200)
        self.assertIn("spreadsheetml", cab.get("Content-Type", ""))
        self.assertTrue(corpo.startswith(b"PK"))
        from openpyxl import load_workbook
        livro = load_workbook(io.BytesIO(corpo))
        self.assertEqual(livro.sheetnames, ["Todos os artigos", "Mais de uma variável"])
        todos = livro["Todos os artigos"]
        multi = livro["Mais de uma variável"]
        self.assertEqual(todos.max_row - 1, 3)
        self.assertLess(multi.max_row, todos.max_row)

    def test_o_link_do_artigo_vira_link_na_planilha(self):
        _, _, corpo = self.buscar("/api/panorama/extracao.xlsx", self.ana, cru=True)
        from openpyxl import load_workbook
        aba = load_workbook(io.BytesIO(corpo))["Todos os artigos"]
        colunas = [c.value for c in aba[1]]
        coluna = colunas.index("Link") + 1
        links = [aba.cell(row=i, column=coluna).hyperlink
                 for i in range(2, aba.max_row + 1)]
        self.assertTrue(any(links), "nenhuma célula de link virou hyperlink")

    def test_sem_entrar_nao_baixa(self):
        self.assertEqual(self.buscar("/api/panorama/extracao.csv")[0], 401)


class TestPaginaDoPanorama(BasePanorama):
    def pagina(self):
        pedido = urllib.request.Request(f"http://127.0.0.1:{self.port}/panorama")
        with urllib.request.urlopen(pedido, timeout=30) as r:
            return r.read().decode("utf-8")

    def test_a_pagina_monta_inteira(self):
        html = self.pagina()
        for marcador in ("__BASE_CSS__", "__ICONS_JS__", "__CHARTS_JS__", "__PANORAMA_JS__"):
            self.assertNotIn(marcador, html, f"marcador não substituído: {marcador}")
        self.assertIn("const Charts", html)
        self.assertIn("const ABAS", html)

    def test_a_pagina_diz_qual_versao_esta_no_ar(self):
        # sem isso "atualizou?" só se responde abrindo um terminal, e
        # quem está no computador do laboratório não vai abrir
        html = self.pagina()
        self.assertIn('class="marca-versao"', html)
        self.assertLess(html.index('class="marca-versao"'), html.index("</body>"))

    def test_a_biblioteca_de_graficos_esta_ligada(self):
        # `C` não definido derrubava todo gráfico da tela, e a página
        # continuava de pé: cartões vazios, nenhum erro visível
        js = (TEMPLATES / "panorama.js").read_text(encoding="utf-8")
        self.assertIn("const C = Charts;", js)

    def test_a_rede_usa_o_campo_que_a_biblioteca_le(self):
        # com `value` em vez de `weight`, o cálculo de posição recebia
        # undefined, virava NaN, e a rede saía sem uma única linha
        js = (TEMPLATES / "panorama.js").read_text(encoding="utf-8")
        trecho = js[js.index("C.network({"):js.index("height: 440")]
        self.assertIn("weight:", trecho)
        self.assertNotIn("value:", trecho)

    def test_a_aba_do_ponto_so_aparece_para_quem_pode_abri_la(self):
        # aba que responde 403 é pior que aba nenhuma: promete e nega
        js = (TEMPLATES / "panorama.js").read_text(encoding="utf-8")
        corpo = js[js.index("function abasVisiveis"):]
        corpo = corpo[:corpo.index("\n}")]
        self.assertIn("coordenacao", corpo)
        self.assertIn("equipe", corpo)
        montar = js[js.index("function montarNav"):js.index("function contaDaAba")
                    if "function contaDaAba" in js else js.index("function montarNav") + 2000]
        self.assertIn("abasVisiveis()", montar)

    def test_toda_aba_declarada_tem_funcao(self):
        js = (TEMPLATES / "panorama.js").read_text(encoding="utf-8")
        bloco = js[js.index("const ABAS = ["):js.index("];", js.index("const ABAS = ["))]
        ids = re.findall(r'id: "(\w+)"', bloco)
        self.assertGreaterEqual(len(ids), 8)
        despacho = js[js.index("({ visao: verVisao"):js.index("[ST.aba] || verVisao)")]
        for identificador in ids:
            with self.subTest(aba=identificador):
                self.assertIn(identificador + ":", despacho)

    def test_todo_icone_da_navegacao_existe(self):
        js = (TEMPLATES / "panorama.js").read_text(encoding="utf-8")
        icones = (TEMPLATES / "icons.js").read_text(encoding="utf-8")
        for nome in re.findall(r'icone: "(\w+)"', js):
            with self.subTest(icone=nome):
                self.assertIn(f"{nome}:", icones)

    def test_nenhum_token_de_tema_indefinido(self):
        tema = (TEMPLATES / "theme.css").read_text(encoding="utf-8")
        definidos = set(re.findall(r"(--[a-z0-9-]+)\s*:", tema))
        pagina = (TEMPLATES / "panorama.html").read_text(encoding="utf-8")
        definidos |= set(re.findall(r"(--[a-z0-9-]+)\s*:", pagina))
        usados = set(re.findall(r"var\((--[a-z0-9-]+)", pagina))
        self.assertEqual(usados - definidos, set())


class TestPecasNovasDaTela(unittest.TestCase):
    """Marcas na curva, tempo real e o mapa-múndi."""

    def js(self, nome):
        return (TEMPLATES / nome).read_text(encoding="utf-8")

    def test_a_biblioteca_desenha_marca_sobre_a_linha(self):
        # o ponto de inflexão só significa algo em cima da curva que o
        # gerou; numa tabela ao lado, é um número
        graficos = self.js("charts.js")
        trecho = graficos[graficos.index("(spec.marks || [])"):]
        trecho = trecho[:trecho.index("/* uma dica por X")]
        self.assertIn("svg.appendChild(g)", trecho)
        self.assertIn("stroke-dasharray", trecho)

    def test_a_marca_vai_depois_das_series(self):
        # marca escondida atrás de uma linha não marca nada
        graficos = self.js("charts.js")
        self.assertLess(graficos.index("serieSpec.values.forEach"),
                        graficos.index("(spec.marks || [])"))

    def test_o_panorama_marca_as_inflexoes_no_grafico(self):
        js = self.js("panorama.js")
        self.assertIn("marks: marcas", js)
        self.assertIn("inflexoes", js[js.index("const marcas = []"):js.index("marks: marcas")])

    def test_o_titulo_leva_a_editora_e_nao_a_copia(self):
        # "leva direto para o site onde foi publicado" é o DOI, que
        # resolve para a editora. O texto livre é outra coisa -- uma
        # cópia -- e fica no botão ao lado, não no título
        js = self.js("panorama.js")
        corpo = js[js.index("function destinoDoTitulo"):]
        corpo = corpo[:corpo.index("\n}")]
        self.assertIn("d.editora", corpo)
        editora = js[js.index("function destinosDoArtigo"):js.index("function destinoDoTitulo")]
        # só o DOI e a página própria do artigo são "editora"
        self.assertEqual(editora.count("editora: true"), 2)
        pmc = editora[editora.index('chave: "pmc"'):]
        self.assertNotIn("editora", pmc[:pmc.index('chave: "pubmed"')])

    def test_toda_base_pedida_tem_botao(self):
        js = self.js("panorama.js")
        editora = js[js.index("function destinosDoArtigo"):js.index("function destinoDoTitulo")]
        for base in ("DOI", "PMC", "PubMed", "Scopus", "Web of Science"):
            with self.subTest(base=base):
                self.assertIn('rotulo: "' + base + '"', editora)

    def test_a_busca_nao_se_disfarca_de_artigo(self):
        # prometer "abre o artigo" e cair numa lista de resultados é pior
        # que avisar antes: sem o id da base não há link direto
        js = self.js("panorama.js")
        editora = js[js.index("function destinosDoArtigo"):js.index("function destinoDoTitulo")]
        self.assertIn("busca: true", editora)
        css = (TEMPLATES / "panorama.html").read_text(encoding="utf-8")
        self.assertIn(".base.busca", css)
        self.assertIn("border-style: dashed", css[css.index(".base.busca"):][:200])

    def test_o_texto_livre_nao_e_so_uma_cor(self):
        css = (TEMPLATES / "panorama.html").read_text(encoding="utf-8")
        self.assertIn(".base.livre", css)
        js = self.js("panorama.js")
        self.assertIn('rotulo: "PMC"', js)   # o rótulo diz qual base é

    def test_o_titulo_e_clicavel_fora_da_tabela_tambem(self):
        # linha do tempo e organograma mostram título; clicar neles tem
        # de levar ao artigo como na tabela
        js = self.js("panorama.js")
        self.assertGreaterEqual(js.count("tituloClicavel("), 3)
        corpo = js[js.index("function tituloClicavel"):]
        corpo = corpo[:corpo.index("\n}")]
        self.assertIn("destino.busca", corpo)
        self.assertIn("stopPropagation", corpo)

    def test_o_raio_x_mostra_a_base_de_cada_medida(self):
        # número sem N é opinião: a tela tem de dizer quantos artigos
        # sustentam cada medida
        js = self.js("panorama.js")
        corpo = js[js.index("function raioX()"):js.index("function formatarMedida")]
        self.assertIn("m.base", corpo)
        self.assertIn("m.leitura", corpo)
        self.assertIn("m.porque", corpo)

    def test_o_raio_x_separa_o_que_tem_base_do_que_nao_tem(self):
        # medida sem base junto das outras vira "—" e parece defeito
        js = self.js("panorama.js")
        corpo = js[js.index("function raioX()"):js.index("function formatarMedida")]
        self.assertIn("m.confiavel", corpo)
        self.assertIn("sem-base", corpo)

    def test_o_selo_da_variavel_principal_se_distingue(self):
        # se o destaque fosse só cor, quem imprime em preto e branco ou
        # não separa bem as cores lê a tabela toda como se fosse igual
        css = (TEMPLATES / "panorama.html").read_text(encoding="utf-8")
        self.assertIn(".selo-var.principal", css)
        self.assertIn(".selo-var.principal::before", css)
        self.assertIn(".selo-var.secundaria", css)

    def test_a_tela_usa_o_campo_principal_que_a_rota_manda(self):
        js = self.js("panorama.js")
        trecho = js[js.index("function pesoDaVariavel"):js.index("function seloVariavel")]
        self.assertIn("v.principal", trecho)
        self.assertIn("principal", js[js.index("function seloVariavel"):][:900])

    def test_o_destaque_vem_com_legenda(self):
        # marca sem legenda é enfeite: ninguém sabe o que ela afirma
        js = self.js("panorama.js")
        self.assertIn("legendaDosSelos()", js)
        legenda = js[js.index("function legendaDosSelos"):]
        legenda = legenda[:legenda.index("\n}")]
        self.assertIn("principal", legenda)
        self.assertIn("secundária", legenda.lower())

    def test_a_cor_nao_e_ciclada_entre_as_variaveis(self):
        # o defeito que este teste guarda: 8 cores para 23 variáveis,
        # cicladas pelo índice — "Exercício" e "Saúde mental" saíram as
        # duas verdes no mesmo gráfico, e a legenda não ajudava
        js = self.js("panorama.js")
        corpo = js[js.index("function montarCores()"):js.index("function corDaVariavel")]
        self.assertIn("i < 8", corpo)
        self.assertIn("--ink-muted", corpo)
        self.assertNotIn("% 8", corpo)

    def test_o_tempo_real_usa_a_mesma_conexao_do_painel(self):
        js = self.js("panorama.js")
        self.assertIn('new EventSource("/api/stream")', js)
        self.assertIn('addEventListener("mudanca"', js)

    def test_a_rajada_de_eventos_vira_uma_recarga_so(self):
        # numa importação de planilha chegam dezenas de eventos; recarregar
        # a cada um faria o painel recalcular tudo dezenas de vezes
        js = self.js("panorama.js")
        trecho = js[js.index('addEventListener("mudanca"'):js.index("function desligarAoVivo")]
        self.assertIn("clearTimeout(recargaMarcada)", trecho)
        self.assertIn("setTimeout(recarregar", trecho)

    def test_a_recarga_preserva_onde_a_pessoa_estava(self):
        # painel que salta para o topo a cada mudança é painel que ninguém
        # consegue ler enquanto a equipe trabalha
        js = self.js("panorama.js")
        corpo = js[js.index("async function recarregar()"):js.index("function marcarPulso")]
        self.assertIn("window.scrollY", corpo)
        self.assertIn("window.scrollTo", corpo)

    def test_o_mapa_mundi_enquadra_o_planeta(self):
        # `geo` enquadra os pontos: cinco países espalhados viram cinco
        # bolhas soltas num plano, e some o que o mapa existe para dar
        graficos = self.js("charts.js")
        corpo = graficos[graficos.index("function mapaMundi(spec)"):]
        corpo = corpo[:corpo.index("\n  }\n")]
        self.assertIn("LON0 = -180", corpo)
        self.assertIn("LON1 = 180", corpo)

    def test_o_contorno_do_mundo_e_de_verdade(self):
        # antes eram dez manchas desenhadas à mão. Continente irreconhecível
        # com bolhas espetadas em cima não diz de onde vem a produção
        mundo = json.loads((ROOT / "data" / "geo" / "mundo.json")
                           .read_text(encoding="utf-8"))["paises"]
        self.assertGreater(len(mundo), 120, "poucos países para um mapa-múndi")
        pontos = 0
        for pais in mundo:
            self.assertTrue(pais["id"] and pais["nome"], pais)
            for anel in pais["d"]:
                self.assertGreaterEqual(len(anel), 4, f"anel degenerado em {pais['nome']}")
                for lon, lat in anel:
                    self.assertTrue(-180 <= lon <= 180, f"{pais['nome']}: longitude {lon}")
                    self.assertTrue(-90 <= lat <= 90, f"{pais['nome']}: latitude {lat}")
                pontos += len(anel)
        self.assertGreater(pontos, 3000, "contorno pobre demais para reconhecer o mundo")

    def test_todo_pais_esperado_esta_no_contorno(self):
        # apaguei a África sem querer ao trocar a Eurásia, e o mapa seguiu
        # desenhando sem reclamar de nada
        mundo = json.loads((ROOT / "data" / "geo" / "mundo.json")
                           .read_text(encoding="utf-8"))["paises"]
        nomes = {p["nome"] for p in mundo}
        for pais in ("Brasil", "Portugal", "Estados Unidos", "Espanha", "Austrália",
                     "China", "Reino Unido", "Noruega", "Canadá", "Nigéria",
                     "Índia", "Nova Zelândia"):
            with self.subTest(pais=pais):
                self.assertIn(pais, nomes)

    def test_o_contorno_cabe_numa_visita(self):
        # é servido à parte da página e guardado pelo navegador; ainda
        # assim, meio megabyte de litoral por causa de cinco países seria
        # pagar caro por detalhe que ninguém vê
        tamanho = (ROOT / "data" / "geo" / "mundo.json").stat().st_size
        self.assertLess(tamanho, 200 * 1024, "contorno pesado demais")

    def test_a_cor_do_pais_e_o_valor(self):
        # bolha espetada em cima do país não é escala: o coroplético diz
        # magnitude com um tom só, do claro ao escuro
        graficos = self.js("charts.js")
        corpo = graficos[graficos.index("function mapaMundi(spec)"):]
        corpo = corpo[:corpo.index("\n  }\n")]
        self.assertIn("PASSOS_MAPA", corpo)
        self.assertNotIn("serie(", corpo, "cor de série é identidade, não magnitude")
        passos = graficos[graficos.index("const PASSOS_MAPA"):]
        passos = passos[:passos.index("\n")]
        self.assertNotIn("--series-", passos, "a escala tem de sair da rampa sequencial")

    def test_a_escala_do_mapa_vem_com_legenda(self):
        graficos = self.js("charts.js")
        self.assertIn("legendaDoMapa", graficos)
        self.assertIn("sem registro", graficos)

    def test_o_mapa_esta_exposto_pela_biblioteca(self):
        graficos = self.js("charts.js")
        self.assertIn("mapaMundi: mapaMundi", graficos)
        self.assertIn("C.mapaMundi(", self.js("panorama.js"))


class TestCurvasAoVivo(unittest.TestCase):
    """Filtro, plotagem quadro a quadro, dendrograma e fluxo.

    O que estes testes guardam não é a aparência: é o cronômetro que
    sobrevive à troca de aba redesenhando um elemento que já saiu da
    página, e o eixo que só sabe subir de zero — dois defeitos que a tela
    não denuncia porque não dão erro nenhum.
    """

    def js(self, nome):
        return (TEMPLATES / nome).read_text(encoding="utf-8")

    # ---- filtro ----
    def test_o_filtro_vive_fora_da_funcao_que_redesenha(self):
        # estado dentro de verCurvas() se perde a cada evento do banco, e
        # um filtro que se apaga sozinho não é filtro
        js = self.js("panorama.js")
        self.assertLess(js.index("const CURVA = {"), js.index("function verCurvas"))
        self.assertIn("escolhidas", js[js.index("const CURVA = {"):][:220])

    def test_os_atalhos_do_filtro_estao_todos_la(self):
        js = self.js("panorama.js")
        corpo = js[js.index("function filtroDeSeries"):js.index("function navegacaoDoTempo")]
        for atalho in ("Todas", "Principais", "Só as confiáveis", "Nenhuma"):
            with self.subTest(atalho=atalho):
                self.assertIn('"' + atalho + '"', corpo)

    def test_variavel_fora_do_recorte_sai_da_selecao(self):
        # marcada ontem, ausente hoje: sem essa limpeza o gráfico pede
        # uma série que não existe mais
        js = self.js("panorama.js")
        corpo = js[js.index("function verCurvas"):js.index("function filtroDeSeries")]
        self.assertIn("CURVA.escolhidas = CURVA.escolhidas.filter", corpo)

    def test_palco_vazio_avisa_em_vez_de_ficar_em_branco(self):
        js = self.js("panorama.js")
        corpo = js[js.index("function desenharCurvas"):]
        self.assertIn("Nenhuma série no gráfico", corpo[:1200])

    # ---- tempo real ----
    def test_o_cronometro_para_quando_o_palco_e_esvaziado(self):
        # este é o defeito silencioso: sair da aba deixava o intervalo
        # vivo, redesenhando um elemento fora da página
        js = self.js("panorama.js")
        corpo = js[js.index("function desenhar()"):]
        corpo = corpo[:corpo.index("palco.innerHTML")]
        self.assertIn("pararACurva()", corpo)
        self.assertIn("clearInterval", js[js.index("function pararACurva"):][:200])

    def test_quem_pediu_menos_movimento_recebe_a_curva_pronta(self):
        js = self.js("panorama.js")
        self.assertIn("prefers-reduced-motion", js)
        corpo = js[js.index("function animarTracado"):]
        self.assertIn("poucoMovimento()", corpo[:200])

    def test_o_eixo_fica_parado_enquanto_a_curva_cresce(self):
        # sem travar o topo, cada ano novo reescala e a linha parece pular
        js = self.js("panorama.js")
        self.assertIn("labels: anos, max: teto", js)
        graficos = self.js("charts.js")
        self.assertIn("spec.max || 0", graficos)

    def test_a_navegacao_tem_os_tres_controles(self):
        js = self.js("panorama.js")
        corpo = js[js.index("function navegacaoDoTempo"):js.index("function tocarACurva")]
        for marca in ('"data-nav": "antes"', '"data-nav": "tocar"',
                      '"data-nav": "depois"', "cursor-ano"):
            with self.subTest(marca=marca):
                self.assertIn(marca, corpo)

    def test_mexer_na_barra_interrompe_a_reproducao(self):
        # o gráfico continuar andando sob o dedo de quem arrasta é briga
        # entre o usuário e a tela
        js = self.js("panorama.js")
        corpo = js[js.index("function navegacaoDoTempo"):js.index("function tocarACurva")]
        # antes, depois, a barra -- e o proprio botao, que pausa
        self.assertEqual(corpo.count("pararACurva()"), 4)
        barra = corpo[corpo.index('type: "range"'):]
        self.assertIn("pararACurva()", barra[:barra.index("caixa.appendChild(cursor)")])

    # ---- o eixo com sinal ----
    def test_a_curva_negativa_fica_dentro_do_grafico(self):
        # velocidade e aceleração têm sinal; com escala só-positiva o
        # trecho abaixo de zero era desenhado por cima da legenda
        graficos = self.js("charts.js")
        self.assertIn("function niceTicksSigned", graficos)
        corpo = graficos[graficos.index("function lines(spec)"):]
        corpo = corpo[:corpo.index("const svg = svgRoot")]
        self.assertIn("niceTicksSigned", corpo)
        self.assertIn("scale.lo", corpo)

    def test_o_rotulo_da_ponta_segue_a_serie_e_nao_o_eixo(self):
        graficos = self.js("charts.js")
        corpo = graficos[graficos.index("function lines(spec)"):]
        corpo = corpo[:corpo.index("(spec.marks || [])")]
        self.assertNotIn("X(labels.length - 1)", corpo)

    def test_grafico_estreito_desbasta_mais_os_rotulos(self):
        # o viewBox tem sempre 760; meia coluna recebe metade disso, e o
        # desbaste de tela cheia deixa os anos encavalados
        graficos = self.js("charts.js")
        self.assertIn("spec.larguraReal", graficos)
        js = self.js("panorama.js")
        self.assertEqual(js.count("larguraReal: 400"), 2)   # velocidade e aceleração

    # ---- dendrograma e fluxo ----
    def test_a_aba_ativa_nao_se_anuncia_so_pela_cor(self):
        # classe "on" pinta o botao e nao diz nada a um leitor de tela
        js = self.js("panorama.js")
        corpo = js[js.index("function montarNav"):]
        self.assertIn('"aria-current"', corpo[:1400])
        self.assertIn('"data-aba"', corpo[:1400])

    def test_a_biblioteca_publica_as_duas_formas_novas(self):
        graficos = self.js("charts.js")
        self.assertIn("dendrograma: dendrograma", graficos)
        self.assertIn("fluxo: fluxo", graficos)

    def test_o_dendrograma_tem_regua_de_distancia(self):
        # sem eixo, a largura do colchete não significa nada
        graficos = self.js("charts.js")
        corpo = graficos[graficos.index("function dendrograma(spec)"):]
        corpo = corpo[:corpo.index("function fluxo(spec)")]
        self.assertIn("distância (1 − Jaccard)", corpo)
        self.assertIn("altura_maxima", corpo)

    def test_o_dendrograma_usa_uma_cor_so(self):
        # altura de fusão é grandeza contínua; cor categórica aqui
        # inventaria grupos que o algoritmo não decidiu
        graficos = self.js("charts.js")
        corpo = graficos[graficos.index("function dendrograma(spec)"):]
        corpo = corpo[:corpo.index("function fluxo(spec)")]
        self.assertNotIn("serie(", corpo)
        self.assertIn("seq(", corpo)

    def test_o_fluxo_nao_encolhe_ate_a_letra_sumir(self):
        graficos = self.js("charts.js")
        corpo = graficos[graficos.index("function fluxo(spec)"):]
        self.assertIn('svg.classList.add("fluxo")', corpo)
        self.assertIn("scrollx", corpo)
        css = (TEMPLATES / "theme.css").read_text(encoding="utf-8")
        self.assertIn("svg.plot.fluxo { width: auto; max-width: none; }", css)

    def test_o_pai_senta_na_media_dos_filhos(self):
        # sem isso a árvore vira lista com setas: o pai fica em cima do
        # primeiro filho e a bifurcação não se vê
        graficos = self.js("charts.js")
        corpo = graficos[graficos.index("function fluxo(spec)"):]
        corpo = corpo[:corpo.index("const colunas = nodes.reduce")]
        self.assertIn("function situar", corpo)
        self.assertIn("/ ys.length", corpo)

    def test_cada_caixa_do_metodo_carrega_um_valor_de_verdade(self):
        # fluxograma que diz só "suavização" é desenho de manual
        js = self.js("panorama.js")
        corpo = js[js.index("function cartaoDoMetodo"):js.index("function cartaoDaDecisao")]
        self.assertIn("p.total_artigos", corpo)
        self.assertIn("todas.length", corpo)
        self.assertIn("inflexoes", corpo)

    def test_a_arvore_de_decisao_usa_o_mesmo_limiar_do_calculo(self):
        # a tela dizer 0,5 e o cálculo usar 0,8 é a tela mentindo sobre
        # a própria regra
        analise_py = (ROOT / "scripts" / "lape" / "analise.py").read_text(encoding="utf-8")
        self.assertIn("RUIDO_ALTO = 0.8", analise_py)
        js = self.js("panorama.js")
        corpo = js[js.index("function cartaoDaDecisao"):]
        self.assertIn("razao_ruido >= 0.8", corpo)
        self.assertIn("razão ≥ 0,8", corpo)
        self.assertIn("(v.anos_com_dado || 0) < 3", corpo)

    def test_as_folhas_da_arvore_somam_o_total(self):
        # cada variável cai em exatamente uma folha; se a soma não fecha,
        # uma condição está errada e ninguém percebe
        js = self.js("panorama.js")
        corpo = js[js.index("function cartaoDaDecisao"):]
        corpo = corpo[:corpo.index("C.fluxo(")]
        self.assertIn("semAnos", corpo)
        self.assertIn("ruidosa", corpo)
        self.assertIn("virou", corpo)
        self.assertIn("lisa", corpo)

if __name__ == "__main__":
    unittest.main(verbosity=2)
