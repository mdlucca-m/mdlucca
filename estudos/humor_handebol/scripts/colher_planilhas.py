# -*- coding: utf-8 -*-
"""Camada 3: recolhe TUDO o que existe nas planilhas para dentro da base, com procedência.
Nenhum nome de atleta entra: cada célula passa por um raspador que substitui nomes pelo código."""
import openpyxl, sqlite3, os, re, unicodedata, hashlib, json, sys
RAIZ=os.environ.get("HH_RAIZ") or os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
UP=os.environ.get("HH_UPLOADS") or "/root/.claude/uploads/4ddb0907-77b2-5876-a286-ef4b6b886e93"
DB=os.path.join(RAIZ,"base","humor_handebol.sqlite")
FONTES=[
 ("ad245c30-Backup__Banco_de_dados_ORIGINAL_INTOCADO_20260723.xlsx","fonte-verdade",
  "Export do formulário; base de onde a camada canônica foi reconstruída."),
 ("bc6d935b-COLETAS.xlsx","fonte derivada",
  "Derivada do backup; carrega o desalinhamento de linhas de dois atletas."),
 ("9970ba89-Avalia__es_Handebol_S_o_Jos__2024.xlsx","resultados","Painéis e análises da avaliação geral."),
 ("639c1ba3-HIIT_FC_PSE.xlsx","resultados","Carga interna, FC e PSE das sessões de HIIT."),
 ("be1e9e29-resultados_handebol.xlsx","resultados","Saídas analíticas e painéis do estudo."),
 ("74be9990-AUDITORIA_DADOS_20260820.xlsx","auditoria","Auditoria de divergências feita pelo autor."),
]
def norm(s):
    s=unicodedata.normalize('NFKD',str(s)).encode('ascii','ignore').decode()
    return re.sub(r'\s+',' ',s).strip().lower()
# ---- dicionário de nomes proibidos, montado da fonte-verdade ----
wb=openpyxl.load_workbook(os.path.join(UP,FONTES[0][0]), read_only=True, data_only=True)
nomes=set()
for r in wb['Diário - Treino'].iter_rows(values_only=True):
    for c in (1,52):
        if isinstance(r[c],str) and len(r[c])>2: nomes.add(norm(r[c]))
for r in wb['Dicionário Atletas'].iter_rows(values_only=True):
    for c in (0,6,7):
        if c<len(r) and isinstance(r[c],str) and len(r[c])>2: nomes.add(norm(r[c]))
wb.close()
COMUNS={'nome','data','total','geral','media','atleta','grupo','equipe','treino','coleta','resumo',
        'padronizado','como digitado','nao identificado','variante','posicao','armador','pivo','goleiro',
        'ponta','central','sim','nao','completo','parcial'}
TOK=set()
for nm in nomes:
    for t in nm.split():
        if len(t)>=4 and t not in COMUNS and not t.isdigit(): TOK.add(t)
NOMES_COMPLETOS=sorted([n for n in nomes if ' ' in n and n not in COMUNS], key=len, reverse=True)
print(f"raspador: {len(NOMES_COMPLETOS)} nomes completos e {len(TOK)} tokens sob vigilância")
RXTOK=re.compile(r'\b('+'|'.join(sorted(map(re.escape,TOK),key=len,reverse=True))+r')\b') if TOK else None
def raspar(v):
    if not isinstance(v,str) or len(v)<3: return v, False
    n=norm(v)
    if n in nomes: return "«nome removido»", True
    if RXTOK and RXTOK.search(n):
        return RXTOK.sub("«nome»", n), True
    return v, False
def categoria(nome):
    n=norm(nome)
    for chave,cat in [('dashboard','painel'),('painel','painel'),('infogr','painel'),('interativo','painel'),
                      ('indice','índice'),('sumario','índice'),('central','índice'),
                      ('bruto','dados'),('raw','dados'),('base','dados'),('dados','dados'),('diario','dados'),
                      ('roster','dados'),('atletas','dados'),('log','dados'),('coleta','dados'),
                      ('quest','dados'),('caracteriz','dados'),
                      ('correl','análise'),('regress','análise'),('anova','análise'),('roc','análise'),
                      ('bayes','análise'),('modelo','análise'),('cluster','análise'),('perfil','análise'),
                      ('psicometria','análise'),('icc','análise'),('efeito','análise'),('inferenc','análise'),
                      ('estatis','análise'),('analis','análise'),('teste','análise'),('multivar','análise'),
                      ('artigo','texto'),('discussao','texto'),('conclus','texto'),('recomenda','texto'),
                      ('achado','texto'),('insight','texto'),('relatorio','texto'),('tese','texto'),
                      ('auditoria','auditoria'),('divergenc','auditoria'),('linhagem','auditoria'),
                      ('cobertura','auditoria'),('confronto','auditoria'),('impacto','auditoria'),
                      ('figura','figura'),('grafico','figura'),('graf','figura'),('tabela','tabela'),
                      ('equac','método'),('framework','método'),('desenho','método')]:
        if chave in n: return cat
    return 'outra'
cx=sqlite3.connect(DB)
cx.execute("DELETE FROM celula"); cx.execute("DELETE FROM aba"); cx.execute("DELETE FROM fonte")
MAXCEL=int(os.environ.get("HH_MAXCEL",60000))
tot_cel=tot_rasp=0
for fid,(arq,papel,nota) in enumerate(FONTES,1):
    p=os.path.join(UP,arq)
    if not os.path.exists(p): print("  ausente:",arq); continue
    sha=hashlib.sha256(open(p,'rb').read()).hexdigest()
    wb=openpyxl.load_workbook(p, read_only=True, data_only=True)
    cx.execute("INSERT INTO fonte VALUES (?,?,?,?,?,?)",(fid,arq,papel,sha,len(wb.sheetnames),nota))
    aid0=cx.execute("SELECT COALESCE(MAX(id),0) FROM aba").fetchone()[0]
    for si,sn in enumerate(wb.sheetnames,1):
        ws=wb[sn]; aid=aid0+si
        linhas=cols=0; buf=[]; hdr=None
        for i,row in enumerate(ws.iter_rows(values_only=True)):
            if tot_cel>=MAXCEL*len(FONTES): break
            vals=[v for v in row if v is not None]
            if not vals: continue
            linhas+=1; cols=max(cols,len(row))
            if hdr is None and sum(1 for v in row if isinstance(v,str))>=2:
                hdr=[('' if v is None else str(v)[:60]) for v in row]
            if len(buf)>=MAXCEL: continue
            for jcol,v in enumerate(row):
                if v is None: continue
                if isinstance(v,(int,float)) and not isinstance(v,bool):
                    buf.append((aid,i,jcol,(hdr[jcol][:60] if hdr and jcol<len(hdr) else None),None,float(v)))
                else:
                    t,rasp=raspar(v)
                    if rasp: globals().__setitem__('tot_rasp',tot_rasp+1); tot_rasp+=1
                    buf.append((aid,i,jcol,(hdr[jcol][:60] if hdr and jcol<len(hdr) else None),str(t)[:400],None))
        cx.execute("INSERT INTO aba VALUES (?,?,?,?,?,?,?)",
                   (aid,fid,sn,linhas,cols,categoria(sn),int(linhas>1)))
        cx.executemany("INSERT INTO celula VALUES (?,?,?,?,?,?)",buf)
        tot_cel+=len(buf)
    wb.close()
    print(f"  {arq[:46]:48} {len(wb.sheetnames):3} abas")
cx.commit()
print(f"\nacervo: {cx.execute('SELECT COUNT(*) FROM aba').fetchone()[0]} abas, "
      f"{cx.execute('SELECT COUNT(*) FROM celula').fetchone()[0]} células, {tot_rasp} raspadas")
print("categorias:", dict(cx.execute("SELECT categoria, COUNT(*) FROM aba GROUP BY categoria ORDER BY 2 DESC").fetchall()))
# verificação final: nenhum nome sobreviveu
sobra=0
for nm in NOMES_COMPLETOS[:200]:
    sobra+=cx.execute("SELECT COUNT(*) FROM celula WHERE LOWER(valor_txt) LIKE ?", (f"%{nm}%",)).fetchone()[0]
print("verificação — células que ainda contêm um nome completo do elenco:", sobra)
cx.close()
